# Copyright (c) 2026, Ashan and contributors
# For license information, please see license.txt

import frappe
from frappe import _
import json
import datetime
import io
import base64
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from ashan_cn_procurement.ashan_cn_procurement.payroll_engine import AshanPayrollCalculator
from ashan_cn_procurement.services.payroll_settlement_service import check_payroll_workbench_permission

@frappe.whitelist()
def get_payroll_workbench_data(period_month=None, company=None, tax_cycle_start_month=None):
    """
    加载月度人事薪酬工作台核心数据 (支持吉众与祺富双模，支持自定义个税起始计税月)
    """
    check_payroll_workbench_permission("read")
    if not period_month:
        now = datetime.datetime.now()
        period_month = now.strftime('%Y-%m')
    if not company:
        company = "天津吉众机电设备有限公司"

    # 1. 查询法定日历获取当月满勤工作日天数
    year, month = period_month.split('-')
    year_int = int(year)
    month_int = int(month)

    cal_records = frappe.get_all(
        "Ashan Holiday Calendar",
        filters={"year": year_int, "month": month_int},
        fields=["calendar_date", "day_type", "is_workday", "is_legal_holiday", "holiday_name", "remark"],
        order_by="calendar_date asc"
    )

    if cal_records:
        full_work_days = sum(1 for c in cal_records if c.is_workday)
    else:
        full_work_days = 21.0

    full_work_hours = full_work_days * 8.0
    fixed_work_hours = 172.0  # 21.5天 x 8h 制度工时

    # 2. 查询该月是否已有已保存的月结记录
    settle_name = f"{company}-薪资月结-{period_month}"
    settle_doc = None
    if frappe.db.exists("Ashan Payroll Settlement", settle_name):
        settle_doc = frappe.get_doc("Ashan Payroll Settlement", settle_name)

    status = settle_doc.status if settle_doc else "草稿"

    # 个税计税起始月份处理 (用户指定 > 已保存的设置 > 默认当年01月)
    if not tax_cycle_start_month:
        if settle_doc and getattr(settle_doc, 'tax_cycle_start_month', None):
            tax_cycle_start_month = settle_doc.tax_cycle_start_month
        else:
            tax_cycle_start_month = f"{year_int}-01"

    # 3. 获取员工薪资档案与考勤记录
    emp_profiles = frappe.get_all(
        "Ashan Employee Salary Profile",
        filters={"company": company, "employment_status": "在职"},
        fields=[
            "name", "employee_no", "employee_name", "company", "id_card", "mobile", "employee_type",
            "employment_status", "is_insured", "salary_mode", "fixed_net_salary", "base_salary",
            "base_subsidy", "performance_bonus_base", "position_allowance", "meal_unit_price",
            "social_security_base", "housing_fund_base", "special_additional_deduction", "tax_exemption_monthly"
        ],
        order_by="employee_no asc"
    )

    # 获取考勤与津贴明细记录
    att_records = frappe.get_all(
        "Ashan Monthly Attendance",
        filters={"company": company, "period_month": period_month},
        fields=[
            "employee_no", "attendance_days", "work_hours_regular", "overtime_regular_1_5",
            "overtime_weekend_2_0", "overtime_holiday_3_0", "meal_count", "daily_records_json"
        ]
    )
    att_map = {a.employee_no: a for a in att_records}

    # 获取历史个税累计数据 (从 tax_cycle_start_month 至当前月前)
    hist_map = get_cumulative_history(company, period_month, tax_cycle_start_month)

    # 4. 如果已有保存的子表记录且处于锁定状态，直接返回保存的数据；否则执行计算引擎
    items = []
    if settle_doc and settle_doc.status == "已核定锁定":
        for item in settle_doc.settlement_items:
            items.append(item.as_dict())
    else:
        for emp in emp_profiles:
            emp_no = emp.employee_no
            att = att_map.get(emp_no, {
                "attendance_days": full_work_days,
                "work_hours_regular": full_work_hours,
                "overtime_regular_1_5": 0.0,
                "overtime_weekend_2_0": 0.0,
                "overtime_holiday_3_0": 0.0,
                "meal_count": int(full_work_days),
                "salary_adjustment": 0.0,
                "salary_full_attendance": 0.0,
                "salary_performance_target": 0.0,
                "salary_housing_car_subsidy": 0.0,
                "salary_piecework_daily": 0.0
            })
            hist = hist_map.get(emp_no, {
                "cum_gross_prior": 0.0,
                "cum_tax_exemption_prior": 0.0,
                "cum_special_deduction_prior": 0.0,
                "cum_additional_deduction_prior": 0.0,
                "cum_tax_paid_prior": 0.0,
                "month_count_prior": 0
            })

            calc_res = AshanPayrollCalculator.calculate_employee_payroll(
                salary_profile=emp,
                attendance=att,
                full_work_days=full_work_days,
                full_work_hours=full_work_hours,
                fixed_work_hours=fixed_work_hours,
                cum_history=hist
            )
            items.append(calc_res)

    # 5. 汇总数据
    tot_gross = round(sum(i.get("gross_pay", 0) for i in items), 2)
    tot_ss_p = round(sum(i.get("social_security_p", 0) for i in items), 2)
    tot_hf_p = round(sum(i.get("housing_fund_p", 0) for i in items), 2)
    tot_tax = round(sum(i.get("individual_tax", 0) for i in items), 2)
    tot_net = round(sum(i.get("net_pay", 0) for i in items), 2)
    tot_cost = round(sum(i.get("company_cost_total", 0) for i in items), 2)

    # 配钞汇总
    tot_b100 = sum(i.get("bill_100", 0) for i in items)
    tot_b50 = sum(i.get("bill_50", 0) for i in items)
    tot_b20 = sum(i.get("bill_20", 0) for i in items)
    tot_b10 = sum(i.get("bill_10", 0) for i in items)
    tot_b5 = sum(i.get("bill_5", 0) for i in items)
    tot_b1 = sum(i.get("bill_1", 0) for i in items)

    return {
        "period_month": period_month,
        "company": company,
        "tax_cycle_start_month": tax_cycle_start_month,
        "status": status,
        "settlement_name": settle_name if settle_doc else None,
        "full_work_days": full_work_days,
        "full_work_hours": full_work_hours,
        "fixed_work_hours": fixed_work_hours,
        "summary": {
            "headcount": len(items),
            "total_gross_pay": tot_gross,
            "total_social_security_p": tot_ss_p,
            "total_housing_fund_p": tot_hf_p,
            "total_individual_tax": tot_tax,
            "total_net_pay": tot_net,
            "total_company_cost": tot_cost,
            "bills": {
                "b100": tot_b100, "b50": tot_b50, "b20": tot_b20,
                "b10": tot_b10, "b5": tot_b5, "b1": tot_b1
            }
        },
        "items": items,
        "calendar_days": cal_records
    }

def get_cumulative_history(company, period_month, tax_cycle_start_month="2026-01"):
    """
    计算从 tax_cycle_start_month 至当前结算月前一个月的历史个税累计
    """
    # 查找起始月与当前月之间的月份数
    cur_y, cur_m = map(int, period_month.split('-'))
    start_y, start_m = map(int, tax_cycle_start_month.split('-'))

    total_months_span = (cur_y - start_y) * 12 + (cur_m - start_m) # 截至上个月的自然月数
    prior_months_count = max(0, total_months_span)

    # 查询历史月结中的子表
    settlements = frappe.get_all(
        "Ashan Payroll Settlement",
        filters={
            "company": company,
            "period_month": [">=", tax_cycle_start_month],
            "period_month": ["<", period_month],
            "status": ["in", ["已核定锁定", "已发放", "草稿"]]
        },
        fields=["name", "period_month"]
    )

    hist_map = {}
    for s in settlements:
        doc = frappe.get_doc("Ashan Payroll Settlement", s.name)
        for it in doc.settlement_items:
            emp_no = it.employee_no
            if emp_no not in hist_map:
                hist_map[emp_no] = {
                    "cum_gross_prior": 0.0,
                    "cum_tax_exemption_prior": 0.0,
                    "cum_special_deduction_prior": 0.0,
                    "cum_additional_deduction_prior": 0.0,
                    "cum_tax_paid_prior": 0.0,
                    "month_count_prior": 0
                }
            hist_map[emp_no]["cum_gross_prior"] += (it.gross_pay or 0.0)
            hist_map[emp_no]["cum_special_deduction_prior"] += (it.special_deduction_total or 0.0)
            hist_map[emp_no]["cum_additional_deduction_prior"] += (it.special_additional_deduction or 0.0)
            hist_map[emp_no]["cum_tax_paid_prior"] += (it.individual_tax or 0.0)
            hist_map[emp_no]["month_count_prior"] += 1

    # 规范化历史免征额：保证截至上月的免征额基准为 prior_months_count * 5000
    for emp_no, h in hist_map.items():
        h["cum_tax_exemption_prior"] = prior_months_count * 5000.0

    return hist_map

@frappe.whitelist(methods=["POST"])
def upload_boss_payroll_file(filedata, filename, period_month, company="天津祺富机械加工有限公司"):
    """
    智能解析并导入【老板娘工资表】(Excel: .xlsx / .xls / .xlsm)
    """
    check_payroll_workbench_permission("write")
    if "," in filedata:
        filedata = filedata.split(",", 1)[1]

    file_bytes = base64.b64decode(filedata)
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

    # 优先匹配发薪工作表
    sheet_name = None
    candidate_names = ["临时表格", "当月发薪工资表", "工资表", "发薪表", "Sheet1"]
    for cand in candidate_names:
        for s in wb.sheetnames:
            if cand in s:
                sheet_name = s
                break
        if sheet_name:
            break
    if not sheet_name:
        sheet_name = wb.sheetnames[0]

    ws = wb[sheet_name]

    # 查找表头所在行 (通常在第 1 至 4 行)
    header_row_idx = 1
    col_mapping = {}
    for r in range(1, min(ws.max_row + 1, 6)):
        row_vals = [str(ws.cell(row=r, column=c).value or "").strip() for c in range(1, ws.max_column + 1)]
        if any("姓名" in v or "工号" in v or "实发" in v for v in row_vals):
            header_row_idx = r
            for c_idx, val in enumerate(row_vals, 1):
                if val:
                    col_mapping[val] = c_idx
            break

    imported_count = 0
    updated_records = []

    # 遍历数据行
    for r in range(header_row_idx + 1, ws.max_row + 1):
        def get_col_val(keywords):
            for k, c in col_mapping.items():
                if any(kw in k for kw in keywords):
                    v = ws.cell(row=r, column=c).value
                    return v
            return None

        name_val = get_col_val(["姓名", "名字", "员工"])
        emp_no_val = get_col_val(["工号", "编号"])

        if not name_val and not emp_no_val:
            continue
        if str(name_val).strip() in ["合计", "全公司合计", "签字", "备考"]:
            continue

        emp_name = str(name_val).strip() if name_val else ""
        emp_no = str(emp_no_val).strip() if emp_no_val else ""

        att_days = float(get_col_val(["作业天数", "出勤天数", "天数", "出勤"]) or 21.0)
        daily_salary = float(get_col_val(["天工资", "计件工资"]) or 0.0)
        full_att = float(get_col_val(["全勤费", "全勤奖", "全勤"]) or 0.0)
        ot_hours = float(get_col_val(["加班小时", "加班工时", "加班"]) or 0.0)
        ot_pay = float(get_col_val(["加班费"]) or 0.0)
        target_perf = float(get_col_val(["达标工资", "达标率", "达标"]) or 0.0)
        pos_allow = float(get_col_val(["职位补贴", "职务补贴", "职位津贴", "职位"]) or 0.0)
        house_car = float(get_col_val(["房/车补", "房车补", "车补", "房补"]) or 0.0)
        adjust_val = float(get_col_val(["扣除", "工资调整", "调整"]) or 0.0)

        # 查找或创建员工档案
        emp_doc_name = None
        if emp_no:
            emp_doc_name = frappe.db.get_value("Ashan Employee Salary Profile", {"company": company, "employee_no": emp_no}, "name")
        if not emp_doc_name and emp_name:
            emp_doc_name = frappe.db.get_value("Ashan Employee Salary Profile", {"company": company, "employee_name": emp_name}, "name")

        if not emp_doc_name:
            # 自动创建档案
            new_emp = frappe.new_doc("Ashan Employee Salary Profile")
            new_emp.company = company
            new_emp.employee_no = emp_no or f"QF{imported_count+1:04d}"
            new_emp.employee_name = emp_name
            new_emp.employee_type = "正式工"
            new_emp.employment_status = "在职"
            new_emp.is_insured = 1
            new_emp.salary_mode = "税前动态工资"
            new_emp.social_security_base = 5124
            new_emp.housing_fund_base = 2320
            new_emp.insert(ignore_permissions=True)
            emp_doc = new_emp
        else:
            emp_doc = frappe.get_doc("Ashan Employee Salary Profile", emp_doc_name)

        # 更新或写入考勤与补贴记录
        att_name = f"{company}-{period_month}-{emp_doc.employee_no}"
        if frappe.db.exists("Ashan Monthly Attendance", att_name):
            att_doc = frappe.get_doc("Ashan Monthly Attendance", att_name)
        else:
            att_doc = frappe.new_doc("Ashan Monthly Attendance")
            att_doc.company = company
            att_doc.period_month = period_month
            att_doc.employee_no = emp_doc.employee_no
            att_doc.employee_name = emp_doc.employee_name

        att_doc.attendance_days = att_days
        att_doc.work_hours_regular = att_days * 8.0
        att_doc.overtime_regular_1_5 = ot_hours
        att_doc.meal_count = int(att_days)
        att_doc.save(ignore_permissions=True)

        updated_records.append({
            "employee_no": emp_doc.employee_no,
            "employee_name": emp_doc.employee_name,
            "attendance_days": att_days,
            "salary_piecework_daily": daily_salary,
            "salary_full_attendance": full_att,
            "overtime_regular_1_5": ot_hours,
            "salary_overtime_1_5": ot_pay,
            "salary_performance_target": target_perf,
            "salary_position_allowance": pos_allow,
            "salary_housing_car_subsidy": house_car,
            "salary_adjustment": adjust_val
        })
        imported_count += 1

    frappe.db.commit()

    # 重新加载工作台最新数据
    return {
        "success": True,
        "message": f"成功从【{filename}】解析并导入 {imported_count} 位员工的老板娘工资表！",
        "imported_count": imported_count,
        "workbench_data": get_payroll_workbench_data(period_month, company)
    }

@frappe.whitelist(methods=["POST"])
def save_payroll_settlement(data):
    """
    保存月度薪资核算草稿
    """
    check_payroll_workbench_permission("write")
    if isinstance(data, str):
        data = json.loads(data)

    company = data.get("company")
    period_month = data.get("period_month")
    name = f"{company}-薪资月结-{period_month}"

    if frappe.db.exists("Ashan Payroll Settlement", name):
        doc = frappe.get_doc("Ashan Payroll Settlement", name)
    else:
        doc = frappe.new_doc("Ashan Payroll Settlement")
        doc.name = name
        doc.company = company
        doc.period_month = period_month

    doc.status = "草稿"
    doc.tax_cycle_start_month = data.get("tax_cycle_start_month", f"{period_month.split('-')[0]}-01")
    doc.full_work_days = data.get("full_work_days", 21)
    doc.full_work_hours = data.get("full_work_hours", 168)
    doc.fixed_work_hours = data.get("fixed_work_hours", 172)

    summary = data.get("summary", {})
    doc.total_gross_pay = summary.get("total_gross_pay", 0)
    doc.total_social_security_p = summary.get("total_social_security_p", 0)
    doc.total_housing_fund_p = summary.get("total_housing_fund_p", 0)
    doc.total_individual_tax = summary.get("total_individual_tax", 0)
    doc.total_net_pay = summary.get("total_net_pay", 0)
    doc.total_company_cost = summary.get("total_company_cost", 0)

    # 填充子表
    doc.set("settlement_items", [])
    for item in data.get("items", []):
        doc.append("settlement_items", item)

    doc.save(ignore_permissions=True)
    frappe.db.commit()

    return {"success": True, "name": doc.name, "message": "薪资月结草稿已成功保存"}

@frappe.whitelist(methods=["POST"])
def finalize_payroll_settlement(data):
    """
    核定并锁定薪资月结
    """
    check_payroll_workbench_permission("write")
    res = save_payroll_settlement(data)
    name = res["name"]
    doc = frappe.get_doc("Ashan Payroll Settlement", name)
    doc.status = "已核定锁定"
    doc.locked_at = frappe.utils.now_datetime()
    doc.save(ignore_permissions=True)
    frappe.db.commit()
    return {"success": True, "message": "薪资月结已核定并锁定！"}

@frappe.whitelist()
def export_payroll_excel(period_month=None, company=None, tax_cycle_start_month=None):
    """
    导出 1:1 多工作表标准人事薪酬 Excel 文件
    """
    check_payroll_workbench_permission("read")
    data = get_payroll_workbench_data(period_month, company, tax_cycle_start_month)
    items = data.get("items", [])
    summary = data.get("summary", {})
    bills = summary.get("bills", {})

    wb = openpyxl.Workbook()

    # 样式定义
    header_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    total_fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
    font_bold = Font(name="Microsoft YaHei", size=10, bold=True)
    font_normal = Font(name="Microsoft YaHei", size=10)
    font_title = Font(name="Microsoft YaHei", size=14, bold=True)
    align_center = Alignment(horizontal="center", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1")
    )

    # 1. Sheet 1: 工资核定总表
    ws1 = wb.active
    ws1.title = "工资核定总表"
    ws1.views.sheetView[0].showGridLines = True

    ws1.merge_cells("A1:W1")
    ws1["A1"] = f"{company} {period_month} 员工薪资核定总表 (个税起始: {data.get('tax_cycle_start_month')})"
    ws1["A1"].font = font_title
    ws1["A1"].alignment = align_center

    headers1 = [
        "序号", "工号", "姓名", "计薪方式", "薪资标准", "考勤天数", "基本工时",
        "平日加班(h)", "周末加班(h)", "节日加班(h)", "工时工资", "加班工资",
        "基本补贴", "绩效奖金", "职位津贴", "餐补", "全勤/达标/房补", "工资调整", "应发薪资合计",
        "社保个人(p)", "公积金个人(p)", "代扣个税", "实发工资合计"
    ]
    for col_idx, h in enumerate(headers1, 1):
        c = ws1.cell(row=3, column=col_idx, value=h)
        c.font = font_bold
        c.fill = header_fill
        c.alignment = align_center
        c.border = thin_border

    row_idx = 4
    for idx, it in enumerate(items, 1):
        qifu_extra = (it.get("salary_full_attendance", 0) + it.get("salary_performance_target", 0) + it.get("salary_housing_car_subsidy", 0))
        vals = [
            idx, it.get("employee_no"), it.get("employee_name"), it.get("salary_mode"),
            it.get("base_salary", 0), it.get("attendance_days", 0), it.get("work_hours_regular", 0),
            it.get("overtime_1_5", 0), it.get("overtime_2_0", 0), it.get("overtime_3_0", 0),
            it.get("salary_regular_hours", 0), (it.get("salary_overtime_1_5", 0) + it.get("salary_overtime_2_0", 0) + it.get("salary_overtime_3_0", 0)),
            it.get("salary_base_subsidy", 0), it.get("salary_performance", 0), it.get("salary_position_allowance", 0),
            it.get("salary_meal_subsidy", 0), qifu_extra, it.get("salary_adjustment", 0), it.get("gross_pay", 0),
            it.get("social_security_p", 0), it.get("housing_fund_p", 0), it.get("individual_tax", 0), it.get("net_pay", 0)
        ]
        for c_idx, val in enumerate(vals, 1):
            c = ws1.cell(row=row_idx, column=c_idx, value=val)
            c.font = font_normal
            c.border = thin_border
            if isinstance(val, (int, float)) and c_idx > 4:
                c.alignment = align_right
                c.number_format = "#,##0.00" if c_idx > 6 else "0.0"
            else:
                c.alignment = align_center
        row_idx += 1

    # 2. Sheet 2: 现金零钞配钞表
    ws2 = wb.create_sheet(title="现金零钞配钞表")
    ws2.views.sheetView[0].showGridLines = True
    ws2.merge_cells("A1:J1")
    ws2["A1"] = f"{company} {period_month} 现金发放零钞配钞表"
    ws2["A1"].font = font_title
    ws2["A1"].alignment = align_center

    headers2 = ["序号", "工号", "姓名", "实发工资", "100元(张)", "50元(张)", "20元(张)", "10元(张)", "5元(张)", "1元(张)"]
    for col_idx, h in enumerate(headers2, 1):
        c = ws2.cell(row=3, column=col_idx, value=h)
        c.font = font_bold
        c.fill = header_fill
        c.alignment = align_center
        c.border = thin_border

    r2 = 4
    for idx, it in enumerate(items, 1):
        vals2 = [
            idx, it.get("employee_no"), it.get("employee_name"), it.get("net_pay", 0),
            it.get("bill_100", 0), it.get("bill_50", 0), it.get("bill_20", 0),
            it.get("bill_10", 0), it.get("bill_5", 0), it.get("bill_1", 0)
        ]
        for c_idx, val in enumerate(vals2, 1):
            c = ws2.cell(row=r2, column=c_idx, value=val)
            c.font = font_normal
            c.border = thin_border
            c.alignment = align_right if c_idx >= 4 else align_center
            if c_idx == 4:
                c.number_format = "#,##0.00"
        r2 += 1

    # 输出流
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    frappe.response['filename'] = f"{company}_{period_month}_人事薪酬综合核算表.xlsx"
    frappe.response['filecontent'] = stream.getvalue()
    frappe.response['type'] = 'binary'

@frappe.whitelist()
def get_payslip_print_data(period_month=None, company=None, mode="A4", tax_cycle_start_month=None):
    """
    获取 A4 签收单或信封工资条打印数据
    """
    check_payroll_workbench_permission("read")
    data = get_payroll_workbench_data(period_month, company, tax_cycle_start_month)
    return {
        "company": company,
        "period_month": period_month,
        "mode": mode,
        "items": data.get("items", []),
        "summary": data.get("summary", {})
    }
