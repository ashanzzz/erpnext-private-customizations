# Copyright (c) 2026, Ashan CN Procurement
"""Semantic importer for historical/current external Qifu payroll workbooks.

The external workbook has evolved over several years.  This module separates
workbook-shape recognition from payroll calculation, normalizes every supported
layout into one canonical row model, and preserves source-only fields for audit.
Preview and final import should call the same parser to avoid inconsistent results.
"""

from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal
import hashlib
import json
import re


PARSER_VERSION = "QIFU-PAYROLL-SEMANTIC-V3"


def normalize_header(value):
    text = str(value if value is not None else "").strip()
    text = re.sub(r"[\s\r\n\t]+", "", text)
    return (
        text.replace("／", "/")
        .replace("（", "(")
        .replace("）", ")")
        .replace("：", ":")
        .replace("－", "-")
        .replace("—", "-")
    )


# Exact aliases are deliberate.  Broad substring matching previously caused
# hours/day columns to overwrite one another.
FIELD_ALIASES = {
    "serial_no": {"编号", "序号"},
    "employee_no": {"工号", "员工工号"},
    "joining_date": {"入职日期", "到职日期"},
    "source_month_date": {"本月日期", "当月日期"},
    "tenure_days": {"入职时间", "在职时间", "在职天数", "入职天数"},
    "name": {"姓名", "员工姓名"},
    "work_days": {"作业天数", "工作天数", "出勤天数"},
    "work_hours": {"作业小时", "作业工时", "工作小时", "工作工时", "出勤小时", "出勤工时"},
    "day_salary": {"天工资", "日工资"},
    "hour_salary": {"小时工资", "时工资"},
    "full_attendance": {"全勤费", "全勤奖", "全勤"},
    # Historical sheets use plain “加班” for overtime hours.
    "overtime_hours": {"加班小时", "加班工时", "加班"},
    "overtime_salary": {"加班费", "加班工资", "加班/抛光", "加班抛光"},
    "national_days": {"国勤天数", "国勤(天)", "国勤天", "法定天数", "法定出勤天数"},
    # Historical sheets use plain “国勤” for the amount.
    "national_salary": {"国勤工资", "国勤", "法定工资", "法定出勤工资"},
    "polishing_salary": {"抛光", "抛光费", "抛光工资"},
    "target_rate": {"达标率", "绩效达标率"},
    "target_salary": {"达标工资", "绩效达标工资"},
    "is_insured": {"是否社保", "是否参保", "社保"},
    "deduction": {"扣除", "扣款", "其他扣除"},
    "source_payable_salary": {"应付工资", "应发工资", "应付工资合计", "应发工资合计"},
    "source_paid_salary": {"已发工资", "已付工资", "已支付工资"},
    "net_salary": {"实发工资", "实发", "实发合计", "实发工资合计"},
    "signature": {"签字", "签名"},
    "remarks": {"备考", "备注", "备注说明"},
}

# Known source-only columns that are intentionally NOT promoted into payroll business fields.
# They stay in source_raw_json for audit/replay. Final cash denomination is generated
# from Tab 6 final net salary, so historical note counts never override settlement.
AUDIT_ONLY_ALIASES = {
    "cash_note_counts": {"100", "100元", "100元张数", "50", "50元", "50元张数", "10", "10元", "10元张数", "5", "5元", "5元张数", "1", "1元", "1元张数"},
}

ALLOWANCE_ALIASES = {
    "name": {"姓名", "员工姓名"},
    "post_allowance": {"职位补贴", "职务补贴", "岗位补贴", "岗位津贴"},
    "house_car_allowance": {"房/车补", "房补/车补", "房车补", "租房/车补", "房补", "车补"},
    "allowance_total": {"合计", "补贴合计"},
}

_ALIAS_TO_FIELD = {}
for _field, _aliases in FIELD_ALIASES.items():
    for _alias in _aliases:
        _ALIAS_TO_FIELD[normalize_header(_alias)] = _field

_AUDIT_ONLY_HEADERS = {normalize_header(alias) for aliases in AUDIT_ONLY_ALIASES.values() for alias in aliases}

_ALLOWANCE_ALIAS_TO_FIELD = {}
for _field, _aliases in ALLOWANCE_ALIASES.items():
    for _alias in _aliases:
        _ALLOWANCE_ALIAS_TO_FIELD[normalize_header(_alias)] = _field


def _cell_to_jsonable(value):
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, (str, int, float, bool)) or value is None:
        return value
    return str(value)


def _to_float(value):
    if value in (None, ""):
        return 0.0
    if isinstance(value, bool):
        return float(int(value))
    if isinstance(value, (int, float, Decimal)):
        return float(value)
    text = str(value).strip().replace(",", "").replace("￥", "").replace("¥", "")
    if not text:
        return 0.0
    if text.endswith("%"):
        try:
            return float(text[:-1]) / 100.0
        except Exception:
            return 0.0
    try:
        return float(text)
    except Exception:
        return 0.0


def _date_to_iso(value):
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value or "").strip()
    if not text:
        return ""
    # Keep non-date source values in raw JSON; only promote clearly parseable dates.
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
        try:
            return datetime.strptime(text[:10], fmt).date().isoformat()
        except Exception:
            continue
    return ""


def build_main_column_map(row_values):
    col_map = {}
    source_headers = {}
    unknown_headers = []
    for c_idx, raw in enumerate(row_values, start=1):
        norm = normalize_header(raw)
        if not norm:
            continue
        field = _ALIAS_TO_FIELD.get(norm)
        if not field:
            if norm in _AUDIT_ONLY_HEADERS:
                continue
            unknown_headers.append({"column": c_idx, "header": str(raw)})
            continue
        # Do not silently overwrite a field.  The first exact semantic match wins;
        # duplicates are reported to the preview diagnostics.
        if field not in col_map:
            col_map[field] = c_idx
            source_headers[field] = str(raw).strip()
        else:
            unknown_headers.append({"column": c_idx, "header": str(raw), "reason": f"duplicate:{field}"})
    return col_map, source_headers, unknown_headers


def score_main_header(row_values):
    col_map, source_headers, unknown = build_main_column_map(row_values)
    score = 0
    if "name" in col_map:
        score += 35
    if "net_salary" in col_map:
        score += 35
    for key in ("work_days", "work_hours", "day_salary", "hour_salary", "full_attendance", "overtime_hours", "overtime_salary", "national_days", "national_salary", "target_rate", "target_salary", "deduction"):
        if key in col_map:
            score += 2
    if "source_payable_salary" in col_map:
        score += 2
    if "joining_date" in col_map:
        score += 2
    return score, col_map, source_headers, unknown


def detect_schema_version(col_map, source_headers=None):
    if "joining_date" in col_map or "polishing_salary" in col_map:
        return "legacy-2023-2024"
    if "source_payable_salary" in col_map or "source_paid_salary" in col_map:
        return "transition-2024"
    if "overtime_hours" in col_map and "national_salary" in col_map:
        return "modern-2026"
    return "generic-compatible"


def _worksheet_rows(ws, max_scan_rows=None):
    end = ws.max_row if max_scan_rows is None else min(ws.max_row, max_scan_rows)
    for r in range(1, end + 1):
        yield r, [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]


def find_best_main_table(workbook):
    best = None
    for ws in workbook.worksheets:
        for r, values in _worksheet_rows(ws, max_scan_rows=30):
            score, col_map, headers, unknown = score_main_header(values)
            if not best or score > best["score"]:
                best = {
                    "sheet": ws,
                    "sheet_name": ws.title,
                    "header_row": r,
                    "score": score,
                    "col_map": col_map,
                    "source_headers": headers,
                    "unknown_headers": unknown,
                }
    if not best or best["score"] < 70 or "name" not in best["col_map"] or "net_salary" not in best["col_map"]:
        raise ValueError("未识别到同时包含【姓名】和【实发工资】的工资主表表头")
    best["schema_version"] = detect_schema_version(best["col_map"], best["source_headers"])
    return best


def find_allowance_table(ws, start_row=1):
    best = None
    for r in range(max(1, int(start_row)), ws.max_row + 1):
        values = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        cols, headers = {}, {}
        for c_idx, raw in enumerate(values, start=1):
            norm = normalize_header(raw)
            field = _ALLOWANCE_ALIAS_TO_FIELD.get(norm)
            if field and field not in cols:
                cols[field] = c_idx
                headers[field] = str(raw).strip()
        if "name" in cols and ("post_allowance" in cols or "house_car_allowance" in cols):
            score = 10 + (3 if "post_allowance" in cols else 0) + (3 if "house_car_allowance" in cols else 0) + (1 if "allowance_total" in cols else 0)
            if not best or score > best["score"]:
                best = {"header_row": r, "col_map": cols, "source_headers": headers, "score": score}
    return best


def find_best_allowance_table(workbook, main_sheet=None, main_header_row=1):
    """Find the strongest allowance table, preferring the main payroll sheet."""
    candidates = []
    if main_sheet is not None:
        info = find_allowance_table(main_sheet, int(main_header_row or 1) + 1)
        if info:
            candidates.append((info.get("score", 0) + 2, main_sheet, info))
    for ws in workbook.worksheets:
        if main_sheet is not None and ws is main_sheet:
            continue
        info = find_allowance_table(ws, 1)
        if info:
            candidates.append((info.get("score", 0), ws, info))
    if not candidates:
        return None, None
    candidates.sort(key=lambda x: x[0], reverse=True)
    _, ws, info = candidates[0]
    return ws, info


def _get(ws, row, col_map, key):
    col = col_map.get(key)
    return ws.cell(row, col).value if col else None


def _row_raw_payload(ws, row_number, header_row, col_map, source_headers):
    """Preserve every source column, including fields unknown to this parser version.

    Canonical fields are still mapped separately for calculation.  The raw snapshot is
    intentionally lossless at the cell-value level so a future parser can reinterpret
    historical rows without requiring the original workbook to be re-uploaded.
    """
    payload = {}
    recognized_by_col = {col: field for field, col in col_map.items()}
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row_number, col).value
        raw_header = ws.cell(header_row, col).value
        header_text = str(raw_header or "").strip()
        if not header_text and value in (None, ""):
            continue
        field = recognized_by_col.get(col)
        key = header_text or (source_headers.get(field) if field else "") or f"COL_{col}"
        if key in payload:
            key = f"{key}__COL_{col}"
        payload[key] = _cell_to_jsonable(value)
    return payload


def _parse_main_rows(ws, header_row, col_map, source_headers, schema_version):
    rows = []
    blank_streak = 0
    for r in range(header_row + 1, ws.max_row + 1):
        raw_name = str(_get(ws, r, col_map, "name") or "").strip()
        norm_name = normalize_header(raw_name)
        if norm_name in {"合计", "总计", "平均", "工资表"}:
            break
        if not raw_name or raw_name == "None":
            blank_streak += 1
            if blank_streak >= 2:
                break
            continue
        blank_streak = 0

        net = _to_float(_get(ws, r, col_map, "net_salary"))
        payable = _to_float(_get(ws, r, col_map, "source_payable_salary"))
        paid = _to_float(_get(ws, r, col_map, "source_paid_salary"))
        # Older workbooks sometimes hold the authoritative amount in 应付/实发 while
        # the component cells are formulas/blank.  Preserve every source amount and
        # use 实发 first, then 应付 as a conservative fallback.
        workshop_net = net if (net != 0 or _get(ws, r, col_map, "net_salary") not in (None, "")) else payable

        raw_payload = _row_raw_payload(ws, r, header_row, col_map, source_headers)
        source_serial = _get(ws, r, col_map, "employee_no") if "employee_no" in col_map else _get(ws, r, col_map, "serial_no")
        source_no_kind = "employee_no" if "employee_no" in col_map else ("serial_no" if "serial_no" in col_map else "")
        rows.append({
            "source_row_number": r,
            "source_schema_version": schema_version,
            "source_serial_no": str(source_serial or "").strip(),
            "source_no_kind": source_no_kind,
            "employee_name": raw_name,
            "source_joining_date": _date_to_iso(_get(ws, r, col_map, "joining_date")),
            "source_month_date": _date_to_iso(_get(ws, r, col_map, "source_month_date")),
            "source_tenure_days": _to_float(_get(ws, r, col_map, "tenure_days")),
            "work_days": _to_float(_get(ws, r, col_map, "work_days")),
            "work_hours": _to_float(_get(ws, r, col_map, "work_hours")),
            "day_salary": _to_float(_get(ws, r, col_map, "day_salary")),
            "hour_salary": _to_float(_get(ws, r, col_map, "hour_salary")),
            "full_attendance": _to_float(_get(ws, r, col_map, "full_attendance")),
            "overtime_hours": _to_float(_get(ws, r, col_map, "overtime_hours")),
            "overtime_salary": _to_float(_get(ws, r, col_map, "overtime_salary")),
            "national_days": _to_float(_get(ws, r, col_map, "national_days")),
            "national_salary": _to_float(_get(ws, r, col_map, "national_salary")),
            "source_polishing_salary": _to_float(_get(ws, r, col_map, "polishing_salary")),
            "target_rate": str(_get(ws, r, col_map, "target_rate") or "").strip(),
            "target_salary": _to_float(_get(ws, r, col_map, "target_salary")),
            "source_is_insured": str(_get(ws, r, col_map, "is_insured") or "").strip(),
            "deduction": _to_float(_get(ws, r, col_map, "deduction")),
            "source_payable_salary": payable,
            "source_paid_salary": paid,
            "workshop_net": workshop_net,
            "source_signature": str(_get(ws, r, col_map, "signature") or "").strip(),
            "remarks": str(_get(ws, r, col_map, "remarks") or "").strip(),
            "post_allowance": 0.0,
            "house_car_allowance": 0.0,
            "source_raw_json": json.dumps(raw_payload, ensure_ascii=False, separators=(",", ":")),
        })
    return rows


def _parse_allowance_rows(ws, table_info):
    if not table_info:
        return {"rows": [], "base_total": None, "grand_total": None}
    cols = table_info["col_map"]
    rows = []
    base_total = None
    grand_total = None
    blank_streak = 0
    for r in range(table_info["header_row"] + 1, ws.max_row + 1):
        raw_name = str(_get(ws, r, cols, "name") or "").strip()
        if not raw_name or raw_name == "None":
            blank_streak += 1
            if blank_streak >= 4 and rows:
                break
            continue
        blank_streak = 0
        norm = normalize_header(raw_name)
        total_value = _to_float(_get(ws, r, cols, "allowance_total"))
        if norm == "工资表":
            base_total = total_value
            continue
        if norm in {"合计", "总计"}:
            grand_total = total_value
            break
        post = _to_float(_get(ws, r, cols, "post_allowance"))
        house = _to_float(_get(ws, r, cols, "house_car_allowance"))
        if post == 0 and house == 0 and total_value == 0:
            continue
        rows.append({
            "source_row_number": r,
            "employee_name": raw_name,
            "post_allowance": post,
            "house_car_allowance": house,
            "allowance_total": total_value if total_value else post + house,
            "source_raw": {
                "姓名": raw_name,
                "职位补贴": _cell_to_jsonable(_get(ws, r, cols, "post_allowance")),
                "房/车补": _cell_to_jsonable(_get(ws, r, cols, "house_car_allowance")),
                "合计": _cell_to_jsonable(_get(ws, r, cols, "allowance_total")),
            },
        })
    return {"rows": rows, "base_total": base_total, "grand_total": grand_total}


def _normalized_name(name):
    return re.sub(r"\s+", "", str(name or "").strip())


def merge_allowances(main_rows, allowance_rows):
    by_name = {_normalized_name(r["employee_name"]): r for r in main_rows}
    for ar in allowance_rows:
        key = _normalized_name(ar["employee_name"])
        target = by_name.get(key)
        if target is None and key:
            # Historical files contain a few one-character spelling variants between
            # the wage table and the allowance table.  Merge only when exactly one
            # candidate has the same length and differs by one character.
            candidates = [
                (name_key, row) for name_key, row in by_name.items()
                if len(name_key) == len(key) and sum(a != b for a, b in zip(name_key, key)) == 1
            ]
            if len(candidates) == 1:
                target = candidates[0][1]
                target.setdefault("source_allowance_name", ar["employee_name"])
        if target is None:
            # Preserve allowance-only people. Employee-master reconciliation later
            # decides whether this is a valid manager/other employee or an unknown.
            target = {
                "source_row_number": ar.get("source_row_number"),
                "source_schema_version": "allowance-only",
                "source_serial_no": "",
                "source_no_kind": "",
                "employee_name": ar["employee_name"],
                "source_joining_date": "",
                "source_month_date": "",
                "source_tenure_days": 0.0,
                "work_days": 0.0,
                "work_hours": 0.0,
                "day_salary": 0.0,
                "hour_salary": 0.0,
                "full_attendance": 0.0,
                "overtime_hours": 0.0,
                "overtime_salary": 0.0,
                "national_days": 0.0,
                "national_salary": 0.0,
                "source_polishing_salary": 0.0,
                "target_rate": "",
                "target_salary": 0.0,
                "source_is_insured": "",
                "deduction": 0.0,
                "source_payable_salary": 0.0,
                "source_paid_salary": 0.0,
                "workshop_net": 0.0,
                "source_signature": "",
                "remarks": "",
                "post_allowance": 0.0,
                "house_car_allowance": 0.0,
                "source_raw_json": "{}",
            }
            main_rows.append(target)
            by_name[key] = target
        target["post_allowance"] = _to_float(ar.get("post_allowance"))
        target["house_car_allowance"] = _to_float(ar.get("house_car_allowance"))
        try:
            raw_snapshot = json.loads(target.get("source_raw_json") or "{}")
        except Exception:
            raw_snapshot = {}
        raw_snapshot["__补贴表__"] = {
            "source_row_number": ar.get("source_row_number"),
            **(ar.get("source_raw") or {}),
        }
        target["source_raw_json"] = json.dumps(raw_snapshot, ensure_ascii=False, separators=(",", ":"))
    return main_rows


def parse_external_payroll_workbook(workbook, file_bytes=None, filename=""):
    main = find_best_main_table(workbook)
    ws = main["sheet"]
    main_rows = _parse_main_rows(ws, main["header_row"], main["col_map"], main["source_headers"], main["schema_version"])
    allowance_ws, allowance_info = find_best_allowance_table(workbook, ws, main["header_row"])
    allowance = _parse_allowance_rows(allowance_ws, allowance_info) if allowance_ws is not None else {"rows": [], "base_total": None, "grand_total": None}
    rows = merge_allowances(main_rows, allowance["rows"])

    main_net_total = round(sum(_to_float(r.get("workshop_net")) for r in rows), 2)
    allowance_total = round(sum(_to_float(r.get("post_allowance")) + _to_float(r.get("house_car_allowance")) for r in rows), 2)
    grand_total = round(main_net_total + allowance_total, 2)
    reported_base = allowance.get("base_total")
    reported_grand = allowance.get("grand_total")
    base_diff = None if reported_base is None else round(main_net_total - _to_float(reported_base), 2)
    grand_diff = None if reported_grand is None else round(grand_total - _to_float(reported_grand), 2)

    recognized_count = len(main["col_map"])
    compatibility_score = min(100, 70 + min(24, recognized_count * 2) + (4 if allowance_info else 0) + (2 if grand_diff in (None, 0) else 0))
    diagnostics = []
    if main["unknown_headers"]:
        diagnostics.append(f"主表有 {len(main['unknown_headers'])} 个未映射/重复表头，已保留原文件但未参与标准字段计算")
    if grand_diff not in (None, 0):
        diagnostics.append(f"源表总计与逐行解析差额 {grand_diff:.2f} 元，需要人工复核")
    if base_diff not in (None, 0):
        diagnostics.append(f"源表‘工资表’基数与主表实发合计差额 {base_diff:.2f} 元")

    file_sha256 = hashlib.sha256(file_bytes).hexdigest() if file_bytes else ""
    return {
        "parser_version": PARSER_VERSION,
        "schema_version": main["schema_version"],
        "sheet_name": main["sheet_name"],
        "header_row": main["header_row"],
        "header_score": main["score"],
        "compatibility_score": compatibility_score,
        "recognized_fields": sorted(main["col_map"].keys()),
        "source_headers": main["source_headers"],
        "unknown_headers": main["unknown_headers"],
        "allowance_table_found": bool(allowance_info),
        "allowance_sheet_name": allowance_ws.title if allowance_ws is not None and allowance_info else "",
        "rows": rows,
        "employee_count": len(rows),
        "main_net_total": main_net_total,
        "allowance_total": allowance_total,
        "grand_total": grand_total,
        "reported_base_total": reported_base,
        "reported_grand_total": reported_grand,
        "base_reconciliation_diff": base_diff,
        "grand_reconciliation_diff": grand_diff,
        "diagnostics": diagnostics,
        "file_sha256": file_sha256,
        "filename": filename or "",
    }


def load_and_parse_external_payroll(file_bytes, filename=""):
    """Load .xlsx/.xlsm bytes with openpyxl and return the semantic parse result."""
    import io
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=False)
    return wb, parse_external_payroll_workbook(wb, file_bytes=file_bytes, filename=filename)
