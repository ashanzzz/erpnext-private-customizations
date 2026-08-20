# Copyright (c) 2026, Ashan CN Procurement and contributors
# For license information, please see license.txt

import io
import json
import calendar
import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

import frappe
from frappe import _
from frappe.utils import flt, cint, getdate, nowdate, now_datetime, formatdate


def check_meal_permission():
	"""
	跨公司权限放宽校验：
	操作人只需要拥有天津祺富或天津吉众任意一家公司的权限（或具有全局财务/人事管理角色），即可通行
	"""
	roles = frappe.get_roles(frappe.session.user)
	allowed_roles = {"System Manager", "Administrator", "HR Manager", "Accounts Manager", "HR User", "Accounts User"}
	if any(r in allowed_roles for r in roles):
		return True

	user = frappe.session.user
	user_perms = frappe.get_all("User Permission", filters={"user": user, "allow": "Company"}, fields=["for_value"])
	user_companies = {p.for_value for p in user_perms}

	# 检查是否包含吉众或祺富
	for comp in user_companies:
		if "吉众" in comp or "祺富" in comp:
			return True

	# 如果没有限定特定公司且能访问 Desk
	if not user_companies and frappe.session.user != "Guest":
		return True

	frappe.throw(_("您需要拥有天津吉众或天津祺富任意一家公司的操作权限"), frappe.PermissionError)


def get_month_days_meta(year, month):
	"""根据年份和月份生成 1~31 日的基础日历元数据（包含星期、是否周末）"""
	year, month = cint(year), cint(month)
	_, last_day = calendar.monthrange(year, month)
	weekdays_cn = ["周一", "周二", "周三", "周四", "周五", "周六", "周日"]

	# 查询法定节假日表（如果有）
	holiday_map = {}
	if frappe.db.exists("DocType", "Ashan Holiday Calendar"):
		holidays = frappe.get_all("Ashan Holiday Calendar", filters={
			"calendar_date": ["between", [f"{year:04d}-{month:02d}-01", f"{year:04d}-{month:02d}-{last_day:02d}"]]
		}, fields=["calendar_date", "is_workday", "is_legal_holiday", "holiday_name"])
		for h in holidays:
			holiday_map[str(h.calendar_date)] = h

	days = []
	for d in range(1, last_day + 1):
		date_str = f"{year:04d}-{month:02d}-{d:02d}"
		dt = datetime.date(year, month, d)
		weekday_idx = dt.weekday()
		day_name = weekdays_cn[weekday_idx]
		is_weekend = weekday_idx in [5, 6]

		# 节假日覆盖判断
		if date_str in holiday_map:
			h_rec = holiday_map[date_str]
			# 如果 is_workday == 0 则为休息/节假日
			is_h = cint(h_rec.is_workday) == 0 or cint(h_rec.is_legal_holiday) == 1
			h_name = h_rec.holiday_name or ("法定节假日" if cint(h_rec.is_legal_holiday) == 1 else "休息日")
		else:
			is_h = is_weekend
			h_name = "周末" if is_weekend else ""

		days.append({
			"dining_date": date_str,
			"day_num": d,
			"day_of_week": day_name,
			"is_holiday": 1 if is_h else 0,
			"holiday_name": h_name
		})
	return days, last_day


@frappe.whitelist()
def get_meal_workbench_data(settlement_month):
	"""
	获取指定月份的工作餐月结工作台数据
	"""
	check_meal_permission()
	settlement_month = str(settlement_month).strip()
	year, month = cint(settlement_month.split("-")[0]), cint(settlement_month.split("-")[1])
	days_meta, last_day = get_month_days_meta(year, month)

	doc_name = f"MEAL-{settlement_month}"
	existing_doc = None
	if frappe.db.exists("Ashan Monthly Meal Settlement", doc_name):
		existing_doc = frappe.get_doc("Ashan Monthly Meal Settlement", doc_name)

	default_price = flt(existing_doc.default_meal_price if existing_doc else 15.0) or 15.0
	status = existing_doc.status if existing_doc else "草稿"
	catering_supplier = existing_doc.catering_supplier if existing_doc else ""
	invoice_status = existing_doc.invoice_status if existing_doc else "未开票"
	matched_tax_invoice = existing_doc.matched_tax_invoice if existing_doc else ""
	remark = existing_doc.remark if existing_doc else ""

	# 汇总已存的每日数据
	item_map = {}
	if existing_doc and existing_doc.daily_items:
		for item in existing_doc.daily_items:
			item_map[str(item.dining_date)] = item

	merged_days = []
	tot_qifu_cnt = 0
	tot_qifu_amt = 0.0
	tot_jizhong_cnt = 0
	tot_jizhong_amt = 0.0

	for meta in days_meta:
		d_str = meta["dining_date"]
		if d_str in item_map:
			it = item_map[d_str]
			q_cnt = cint(it.qifu_count or 0)
			j_cnt = cint(it.jizhong_count or 0)
			price = flt(it.meal_price or default_price)
			r_note = it.remark or ""
		else:
			q_cnt = 0
			j_cnt = 0
			price = default_price
			r_note = ""

		q_amt = flt(round(q_cnt * price, 2))
		j_amt = flt(round(j_cnt * price, 2))
		t_cnt = q_cnt + j_cnt
		t_amt = flt(round(q_amt + j_amt, 2))

		tot_qifu_cnt += q_cnt
		tot_qifu_amt += q_amt
		tot_jizhong_cnt += j_cnt
		tot_jizhong_amt += j_amt

		merged_days.append({
			"dining_date": d_str,
			"day_num": meta["day_num"],
			"day_of_week": meta["day_of_week"],
			"is_holiday": meta["is_holiday"],
			"holiday_name": meta["holiday_name"],
			"qifu_count": q_cnt,
			"jizhong_count": j_cnt,
			"meal_price": price,
			"qifu_amount": q_amt,
			"jizhong_amount": j_amt,
			"total_count": t_cnt,
			"total_amount": t_amt,
			"remark": r_note
		})

	return {
		"settlement_month": settlement_month,
		"status": status,
		"default_meal_price": default_price,
		"catering_supplier": catering_supplier,
		"invoice_status": invoice_status,
		"matched_tax_invoice": matched_tax_invoice,
		"remark": remark,
		"kpis": {
			"qifu_total_count": tot_qifu_cnt,
			"qifu_total_amount": flt(round(tot_qifu_amt, 2)),
			"jizhong_total_count": tot_jizhong_cnt,
			"jizhong_total_amount": flt(round(tot_jizhong_amt, 2)),
			"grand_total_count": tot_qifu_cnt + tot_jizhong_cnt,
			"grand_total_amount": flt(round(tot_qifu_amt + tot_jizhong_amt, 2)),
			"average_daily_amount": flt(round((tot_qifu_amt + tot_jizhong_amt) / max(1, last_day), 2))
		},
		"daily_records": merged_days
	}


@frappe.whitelist()
def save_meal_workbench_data(settlement_month, records, default_meal_price=15.0, catering_supplier=None, remark=None):
	"""
	保存每日订餐数据 (支持内联表格自动保存)
	"""
	check_meal_permission()
	settlement_month = str(settlement_month).strip()
	default_meal_price = flt(default_meal_price or 15.0)

	if isinstance(records, str):
		records = json.loads(records)

	doc_name = f"MEAL-{settlement_month}"
	if frappe.db.exists("Ashan Monthly Meal Settlement", doc_name):
		doc = frappe.get_doc("Ashan Monthly Meal Settlement", doc_name)
	else:
		doc = frappe.new_doc("Ashan Monthly Meal Settlement")
		doc.settlement_month = settlement_month

	doc.default_meal_price = default_meal_price
	if catering_supplier is not None:
		doc.catering_supplier = catering_supplier
	if remark is not None:
		doc.remark = remark

	doc.daily_items = []
	for r in records:
		doc.append("daily_items", {
			"dining_date": r.get("dining_date"),
			"day_of_week": r.get("day_of_week"),
			"is_holiday": cint(r.get("is_holiday") or 0),
			"qifu_count": cint(r.get("qifu_count") or 0),
			"jizhong_count": cint(r.get("jizhong_count") or 0),
			"meal_price": flt(r.get("meal_price") or default_meal_price),
			"remark": r.get("remark") or ""
		})

	doc.save(ignore_permissions=True)
	frappe.db.commit()

	return get_meal_workbench_data(settlement_month)


@frappe.whitelist()
def upload_and_parse_meal_excel():
	"""
	上传 Excel 订餐记录并智能解析
	"""
	check_meal_permission()

	if "file" not in frappe.request.files:
		frappe.throw(_("请选择要上传的 Excel 文件 (.xlsx / .xls)"))

	uploaded_file = frappe.request.files["file"]
	filename = uploaded_file.filename
	settlement_month = frappe.form_dict.get("settlement_month") or ""
	settlement_month = settlement_month.strip()

	if not (filename.lower().endswith(".xlsx") or filename.lower().endswith(".xls")):
		frappe.throw(_("仅支持上传 Excel 格式文件 (.xlsx / .xls)"))

	content = uploaded_file.read()
	file_io = io.BytesIO(content)

	try:
		wb = openpyxl.load_workbook(file_io, data_only=True)
	except Exception as e:
		frappe.throw(_("无法读取 Excel 文件，请确认文件未损坏且格式正确。错误: {0}").format(str(e)))

	# 1. 寻找匹配的 Sheet
	sheet_names = wb.sheetnames
	matched_sheet = None

	if settlement_month:
		# 例如 2026-06 -> 年: 2026, 月: 06 或 6
		parts = settlement_month.split("-")
		s_year, s_month = parts[0], parts[1]
		s_month_int = str(cint(s_month))

		candidates = [
			f"{s_year}年订餐{s_month_int}月",
			f"{s_year}年订餐{s_month}月",
			f"{s_year}订餐{s_month_int}月",
			f"{s_year}订餐{s_month}月",
			f"{s_year}年{s_month_int}月",
			f"{s_year}年{s_month}月",
			f"{s_year}-{s_month}",
			f"{s_year}{s_month}",
			f"{s_month_int}月",
			f"{s_month}月"
		]

		for cand in candidates:
			for sn in sheet_names:
				if cand in sn.replace(" ", ""):
					matched_sheet = wb[sn]
					break
			if matched_sheet:
				break

	if not matched_sheet:
		# 默认取最后一个工作表或者第一个
		# 如果 Sheet 名称含有当前年份或月份
		for sn in reversed(sheet_names):
			if "订餐" in sn:
				matched_sheet = wb[sn]
				break

	if not matched_sheet:
		matched_sheet = wb.active

	if not matched_sheet:
		frappe.throw(_("未能从 Excel 文件中找到有效的订餐工作表，请检查 Sheet 命名是否包含年月（例如：2026年订餐6月）"))

	# 2. 识别表头行与列索引
	ws = matched_sheet
	header_row_idx = None
	col_map = {}

	for r in range(1, min(15, ws.max_row + 1)):
		row_vals = [str(ws.cell(r, c).value or "").strip() for c in range(1, ws.max_column + 1)]
		# 检查是否包含关键列名
		has_date = any("日期" in v for v in row_vals)
		has_qifu = any("祺富" in v and ("数" in v or "量" in v) for v in row_vals)
		has_jizhong = any("吉众" in v and ("数" in v or "量" in v) for v in row_vals)

		if has_date and (has_qifu or has_jizhong):
			header_row_idx = r
			for c_idx, val in enumerate(row_vals, 1):
				if "日期" in val:
					col_map["date"] = c_idx
				elif "祺富" in val and ("数" in val or "量" in val):
					col_map["qifu"] = c_idx
				elif "吉众" in val and ("数" in val or "量" in val):
					col_map["jizhong"] = c_idx
				elif "单价" in val:
					col_map["price"] = c_idx
				elif "备注" in val:
					col_map["remark"] = c_idx
			break

	if not header_row_idx or "date" not in col_map:
		frappe.throw(_("未能识别工作表【{0}】的表头结构。请确保表头包含【日期】、【祺富数量】、【吉众数量】、【餐费单价】等标准列。").format(ws.title))

	# 3. 逐行解析数据
	parsed_records = []
	excel_base_date = datetime.date(1899, 12, 30)

	for r in range(header_row_idx + 1, ws.max_row + 1):
		raw_date = ws.cell(r, col_map["date"]).value
		if raw_date is None:
			continue

		# 处理日期 (可能是整数序号、datetime、或者字符串)
		date_obj = None
		if isinstance(raw_date, (int, float)):
			try:
				date_obj = excel_base_date + datetime.timedelta(days=int(raw_date))
			except Exception:
				pass
		elif isinstance(raw_date, datetime.datetime):
			date_obj = raw_date.date()
		elif isinstance(raw_date, datetime.date):
			date_obj = raw_date
		elif isinstance(raw_date, str):
			raw_str = raw_date.strip()
			if "合计" in raw_str or "总计" in raw_str:
				break  # 到达汇总行
			try:
				date_obj = getdate(raw_str)
			except Exception:
				pass

		if not date_obj:
			continue

		date_str = str(date_obj)

		# 份数与单价
		qifu_cnt = cint(ws.cell(r, col_map.get("qifu", 0)).value or 0) if "qifu" in col_map else 0
		jizhong_cnt = cint(ws.cell(r, col_map.get("jizhong", 0)).value or 0) if "jizhong" in col_map else 0
		price_val = ws.cell(r, col_map.get("price", 0)).value if "price" in col_map else 15.0
		meal_price = flt(price_val or 15.0)
		remark_val = str(ws.cell(r, col_map.get("remark", 0)).value or "").strip() if "remark" in col_map else ""

		parsed_records.append({
			"dining_date": date_str,
			"qifu_count": qifu_cnt,
			"jizhong_count": jizhong_cnt,
			"meal_price": meal_price,
			"remark": remark_val
		})

	if not parsed_records:
		frappe.throw(_("工作表【{0}】未提取到任何有效订餐数据行，请检查数据区域。").format(ws.title))

	# 确定结算月份
	if not settlement_month:
		settlement_month = parsed_records[0]["dining_date"][:7]

	# 保存并返回
	res = save_meal_workbench_data(
		settlement_month=settlement_month,
		records=parsed_records,
		default_meal_price=parsed_records[0].get("meal_price", 15.0)
	)

	return {
		"ok": True,
		"sheet_name": ws.title,
		"parsed_count": len(parsed_records),
		"workbench_data": res
	}


@frappe.whitelist()
def export_meal_settlement_excel(settlement_month):
	"""
	1:1 导出高品质 Excel 订餐记录表
	"""
	check_meal_permission()
	data = get_meal_workbench_data(settlement_month)
	records = data.get("daily_records") or []
	kpis = data.get("kpis") or {}

	wb = openpyxl.Workbook()
	ws = wb.active
	ws.title = f"{settlement_month}订餐记录"
	ws.views.sheetView[0].showGridLines = True

	# 样式常量
	font_title = Font(name="等线", size=16, bold=True)
	font_hdr = Font(name="等线", size=11, bold=True)
	font_data = Font(name="等线", size=11, bold=False)
	font_total = Font(name="等线", size=11, bold=True)
	align_center = Alignment(horizontal="center", vertical="center")
	align_right = Alignment(horizontal="right", vertical="center")

	thin_side = Side(style="thin", color="000000")
	med_side = Side(style="medium", color="000000")
	b_all_thin = Border(top=thin_side, bottom=thin_side, left=thin_side, right=thin_side)

	# 列宽 (9列)
	col_widths = {1: 14.0, 2: 12.0, 3: 12.0, 4: 12.0, 5: 14.0, 6: 14.0, 7: 15.0, 8: 15.0, 9: 20.0}
	for c_idx, w in col_widths.items():
		ws.column_dimensions[get_column_letter(c_idx)].width = w

	# 标题行 (Row 1)
	ws.row_dimensions[1].height = 32.0
	ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)
	c_title = ws.cell(1, 1, "吉众 & 祺富公司员工订餐记录表")
	c_title.font = font_title
	c_title.alignment = align_center

	# 表头行 (Row 2)
	ws.row_dimensions[2].height = 26.0
	headers = ["日期", "祺富数量", "吉众数量", "餐费单价", "祺富金额", "吉众金额", "用餐数量合计", "金额合计", "备注"]
	for idx, h in enumerate(headers, 1):
		c = ws.cell(2, idx, h)
		c.font = font_hdr
		c.alignment = align_center
		c.border = Border(top=med_side, bottom=med_side, left=thin_side, right=thin_side)

	# 数据行 (Row 3 ~ N)
	curr_row = 3
	for r in records:
		ws.row_dimensions[curr_row].height = 22.0
		ws.cell(curr_row, 1, r["dining_date"])
		ws.cell(curr_row, 2, r["qifu_count"])
		ws.cell(curr_row, 3, r["jizhong_count"])
		ws.cell(curr_row, 4, r["meal_price"])
		ws.cell(curr_row, 5, f"=B{curr_row}*D{curr_row}")
		ws.cell(curr_row, 6, f"=C{curr_row}*D{curr_row}")
		ws.cell(curr_row, 7, f"=B{curr_row}+C{curr_row}")
		ws.cell(curr_row, 8, f"=E{curr_row}+F{curr_row}")
		ws.cell(curr_row, 9, r["remark"] or "")

		for c in range(1, 10):
			cell = ws.cell(curr_row, c)
			cell.font = font_data
			cell.alignment = align_center
			cell.border = b_all_thin
			if c in [2, 3, 7]:
				cell.number_format = "#,##0"
			elif c in [4, 5, 6, 8]:
				cell.number_format = "#,##0.00"
		curr_row += 1

	# 合计行
	ws.row_dimensions[curr_row].height = 28.0
	ws.cell(curr_row, 1, "合计").font = font_total
	ws.cell(curr_row, 2, f"=SUM(B3:B{curr_row-1})").font = font_total
	ws.cell(curr_row, 3, f"=SUM(C3:C{curr_row-1})").font = font_total
	ws.cell(curr_row, 4, "").font = font_total
	ws.cell(curr_row, 5, f"=SUM(E3:E{curr_row-1})").font = font_total
	ws.cell(curr_row, 6, f"=SUM(F3:F{curr_row-1})").font = font_total
	ws.cell(curr_row, 7, f"=SUM(G3:G{curr_row-1})").font = font_total
	ws.cell(curr_row, 8, f"=SUM(H3:H{curr_row-1})").font = font_total
	ws.cell(curr_row, 9, "").font = font_total

	for c in range(1, 10):
		cell = ws.cell(curr_row, c)
		cell.alignment = align_center
		cell.border = Border(top=thin_side, bottom=med_side, left=thin_side, right=thin_side)
		if c in [2, 3, 7]:
			cell.number_format = "#,##0"
		elif c in [5, 6, 8]:
			cell.number_format = "#,##0.00"

	bio = io.BytesIO()
	wb.save(bio)
	bio.seek(0)

	fname = f"吉众_祺富_{settlement_month}_员工订餐记录表.xlsx"
	frappe.response["filename"] = fname
	frappe.response["filecontent"] = bio.getvalue()
	frappe.response["type"] = "binary"


@frappe.whitelist()
def finalize_meal_settlement(settlement_month):
	"""核定锁定当月餐费账目"""
	check_meal_permission()
	doc_name = f"MEAL-{settlement_month}"
	if not frappe.db.exists("Ashan Monthly Meal Settlement", doc_name):
		frappe.throw(_("请先录入并保存订餐数据"))

	doc = frappe.get_doc("Ashan Monthly Meal Settlement", doc_name)
	doc.status = "已核定"
	doc.save(ignore_permissions=True)
	frappe.db.commit()

	return get_meal_workbench_data(settlement_month)


@frappe.whitelist()
def revert_finalize_meal_settlement(settlement_month):
	"""取消核定 (重新录入)"""
	check_meal_permission()
	doc_name = f"MEAL-{settlement_month}"
	if not frappe.db.exists("Ashan Monthly Meal Settlement", doc_name):
		frappe.throw(_("未找到该月餐费核定记录"))

	doc = frappe.get_doc("Ashan Monthly Meal Settlement", doc_name)
	doc.status = "草稿"
	doc.save(ignore_permissions=True)
	frappe.db.commit()

	return get_meal_workbench_data(settlement_month)


@frappe.whitelist()
def clear_meal_workbench_data(settlement_month):
	"""清空当月所有订餐明细（重置为0）"""
	check_meal_permission()
	doc_name = f"MEAL-{settlement_month}"
	if not frappe.db.exists("Ashan Monthly Meal Settlement", doc_name):
		return get_meal_workbench_data(settlement_month)

	doc = frappe.get_doc("Ashan Monthly Meal Settlement", doc_name)
	if doc.status == "已核定":
		frappe.throw(_("当前月份已核定锁定，请先取消核定后再执行清空！"))

	for item in doc.daily_items:
		item.qifu_count = 0
		item.jizhong_count = 0
		item.remark = ""

	doc.save(ignore_permissions=True)
	frappe.db.commit()

	return get_meal_workbench_data(settlement_month)
