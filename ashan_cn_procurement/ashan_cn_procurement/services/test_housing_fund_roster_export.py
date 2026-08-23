"""Focused regression tests for housing-fund increase and sealing roster logic."""

import sys
import types
import unittest
from datetime import date, datetime
from io import BytesIO


def _install_service_fakes():
    """Provide the small Frappe surface needed by the standalone service test."""
    frappe = types.ModuleType("frappe")
    frappe.throw = lambda message: (_ for _ in ()).throw(ValueError(message))
    frappe.whitelist = lambda *args, **kwargs: (
        lambda fn: fn
    ) if not (args and callable(args[0])) else args[0]
    frappe.get_all = lambda *args, **kwargs: []
    sys.modules["frappe"] = frappe

    utils = types.ModuleType("frappe.utils")
    utils.flt = lambda value=0: float(value or 0)
    utils.getdate = lambda value: (
        value if isinstance(value, date) else datetime.strptime(str(value)[:10], "%Y-%m-%d").date()
    )
    sys.modules["frappe.utils"] = utils

    employee_salary = types.ModuleType(
        "ashan_cn_procurement.services.employee_salary_service"
    )
    employee_salary.get_insurance_setting = lambda company, year: {
        "hf_company_rate": 5,
        "hf_person_rate": 5,
        "hf_auto_rule_enabled": 1,
        "hf_contribution_months": "1,4,7,10",
        "hf_off_month_action": "停缴",
    }
    sys.modules[employee_salary.__name__] = employee_salary

    policy = types.ModuleType(
        "ashan_cn_procurement.services.housing_fund_policy_service"
    )
    policy.EXCLUDED_EMPLOYEE_TYPES = {"临时工"}
    policy.POLICY_FIXED_OFF = "固定停缴"
    policy.normalize_period_month = lambda value: (
        f"{str(value)[:4]}-{str(value)[5:7]}", int(str(value)[:4]), int(str(value)[5:7])
    )
    policy.normalize_policy_setting = lambda setting: {
        "enabled": True,
        "months": [1, 4, 7, 10],
        "off_action": "停缴",
    }
    sys.modules[policy.__name__] = policy

    proof = types.ModuleType(
        "ashan_cn_procurement.services.payroll_proof_validation"
    )
    proof.expected_proof_period = lambda period: (
        f"{int(period[:4]) + (period[5:7] == '12'):04d}-"
        f"{(int(period[5:7]) % 12) + 1:02d}"
    )
    sys.modules[proof.__name__] = proof

    payroll = types.ModuleType(
        "ashan_cn_procurement.services.payroll_settlement_service"
    )
    payroll.check_payroll_workbench_permission = lambda *args, **kwargs: None
    sys.modules[payroll.__name__] = payroll


_install_service_fakes()

from openpyxl import load_workbook  # noqa: E402
from ashan_cn_procurement.services import housing_fund_roster_export_service as roster  # noqa: E402


class TestHousingFundRosterExport(unittest.TestCase):
    def setUp(self):
        self.original_get_all = roster.frappe.get_all

        def get_all(doctype, filters=None, **kwargs):
            if doctype == "Ashan Employee Salary Profile":
                return [
                    {"employee_no": "A", "employee_name": "甲", "id_card": "1", "employee_type": "正式工", "housing_fund_base": 2320, "housing_fund_policy": "跟随公司规则"},
                    {"employee_no": "B", "employee_name": "乙", "id_card": "2", "employee_type": "正式工", "housing_fund_base": 0, "housing_fund_policy": "跟随公司规则"},
                    {"employee_no": "C", "employee_name": "丙", "id_card": "3", "employee_type": "正式工", "housing_fund_base": 2320, "housing_fund_policy": "跟随公司规则"},
                ]
            if doctype == "Ashan Monthly Payroll Settlement":
                return [
                    {"name": "P-2026-07", "period_month": "2026-07"},
                    {"name": "P-2026-06", "period_month": "2026-06"},
                ]
            if doctype == "Ashan Monthly Payroll Item":
                if filters["parent"] == "P-2026-07":
                    return [{"employee_no": "A", "employee_name": "甲", "id_card": "1", "hf_base": 2320}]
                return [
                    {"employee_no": "A", "employee_name": "甲", "id_card": "1", "hf_base": 2320},
                    {"employee_no": "B", "employee_name": "乙", "id_card": "2", "hf_base": 2320},
                ]
            return []

        roster.frappe.get_all = get_all

    def tearDown(self):
        roster.frappe.get_all = self.original_get_all

    def test_deleted_member_seals_against_latest_scheduled_material(self):
        changes = roster._get_change_rows("测试公司", "2026-08")

        self.assertEqual(changes["comparison_mode"], "material_delta")
        self.assertEqual(changes["reference_period_month"], "2026-06")
        self.assertEqual([row["employee_no"] for row in changes[roster.CHANGE_INCREASE]], ["C"])
        self.assertEqual([row["employee_no"] for row in changes[roster.CHANGE_SEAL]], ["B"])

    def test_sealing_workbook_keeps_reason_code_four(self):
        content = roster._build_roster_workbook(
            roster.CHANGE_SEAL,
            [{"employee_no": "B", "employee_name": "乙", "id_card": "2", "base": 2320, "amount": 232}],
        )
        sheet = load_workbook(BytesIO(content)).active

        self.assertEqual(sheet.title, "Sheet0")
        self.assertEqual(sheet.cell(1, 1).value, "清册类别(必须填'fc')")
        self.assertEqual(sheet.cell(2, 1).value, "fc")
        self.assertEqual(sheet.cell(2, 4).value, "4")


if __name__ == "__main__":
    unittest.main()
