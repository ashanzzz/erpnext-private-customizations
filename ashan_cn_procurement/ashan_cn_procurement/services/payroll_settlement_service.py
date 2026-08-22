import pypdf
import zipfile
# Copyright (c) 2026, Ashan CN Procurement
import json
import re
from decimal import Decimal, ROUND_HALF_UP

import frappe
from frappe.utils import flt, cint, now_datetime, getdate
from ashan_cn_procurement.services.employee_salary_service import get_insurance_setting

# 个税台账参与规则：临时工/零工不进入个税申报台账；返聘类（含“其他-返聘工”）参与个税。
TAX_LEDGER_EXCLUDED_EMPLOYEE_TYPES = {"临时工", "零工"}
TAX_REHIRE_EMPLOYEE_TYPES = {"返聘工", "退休返聘", "其他-返聘工"}
TAX_AFTER_SALARY_MODES = {"税后", "税后倒推", "税后管理工资"}

def is_tax_ledger_employee(employee_type):
    """是否进入个人所得税申报台账/累计预扣计算。"""
    return str(employee_type or "").strip() not in TAX_LEDGER_EXCLUDED_EMPLOYEE_TYPES

def is_tax_after_salary_mode(salary_mode):
    return str(salary_mode or "").strip() in TAX_AFTER_SALARY_MODES


def _normalize_period_month(period_month):
    """Normalize period month to YYYY-MM format."""
    pm_str = str(period_month or "").strip()
    if len(pm_str) == 6 and pm_str.isdigit():
        return f"{pm_str[:4]}-{pm_str[4:]}"
    return pm_str or "2026-07"


def _salary_profiles_for_period(company, period_month, fields=None, order_by="employee_no asc"):
    """Return employees who actually belong to a payroll month.

    A current ``employment_status='离职'`` must not erase the employee from an earlier
    payroll month. Conversely, a future joiner must not appear in an older month.  The
    period boundaries therefore come from joining/relieving dates, with current status
    used only as a fallback when no relieving date exists.
    """
    from calendar import monthrange
    from datetime import date

    pm_str = _normalize_period_month(period_month)
    try:
        year, month = [int(part) for part in pm_str.split("-")[:2]]
        month_start = date(year, month, 1)
        month_end = date(year, month, monthrange(year, month)[1])
    except Exception:
        frappe.throw("账期格式必须为 YYYY-MM 或 YYYYMM，例如 2026-07。")

    wanted = list(fields or ["name", "employee_no", "employee_name"])
    required = ["employment_status", "date_of_joining", "relieving_date"]
    query_fields = list(dict.fromkeys(wanted + required))
    rows = frappe.get_all(
        "Ashan Employee Salary Profile",
        filters={"company": company},
        fields=query_fields,
        order_by=order_by,
    )

    result = []
    for row in rows:
        joined = getdate(row.get("date_of_joining")) if row.get("date_of_joining") else None
        relieved = getdate(row.get("relieving_date")) if row.get("relieving_date") else None
        status = str(row.get("employment_status") or "在职").strip()
        if joined and joined > month_end:
            continue
        if relieved and relieved < month_start:
            continue
        if status in {"离职", "已离职"} and not relieved:
            continue
        result.append(row)
    return result


def check_payroll_workbench_permission(perm_type="read"):
	"""
	检查当前登录用户是否具备人事薪酬工作台及凭证的访问/操作权限
	"""
	user = frappe.session.user
	if user == "Administrator":
		return True
	
	user_roles = set(frappe.get_roles(user))
	allowed_roles = {"System Manager", "HR Manager", "HR User", "Accounts Manager", "Accounts User"}
	if user_roles & allowed_roles:
		return True
	
	if frappe.has_permission("Ashan Monthly Payroll Settlement", perm_type):
		return True
		
	action_name = "上传或修改凭证数据" if perm_type in ["write", "create"] else "查看或下载凭证原件"
	frappe.throw(f"【⛔ 权限拦截】您的账号 ({user}) 未被授予人事或财务权限，禁止{action_name}！", frappe.PermissionError)

# 个税统一由累计预扣/税后反推引擎处理；不再保留旧“单月税率表”算法，避免后续入口误用。
@frappe.whitelist()
def get_payroll_settlement_detail(company, period_month):
	check_payroll_workbench_permission("read")
	# 缴纳期计算
	parts = period_month.split("-")
	year_int = int(parts[0]) if len(parts) > 0 else 2026
	month_int = int(parts[1]) if len(parts) > 1 else 7
	if month_int == 12:
		payment_year = year_int + 1
		payment_month = 1
	else:
		payment_year = year_int
		payment_month = month_int + 1
	payment_month_name = f"{payment_year}年{payment_month}月"
	period_month_str = period_month.replace("-", "")

	# 母表人员结构统计：按选择账期还原实际在册边界，避免今天的离职状态污染历史月份。
	all_profiles = _salary_profiles_for_period(
		company, period_month,
		fields=["employee_no", "employee_name", "employee_type", "employment_status", "relieving_date", "fixed_salary", "base_salary", "social_security_base", "housing_fund_base"],
	)
	total_profile_count = len(all_profiles)
	regular_count = 0
	regular_zero_count = 0
	rehire_count = 0
	rehire_zero_count = 0
	temp_count = 0
	temp_zero_count = 0
	other_count = 0
	other_zero_count = 0
	resigned_count = 0
	resigned_zero_count = 0

	# 建立当期实发/税前/固定薪资映射
	item_sal_map = {}
	doc_name = f"{company}-{period_month}"
	if frappe.db.exists("Ashan Monthly Payroll Settlement", doc_name):
		settlement_doc = frappe.get_doc("Ashan Monthly Payroll Settlement", doc_name)
		for it in settlement_doc.items:
			sal_val = max(flt(it.get("net_salary")), flt(it.get("gross_salary")), flt(it.get("fixed_salary")), flt(it.get("target_salary")))
			item_sal_map[it.employee_no] = sal_val

	for p in all_profiles:
		etype = (p.get("employee_type") or "正式工").strip()
		emp_no = p.get("employee_no")
		rel_d = str(p.get("relieving_date") or "")
		is_resigned = bool(rel_d and rel_d.startswith(period_month))
		
		emp_cur_sal = item_sal_map.get(emp_no, flt(p.get("fixed_salary")) or flt(p.get("base_salary")))
		is_zero_sal = (emp_cur_sal <= 0.001)

		if is_resigned:
			resigned_count += 1
			if is_zero_sal:
				resigned_zero_count += 1
		elif etype == "正式工":
			regular_count += 1
			if is_zero_sal:
				regular_zero_count += 1
		elif etype in TAX_REHIRE_EMPLOYEE_TYPES:
			rehire_count += 1
			if is_zero_sal:
				rehire_zero_count += 1
		elif etype in ["临时工", "零工"]:
			temp_count += 1
			if is_zero_sal:
				temp_zero_count += 1
		else:
			# 外籍工, 其他-管理, 其他-正式工, 其他-外籍工, 其他-返聘工, 其他等
			other_count += 1
			if is_zero_sal:
				other_zero_count += 1

	total_zero_count = regular_zero_count + rehire_zero_count + temp_zero_count + other_zero_count + resigned_zero_count
	insured_count = regular_count  # 兼容旧字段
	insured_zero_count = regular_zero_count

	# 社保与公积金统计 (使用现有的明细计算函数)
	ss_data = get_social_insurance_sheet(company, period_month)
	ss_totals = ss_data.get("totals", {})
	hf_data = get_housing_fund_sheet(company, period_month)
	hf_totals = hf_data.get("totals", {})

	if not frappe.db.exists("Ashan Monthly Payroll Settlement", doc_name):
		return {
			"exists": False,
			"company": company,
			"period_month": period_month,
			"status": "未创建",
			"locked": 0,
			"items": [],
			"kpi_summary": {
				"total_profile_count": total_profile_count,
				"total_zero_count": total_zero_count,
				"regular_count": regular_count,
				"regular_zero_count": regular_zero_count,
				"insured_count": regular_count,
				"insured_zero_count": regular_zero_count,
				"rehire_count": rehire_count,
				"rehire_zero_count": rehire_zero_count,
				"temp_count": temp_count,
				"temp_zero_count": temp_zero_count,
				"other_count": other_count,
				"other_zero_count": other_zero_count,
				"resigned_count": resigned_count,
				"resigned_zero_count": resigned_zero_count,
				"ss_payment_month_name": payment_month_name,
				"ss_period_month_str": period_month_str,
				"ss_grand_total": ss_totals.get("grand_total", 0.0),
				"ss_comp_total": ss_totals.get("comp_total", 0.0),
				"ss_pers_total": ss_totals.get("pers_total", 0.0),
				"ss_base_total": ss_totals.get("ss_base", 0.0),
				"ss_count": len(ss_data.get("rows", [])),
				"hf_payment_month_name": payment_month_name,
				"hf_period_month_str": period_month_str,
				"hf_grand_total": hf_totals.get("total_amount", 0.0),
				"hf_comp_total": hf_totals.get("comp_amount", 0.0),
				"hf_pers_total": hf_totals.get("pers_amount", 0.0),
				"hf_base_total": hf_totals.get("hf_base", 0.0),
				"hf_count": len(hf_data.get("rows", [])),
				"payroll_period_month": period_month,
				"payroll_emp_count": 0,
				"total_net_salary": 0.0,
				"total_gross_salary": 0.0,
				"total_tax": 0.0,
				"total_person_deductions": ss_totals.get("pers_total", 0.0) + hf_totals.get("pers_amount", 0.0)
			}
		}
	
	doc = frappe.get_doc("Ashan Monthly Payroll Settlement", doc_name)
	items = []
	for it in doc.items:
		items.append(it.as_dict())

	total_gross = flt(doc.total_gross_salary) or sum(flt(it.get("gross_salary")) for it in items)
	total_net = flt(doc.total_net_salary) or sum(flt(it.get("net_salary")) for it in items)
	total_tax = flt(doc.total_tax) or sum(flt(it.get("tax_amount")) for it in items)
	pers_ded_total = ss_totals.get("pers_total", 0.0) + hf_totals.get("pers_amount", 0.0) + total_tax

	kpi_summary = {
		"total_profile_count": total_profile_count,
		"total_zero_count": total_zero_count,
		"regular_count": regular_count,
		"regular_zero_count": regular_zero_count,
		"insured_count": regular_count,
		"insured_zero_count": regular_zero_count,
		"rehire_count": rehire_count,
		"rehire_zero_count": rehire_zero_count,
		"temp_count": temp_count,
		"temp_zero_count": temp_zero_count,
		"other_count": other_count,
		"other_zero_count": other_zero_count,
		"resigned_count": resigned_count,
		"resigned_zero_count": resigned_zero_count,
		"ss_payment_month_name": payment_month_name,
		"ss_period_month_str": period_month_str,
		"ss_grand_total": ss_totals.get("grand_total", 0.0),
		"ss_comp_total": ss_totals.get("comp_total", 0.0),
		"ss_pers_total": ss_totals.get("pers_total", 0.0),
		"ss_base_total": ss_totals.get("ss_base", 0.0),
		"ss_count": len(ss_data.get("rows", [])),
		"hf_payment_month_name": payment_month_name,
		"hf_period_month_str": period_month_str,
		"hf_grand_total": hf_totals.get("total_amount", 0.0),
		"hf_comp_total": hf_totals.get("comp_amount", 0.0),
		"hf_pers_total": hf_totals.get("pers_amount", 0.0),
		"hf_base_total": hf_totals.get("hf_base", 0.0),
		"hf_count": len(hf_data.get("rows", [])),
		"payroll_period_month": period_month,
		"payroll_emp_count": len(items),
		"total_net_salary": total_net,
		"total_gross_salary": total_gross,
		"total_tax": total_tax,
		"total_person_deductions": pers_ded_total
	}

	return {
		"exists": True,
		"name": doc.name,
		"company": doc.company,
		"period_month": doc.period_month,
		"status": doc.status,
		"locked": cint(doc.locked),
		"confirmed_by": doc.confirmed_by or "",
		"confirmed_date": str(doc.confirmed_date or ""),
		"unlock_reason": doc.unlock_reason or "",
		"total_employees": doc.total_employees or len(items),
		"total_gross_salary": total_gross,
		"total_net_salary": total_net,
		"total_social_security_company": flt(doc.total_social_security_company),
		"total_social_security_person": flt(doc.total_social_security_person),
		"total_housing_fund_company": flt(doc.total_housing_fund_company),
		"total_housing_fund_person": flt(doc.total_housing_fund_person),
		"total_tax": total_tax,
		"items": items,
		"kpi_summary": kpi_summary
	}

@frappe.whitelist(methods=["POST"])
def calculate_and_generate_payroll(company, period_month):
    """
    按当前员工母表、社保/公积金配置与历史工资快照重新融合核算。

    个税统一使用与 Excel VBA 相同的累计预扣口径：
    - 税后工资：按 7 级累计税率闭式反推税前工资，并反向验算到分；
    - 税前工资：按累计应纳税所得额正向计算本月应扣税额；
    - 临时工/零工：保留薪酬核算，但不进入个税台账，本月个税固定为 0；
    - 返聘工/退休返聘/其他-返聘工：参与个税累计预扣，但通常不缴社保公积金。
    """
    check_payroll_workbench_permission("write")
    doc_name = f"{company}-{period_month}"
    if frappe.db.exists("Ashan Monthly Payroll Settlement", doc_name):
        existing = frappe.get_doc("Ashan Monthly Payroll Settlement", doc_name)
        if existing.locked:
            frappe.throw(f"【{company}】{period_month} 月度薪酬核定表已被锁定！如需重新核算请先执行【反审核解锁】。")

    year = _period_year(period_month)
    setting_name = f"{company}-{year}"

    # 社保、公积金与个税基础参数
    ss_company_rate = 27.05 if "祺富" in company else 26.85
    ss_person_rate = 10.5
    hf_company_rate = 5.0
    hf_person_rate = 5.0
    current_month_num = cint(period_month.split("-")[1]) if "-" in period_month else 6
    big_med_amount = 22.0
    pension_person_rate = 8.0
    medical_person_rate = 2.0
    unemployment_person_rate = 0.5

    if frappe.db.exists("Ashan Insurance Setting", setting_name):
        ins = frappe.get_doc("Ashan Insurance Setting", setting_name)
        ss_company_rate = flt(ins.ss_company_pension) + flt(ins.ss_company_unemployment) + flt(ins.ss_company_medical) + flt(ins.ss_company_other_medical) + flt(ins.ss_company_injury)
        ss_person_rate = flt(ins.ss_person_pension) + flt(ins.ss_person_unemployment) + flt(ins.ss_person_medical)
        pension_person_rate = flt(ins.ss_person_pension) or 8.0
        medical_person_rate = flt(ins.ss_person_medical) or 2.0
        unemployment_person_rate = flt(ins.ss_person_unemployment) or 0.5
        hf_company_rate = flt(ins.hf_company_rate) or 5.0
        hf_person_rate = flt(ins.hf_person_rate) or 5.0
        special_months_str = str(ins.get("big_medical_special_months") or "3,12")
        special_months = [cint(m.strip()) for m in special_months_str.split(",") if m.strip().isdigit()]
        if current_month_num in special_months:
            big_med_amount = flt(ins.get("big_medical_amount_special")) or 21.0
        else:
            big_med_amount = flt(ins.get("big_medical_amount_default")) or 22.0

    tax_params = get_effective_tax_parameters(company, period_month)
    tax_threshold = tax_params["tax_threshold"]
    cinfo = get_tax_cycle_info(period_month, tax_threshold, tax_params["tax_cycle_start_month"])

    employees = _salary_profiles_for_period(
        company,
        period_month,
        fields=[
            "name", "employee_no", "employee_name", "id_card", "gender", "mobile", "birth_date",
            "department", "job_title", "employee_type", "salary_mode", "fixed_salary", "base_salary",
            "post_allowance", "performance_base", "meal_allowance", "traffic_allowance",
            "communication_allowance", "other_allowance", "social_security_base", "housing_fund_base",
            "deduction_child_education", "deduction_continuing_education", "deduction_housing_loan",
            "deduction_housing_rent", "deduction_elderly_care", "deduction_infant_care",
            "deduction_serious_illness"
        ],
    )

    attendances = {}
    att_records = frappe.get_all(
        "Ashan Monthly Attendance",
        filters={"company": company, "period_month": period_month},
        fields=["employee_no", "attendance_days", "work_hours_regular", "overtime_regular_1_5", "overtime_weekend_2_0", "meal_count"],
    )
    for a in att_records:
        attendances[a.employee_no] = a

    if frappe.db.exists("Ashan Monthly Payroll Settlement", doc_name):
        doc = frappe.get_doc("Ashan Monthly Payroll Settlement", doc_name)
        doc.items = []
    else:
        doc = frappe.new_doc("Ashan Monthly Payroll Settlement")
        doc.company = company
        doc.period_month = period_month

    doc.status = "草稿"
    doc.locked = 0

    total_gross = total_net = total_ss_comp = total_ss_pers = 0.0
    total_hf_comp = total_hf_pers = total_tax = 0.0
    tax_participant_count = 0

    for emp in employees:
        emp_no = emp.employee_no
        emp_type = str(emp.employee_type or "正式工").strip()
        salary_mode = str(emp.salary_mode or "税后").strip()
        tax_eligible = is_tax_ledger_employee(emp_type)
        if tax_eligible:
            tax_participant_count += 1

        att = attendances.get(emp_no)
        att_days = flt(att.attendance_days) if att else 21.75
        work_hrs = flt(att.work_hours_regular) if att else 174.0
        ot_hrs = flt(att.overtime_regular_1_5 + att.overtime_weekend_2_0) if att else 0.0
        meal_cnt = cint(att.meal_count) if att else 0

        fixed_sal = flt(emp.fixed_salary)
        base_sal = flt(emp.base_salary)
        post_allow = flt(emp.post_allowance)
        perf_base = flt(emp.performance_base)
        allow_tot = flt(emp.meal_allowance) + flt(emp.traffic_allowance) + flt(emp.communication_allowance) + flt(emp.other_allowance)
        structured_salary = round(base_sal + post_allow + perf_base + allow_tot, 2)
        target_or_gross = round(fixed_sal if fixed_sal > 0 else structured_salary, 2)

        ss_base = flt(emp.social_security_base)
        hf_base = flt(emp.housing_fund_base)
        # 返聘、临时、外籍、实习通常不在本模块缴纳五险一金；强制按现有业务边界归零，避免误设基数带入扣款。
        no_insurance_types = TAX_REHIRE_EMPLOYEE_TYPES | {"临时工", "零工", "外籍工", "实习生"}
        if emp_type in no_insurance_types:
            ss_base = 0.0
            hf_base = 0.0

        ss_pers = round(ss_base * (ss_person_rate / 100.0) + (big_med_amount if ss_base > 0 else 0), 2)
        ss_comp = round(ss_base * (ss_company_rate / 100.0), 2)
        hf_pers = round(hf_base * (hf_person_rate / 100.0), 2)
        hf_comp = round(hf_base * (hf_company_rate / 100.0), 2)
        deduction_cur = round(ss_pers + hf_pers, 2)

        pension_person = round(ss_base * pension_person_rate / 100.0, 2) if ss_base > 0 else 0.0
        medical_person = round(ss_base * medical_person_rate / 100.0, 2) if ss_base > 0 else 0.0
        unemployment_person = round(ss_base * unemployment_person_rate / 100.0, 2) if ss_base > 0 else 0.0
        large_medical_person = round(ss_pers - pension_person - medical_person - unemployment_person, 2) if ss_pers > 0 else 0.0

        # 7 项专项附加扣除，完整与 VBA 对齐。
        add_child = flt(emp.deduction_child_education)
        add_edu = flt(emp.deduction_continuing_education)
        add_med = flt(emp.deduction_serious_illness)
        add_loan = flt(emp.deduction_housing_loan)
        add_rent = flt(emp.deduction_housing_rent)
        add_elder = flt(emp.deduction_elderly_care)
        add_baby = flt(emp.deduction_infant_care)
        spec_add_cur = round(add_child + add_edu + add_med + add_loan + add_rent + add_elder + add_baby, 2)

        gross = target_or_gross
        tax = 0.0
        taxable = 0.0
        net = target_or_gross

        if not tax_eligible:
            # 临时工/零工：不进入个税表，也不产生个税；薪酬仍按税前/税后模式正常保留。
            if is_tax_after_salary_mode(salary_mode):
                net = target_or_gross
                gross = round(net + deduction_cur, 2)
            else:
                gross = target_or_gross
                net = round(gross - deduction_cur, 2)
            taxable = 0.0
            tax = 0.0
        else:
            pdata = get_employee_prior_tax_data(company, emp_no, cinfo["prior_months"], tax_threshold)
            if is_tax_after_salary_mode(salary_mode):
                calc = derive_gross_from_net_vba(
                    net_salary=target_or_gross,
                    deduction_cur=deduction_cur,
                    gross_prior=pdata["gross_prior"],
                    threshold_cur=tax_threshold,
                    threshold_prior=pdata["threshold_prior"],
                    spec_ded_cur=deduction_cur,
                    spec_ded_prior=pdata["spec_ded_prior"],
                    spec_add_cur=spec_add_cur,
                    spec_add_prior=pdata["spec_add_prior"],
                    paid_tax_prior=pdata["paid_tax_prior"],
                )
                gross = calc["gross_salary"]
                taxable = calc["taxable_income"]
                tax = calc["tax_amount_cur"]
                net = calc["net_verified"]
            else:
                gross = target_or_gross
                gross_all = round(pdata["gross_prior"] + gross, 2)
                thresh_all = round(pdata["threshold_prior"] + tax_threshold, 2)
                spec_ded_all = round(pdata["spec_ded_prior"] + deduction_cur, 2)
                spec_add_all = round(pdata["spec_add_prior"] + spec_add_cur, 2)
                taxable_raw = round(gross_all - thresh_all - spec_ded_all - spec_add_all, 2)
                taxable_for_tax = max(0.0, taxable_raw)
                rate, quick = 0.03, 0.0
                for _lower, upper, r_val, q_val in TAX_BRACKETS:
                    if taxable_for_tax <= upper:
                        rate, quick = r_val, q_val
                        break
                cum_tax = round(taxable_for_tax * rate - quick, 2)
                tax = max(0.0, round(cum_tax - pdata["paid_tax_prior"], 2))
                taxable = max(0.0, taxable_raw)
                net = round(gross - deduction_cur - tax, 2)

        total_gross += gross
        total_net += net
        total_ss_comp += ss_comp
        total_ss_pers += ss_pers
        total_hf_comp += hf_comp
        total_hf_pers += hf_pers
        total_tax += tax

        doc.append("items", {
            "employee_no": emp_no,
            "employee_name": emp.employee_name,
            "id_card": emp.id_card or "",
            "gender": emp.gender or "",
            "mobile": emp.mobile or "",
            "birth_date": emp.birth_date,
            "department": emp.department or ("生产部" if "祺富" in company else "技术部"),
            "job_title": emp.job_title or ("操作工" if "祺富" in company else "工程师"),
            "employee_type": emp_type,
            "salary_mode": salary_mode,
            "attendance_days": att_days,
            "work_hours": work_hrs,
            "overtime_hours": ot_hrs,
            "meal_count": meal_cnt,
            "fixed_salary": fixed_sal,
            "base_salary": base_sal,
            "post_allowance": post_allow,
            "performance_salary": perf_base,
            "allowances_total": allow_tot,
            "gross_salary": round(gross, 2),
            "ss_base": ss_base,
            "ss_person_total": ss_pers,
            "ss_company_total": ss_comp,
            "hf_base": hf_base,
            "hf_person_total": hf_pers,
            "hf_company_total": hf_comp,
            "pension_person": pension_person,
            "medical_person": medical_person,
            "large_medical_person": large_medical_person,
            "unemployment_person": unemployment_person,
            "housing_fund_person": hf_pers,
            "tax_threshold": tax_threshold if tax_eligible else 0.0,
            "deduction_child_education": add_child if tax_eligible else 0.0,
            "deduction_continuing_education": add_edu if tax_eligible else 0.0,
            "deduction_serious_illness": add_med if tax_eligible else 0.0,
            "deduction_housing_loan": add_loan if tax_eligible else 0.0,
            "deduction_housing_rent": add_rent if tax_eligible else 0.0,
            "deduction_elderly_care": add_elder if tax_eligible else 0.0,
            "deduction_infant_care": add_baby if tax_eligible else 0.0,
            "special_deductions_total": spec_add_cur if tax_eligible else 0.0,
            "taxable_income": round(taxable, 2),
            "tax_amount": round(tax, 2),
            "net_salary": round(net, 2),
            "remarks": "临时工/零工不进入个税申报台账" if not tax_eligible else "",
        })

    doc.total_employees = len(employees)
    doc.total_gross_salary = round(total_gross, 2)
    doc.total_net_salary = round(total_net, 2)
    doc.total_social_security_company = round(total_ss_comp, 2)
    doc.total_social_security_person = round(total_ss_pers, 2)
    doc.total_housing_fund_company = round(total_hf_comp, 2)
    doc.total_housing_fund_person = round(total_hf_pers, 2)
    doc.total_tax = round(total_tax, 2)

    frappe.flags.ignore_lock = True
    doc.save(ignore_permissions=True)
    frappe.db.commit()

    return {
        "success": True,
        "message": f"【{company}】{period_month} 月度薪酬已按 VBA 同口径累计预扣逻辑融合核算，共 {len(employees)} 人；个税台账参与 {tax_participant_count} 人。",
        "doc": get_payroll_settlement_detail(company, period_month),
    }

@frappe.whitelist(methods=["POST"])
def confirm_and_lock_payroll(company, period_month):
	check_payroll_workbench_permission("write")
	doc_name = f"{company}-{period_month}"
	if not frappe.db.exists("Ashan Monthly Payroll Settlement", doc_name):
		frappe.throw(f"请先生成【{company}】{period_month} 的月度薪酬测算表！")

	doc = frappe.get_doc("Ashan Monthly Payroll Settlement", doc_name)
	doc.status = "已核定锁定"
	doc.locked = 1
	doc.confirmed_by = frappe.session.user or "系统管理员"
	doc.confirmed_date = now_datetime()
	
	frappe.flags.ignore_lock = True
	doc.save(ignore_permissions=True)
	frappe.db.commit()

	return {
		"success": True,
		"message": f"🎉【{company}】{period_month} 月度薪酬已成功核定并锁定！参数已冻结，防止误修改。",
		"doc": get_payroll_settlement_detail(company, period_month)
	}

@frappe.whitelist(methods=["POST"])
def unlock_payroll(company, period_month, reason=""):
	check_payroll_workbench_permission("write")
	doc_name = f"{company}-{period_month}"
	if not frappe.db.exists("Ashan Monthly Payroll Settlement", doc_name):
		frappe.throw("单据不存在！")

	doc = frappe.get_doc("Ashan Monthly Payroll Settlement", doc_name)
	doc.status = "草稿"
	doc.locked = 0
	doc.unlock_reason = f"[{str(now_datetime())[:19]}] 由 {frappe.session.user} 反审核解锁。原因: {reason or '重新测算核对'}"

	frappe.flags.ignore_lock = True
	doc.save(ignore_permissions=True)
	frappe.db.commit()

	return {
		"success": True,
		"message": f"🔓【{company}】{period_month} 月度薪酬表已解锁为草稿状态，允许重新调整与测算！",
		"doc": get_payroll_settlement_detail(company, period_month)
	}


# 7 级综合所得年度税率表 (严格还原中国税法与 VBA 个人所得税_税率表)
TAX_BRACKETS = [
    (0.0, 36000.0, 0.03, 0.0),
    (36000.0, 144000.0, 0.10, 2520.0),
    (144000.0, 300000.0, 0.20, 16920.0),
    (300000.0, 420000.0, 0.25, 31920.0),
    (420000.0, 660000.0, 0.30, 52920.0),
    (660000.0, 960000.0, 0.35, 85920.0),
    (960000.0, 999999999.0, 0.45, 181920.0)
]


def _period_year(period_month):
    try:
        return int(str(period_month).split("-")[0])
    except Exception:
        return 2026


def get_effective_tax_parameters(company, period_month):
    """统一读取当前账期使用的个税基本参数，避免计算、台账、历史视图各自取值。"""
    year = _period_year(period_month)
    setting = frappe.db.get_value(
        "Ashan Insurance Setting",
        {"company": company, "effective_year": year},
        ["tax_threshold", "tax_cycle_start_month"],
        as_dict=True,
    ) or {}
    if not setting:
        # 兼容旧数据：按公司取一条，但不覆盖已有历史记录
        setting = frappe.db.get_value(
            "Ashan Insurance Setting",
            {"company": company},
            ["tax_threshold", "tax_cycle_start_month"],
            as_dict=True,
        ) or {}
    threshold = flt(setting.get("tax_threshold")) or 5000.0
    cycle_start_month = cint(setting.get("tax_cycle_start_month")) or 12
    if cycle_start_month < 1 or cycle_start_month > 12:
        cycle_start_month = 12
    return {
        "tax_threshold": threshold,
        "tax_cycle_start_month": cycle_start_month,
        "effective_year": year,
    }


def _add_month(year, month, offset):
    total = year * 12 + (month - 1) + offset
    return total // 12, total % 12 + 1


def get_tax_cycle_info(period_month, tax_threshold=5000.0, cycle_start_month=12):
    """
    统一计算个税累计申报周期。默认与现有 VBA 一致：12 月开始、次年 11 月结束。
    例如 2026-07 -> 202512-202611，往期为 2025-12 至 2026-06。
    """
    parts = str(period_month).split("-")
    year = int(parts[0])
    month = int(parts[1])
    cycle_start_month = cint(cycle_start_month) or 12
    if cycle_start_month < 1 or cycle_start_month > 12:
        cycle_start_month = 12
    tax_threshold = flt(tax_threshold) or 5000.0

    cycle_start_year = year if month >= cycle_start_month else year - 1
    prior_months = []
    offset = 0
    while True:
        y, m = _add_month(cycle_start_year, cycle_start_month, offset)
        if y == year and m == month:
            break
        prior_months.append(f"{y:04d}-{m:02d}")
        offset += 1
        if offset > 11:
            break

    end_y, end_m = _add_month(cycle_start_year, cycle_start_month, 11)
    cycle_name = f"{cycle_start_year:04d}{cycle_start_month:02d}-{end_y:04d}{end_m:02d}"
    cur_month_index = len(prior_months) + 1
    return {
        "cycle_name": cycle_name,
        "cycle_start_year": cycle_start_year,
        "cycle_start_month": cycle_start_month,
        "cycle_end_year": end_y,
        "cycle_end_month": end_m,
        "prior_months": prior_months,
        "cur_month_index": cur_month_index,
        "cum_threshold": cur_month_index * tax_threshold,
        "prior_threshold": len(prior_months) * tax_threshold,
        "cur_threshold": tax_threshold,
    }


def get_employee_prior_tax_data(company, employee_no, prior_months, tax_threshold=5000.0):
    """
    从历史薪酬明细汇总员工在本申报周期内的真实往期累计。
    起征点按“实际存在的历史月份记录”累计，而不是按周期月份数机械乘 5000，
    以匹配 VBA 对新入职/中途纳入人员的处理。
    """
    empty = {
        "gross_prior": 0.0,
        "threshold_prior": 0.0,
        "spec_ded_prior": 0.0,
        "spec_add_prior": 0.0,
        "paid_tax_prior": 0.0,
        "pension_prior": 0.0,
        "medical_prior": 0.0,
        "large_medical_prior": 0.0,
        "unemployment_prior": 0.0,
        "housing_fund_prior": 0.0,
        "add_child_prior": 0.0,
        "add_edu_prior": 0.0,
        "add_med_prior": 0.0,
        "add_loan_prior": 0.0,
        "add_rent_prior": 0.0,
        "add_elder_prior": 0.0,
        "add_baby_prior": 0.0,
        "record_month_count": 0,
    }
    if not prior_months:
        return empty

    parent_names = [f"{company}-{m}" for m in prior_months]
    fields = [
        "parent", "gross_salary", "tax_threshold",
        "ss_person_total", "hf_person_total",
        "pension_person", "medical_person", "large_medical_person", "unemployment_person", "housing_fund_person",
        "special_deductions_total", "deduction_child_education", "deduction_continuing_education",
        "deduction_serious_illness", "deduction_housing_loan", "deduction_housing_rent",
        "deduction_elderly_care", "deduction_infant_care", "tax_amount",
    ]
    items = frappe.get_all(
        "Ashan Monthly Payroll Item",
        filters={"employee_no": employee_no, "parent": ["in", parent_names]},
        fields=fields,
        order_by="parent asc",
    )
    if not items:
        return empty

    out = dict(empty)
    out["record_month_count"] = len(items)
    for it in items:
        out["gross_prior"] += flt(it.gross_salary)
        out["threshold_prior"] += flt(it.tax_threshold) or flt(tax_threshold) or 5000.0
        out["spec_ded_prior"] += flt(it.ss_person_total) + flt(it.hf_person_total)
        out["spec_add_prior"] += flt(it.special_deductions_total)
        out["paid_tax_prior"] += flt(it.tax_amount)
        out["pension_prior"] += flt(it.pension_person)
        out["medical_prior"] += flt(it.medical_person)
        out["large_medical_prior"] += flt(it.large_medical_person)
        out["unemployment_prior"] += flt(it.unemployment_person)
        out["housing_fund_prior"] += flt(it.housing_fund_person) or flt(it.hf_person_total)
        out["add_child_prior"] += flt(it.deduction_child_education)
        out["add_edu_prior"] += flt(it.deduction_continuing_education)
        out["add_med_prior"] += flt(it.deduction_serious_illness)
        out["add_loan_prior"] += flt(it.deduction_housing_loan)
        out["add_rent_prior"] += flt(it.deduction_housing_rent)
        out["add_elder_prior"] += flt(it.deduction_elderly_care)
        out["add_baby_prior"] += flt(it.deduction_infant_care)

    for key in out:
        if key != "record_month_count":
            out[key] = round(flt(out[key]), 2)
    return out


def derive_gross_from_net_vba(net_salary, deduction_cur, gross_prior,
                              threshold_cur, threshold_prior,
                              spec_ded_cur, spec_ded_prior,
                              spec_add_cur, spec_add_prior,
                              paid_tax_prior):
	"""
	闭式反推税前工资 (1:1 严格还原 VBA 个税_反推_税后反推税前工资)
	"""
	eps = 0.005
	d_all = (threshold_cur + threshold_prior) + (spec_ded_cur + spec_ded_prior) + (spec_add_cur + spec_add_prior)

	# Step A: 零扣缴判定
	x0 = net_salary + deduction_cur
	t0 = gross_prior + x0 - d_all

	if t0 <= 0.0 + eps:
		r0, q0 = 0.03, 0.0
	else:
		r0, q0 = 0.45, 181920.0
		for l_val, u_val, r_val, q_val in TAX_BRACKETS:
			if t0 >= l_val and (t0 < u_val or u_val > 90000000.0):
				r0, q0 = r_val, q_val
				break

	cum_tax0 = round(r0 * t0 - q0, 2)
	cur_tax0 = round(cum_tax0 - paid_tax_prior, 2)

	if cur_tax0 <= 0.0 + eps:
		# 命中零扣缴
		gross_res = round(x0, 2)
		return {
			"gross_salary": gross_res,
			"taxable_income": round(t0, 2),
			"tax_rate": r0 * 100,
			"quick_deduction": q0,
			"cum_tax_amount": cum_tax0,
			"tax_amount_cur": 0.0,
			"net_verified": round(gross_res - deduction_cur, 2)
		}

	# Step B: 需扣税判定 (逐档闭式反推)
	for l_val, u_val, r_val, q_val in TAX_BRACKETS:
		if r_val >= 1.0 - 1e-9: continue
		x = (net_salary + deduction_cur + r_val * gross_prior - r_val * d_all - q_val - paid_tax_prior) / (1.0 - r_val)
		t = gross_prior + x - d_all
		if t < l_val - eps or t > u_val + eps:
			continue

		cum_tax = round(r_val * t - q_val, 2)
		cur_tax = round(cum_tax - paid_tax_prior, 2)
		if cur_tax < 0:
			continue

		gross_res = round(x, 2)
		cur_tax_res = max(0.0, cur_tax)
		net_ver = round(gross_res - deduction_cur - cur_tax_res, 2)

		if abs(net_ver - round(net_salary, 2)) < 0.02:
			return {
				"gross_salary": gross_res,
				"taxable_income": round(t, 2),
				"tax_rate": r_val * 100,
				"quick_deduction": q_val,
				"cum_tax_amount": cum_tax,
				"tax_amount_cur": cur_tax_res,
				"net_verified": net_ver
			}

	# 保底
	return {
		"gross_salary": round(x0, 2),
		"taxable_income": round(t0, 2),
		"tax_rate": 3.0,
		"quick_deduction": 0.0,
		"cum_tax_amount": 0.0,
		"tax_amount_cur": 0.0,
		"net_verified": round(x0 - deduction_cur, 2)
	}

def derive_gross_from_net(net_salary, ss_person, hf_person, spec_deduction, tax_threshold=5000.0,
                          company="天津祺富机械加工有限公司", employee_no="", period_month="2026-07"):
    """按与 VBA 相同的累计预扣口径，根据税后实发反推税前工资。"""
    if net_salary <= 0:
        return 0.0, 0.0

    params = get_effective_tax_parameters(company, period_month)
    effective_threshold = flt(tax_threshold) or params["tax_threshold"]
    cinfo = get_tax_cycle_info(period_month, effective_threshold, params["tax_cycle_start_month"])
    pdata = get_employee_prior_tax_data(company, employee_no, cinfo["prior_months"], effective_threshold)
    ded_cur = flt(ss_person) + flt(hf_person)

    res = derive_gross_from_net_vba(
        net_salary=flt(net_salary),
        deduction_cur=ded_cur,
        gross_prior=pdata["gross_prior"],
        threshold_cur=effective_threshold,
        threshold_prior=pdata["threshold_prior"],
        spec_ded_cur=ded_cur,
        spec_ded_prior=pdata["spec_ded_prior"],
        spec_add_cur=flt(spec_deduction),
        spec_add_prior=pdata["spec_add_prior"],
        paid_tax_prior=pdata["paid_tax_prior"],
    )
    return res["gross_salary"], res["tax_amount_cur"]


def extract_year_month_from_text(text):
	"""
	从任意文本（工作表名/单元格标题/文件名）中高兼容性智能识别年月：
	支持：202607 / 202606 / 2026年7月 / 2026年07月 / 2026-07 / 2026.07 / 26.7 / 26年7月 / 2026/07 等
	"""
	import re
	if not text:
		return None
	t = str(text).strip().replace(" ", "")

	# 1. 202607 或 202606 纯6位连续数字 (20YYMM)
	m1 = re.search(r'(20\d{2})(0[1-9]|1[0-2])', t)
	if m1:
		y = int(m1.group(1))
		m = int(m1.group(2))
		return f"{y}-{m:02d}"

	# 2. 中文年月: 2026年7月, 2026年07月, 26年7月, 2026年7
	m2 = re.search(r'(?:20)?(\d{2})年(0?[1-9]|1[0-2])月?', t)
	if m2:
		yy = int(m2.group(1))
		y = 2000 + yy if yy < 100 else yy
		m = int(m2.group(2))
		return f"{y}-{m:02d}"

	# 3. 常见分隔符: 2026-07, 2026.07, 2026/7, 26.7, 2026_07, 26-7
	m3 = re.search(r'(?:20)?(\d{2})[\.\-\/_](0?[1-9]|1[0-2])(?!\d)', t)
	if m3:
		yy = int(m3.group(1))
		y = 2000 + yy if yy < 100 else yy
		m = int(m3.group(2))
		return f"{y}-{m:02d}"

	return None

def detect_payroll_period_month(wb, filename=""):
	"""
	三重智能高兼容性探测与校准工资账期月份：
	1. 工作表名 (Sheet Name，如 26.7 / 202607 / 2026年7月 -> 2026-07)
	2. 工作表前5行单元格内容 (如 2026年7月 工资明细 -> 2026-07)
	3. 文件名 (如 祺富工资2026.7.xlsx / 祺富202607.xlsm -> 2026-07)
	"""
	# 1. 优先扫描工作表名称
	for sname in wb.sheetnames:
		res = extract_year_month_from_text(sname)
		if res:
			return res, f"工作表名 ({sname})"

	# 2. 扫描首行/前5行单元格内容
	ws = wb.active
	for r in range(1, min(6, ws.max_row + 1)):
		for c in range(1, min(12, ws.max_column + 1)):
			val = str(ws.cell(r, c).value or "").strip()
			if not val: continue
			res = extract_year_month_from_text(val)
			if res:
				return res, f"表格标题内容 (R{r}C{c}: {val})"

	# 3. 扫描文件名
	if filename:
		res = extract_year_month_from_text(filename)
		if res:
			return res, f"文件名 ({filename})"

	return None, "未识别"

def get_next_month_str(period_month):
	"""计算下一个连续自然月 (YYYY-MM)"""
	if not period_month or "-" not in period_month:
		return "2026-07"
	parts = period_month.split("-")
	y = int(parts[0])
	m = int(parts[1])
	if m == 12:
		return f"{y+1}-01"
	else:
		return f"{y}-{m+1:02d}"

def get_months_between(start_month, end_month):
	"""获取介于 start_month 与 end_month 之间跳过的所有自然月列表 (不包含两端)"""
	res = []
	curr = get_next_month_str(start_month)
	while curr < end_month:
		res.append(curr)
		curr = get_next_month_str(curr)
	return res

@frappe.whitelist()
def get_payroll_periods_summary(company="天津祺富机械加工有限公司"):
	"""
	获取企业已核定/已创建的账期汇总列表，并计算当前最新账期与下一连续应导入账期
	"""
	check_payroll_workbench_permission("read")
	records = frappe.get_all(
		"Ashan Monthly Payroll Settlement",
		filters={"company": company},
		fields=["period_month", "total_employees", "total_net_salary", "total_gross_salary", "status"],
		order_by="period_month asc"
	)
	periods = [r.period_month for r in records if r.period_month]
	latest_period = periods[-1] if periods else "2026-06"
	expected_next_period = get_next_month_str(latest_period)

	return {
		"company": company,
		"periods": periods,
		"records": records,
		"latest_period": latest_period,
		"expected_next_period": expected_next_period
	}

@frappe.whitelist(methods=["POST"])
def create_blank_payroll_period(company, period_month):
	"""
	为指定月份创建【空白/零工资核定账期】（用于停产月或跳过月份的自动补齐，确保账期连续）
	在保人员正常代扣社保公积金并生成企业统筹，实发/应发为0
	"""
	check_payroll_workbench_permission("write")
	doc_name = f"{company}-{period_month}"
	year = period_month.split("-")[0] if "-" in period_month else "2026"
	setting_name = f"{company}-{year}"

	ss_comp_rate = 27.55
	ss_pers_rate = 10.50
	ss_person_pension_rate = 8.0
	ss_person_unemployment_rate = 0.5
	ss_person_medical_rate = 2.0
	ss_company_pension_rate = 16.0
	ss_company_unemployment_rate = 0.5
	ss_company_medical_rate = 10.0
	ss_company_other_medical_rate = 0.5
	ss_company_injury_rate = 0.55
	hf_comp_rate = 5.0
	hf_pers_rate = 5.0
	cur_m = cint(period_month.split("-")[1]) if "-" in period_month else 6
	big_med_amount = 22.0

	if frappe.db.exists("Ashan Insurance Setting", setting_name):
		ins = frappe.get_doc("Ashan Insurance Setting", setting_name)
		ss_company_pension_rate = flt(ins.ss_company_pension)
		ss_company_unemployment_rate = flt(ins.ss_company_unemployment)
		ss_company_medical_rate = flt(ins.ss_company_medical)
		ss_company_other_medical_rate = flt(ins.ss_company_other_medical)
		ss_company_injury_rate = flt(ins.ss_company_injury)
		ss_person_pension_rate = flt(ins.ss_person_pension)
		ss_person_unemployment_rate = flt(ins.ss_person_unemployment)
		ss_person_medical_rate = flt(ins.ss_person_medical)
		ss_comp_rate = ss_company_pension_rate + ss_company_unemployment_rate + ss_company_medical_rate + ss_company_other_medical_rate + ss_company_injury_rate
		ss_pers_rate = ss_person_pension_rate + ss_person_unemployment_rate + ss_person_medical_rate
		hf_comp_rate = flt(ins.hf_company_rate) or 5.0
		hf_pers_rate = flt(ins.hf_person_rate) or 5.0
		spec_m_str = str(ins.get("big_medical_special_months") or "3,12")
		spec_months = [cint(x.strip()) for x in spec_m_str.split(",") if x.strip().isdigit()]
		if cur_m in spec_months:
			big_med_amount = flt(ins.get("big_medical_amount_special")) or 21.0
		else:
			big_med_amount = flt(ins.get("big_medical_amount_default")) or 22.0

	# 按账期提取实际在册员工；历史月份不能被当前“离职”状态反向抹除。
	employees = _salary_profiles_for_period(
		company,
		period_month,
		fields=["employee_no", "employee_name", "department", "job_title", "employee_type", "social_security_base", "housing_fund_base"],
	)

	items_data = []
	total_ss_comp = 0.0
	total_ss_pers = 0.0
	total_hf_comp = 0.0
	total_hf_pers = 0.0

	for emp in employees:
		ss_base = flt(emp.social_security_base)
		hf_base = flt(emp.housing_fund_base)

		ss_p = round(ss_base * (ss_pers_rate / 100.0) + (big_med_amount if ss_base > 0 else 0), 2)
		ss_c = round(ss_base * (ss_comp_rate / 100.0), 2)
		hf_p = round(hf_base * (hf_pers_rate / 100.0), 2)
		hf_c = round(hf_base * (hf_comp_rate / 100.0), 2)
		pension_p = round(ss_base * (ss_person_pension_rate / 100.0), 2)
		unemployment_p = round(ss_base * (ss_person_unemployment_rate / 100.0), 2)
		medical_p = round(ss_base * (ss_person_medical_rate / 100.0), 2)
		large_med_p = round(big_med_amount if ss_base > 0 else 0.0, 2)
		pension_c = round(ss_base * (ss_company_pension_rate / 100.0), 2)
		unemployment_c = round(ss_base * (ss_company_unemployment_rate / 100.0), 2)
		medical_c = round(ss_base * (ss_company_medical_rate / 100.0), 2)
		other_medical_c = round(ss_base * (ss_company_other_medical_rate / 100.0), 2)
		injury_c = round(ss_base * (ss_company_injury_rate / 100.0), 2)

		total_ss_comp += ss_c
		total_ss_pers += ss_p
		total_hf_comp += hf_c
		total_hf_pers += hf_p

		items_data.append({
			"employee_no": emp.employee_no,
			"employee_name": emp.employee_name,
			"department": emp.department or "生产部",
			"job_title": emp.job_title or "操作工",
			"employee_type": emp.employee_type or "正式工",
			"salary_mode": "空白账期",
			"attendance_days": 0,
			"work_hours": 0,
			"fixed_salary": 0.0,
			"gross_salary": 0.0,
			"ss_base": ss_base,
			"hf_base": hf_base,
			"ss_person_total": ss_p,
			"ss_company_total": ss_c,
			"hf_person_total": hf_p,
			"hf_company_total": hf_c,
			"taxable_income": 0.0,
			"tax_amount": 0.0,
			"net_salary": 0.0,
			"remarks": ""
		})

	if frappe.db.exists("Ashan Monthly Payroll Settlement", doc_name):
		settle_doc = frappe.get_doc("Ashan Monthly Payroll Settlement", doc_name)
		settle_doc.items = []
	else:
		settle_doc = frappe.new_doc("Ashan Monthly Payroll Settlement")
		settle_doc.company = company
		settle_doc.period_month = period_month

	settle_doc.status = "草稿"
	settle_doc.locked = 0
	settle_doc.total_employees = len(items_data)
	settle_doc.total_gross_salary = 0.0
	settle_doc.total_net_salary = 0.0
	settle_doc.total_social_security_company = round(total_ss_comp, 2)
	settle_doc.total_social_security_person = round(total_ss_pers, 2)
	settle_doc.total_housing_fund_company = round(total_hf_comp, 2)
	settle_doc.total_housing_fund_person = round(total_hf_pers, 2)
	settle_doc.total_tax = 0.0

	# 保存原始车间实发表附件
	if file_bytes:
		save_excel_name = f"{period_month}_{company}_外部车间实发表.xlsx"
		_file = frappe.get_doc({
			"doctype": "File",
			"file_name": save_excel_name,
			"attached_to_doctype": "Ashan Monthly Payroll Settlement",
			"attached_to_name": doc_name,
			"content": file_bytes,
			"is_private": 1
		})
		_file.save(ignore_permissions=True)
		settle_doc.imported_excel_file = _file.file_url

	for item in items_data:
		settle_doc.append("items", item)

	frappe.flags.ignore_lock = True
	settle_doc.save(ignore_permissions=True)
	from ashan_cn_procurement.services.payroll_recalculation_service import queue_recalculation_after_change
	queue_recalculation_after_change(
		company=company,
		period_month=period_month,
		employee_no=None,
		trigger_source="外部实发与发放表",
		trigger_detail="导入/覆盖外部实发与发放表后统一服务器复核",
	)
	frappe.db.commit()

	return {
		"success": True,
		"period_month": period_month,
		"message": f"✅ 已成功为【{company}】创建 {period_month} 空白零工资核定账期！共登记 {len(items_data)} 人。"
	}

@frappe.whitelist()
def detect_qifu_excel_info(file_data=None, file_url=None, server_file_path=None, filename=None):
	"""
	预检上传的 Excel 文件，返回智能识别的核定月份与连续性校验状态
	"""
	check_payroll_workbench_permission("read")
	import io
	import openpyxl
	import base64
	import os

	company = "天津祺富机械加工有限公司"

	wb = None
	if server_file_path and os.path.exists(server_file_path):
		wb = openpyxl.load_workbook(server_file_path, data_only=True)
		filename = filename or os.path.basename(server_file_path)
	elif file_data:
		if "," in file_data:
			file_data = file_data.split(",")[1]
		file_bytes = base64.b64decode(file_data)
		wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
	elif file_url:
		file_doc = frappe.get_doc("File", {"file_url": file_url})
		file_path = file_doc.get_full_path()
		wb = openpyxl.load_workbook(file_path, data_only=True)
		filename = filename or file_doc.file_name

	if not wb:
		return {"success": False, "message": "无法加载工作簿"}

	detected_month, source = detect_payroll_period_month(wb, filename or "")
	ws = wb.active

	# 连续性校验
	summary = get_payroll_periods_summary(company)
	latest_p = summary.get("latest_period") or "2026-06"
	expected_next_p = summary.get("expected_next_period") or "2026-07"
	target_m = detected_month or expected_next_p

	rule_status = "valid_next"
	rule_msg = f"符合连续账期推进 (最新已核定: {latest_p} -> 本次导入: {target_m})"
	skipped_months = []

	if target_m == expected_next_p:
		rule_status = "valid_next"
		rule_msg = f"✅ 符合连续账期推进标准！(当前最大账期: {latest_p} ➡️ 本次导入: {target_m})"
	elif target_m == latest_p:
		rule_status = "overwrite_latest"
		rule_msg = f"🔄 重新上传并覆盖当前最新账期 ({target_m})"
	elif target_m < latest_p:
		rule_status = "overwrite_history"
		rule_msg = f"⚠️ 正在覆盖历史已归档账期 ({target_m})，请谨慎操作"
	else:
		# target_m > expected_next_p
		rule_status = "skipped_months"
		skipped_months = get_months_between(latest_p, target_m)
		rule_msg = f"⚠️ 检测到跨月跳跃！缺失中间账期: {', '.join(skipped_months)}。系统将自动先为缺失月份补齐【空白零工资账期】，保证账期连续！"

	return {
		"success": True,
		"detected_period_month": target_m,
		"detection_source": source,
		"sheet_names": wb.sheetnames,
		"max_rows": ws.max_row,
		"latest_period": latest_p,
		"expected_next_period": expected_next_p,
		"rule_status": rule_status,
		"rule_msg": rule_msg,
		"skipped_months": skipped_months
	}


@frappe.whitelist()
def preview_import_excel_data(file_base64=None, file_name=None, file_data=None, filename=None, company="天津祺富机械加工有限公司", period_month=None):
	"""
	预检并解析上传的 Excel 文件，返回智能识别的账期、工资人数、考勤工资合计、考勤加补贴合计以及前 10 行预览数据
	"""
	check_payroll_workbench_permission("read")
	import io
	import openpyxl
	import base64

	raw_data = file_base64 or file_data
	fname = file_name or filename or ""

	if not raw_data:
		frappe.throw("未提供有效的 Excel 文件数据！")

	if "," in raw_data:
		raw_data = raw_data.split(",")[1]

	try:
		file_bytes = base64.b64decode(raw_data)
		wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
	except Exception as e:
		err_str = str(e)
		if "not a zip file" in err_str.lower() or "invalidfile" in err_str.lower() or (fname and fname.lower().endswith(".xls")):
			frappe.throw(f"【❌ 文件格式不支持】系统仅支持标准 Office OpenXML 格式（.xlsx / .xlsm）。您上传的文件【{fname}】疑似为旧版二进制 .xls 或损坏文件，请使用 Excel 打开并【另存为 .xlsx 格式】后再上传！")
		else:
			frappe.throw(f"【❌ Excel 读取失败】解析文件【{fname}】时发生错误: {err_str}")

	ws = wb.active

	# 1. 智能探测与严格校准账期
	detected_month, source = detect_payroll_period_month(wb, fname)
	current_target_month = period_month or "2026-07"
	is_period_matched = True
	mismatch_message = ""

	if detected_month:
		if detected_month != current_target_month:
			is_period_matched = False
			mismatch_message = f"❌ 账期不匹配拦截！当前工作台发薪账期为【{current_target_month}】，但上传文件识别为【{detected_month}】（来源: {source}）。系统禁止将【{detected_month}】的实发表导入到【{current_target_month}】！请核对文件或在顶部切换发薪月份。"
	else:
		detected_month = current_target_month

	# 2. 表头定位；姓名不再硬编码个人别名，优先使用工号并要求姓名精确匹配。
	name_alias_map = {}

	header_row = -1
	col_map = {}
	for r in range(1, min(10, ws.max_row + 1)):
		row_texts = [str(ws.cell(r, c).value or "").strip().replace(" ", "") for c in range(1, ws.max_column + 1)]
		if any("姓名" in t for t in row_texts) and any("实发" in t or "工资" in t for t in row_texts):
			header_row = r
			for c_idx, text in enumerate(row_texts, start=1):
				if not text: continue
				if "姓名" in text: col_map["name"] = c_idx
				elif "工号" in text or "编号" in text: col_map["no"] = c_idx
				elif "实发" in text: col_map["net_salary"] = c_idx
				elif "作业天数" in text or "天数" in text: col_map["work_days"] = c_idx
				elif "作业小时" in text or "小时" in text: col_map["work_hours"] = c_idx
				elif "天工资" in text: col_map["day_salary"] = c_idx
				elif "小时工资" in text: col_map["hour_salary"] = c_idx
				elif "全勤" in text: col_map["full_attendance"] = c_idx
				elif "加班小时" in text: col_map["overtime_hours"] = c_idx
				elif "加班费" in text: col_map["overtime_salary"] = c_idx
				elif "国勤天数" in text: col_map["national_days"] = c_idx
				elif "国勤工资" in text: col_map["national_salary"] = c_idx
				elif "达标率" in text: col_map["target_rate"] = c_idx
				elif "达标工资" in text: col_map["target_salary"] = c_idx
				elif "扣除" in text: col_map["deduction"] = c_idx
				elif "职位补贴" in text: col_map["post_allowance"] = c_idx
				elif "房" in text or "车补" in text: col_map["house_allowance"] = c_idx
			break

	parsed_rows = []
	tot_workshop = 0.0
	tot_allowance = 0.0
	tot_payable = 0.0
	tot_net = 0.0

	if header_row != -1 and "name" in col_map:
		r = header_row + 1
		idx = 1
		while r <= ws.max_row:
			raw_name = str(ws.cell(r, col_map["name"]).value or "").strip()
			if not raw_name or raw_name in ["合计", "总计", "平均", "None"]:
				break
			std_name = name_alias_map.get(raw_name, raw_name)
			net_val = flt(ws.cell(r, col_map["net_salary"]).value) if "net_salary" in col_map else 0.0
			days = flt(ws.cell(r, col_map["work_days"]).value) if "work_days" in col_map else 0
			hours = flt(ws.cell(r, col_map["work_hours"]).value) if "work_hours" in col_map else 0
			day_sal = flt(ws.cell(r, col_map["day_salary"]).value) if "day_salary" in col_map else 0.0
			full_att = flt(ws.cell(r, col_map["full_attendance"]).value) if "full_attendance" in col_map else 0.0
			ot_sal = flt(ws.cell(r, col_map["overtime_salary"]).value) if "overtime_salary" in col_map else 0.0
			nat_sal = flt(ws.cell(r, col_map["national_salary"]).value) if "national_salary" in col_map else 0.0
			t_sal = flt(ws.cell(r, col_map["target_salary"]).value) if "target_salary" in col_map else 0.0
			ded_val = flt(ws.cell(r, col_map["deduction"]).value) if "deduction" in col_map else 0.0
			post_all = flt(ws.cell(r, col_map["post_allowance"]).value) if "post_allowance" in col_map else 0.0
			house_all = flt(ws.cell(r, col_map["house_allowance"]).value) if "house_allowance" in col_map else 0.0

			# 考勤绩效工资小计 = 天工资 + 小时工资 + 全勤 + 加班费 + 国勤 + 达标 - 扣除
			workshop_sub = day_sal + full_att + ot_sal + nat_sal + t_sal - ded_val
			if workshop_sub <= 0 and net_val > 0:
				workshop_sub = net_val

			allowance_sub = post_all + house_all
			payable_sub = workshop_sub + allowance_sub

			parsed_rows.append({
				"seq": idx,
				"employee_name": std_name,
				"work_days": days,
				"work_hours": hours,
				"workshop_subtotal": workshop_sub,
				"post_allowance": post_all,
				"house_rent_allowance": house_all,
				"allowance_subtotal": allowance_sub,
				"payable_salary": payable_sub,
				"net_salary": net_val
			})

			tot_workshop += workshop_sub
			tot_allowance += allowance_sub
			tot_payable += payable_sub
			tot_net += net_val
			idx += 1
			r += 1

	# 判断当前账期是否已经导入过
	doc_name = f"{company}-{detected_month}"
	is_already_imported = frappe.db.exists("Ashan Monthly Payroll Settlement", doc_name)

	return {
		"success": True,
		"detected_period_month": detected_month,
		"current_target_month": current_target_month,
		"detection_source": source,
		"is_period_matched": is_period_matched,
		"mismatch_message": mismatch_message,
		"is_already_imported": bool(is_already_imported),
		"employee_count": len(parsed_rows),
		"attendance_salary_total": round(tot_workshop, 2),
		"allowance_total": round(tot_allowance, 2),
		"attendance_allowance_total": round(tot_payable, 2),
		"net_salary_total": round(tot_net, 2),
		"preview_rows": parsed_rows[:15],
		"filename": fname,
		"message": f"成功解析 Excel: 识别账期【{detected_month}】，共 {len(parsed_rows)} 位员工发放记录！"
	}


@frappe.whitelist(methods=["POST"])
def import_and_calculate_payroll_excel(file_base64=None, file_name=None, file_data=None, filename=None, company="天津祺富机械加工有限公司", period_month=None):
	"""
	导入外部实发表并执行融合计算：
	如果当前账期已存在，自动清空旧数据并重新导入覆盖！
	"""
	check_payroll_workbench_permission("write")
	return upload_and_import_qifu_salary(
		file_data=file_base64 or file_data,
		filename=file_name or filename,
		period_month=period_month
	)


@frappe.whitelist(methods=["POST"])
def upload_and_import_qifu_salary(file_url=None, file_data=None, filename=None, period_month=None, server_file_path=None):
	"""
	上传/解析祺富外部实发工资表 Excel，智能核定月份与匹配人员信息，并执行【税后实发倒推税前应发与个税】
	"""
	check_payroll_workbench_permission("write")
	import io
	import openpyxl
	import base64
	import os

	company = "天津祺富机械加工有限公司"

	# 1. 加载工作簿并妥善保存原始凭证文件
	wb = None
	raw_file_bytes = None
	try:
		if server_file_path and os.path.exists(server_file_path):
			with open(server_file_path, "rb") as f:
				raw_file_bytes = f.read()
			wb = openpyxl.load_workbook(server_file_path, data_only=True)
			filename = filename or os.path.basename(server_file_path)
		elif file_data:
			# Base64 编码的 Excel
			if "," in file_data:
				file_data = file_data.split(",")[1]
			raw_file_bytes = base64.b64decode(file_data)
			wb = openpyxl.load_workbook(io.BytesIO(raw_file_bytes), data_only=True)
		elif file_url:
			file_doc = frappe.get_doc("File", {"file_url": file_url})
			file_path = file_doc.get_full_path()
			with open(file_path, "rb") as f:
				raw_file_bytes = f.read()
			wb = openpyxl.load_workbook(file_path, data_only=True)
			filename = filename or file_doc.file_name
		else:
			frappe.throw("未提供有效的 Excel 文件或路径！")
	except Exception as e:
		err_str = str(e)
		if "not a zip file" in err_str.lower() or "invalidfile" in err_str.lower() or (filename and filename.lower().endswith(".xls")):
			frappe.throw(f"【❌ 文件格式不支持】系统仅支持标准 Office OpenXML 格式（.xlsx / .xlsm）。您上传的文件【{filename or 'Excel'}】疑似为旧版二进制 .xls 或损坏文件，请使用 Excel 打开并【另存为 .xlsx 格式】后再上传！")
		else:
			frappe.throw(f"【❌ Excel 读取失败】解析文件【{filename or 'Excel'}】时发生错误: {err_str}")

	ws = wb.active

	# 2. 智能核定与严格校准工资账期月份
	detected_month, det_source = detect_payroll_period_month(wb, filename or "")
	if period_month and period_month != "auto" and detected_month and detected_month != period_month:
		frappe.throw(f"【❌ 账期校准拦截】当前发薪账期为【{period_month}】，但上传的 Excel 文件识别为【{detected_month}】（来源: {det_source}）。系统禁止将【{detected_month}】的外部工资表导入到【{period_month}】！请核对文件或在工作台顶部切换账期。")

	if not period_month or period_month == "auto":
		if detected_month:
			period_month = detected_month
		else:
			period_month = "2026-07"

	# 检查是否存在跳月：若存在中间未核定的月份，自动补齐空白零工资账期
	summary = get_payroll_periods_summary(company)
	latest_p = summary.get("latest_period") or "2026-06"
	expected_next_p = summary.get("expected_next_period") or "2026-07"
	if period_month > expected_next_p:
		skipped = get_months_between(latest_p, period_month)
		for sm in skipped:
			if not frappe.db.exists("Ashan Monthly Payroll Settlement", f"{company}-{sm}"):
				create_blank_payroll_period(company, sm)

	# 不在源码中硬编码具体员工别名。
	name_alias_map = {}

	# 3. 提取老板娘表中的【主表车间实发】
	header_row = -1
	col_map = {}
	for r in range(1, min(10, ws.max_row + 1)):
		row_texts = [str(ws.cell(r, c).value or "").strip().replace(" ", "") for c in range(1, ws.max_column + 1)]
		if any("姓名" in t for t in row_texts) and any("实发" in t or "工资" in t for t in row_texts):
			header_row = r
			for c_idx, text in enumerate(row_texts, start=1):
				if not text: continue
				if "姓名" in text: col_map["name"] = c_idx
				elif "工号" in text or "编号" in text: col_map["no"] = c_idx
				elif "实发" in text: col_map["net_salary"] = c_idx
				elif "作业天数" in text or "天数" in text: col_map["work_days"] = c_idx
				elif "作业小时" in text or "小时" in text: col_map["work_hours"] = c_idx
				elif "天工资" in text: col_map["day_salary"] = c_idx
				elif "小时工资" in text: col_map["hour_salary"] = c_idx
				elif "全勤" in text: col_map["full_attendance"] = c_idx
				elif "加班小时" in text: col_map["overtime_hours"] = c_idx
				elif "加班费" in text: col_map["overtime_salary"] = c_idx
				elif "国勤天数" in text: col_map["national_days"] = c_idx
				elif "国勤工资" in text: col_map["national_salary"] = c_idx
				elif "达标率" in text: col_map["target_rate"] = c_idx
				elif "达标工资" in text: col_map["target_salary"] = c_idx
				elif "扣除" in text: col_map["deduction"] = c_idx
				elif "是否社保" in text or "社保" in text: col_map["is_insured"] = c_idx
				elif "备考" in text or "备注" in text: col_map["remarks"] = c_idx
			break

	wife_payroll_dict = {} # std_name -> data

	if header_row != -1 and "name" in col_map and "net_salary" in col_map:
		r = header_row + 1
		while r <= ws.max_row:
			raw_name = str(ws.cell(r, col_map["name"]).value or "").strip()
			if not raw_name or raw_name in ["合计", "总计", "平均", "None"]:
				break
			std_name = name_alias_map.get(raw_name, raw_name)
			net_val = flt(ws.cell(r, col_map["net_salary"]).value)
			days = flt(ws.cell(r, col_map["work_days"]).value) if "work_days" in col_map else 0
			hours = flt(ws.cell(r, col_map["work_hours"]).value) if "work_hours" in col_map else 0
			day_sal = flt(ws.cell(r, col_map["day_salary"]).value) if "day_salary" in col_map else 0.0
			hour_sal = flt(ws.cell(r, col_map["hour_salary"]).value) if "hour_salary" in col_map else 0.0
			full_att = flt(ws.cell(r, col_map["full_attendance"]).value) if "full_attendance" in col_map else 0.0
			ot_hours = flt(ws.cell(r, col_map["overtime_hours"]).value) if "overtime_hours" in col_map else 0.0
			ot_sal = flt(ws.cell(r, col_map["overtime_salary"]).value) if "overtime_salary" in col_map else 0.0
			nat_days = flt(ws.cell(r, col_map["national_days"]).value) if "national_days" in col_map else 0.0
			nat_sal = flt(ws.cell(r, col_map["national_salary"]).value) if "national_salary" in col_map else 0.0
			t_rate = str(ws.cell(r, col_map["target_rate"]).value or "") if "target_rate" in col_map else ""
			t_sal = flt(ws.cell(r, col_map["target_salary"]).value) if "target_salary" in col_map else 0.0
			ded_val = flt(ws.cell(r, col_map["deduction"]).value) if "deduction" in col_map else 0.0
			is_ss = str(ws.cell(r, col_map["is_insured"]).value or "").strip() if "is_insured" in col_map else "是"
			rem_val = str(ws.cell(r, col_map["remarks"]).value or "").strip() if "remarks" in col_map else ""

			wife_payroll_dict[std_name] = {
				"raw_name": raw_name,
				"workshop_net": net_val,
				"work_days": days,
				"work_hours": hours,
				"day_salary": day_sal,
				"hour_salary": hour_sal,
				"full_attendance": full_att,
				"overtime_hours": ot_hours,
				"overtime_salary": ot_sal,
				"national_days": nat_days,
				"national_salary": nat_sal,
				"target_rate": t_rate,
				"target_salary": t_sal,
				"deduction": ded_val,
				"is_insured": is_ss,
				"remarks": rem_val,
				"post_allowance": 0.0,
				"house_car_allowance": 0.0
			}
			r += 1

	# 4. 扫描老板娘表底部的【职位补贴与房/车补】
	sub_header_row = -1
	sub_cols = {}
	for r in range((header_row if header_row != -1 else 1) + 5, ws.max_row + 1):
		row_texts = [str(ws.cell(r, c).value or "").strip().replace(" ", "") for c in range(1, ws.max_column + 1)]
		if any("职位补贴" in t for t in row_texts) or any("房/车补" in t for t in row_texts) or any("车补" in t for t in row_texts):
			sub_header_row = r
			for c_idx, text in enumerate(row_texts, start=1):
				if "姓名" in text: sub_cols["name"] = c_idx
				elif "职位补贴" in text or "职务补贴" in text: sub_cols["post_allowance"] = c_idx
				elif "房/车补" in text or "车补" in text or "房补" in text: sub_cols["house_car_allowance"] = c_idx
			break

	if sub_header_row != -1:
		for r in range(sub_header_row + 1, ws.max_row + 1):
			raw_name = str(ws.cell(r, sub_cols.get("name", 13)).value or "").strip()
			if not raw_name or raw_name in ["合计", "总计", "工资表", "None"]:
				continue
			std_name = name_alias_map.get(raw_name, raw_name)
			post_all = flt(ws.cell(r, sub_cols["post_allowance"]).value) if "post_allowance" in sub_cols else 0.0
			hc_all = flt(ws.cell(r, sub_cols["house_car_allowance"]).value) if "house_car_allowance" in sub_cols else 0.0

			if std_name not in wife_payroll_dict:
				wife_payroll_dict[std_name] = {
					"raw_name": raw_name,
					"workshop_net": 0.0,
					"work_days": 0.0,
					"work_hours": 0.0,
					"is_insured": "是",
					"post_allowance": post_all,
					"house_car_allowance": hc_all
				}
			else:
				wife_payroll_dict[std_name]["post_allowance"] = post_all
				wife_payroll_dict[std_name]["house_car_allowance"] = hc_all

	# 5. 加载社保公积金费率与规则
	year = period_month.split("-")[0] if "-" in period_month else "2026"
	setting_name = f"{company}-{year}"
	ss_comp_rate = 27.55
	ss_pers_rate = 10.50
	ss_person_pension_rate = 8.0
	ss_person_unemployment_rate = 0.5
	ss_person_medical_rate = 2.0
	ss_company_pension_rate = 16.0
	ss_company_unemployment_rate = 0.5
	ss_company_medical_rate = 10.0
	ss_company_other_medical_rate = 0.5
	ss_company_injury_rate = 0.55
	hf_comp_rate = 5.0
	hf_pers_rate = 5.0
	tax_thresh = 5000.0
	big_med_amount = 22.0
	cur_m = cint(period_month.split("-")[1]) if "-" in period_month else 6

	if frappe.db.exists("Ashan Insurance Setting", setting_name):
		ins = frappe.get_doc("Ashan Insurance Setting", setting_name)
		ss_company_pension_rate = flt(ins.ss_company_pension)
		ss_company_unemployment_rate = flt(ins.ss_company_unemployment)
		ss_company_medical_rate = flt(ins.ss_company_medical)
		ss_company_other_medical_rate = flt(ins.ss_company_other_medical)
		ss_company_injury_rate = flt(ins.ss_company_injury)
		ss_person_pension_rate = flt(ins.ss_person_pension)
		ss_person_unemployment_rate = flt(ins.ss_person_unemployment)
		ss_person_medical_rate = flt(ins.ss_person_medical)
		ss_comp_rate = ss_company_pension_rate + ss_company_unemployment_rate + ss_company_medical_rate + ss_company_other_medical_rate + ss_company_injury_rate
		ss_pers_rate = ss_person_pension_rate + ss_person_unemployment_rate + ss_person_medical_rate
		hf_comp_rate = flt(ins.hf_company_rate) or 5.0
		hf_pers_rate = flt(ins.hf_person_rate) or 5.0
		tax_thresh = flt(ins.tax_threshold) or 5000.0
		spec_m_str = str(ins.get("big_medical_special_months") or "3,12")
		spec_months = [cint(x.strip()) for x in spec_m_str.split(",") if x.strip().isdigit()]
		if cur_m in spec_months:
			big_med_amount = flt(ins.get("big_medical_amount_special")) or 21.0
		else:
			big_med_amount = flt(ins.get("big_medical_amount_default")) or 22.0

	# 6. 加载该账期实际在册母表档案 (Master Data)
	master_employees = _salary_profiles_for_period(
		company,
		period_month,
		fields=[
			"name", "employee_no", "employee_name", "department", "job_title",
			"employee_type", "salary_mode", "fixed_salary", "social_security_base",
			"housing_fund_base", "deduction_child_education", "deduction_continuing_education", "deduction_housing_loan",
			"deduction_housing_rent", "deduction_elderly_care", "deduction_infant_care",
			"deduction_serious_illness"
		],
		order_by="employee_no asc",
	)

	master_by_name = {emp.employee_name.strip(): emp for emp in master_employees}

	# 外部实发表是月度输入，不得反向猜测并创建权威员工母表。
	# 如发现未建档人员，先在 Tab 1 完成人员属性/计薪方式/参保信息，再重新导入。
	unknown_names = [name for name in wife_payroll_dict if name not in master_by_name]
	if unknown_names:
		preview = "、".join(unknown_names[:8])
		more = f" 等 {len(unknown_names)} 人" if len(unknown_names) > 8 else ""
		frappe.throw(
			f"外部实发表存在未录入【员工薪酬档案（母表底册）】的人员：{preview}{more}。"
			"为避免系统自动猜测员工类型、工资类型或社保公积金基数，已停止导入。请先在 Tab 1 建档后再试。"
		)

	# 7. 以母表为基准，融合老板娘表并全员执行【税后倒推税前】
	items_data = []
	total_net = 0.0
	total_gross = 0.0
	total_ss_comp = 0.0
	total_ss_pers = 0.0
	total_hf_comp = 0.0
	total_hf_pers = 0.0
	total_tax = 0.0

	for emp in master_employees:
		emp_name = emp.employee_name.strip()
		wdata = wife_payroll_dict.get(emp_name)

		if wdata:
			workshop_net = wdata["workshop_net"]
			post_allowance = wdata["post_allowance"]
			hc_allowance = wdata["house_car_allowance"]
			work_days = wdata["work_days"]
			work_hours = wdata["work_hours"]
			is_insured_flag = wdata["is_insured"]
		else:
			# 未在老板娘表中出现 (管理/未出勤人员)
			workshop_net = 0.0
			post_allowance = 0.0
			hc_allowance = 0.0
			work_days = 0.0
			work_hours = 0.0
			is_insured_flag = "是" if flt(emp.social_security_base) > 0 else "否"

		# 当月实发总额 = 车间实发 + 职位补贴 + 房车补
		net_salary = round(workshop_net + post_allowance + hc_allowance, 2)

		# 社保与公积金基数
		ss_base = flt(emp.social_security_base) if is_insured_flag != "否" else 0.0
		hf_base = flt(emp.housing_fund_base) if is_insured_flag != "否" else 0.0

		ss_p = round(ss_base * (ss_pers_rate / 100.0) + (big_med_amount if ss_base > 0 else 0), 2)
		ss_c = round(ss_base * (ss_comp_rate / 100.0), 2)
		hf_p = round(hf_base * (hf_pers_rate / 100.0), 2)
		hf_c = round(hf_base * (hf_comp_rate / 100.0), 2)
		pension_p = round(ss_base * (ss_person_pension_rate / 100.0), 2)
		unemployment_p = round(ss_base * (ss_person_unemployment_rate / 100.0), 2)
		medical_p = round(ss_base * (ss_person_medical_rate / 100.0), 2)
		large_med_p = round(big_med_amount if ss_base > 0 else 0.0, 2)
		pension_c = round(ss_base * (ss_company_pension_rate / 100.0), 2)
		unemployment_c = round(ss_base * (ss_company_unemployment_rate / 100.0), 2)
		medical_c = round(ss_base * (ss_company_medical_rate / 100.0), 2)
		other_medical_c = round(ss_base * (ss_company_other_medical_rate / 100.0), 2)
		injury_c = round(ss_base * (ss_company_injury_rate / 100.0), 2)

		# 7项专项附加扣除
		spec_d = (
			flt(emp.get("deduction_child_education")) +
			flt(emp.get("deduction_continuing_education")) +
			flt(emp.get("deduction_housing_loan")) +
			flt(emp.get("deduction_housing_rent")) +
			flt(emp.get("deduction_elderly_care")) +
			flt(emp.get("deduction_infant_care")) +
			flt(emp.get("deduction_serious_illness"))
		)

		# 核心：税后实发倒推税前应发与个税预扣！
		if net_salary > 0:
			gross_salary, tax_amount = derive_gross_from_net(net_salary=net_salary, ss_person=ss_p, hf_person=hf_p, spec_deduction=spec_d, tax_threshold=tax_thresh, company=company, employee_no=emp.employee_no, period_month=period_month)
		else:
			gross_salary = 0.0
			tax_amount = 0.0

		total_net += net_salary
		total_gross += gross_salary
		total_ss_comp += ss_c
		total_ss_pers += ss_p
		total_hf_comp += hf_c
		total_hf_pers += hf_p
		total_tax += tax_amount

		# 备注说明
		remark_parts = []
		if post_allowance > 0: remark_parts.append(f"职位补贴¥{post_allowance:,.0f}")
		if hc_allowance > 0: remark_parts.append(f"房车补¥{hc_allowance:,.0f}")
		if work_days > 0: remark_parts.append(f"出勤{work_days}天")
		if not wdata: remark_parts.append("")

		items_data.append({
			"employee_no": emp.employee_no,
			"employee_name": emp_name,
			"department": emp.department or "生产部",
			"job_title": emp.job_title or "操作工",
			"employee_type": emp.employee_type or "正式工",
			"salary_mode": "税后倒推",
			"attendance_days": work_days,
			"work_hours": work_hours,
			"day_salary": wdata.get("day_salary", 0.0) if wdata else 0.0,
			"hour_salary": wdata.get("hour_salary", 0.0) if wdata else 0.0,
			"full_attendance": wdata.get("full_attendance", 0.0) if wdata else 0.0,
			"overtime_hours": wdata.get("overtime_hours", 0.0) if wdata else 0.0,
			"overtime_salary": wdata.get("overtime_salary", 0.0) if wdata else 0.0,
			"national_days": wdata.get("national_days", 0.0) if wdata else 0.0,
			"national_salary": wdata.get("national_salary", 0.0) if wdata else 0.0,
			"target_rate": wdata.get("target_rate", "") if wdata else "",
			"target_salary": wdata.get("target_salary", 0.0) if wdata else 0.0,
			"deduction": wdata.get("deduction", 0.0) if wdata else 0.0,
			"fixed_salary": workshop_net,
			"post_allowance": post_allowance,
			"house_rent_allowance": hc_allowance,
			"allowances_total": post_allowance + hc_allowance,
			"gross_salary": gross_salary,
			"ss_base": ss_base,
			"hf_base": hf_base,
			"ss_person_total": ss_p,
			"ss_company_total": ss_c,
			"hf_person_total": hf_p,
			"hf_company_total": hf_c,
			"tax_threshold": tax_thresh,
			"pension_person": pension_p,
			"medical_person": medical_p,
			"large_medical_person": large_med_p,
			"unemployment_person": unemployment_p,
			"housing_fund_person": hf_p,
			"pension_company": pension_c,
			"unemployment_company": unemployment_c,
			"medical_company": medical_c,
			"other_medical_company": other_medical_c,
			"work_injury_company": injury_c,
			"housing_fund_company": hf_c,
			"deduction_child_education": flt(emp.get("deduction_child_education")),
			"deduction_continuing_education": flt(emp.get("deduction_continuing_education")),
			"deduction_serious_illness": flt(emp.get("deduction_serious_illness")),
			"deduction_housing_loan": flt(emp.get("deduction_housing_loan")),
			"deduction_housing_rent": flt(emp.get("deduction_housing_rent")),
			"deduction_elderly_care": flt(emp.get("deduction_elderly_care")),
			"deduction_infant_care": flt(emp.get("deduction_infant_care")),
			"special_deductions_total": round(spec_d, 2),
			"taxable_income": round(gross_salary - ss_p - hf_p - spec_d - tax_thresh, 2),
			"tax_amount": tax_amount,
			"net_salary": net_salary,
			"remarks": (str(wdata.get("remarks") or "").strip() if (wdata and wdata.get("remarks")) else "")
		})

		# 同步考勤
		att_name = f"{company}-{period_month}-{emp.employee_no}"
		if frappe.db.exists("Ashan Monthly Attendance", att_name):
			att = frappe.get_doc("Ashan Monthly Attendance", att_name)
		else:
			att = frappe.new_doc("Ashan Monthly Attendance")
			att.company = company
			att.employee_no = emp.employee_no
			att.employee_name = emp_name
			att.period_month = period_month
		att.attendance_days = work_days
		att.work_hours_regular = work_hours
		att.save(ignore_permissions=True)

	# 8. 保存月度薪酬核定表
	doc_name = f"{company}-{period_month}"
	if frappe.db.exists("Ashan Monthly Payroll Settlement", doc_name):
		settle_doc = frappe.get_doc("Ashan Monthly Payroll Settlement", doc_name)
		settle_doc.items = []
	else:
		settle_doc = frappe.new_doc("Ashan Monthly Payroll Settlement")
		settle_doc.company = company
		settle_doc.period_month = period_month

	# 汇总分类
	tot_workshop = sum(flt(it.get("fixed_salary")) for it in items_data)
	tot_subsidies = sum(flt(it.get("post_allowance")) + flt(it.get("house_rent_allowance")) for it in items_data)
	count_workshop = sum(1 for it in items_data if flt(it.get("fixed_salary")) > 0 or flt(it.get("attendance_days")) > 0)
	count_non_workshop = sum(1 for it in items_data if flt(it.get("fixed_salary")) == 0 and flt(it.get("post_allowance")) == 0 and flt(it.get("house_rent_allowance")) == 0)

	settle_doc.status = "草稿"
	settle_doc.locked = 0
	settle_doc.total_employees = len(items_data)
	settle_doc.total_gross_salary = round(total_gross, 2)
	settle_doc.total_net_salary = round(total_net, 2)
	settle_doc.total_social_security_company = round(total_ss_comp, 2)
	settle_doc.total_social_security_person = round(total_ss_pers, 2)
	settle_doc.total_housing_fund_company = round(total_hf_comp, 2)
	settle_doc.total_housing_fund_person = round(total_hf_pers, 2)
	settle_doc.total_tax = round(total_tax, 2)

	# 统一规范中文命名并私有化存储原始实发表 Excel
	ext = ".xls" if (filename and filename.lower().endswith(".xls")) else ".xlsx"
	std_sal_fn = f"{period_month}_{company}_车间实发工资表_原始凭证{ext}"
	if raw_file_bytes and not file_url:
		saved_file = frappe.get_doc({
			"doctype": "File",
			"file_name": std_sal_fn,
			"attached_to_doctype": "Ashan Monthly Payroll Settlement",
			"attached_to_name": doc_name,
			"content": raw_file_bytes,
			"is_private": 1
		})
		saved_file.save(ignore_permissions=True)
		file_url = saved_file.file_url

	if file_url:
		settle_doc.imported_excel_file = file_url

	for item in items_data:
		settle_doc.append("items", item)

	frappe.flags.ignore_lock = True
	settle_doc.save(ignore_permissions=True)
	from ashan_cn_procurement.services.payroll_recalculation_service import queue_recalculation_after_change
	queue_recalculation_after_change(
		company=company,
		period_month=period_month,
		employee_no=None,
		trigger_source="外部实发与发放表",
		trigger_detail="导入/覆盖24列外部实发后，服务器统一按VBA累计口径复核全月",
	)
	frappe.db.commit()

	return {
		"success": True,
		"period_month": period_month,
		"total_imported": len(items_data),
		"wife_matched_count": len(wife_payroll_dict),
		"count_workshop": count_workshop,
		"count_non_workshop": count_non_workshop,
		"total_workshop_net": round(tot_workshop, 2),
		"total_subsidies": round(tot_subsidies, 2),
		"total_net_salary": round(total_net, 2),
		"total_gross_salary": round(total_gross, 2),
		"total_tax": round(total_tax, 2),
		"total_ss_comp": round(total_ss_comp, 2),
		"total_ss_pers": round(total_ss_pers, 2),
		"total_hf_comp": round(total_hf_comp, 2),
		"total_hf_pers": round(total_hf_pers, 2),
		"total_comp_cost": round(total_ss_comp + total_hf_comp, 2),
		"message": f"✅ 已以员工母表为基准导入 {len(items_data)} 人的24列外部实发数据，实发总盘 ¥{total_net:,.2f}。服务器后台任务已提交，将统一按 VBA 累计预扣/税后反推口径复核工资、社保、公积金与个税；计算中心可查看状态和完成时间。",
		"doc": get_payroll_settlement_detail(company, period_month)
	}


def _cash_breakdown_from_net(net_salary):
	"""Return the XLSM-compatible cash-note breakdown for an employee net pay.

	The reference workbook keeps cash helper columns outside the 24-column print area:
	100 / 50 / 10 / 5 / 1 yuan, cash total, and a verification difference.  Excel
	uses ``ROUND(net_salary, 0) - cash_total`` for the final verification column, so
	the server follows the same half-up rounding semantics rather than Python's
	banker's rounding.
	"""
	target = int(Decimal(str(flt(net_salary))).quantize(Decimal("1"), rounding=ROUND_HALF_UP))
	remaining = max(target, 0)
	counts = {}
	for note in (100, 50, 10, 5, 1):
		counts[note], remaining = divmod(remaining, note)
	cash_total = sum(note * counts[note] for note in counts)
	return {
		"cash_100": counts[100],
		"cash_50": counts[50],
		"cash_10": counts[10],
		"cash_5": counts[5],
		"cash_1": counts[1],
		"cash_total": cash_total,
		"cash_check": round(target - cash_total, 2),
		"cash_target": target,
	}


@frappe.whitelist()
def get_salary_distribution_sheet(company="天津祺富机械加工有限公司", period_month="2026-07"):
	"""
	获取《薪资发放表》数据 (精准 24 列)：
	序号, 工号, 姓名, 作业天数, 作业小时, 天工资, 小时工资, 全勤费, 加班小时, 加班费,
	国勤天数, 国勤工资, 达标率, 达标工资, 扣除, 考勤绩效工资合计, 职位补贴, 房/车补,
	补贴工资合计, 应发工资合计, 工资调整, 实发工资合计, 签字, 备考
	"""
	check_payroll_workbench_permission("read")
	detail = get_payroll_settlement_detail(company, period_month)
	items = detail.get("items", [])

	rows = []
	tot_workshop = 0.0
	tot_post = 0.0
	tot_house = 0.0
	tot_subsidies = 0.0
	tot_payable = 0.0
	tot_net = 0.0

	seq_num = 1
	for it in items:
		rem = it.get("remarks") or ""
		# 过滤非车间出勤人员（如外籍工/非车间在册人员即便直接输入工资，也绝不混入车间24列实发表）
		if "非车间出勤" in rem or it.get("employee_type") == "外籍工" or (flt(it.get("work_hours")) == 0 and flt(it.get("attendance_days")) == 0 and not str(it.get("employee_no", "")).startswith("A")):
			continue

		workshop_net = flt(it.get("fixed_salary"))
		post_all = flt(it.get("post_allowance"))
		house_all = flt(it.get("house_rent_allowance"))
		subsidies_tot = post_all + house_all
		payable_tot = workshop_net + subsidies_tot
		net_salary = flt(it.get("net_salary"))
		adjust_val = round(net_salary - payable_tot, 2)
		if abs(adjust_val) < 0.01: adjust_val = 0.0

		cash = _cash_breakdown_from_net(net_salary)
		row = {
			"seq": seq_num,
			"employee_no": it.get("employee_no"),
			"employee_name": it.get("employee_name"),
			"work_days": flt(it.get("attendance_days")),
			"work_hours": flt(it.get("work_hours")),
			"day_salary": flt(it.get("day_salary")),
			"hour_salary": flt(it.get("hour_salary")),
			"full_attendance": flt(it.get("full_attendance")),
			"overtime_hours": flt(it.get("overtime_hours")),
			"overtime_salary": flt(it.get("overtime_salary")),
			"national_days": flt(it.get("national_days")),
			"national_salary": flt(it.get("national_salary")),
			"target_rate": str(it.get("target_rate") or ""),
			"target_salary": flt(it.get("target_salary")),
			"deduction": flt(it.get("deduction")),
			"workshop_net": workshop_net,
			"post_allowance": post_all,
			"house_rent_allowance": house_all,
			"subsidies_total": subsidies_tot,
			"payable_total": payable_tot,
			"salary_adjust": adjust_val,
			"net_salary": net_salary,
			"sign": "",
			"remarks": it.get("remarks") or ""
		}
		row.update(cash)
		# Compatibility aliases used by the legacy modal/UI.  Keeping them here avoids
		# silent blank cells for older clients while the canonical field names remain above.
		row.update({
			"attendance_days": row["work_days"],
			"workshop_subtotal": row["workshop_net"],
			"allowance_subtotal": row["subsidies_total"],
			"payable_salary": row["payable_total"],
			"salary_adjustment": row["salary_adjust"],
		})
		rows.append(row)

		tot_workshop += workshop_net
		tot_post += post_all
		tot_house += house_all
		tot_subsidies += subsidies_tot
		tot_payable += payable_tot
		tot_net += net_salary
		seq_num += 1

	totals = {
		"seq": "合计",
		"employee_no": f"共 {len(rows)} 人",
		"employee_name": "",
		"work_days": sum(r["work_days"] for r in rows),
		"work_hours": sum(r["work_hours"] for r in rows),
		"day_salary": sum(r["day_salary"] for r in rows),
		"hour_salary": sum(r["hour_salary"] for r in rows),
		"full_attendance": sum(r["full_attendance"] for r in rows),
		"overtime_hours": sum(r["overtime_hours"] for r in rows),
		"overtime_salary": sum(r["overtime_salary"] for r in rows),
		"national_days": sum(r["national_days"] for r in rows),
		"national_salary": sum(r["national_salary"] for r in rows),
		"target_rate": "",
		"target_salary": sum(r["target_salary"] for r in rows),
		"deduction": sum(r["deduction"] for r in rows),
		"workshop_net": tot_workshop,
		"post_allowance": tot_post,
		"house_rent_allowance": tot_house,
		"subsidies_total": tot_subsidies,
		"payable_total": tot_payable,
		"salary_adjust": sum(r["salary_adjust"] for r in rows),
		"net_salary": tot_net,
		"sign": "",
		"remarks": ""
	}

	# Compatibility aliases for older UI/modal clients.
	totals.update({
		"attendance_days": totals["work_days"],
		"workshop_subtotal": totals["workshop_net"],
		"allowance_subtotal": totals["subsidies_total"],
		"payable_salary": totals["payable_total"],
		"salary_adjustment": totals["salary_adjust"],
	})

	# XLSM 参考中的隐藏发放辅助列（Y:AE）汇总。核定列为 0 表示现金张数与 ROUND(实发,0) 一致。
	totals.update({
		"cash_100": sum(r.get("cash_100", 0) for r in rows),
		"cash_50": sum(r.get("cash_50", 0) for r in rows),
		"cash_10": sum(r.get("cash_10", 0) for r in rows),
		"cash_5": sum(r.get("cash_5", 0) for r in rows),
		"cash_1": sum(r.get("cash_1", 0) for r in rows),
		"cash_total": sum(r.get("cash_total", 0) for r in rows),
		"cash_check": round(sum(r.get("cash_check", 0) for r in rows), 2),
	})

	# 判断当前账期是否已导入过外部工资表
	is_imported = any(r["work_days"] > 0 or r["workshop_net"] > 0 for r in rows) if rows else False

	return {
		"company": company,
		"period_month": period_month,
		"is_imported": is_imported,
		"salary_people_count": len(rows),
		"attendance_salary_total": round(tot_workshop, 2),
		"allowance_total": round(tot_subsidies, 2),
		"attendance_allowance_total": round(tot_payable, 2),
		"net_salary_total": round(tot_net, 2),
		"rows": rows,
		"totals": totals
	}


@frappe.whitelist()
def get_salary_cash_count_sheet(company="天津祺富机械加工有限公司", period_month="2026-07"):
	"""Return the cash-counting view derived from the 24-column external pay sheet."""
	data = get_salary_distribution_sheet(company, period_month)
	rows = [
		{
			"seq": r.get("seq"),
			"employee_no": r.get("employee_no"),
			"employee_name": r.get("employee_name"),
			"cash_100": r.get("cash_100", 0),
			"cash_50": r.get("cash_50", 0),
			"cash_10": r.get("cash_10", 0),
			"cash_5": r.get("cash_5", 0),
			"cash_1": r.get("cash_1", 0),
			"cash_total": r.get("cash_total", 0),
			"cash_check": r.get("cash_check", 0),
		}
		for r in data.get("rows", [])
	]
	return {
		"company": company,
		"period_month": period_month,
		"rows": rows,
		"totals": {key: data.get("totals", {}).get(key, 0) for key in (
			"cash_100", "cash_50", "cash_10", "cash_5", "cash_1", "cash_total", "cash_check"
		)},
	}


@frappe.whitelist()
def get_accounting_payroll_sheet(company="天津祺富机械加工有限公司", period_month="2026-07"):
	"""Return the XLSM-aligned accounting payroll ledger.

	The reference workbook has one 11-column sheet for domestic/rehired staff and a
	separate ``记账工资表外籍`` sheet.  The API therefore returns ``rows`` and
	``foreign_rows`` separately so the UI/export can preserve that accounting boundary.
	"""
	check_payroll_workbench_permission("read")
	detail = get_payroll_settlement_detail(company, period_month)
	items = detail.get("items", [])

	def make_row(it):
		gross = flt(it.get("gross_salary"))
		post_all = flt(it.get("post_allowance"))
		house_all = flt(it.get("house_rent_allowance"))
		base_perf = max(0.0, round(gross - post_all - house_all, 2))
		ss_p = flt(it.get("ss_person_total"))
		hf_p = flt(it.get("hf_person_total"))
		tax = flt(it.get("tax_amount"))
		total_ded = round(ss_p + hf_p + tax, 2)
		return {
			"employee_no": it.get("employee_no"),
			"employee_name": it.get("employee_name"),
			"employee_type": it.get("employee_type") or "",
			"base_perf_salary": base_perf,
			# compatibility alias used by the existing modal
			"base_performance_salary": base_perf,
			"post_allowance": post_all,
			"house_rent_allowance": house_all,
			"gross_salary": gross,
			"hf_person_total": hf_p,
			"ss_person_total": ss_p,
			"tax_amount": tax,
			"total_deduction": total_ded,
			"net_salary": flt(it.get("net_salary")),
		}

	all_rows = [make_row(it) for it in items]
	rows = [r for r in all_rows if "外籍" not in str(r.get("employee_type") or "")]
	foreign_rows = [r for r in all_rows if "外籍" in str(r.get("employee_type") or "")]
	for idx, row in enumerate(rows, start=1):
		row["seq"] = idx
	for idx, row in enumerate(foreign_rows, start=1):
		row["seq"] = idx

	def make_totals(source_rows):
		keys = (
			"base_perf_salary", "post_allowance", "house_rent_allowance", "gross_salary",
			"hf_person_total", "ss_person_total", "tax_amount", "total_deduction", "net_salary",
		)
		result = {key: round(sum(flt(r.get(key)) for r in source_rows), 2) for key in keys}
		result["base_performance_salary"] = result["base_perf_salary"]
		result["employee_no"] = "合计"
		result["employee_name"] = f"共 {len(source_rows)} 人"
		return result

	return {
		"company": company,
		"period_month": period_month,
		"rows": rows,
		"foreign_rows": foreign_rows,
		"all_rows": all_rows,
		"totals": make_totals(rows),
		"foreign_totals": make_totals(foreign_rows),
		"all_totals": make_totals(all_rows),
	}



# ==========================================
# 3. 社保台账服务 (19列)
# ==========================================
@frappe.whitelist()
def get_social_insurance_sheet(company="天津祺富机械加工有限公司", period_month="2026-07"):
	check_payroll_workbench_permission("read")
	parent_name = f"{company}-{period_month}"
	locked_snapshot = False
	if frappe.db.exists("Ashan Monthly Payroll Settlement", parent_name):
		state = frappe.db.get_value("Ashan Monthly Payroll Settlement", parent_name, ["locked", "status"], as_dict=True) or {}
		locked_snapshot = bool(cint(state.get("locked"))) or state.get("status") in ["已核定锁定", "已归档发放"]
	if locked_snapshot:
		snapshot_rows = frappe.get_all(
			"Ashan Monthly Payroll Item", filters={"parent": parent_name},
			fields=[
				"employee_no", "employee_name", "id_card", "employee_type", "ss_base",
				"pension_company", "unemployment_company", "medical_company", "other_medical_company", "work_injury_company", "ss_company_total",
				"pension_person", "unemployment_person", "medical_person", "large_medical_person", "ss_person_total",
			], order_by="idx asc",
		)
	else:
		snapshot_rows = []

	items = _salary_profiles_for_period(
		company, period_month,
		fields=["employee_no", "employee_name", "id_card", "employee_type", "employment_status", "relieving_date", "social_security_base"],
		order_by="employee_no asc",
	)
	ss_setting = get_insurance_setting(company, period_month.split("-")[0] if "-" in period_month else 2026)

	rows = []
	seq_idx = 1
	if locked_snapshot:
		for it in snapshot_rows:
			ss_base = flt(it.get("ss_base"))
			if ss_base <= 0:
				continue
			comp_tot = flt(it.get("ss_company_total"))
			pers_tot = flt(it.get("ss_person_total"))
			rows.append({
				"seq": seq_idx, "employee_no": it.get("employee_no"), "employee_name": it.get("employee_name"),
				"id_card": it.get("id_card") or "-", "period_month_str": period_month.replace("-", ""),
				"employee_type": it.get("employee_type") or "正式工", "ss_base": ss_base,
				"comp_pension": flt(it.get("pension_company")), "comp_unemp": flt(it.get("unemployment_company")),
				"comp_med": flt(it.get("medical_company")), "comp_other_med": flt(it.get("other_medical_company")),
				"comp_injury": flt(it.get("work_injury_company")), "comp_total": comp_tot,
				"pers_pension": flt(it.get("pension_person")), "pers_unemp": flt(it.get("unemployment_person")),
				"pers_med": flt(it.get("medical_person")), "pers_large_med": flt(it.get("large_medical_person")),
				"pers_total": pers_tot, "grand_total": round(comp_tot + pers_tot, 2),
			})
			seq_idx += 1
	for it in ([] if locked_snapshot else items):
		ss_base = flt(it.get("social_security_base"))
		emp_type = it.get("employee_type") or "正式工"
		emp_status = it.get("employment_status") or "在职"
		rel_d = str(it.get("relieving_date") or "")
		is_retired = (emp_type in TAX_REHIRE_EMPLOYEE_TYPES)
		is_other = (emp_type in ["临时工", "零工", "外籍工", "实习生", "劳务派遣"])
		is_resigned = bool(rel_d and rel_d.startswith(period_month))

		# 纯净权责边界：不缴纳社保人员（退休返聘、临时工、外籍工、实习生、本月离职减员或基数<=0）自动过滤排除
		if is_retired or is_other or is_resigned or ss_base <= 0:
			continue

		comp_pension = round(ss_base * (flt(ss_setting.get("ss_company_pension", 16.0)) / 100.0), 2)
		comp_unemp = round(ss_base * (flt(ss_setting.get("ss_company_unemployment", 0.5)) / 100.0), 2)
		comp_med = round(ss_base * (flt(ss_setting.get("ss_company_medical", 10.0)) / 100.0), 2)
		comp_other_med = round(ss_base * (flt(ss_setting.get("ss_company_other_medical", 0.5)) / 100.0), 2)
		comp_injury = round(ss_base * (flt(ss_setting.get("ss_company_injury", 0.55)) / 100.0), 2)
		comp_tot = round(comp_pension + comp_unemp + comp_med + comp_other_med + comp_injury, 2)

		pers_pension = round(ss_base * (flt(ss_setting.get("ss_person_pension", 8.0)) / 100.0), 2)
		pers_unemp = round(ss_base * (flt(ss_setting.get("ss_person_unemployment", 0.5)) / 100.0), 2)
		pers_med = round(ss_base * (flt(ss_setting.get("ss_person_medical", 2.0)) / 100.0), 2)

		cur_m_num = int(period_month.split("-")[1]) if "-" in period_month else 7
		spec_months = [int(m.strip()) for m in str(ss_setting.get("big_medical_special_months", "3,12")).split(",") if m.strip().isdigit()]
		pers_large_med = flt(ss_setting.get("big_medical_amount_special", 21.0)) if cur_m_num in spec_months else flt(ss_setting.get("big_medical_amount_default", 22.0))

		pers_tot = round(pers_pension + pers_unemp + pers_med + pers_large_med, 2)
		grand_tot = round(comp_tot + pers_tot, 2)

		rows.append({
			"seq": seq_idx,
			"employee_no": it.get("employee_no"),
			"employee_name": it.get("employee_name"),
			"id_card": it.get("id_card") or "-",
			"period_month_str": period_month.replace("-", ""),
			"employee_type": emp_type,
			"ss_base": ss_base,
			"comp_pension": comp_pension,
			"comp_unemp": comp_unemp,
			"comp_med": comp_med,
			"comp_other_med": comp_other_med,
			"comp_injury": comp_injury,
			"comp_total": comp_tot,
			"pers_pension": pers_pension,
			"pers_unemp": pers_unemp,
			"pers_med": pers_med,
			"pers_large_med": pers_large_med,
			"pers_total": pers_tot,
			"grand_total": grand_tot
		})
		seq_idx += 1

	# 特殊补缴/滞纳金作为独立财务快照行，不回写员工母表基数。
	if frappe.db.exists("DocType", "Ashan Social Insurance Adjustment"):
		adjustments = frappe.get_all(
			"Ashan Social Insurance Adjustment",
			filters={"company": company, "payroll_period": period_month},
			fields=[
				"name", "employee_no", "employee_name", "id_card", "employee_type", "adjustment_period", "biz_type",
				"ss_base", "comp_pension", "comp_unemp", "comp_med", "comp_other_med", "comp_injury", "comp_total",
				"pers_pension", "pers_unemp", "pers_med", "pers_large_med", "pers_total", "late_fee", "grand_total", "remarks",
			],
			order_by="creation asc",
		)
		for adj in adjustments:
			rows.append({
				"seq": seq_idx, "adj_id": adj.name, "employee_no": adj.employee_no, "employee_name": adj.employee_name,
				"id_card": adj.id_card or "-", "period_month_str": adj.adjustment_period, "employee_type": adj.employee_type or "调整项",
				"biz_type": adj.biz_type, "ss_base": flt(adj.ss_base), "comp_pension": flt(adj.comp_pension),
				"comp_unemp": flt(adj.comp_unemp), "comp_med": flt(adj.comp_med), "comp_other_med": flt(adj.comp_other_med),
				"comp_injury": flt(adj.comp_injury), "comp_total": flt(adj.comp_total), "pers_pension": flt(adj.pers_pension),
				"pers_unemp": flt(adj.pers_unemp), "pers_med": flt(adj.pers_med), "pers_large_med": flt(adj.pers_large_med),
				"pers_total": flt(adj.pers_total), "late_fee": flt(adj.late_fee), "grand_total": flt(adj.grand_total),
				"remarks": adj.remarks or "",
			})
			seq_idx += 1

	totals = {
		"seq": "合计",
		"employee_no": f"台账共 {len(rows)} 行",
		"ss_base": sum(r["ss_base"] for r in rows),
		"comp_pension": sum(r["comp_pension"] for r in rows),
		"comp_unemp": sum(r["comp_unemp"] for r in rows),
		"comp_med": sum(r["comp_med"] for r in rows),
		"comp_other_med": sum(r["comp_other_med"] for r in rows),
		"comp_injury": sum(r["comp_injury"] for r in rows),
		"comp_total": sum(r["comp_total"] for r in rows),
		"pers_pension": sum(r["pers_pension"] for r in rows),
		"pers_unemp": sum(r["pers_unemp"] for r in rows),
		"pers_med": sum(r["pers_med"] for r in rows),
		"pers_large_med": sum(r["pers_large_med"] for r in rows),
		"pers_total": sum(r["pers_total"] for r in rows),
		"grand_total": sum(r["grand_total"] for r in rows)
	}

	return {
		"company": company,
		"period_month": period_month,
		"report_title": f"{period_month} 社会保险缴费明细表",
		"rows": rows,
		"totals": totals
	}


@frappe.whitelist(methods=["POST"])
def save_social_insurance_adjustment(company, period_month, adjustment_json):
	"""保存社保特殊补缴/滞纳金快照。仅影响社保台账与凭证核验，不静默改写历史工资。"""
	check_payroll_workbench_permission("write")
	parent_name = f"{company}-{period_month}"
	if frappe.db.exists("Ashan Monthly Payroll Settlement", parent_name):
		parent_state = frappe.db.get_value("Ashan Monthly Payroll Settlement", parent_name, ["locked", "status"], as_dict=True) or {}
		if cint(parent_state.get("locked")) or parent_state.get("status") in ["已核定锁定", "已归档发放"]:
			frappe.throw("当前薪酬账期已冻结，不能新增社保调整。请先反审核解锁。")
	payload = json.loads(adjustment_json) if isinstance(adjustment_json, str) else (adjustment_json or {})
	emp_no = str(payload.get("employee_no") or "").strip()
	adjustment_period = str(payload.get("period_month_str") or "").replace("-", "").strip()
	if len(adjustment_period) != 6 or not adjustment_period.isdigit():
		frappe.throw("补缴/调整所属期必须为 YYYYMM，例如 202605。")
	if not emp_no:
		frappe.throw("请选择员工。")
	emp = frappe.db.get_value(
		"Ashan Employee Salary Profile", {"company": company, "employee_no": emp_no},
		["employee_name", "employee_type", "id_card"], as_dict=True,
	)
	if not emp:
		frappe.throw(f"未找到员工 {emp_no} 的薪酬档案。")
	ss_base = round(flt(payload.get("ss_base")), 2)
	late_fee = round(flt(payload.get("late_fee")), 2)
	if ss_base < 0 or late_fee < 0:
		frappe.throw("社保基数与滞纳金不能为负数。")
	year = cint(adjustment_period[:4])
	month = cint(adjustment_period[4:6])
	setting = get_insurance_setting(company, year) or {}
	comp_pension = round(ss_base * flt(setting.get("ss_company_pension", 16.0)) / 100.0, 2)
	comp_unemp = round(ss_base * flt(setting.get("ss_company_unemployment", 0.5)) / 100.0, 2)
	comp_med = round(ss_base * flt(setting.get("ss_company_medical", 10.0)) / 100.0, 2)
	comp_other_med = round(ss_base * flt(setting.get("ss_company_other_medical", 0.5)) / 100.0, 2)
	comp_injury = round(ss_base * flt(setting.get("ss_company_injury", 0.55)) / 100.0, 2)
	comp_contrib = round(comp_pension + comp_unemp + comp_med + comp_other_med + comp_injury, 2)
	pers_pension = round(ss_base * flt(setting.get("ss_person_pension", 8.0)) / 100.0, 2)
	pers_unemp = round(ss_base * flt(setting.get("ss_person_unemployment", 0.5)) / 100.0, 2)
	pers_med = round(ss_base * flt(setting.get("ss_person_medical", 2.0)) / 100.0, 2)
	special_months = [cint(x.strip()) for x in str(setting.get("big_medical_special_months") or "3,12").split(",") if x.strip().isdigit()]
	pers_large = flt(setting.get("big_medical_amount_special", 21.0)) if month in special_months else flt(setting.get("big_medical_amount_default", 22.0))
	pers_total = round(pers_pension + pers_unemp + pers_med + pers_large, 2) if ss_base > 0 else 0.0
	# 滞纳金作为企业侧额外成本计入单位合计，保留独立字段以便审计。
	comp_total = round(comp_contrib + late_fee, 2)
	grand_total = round(comp_total + pers_total, 2)
	doc = frappe.new_doc("Ashan Social Insurance Adjustment")
	doc.company = company
	doc.payroll_period = period_month
	doc.adjustment_period = adjustment_period
	doc.employee_no = emp_no
	doc.employee_name = emp.employee_name
	doc.employee_type = emp.employee_type
	doc.id_card = emp.id_card
	doc.biz_type = payload.get("biz_type") or "历史补缴"
	doc.ss_base = ss_base
	doc.comp_pension = comp_pension
	doc.comp_unemp = comp_unemp
	doc.comp_med = comp_med
	doc.comp_other_med = comp_other_med
	doc.comp_injury = comp_injury
	doc.comp_total = comp_total
	doc.pers_pension = pers_pension
	doc.pers_unemp = pers_unemp
	doc.pers_med = pers_med
	doc.pers_large_med = pers_large
	doc.pers_total = pers_total
	doc.late_fee = late_fee
	doc.grand_total = grand_total
	doc.remarks = payload.get("remarks") or ""
	doc.insert(ignore_permissions=True)
	return {"success": True, "name": doc.name, "message": f"社保调整 {doc.name} 已登记，台账增加 ¥{grand_total:,.2f}。"}


@frappe.whitelist(methods=["POST"])
def delete_social_insurance_adjustment(company, period_month, adj_id):
	"""删除未冻结账期中的社保特殊调整记录。"""
	check_payroll_workbench_permission("write")
	parent_name = f"{company}-{period_month}"
	if frappe.db.exists("Ashan Monthly Payroll Settlement", parent_name):
		state = frappe.db.get_value("Ashan Monthly Payroll Settlement", parent_name, ["locked", "status"], as_dict=True) or {}
		if cint(state.get("locked")) or state.get("status") in ["已核定锁定", "已归档发放"]:
			frappe.throw("当前薪酬账期已冻结，不能删除社保调整。请先反审核解锁。")
	row = frappe.db.get_value("Ashan Social Insurance Adjustment", adj_id, ["company", "payroll_period"], as_dict=True)
	if not row or row.company != company or row.payroll_period != period_month:
		frappe.throw("未找到当前公司/账期对应的社保调整记录。")
	frappe.delete_doc("Ashan Social Insurance Adjustment", adj_id, ignore_permissions=True)
	return {"success": True, "message": f"社保调整 {adj_id} 已删除。"}


# ==========================================
# 4. 公积金台账服务 (12列)
# ==========================================
@frappe.whitelist()
def get_housing_fund_sheet(company="天津祺富机械加工有限公司", period_month="2026-07"):
	check_payroll_workbench_permission("read")
	parent_name = f"{company}-{period_month}"
	locked_snapshot = False
	if frappe.db.exists("Ashan Monthly Payroll Settlement", parent_name):
		state = frappe.db.get_value("Ashan Monthly Payroll Settlement", parent_name, ["locked", "status"], as_dict=True) or {}
		locked_snapshot = bool(cint(state.get("locked"))) or state.get("status") in ["已核定锁定", "已归档发放"]
	if locked_snapshot:
		snapshot_rows = frappe.get_all(
			"Ashan Monthly Payroll Item", filters={"parent": parent_name},
			fields=["employee_no", "employee_name", "id_card", "employee_type", "hf_base", "hf_company_total", "hf_person_total"],
			order_by="idx asc",
		)
	else:
		snapshot_rows = []
	items = _salary_profiles_for_period(
		company, period_month,
		fields=["employee_no", "employee_name", "id_card", "employee_type", "employment_status", "relieving_date", "housing_fund_base"],
		order_by="employee_no asc",
	)
	ss_setting = get_insurance_setting(company, period_month.split("-")[0] if "-" in period_month else 2026)

	comp_rate = flt(ss_setting.get("hf_company_rate", 5.0))
	pers_rate = flt(ss_setting.get("hf_person_rate", 5.0))

	rows = []
	seq_idx = 1
	if locked_snapshot:
		for it in snapshot_rows:
			hf_base = flt(it.get("hf_base"))
			if hf_base <= 0:
				continue
			c_amt = flt(it.get("hf_company_total"))
			p_amt = flt(it.get("hf_person_total"))
			rows.append({
				"seq": seq_idx, "employee_no": it.get("employee_no"), "employee_name": it.get("employee_name"),
				"id_card": it.get("id_card") or "-", "period_month_str": period_month.replace("-", ""),
				"employee_type": it.get("employee_type") or "正式工", "hf_base": hf_base,
				"comp_rate": round(c_amt / hf_base * 100, 4) if hf_base else 0.0, "comp_amount": c_amt,
				"pers_rate": round(p_amt / hf_base * 100, 4) if hf_base else 0.0, "pers_amount": p_amt,
				"total_amount": round(c_amt + p_amt, 2),
			})
			seq_idx += 1
	for it in ([] if locked_snapshot else items):
		hf_base = flt(it.get("housing_fund_base"))
		emp_type = it.get("employee_type") or "正式工"
		emp_status = it.get("employment_status") or "在职"
		rel_d = str(it.get("relieving_date") or "")
		emp_name = it.get("employee_name") or ""
		is_resigned = bool(rel_d and rel_d.startswith(period_month))

		# 纯净权责边界：不缴纳公积金人员（退休返聘、临时工、外籍工、实习生、本月离职减员或基数<=0）自动过滤排除
		if emp_type in TAX_REHIRE_EMPLOYEE_TYPES or emp_type in ["临时工", "零工", "外籍工", "实习生"] or is_resigned or hf_base <= 0:
			continue

		c_amt = round(hf_base * (comp_rate / 100.0), 2)
		p_amt = round(hf_base * (pers_rate / 100.0), 2)
		tot_amt = round(c_amt + p_amt, 2)

		rows.append({
			"seq": seq_idx,
			"employee_no": it.get("employee_no"),
			"employee_name": emp_name,
			"id_card": it.get("id_card") or "-",
			"period_month_str": period_month.replace("-", ""),
			"employee_type": emp_type,
			"hf_base": hf_base,
			"comp_rate": comp_rate,
			"comp_amount": c_amt,
			"pers_rate": pers_rate,
			"pers_amount": p_amt,
			"total_amount": tot_amt
		})
		seq_idx += 1

	totals = {
		"seq": "合计",
		"employee_no": f"参缴共 {len(rows)} 人",
		"hf_base": sum(r["hf_base"] for r in rows),
		"comp_amount": sum(r["comp_amount"] for r in rows),
		"pers_amount": sum(r["pers_amount"] for r in rows),
		"total_amount": sum(r["total_amount"] for r in rows)
	}

	return {
		"company": company,
		"period_month": period_month,
		"report_title": f"{period_month} 住房公积金缴存明细表",
		"rows": rows,
		"totals": totals
	}


# ==========================================
# 5. 个人所得税法定 68 列大宽表与科学精简版服务
# ==========================================
@frappe.whitelist()
def get_tax_settlement_sheet(company="天津祺富机械加工有限公司", period_month="2026-07"):
	"""
	获取个人所得税 5 大逻辑分组科学精简版数据 (17列)
	"""
	return get_tax_settlement_full_sheet(company, period_month)


@frappe.whitelist()
def get_tax_settlement_full_sheet(company="天津祺富机械加工有限公司", period_month="2026-07"):
    """
    组装与现行 VBA《个人所得税》一致的 68 列累计预扣申报台账。
    核心原则：历史月份使用已落库快照；累计起征点按员工实际存在的月份记录累计；
    负累计应纳税所得额不截断，以便与 VBA 核对。
    """
    check_payroll_workbench_permission("read")
    params = get_effective_tax_parameters(company, period_month)
    tax_thresh = params["tax_threshold"]
    cycle_start_m = params["tax_cycle_start_month"]
    cinfo = get_tax_cycle_info(period_month, tax_thresh, cycle_start_m)
    prior_months = cinfo["prior_months"]

    detail = get_payroll_settlement_detail(company, period_month)
    items = detail.get("items", [])

    # 一次性读取往期明细，保留 VBA 需要的专项扣除与 7 项专项附加扣除分项。
    prior_parent_names = [f"{company}-{m}" for m in prior_months]
    prior_fields = [
        "employee_no", "parent", "gross_salary", "tax_threshold",
        "ss_person_total", "hf_person_total",
        "pension_person", "medical_person", "large_medical_person", "unemployment_person", "housing_fund_person",
        "special_deductions_total", "deduction_child_education", "deduction_continuing_education",
        "deduction_serious_illness", "deduction_housing_loan", "deduction_housing_rent",
        "deduction_elderly_care", "deduction_infant_care", "tax_amount",
    ]
    prior_records = []
    if prior_parent_names:
        prior_records = frappe.get_all(
            "Ashan Monthly Payroll Item",
            filters={"parent": ["in", prior_parent_names]},
            fields=prior_fields,
            order_by="parent asc",
        )
    prior_by_emp = {}
    for r in prior_records:
        prior_by_emp.setdefault(r.employee_no, []).append(r)

    # 当前年度费率只用于“旧数据缺少分项字段”时的可审计回填，不改变已保存的社保/公积金总额。
    year = _period_year(period_month)
    ins = get_insurance_setting(company, year) or {}
    p_rate = flt(ins.get("ss_person_pension")) or 8.0
    m_rate = flt(ins.get("ss_person_medical")) or 2.0
    u_rate = flt(ins.get("ss_person_unemployment")) or 0.5

    # 临时工/零工不进入个人所得税申报台账；返聘类（含“其他-返聘工”）正常进入。
    items = [it for it in items if is_tax_ledger_employee(it.get("employee_type"))]

    rows = []
    for idx, it in enumerate(items, start=1):
        emp_no = it.get("employee_no")
        gender = it.get("gender") or "-"
        id_card = it.get("id_card") or "-"
        salary_mode = it.get("salary_mode") or "税后"

        gross_cur = flt(it.get("gross_salary"))
        ss_p = flt(it.get("ss_person_total"))
        hf_p = flt(it.get("hf_person_total"))
        threshold_cur = flt(it.get("tax_threshold")) or tax_thresh
        wage_ded_cur = round(ss_p + hf_p, 2)

        # 本月专项扣除分项。若旧记录尚未写入分项，但有基数与总额，则按当期配置回填前三项，
        # 大额医疗使用“总额减比例项”的残差，确保与真实已扣社保总额对齐。
        pension_cur = flt(it.get("pension_person"))
        medical_cur = flt(it.get("medical_person"))
        large_med_cur = flt(it.get("large_medical_person"))
        unemp_cur = flt(it.get("unemployment_person"))
        hf_stat_cur = flt(it.get("housing_fund_person")) or hf_p
        ss_base = flt(it.get("ss_base"))
        if ss_p > 0 and abs(pension_cur) + abs(medical_cur) + abs(large_med_cur) + abs(unemp_cur) < 0.005 and ss_base > 0:
            pension_cur = round(ss_base * p_rate / 100.0, 2)
            medical_cur = round(ss_base * m_rate / 100.0, 2)
            unemp_cur = round(ss_base * u_rate / 100.0, 2)
            large_med_cur = round(ss_p - pension_cur - medical_cur - unemp_cur, 2)
        spec_ded_cur = round(pension_cur + medical_cur + large_med_cur + unemp_cur + hf_stat_cur, 2)
        # 如分项历史不完整，专项扣除合计仍以真实工资扣款总额为准。
        if abs(spec_ded_cur - wage_ded_cur) > 0.02:
            spec_ded_cur = wage_ded_cur

        add_cur = {
            "child": flt(it.get("deduction_child_education")),
            "edu": flt(it.get("deduction_continuing_education")),
            "med": flt(it.get("deduction_serious_illness")),
            "loan": flt(it.get("deduction_housing_loan")),
            "rent": flt(it.get("deduction_housing_rent")),
            "elder": flt(it.get("deduction_elderly_care")),
            "baby": flt(it.get("deduction_infant_care")),
        }
        add_cur_total = round(sum(add_cur.values()), 2)
        stored_add_total = flt(it.get("special_deductions_total"))
        if abs(add_cur_total - stored_add_total) > 0.02 and stored_add_total:
            # 组件缺失时保留真实总额；重新核定后组件会被完整写入。
            add_cur_total = stored_add_total

        plist = prior_by_emp.get(emp_no, [])
        gross_prior = round(sum(flt(x.gross_salary) for x in plist), 2)
        threshold_prior = round(sum((flt(x.tax_threshold) or tax_thresh) for x in plist), 2)
        pension_prior = round(sum(flt(x.pension_person) for x in plist), 2)
        medical_prior = round(sum(flt(x.medical_person) for x in plist), 2)
        large_med_prior = round(sum(flt(x.large_medical_person) for x in plist), 2)
        unemp_prior = round(sum(flt(x.unemployment_person) for x in plist), 2)
        hf_prior = round(sum((flt(x.housing_fund_person) or flt(x.hf_person_total)) for x in plist), 2)
        spec_ded_prior = round(sum(flt(x.ss_person_total) + flt(x.hf_person_total) for x in plist), 2)
        add_prior = {
            "child": round(sum(flt(x.deduction_child_education) for x in plist), 2),
            "edu": round(sum(flt(x.deduction_continuing_education) for x in plist), 2),
            "med": round(sum(flt(x.deduction_serious_illness) for x in plist), 2),
            "loan": round(sum(flt(x.deduction_housing_loan) for x in plist), 2),
            "rent": round(sum(flt(x.deduction_housing_rent) for x in plist), 2),
            "elder": round(sum(flt(x.deduction_elderly_care) for x in plist), 2),
            "baby": round(sum(flt(x.deduction_infant_care) for x in plist), 2),
        }
        add_prior_total = round(sum(flt(x.special_deductions_total) for x in plist), 2)
        paid_tax_prior = round(sum(flt(x.tax_amount) for x in plist), 2)

        gross_all = round(gross_prior + gross_cur, 2)
        threshold_all = round(threshold_prior + threshold_cur, 2)
        pension_all = round(pension_prior + pension_cur, 2)
        medical_all = round(medical_prior + medical_cur, 2)
        large_med_all = round(large_med_prior + large_med_cur, 2)
        unemp_all = round(unemp_prior + unemp_cur, 2)
        hf_all = round(hf_prior + hf_stat_cur, 2)
        spec_ded_all = round(spec_ded_prior + spec_ded_cur, 2)
        add_all = {k: round(add_prior[k] + add_cur[k], 2) for k in add_cur}
        add_all_total = round(add_prior_total + add_cur_total, 2)

        taxable_all = round(gross_all - threshold_all - spec_ded_all - add_all_total, 2)
        rate = 0.03
        quick = 0.0
        for lower, upper, r_pct, q_val in TAX_BRACKETS:
            if taxable_all <= upper:
                rate, quick = r_pct, q_val
                break
        cumulative_tax = round(taxable_all * rate - quick, 2)
        tax_cur = flt(it.get("tax_amount"))
        net_cur = flt(it.get("net_salary"))
        target_salary = net_cur if salary_mode in ["税后", "税后倒推"] else (flt(it.get("fixed_salary")) or gross_cur)

        row = {
            "seq": idx, "employee_no": emp_no, "employee_name": it.get("employee_name"),
            "id_card": id_card, "gender": gender, "period_month_str": period_month.replace("-", ""),
            "employee_type": it.get("employee_type") or "正式工", "target_salary": target_salary,
            "salary_mode": salary_mode,
            "gross_salary": gross_cur, "thresh_cur": threshold_cur, "hf_person": hf_p, "ss_person": ss_p,
            "deduct_cur_tot": wage_ded_cur,
            "ss_pension": pension_cur, "ss_med": medical_cur, "ss_large_med": large_med_cur,
            "ss_unemp": unemp_cur, "hf_spec": hf_stat_cur, "spec_tot_cur": spec_ded_cur,
            "spec_add_child": add_cur["child"], "spec_add_edu": add_cur["edu"], "spec_add_med": add_cur["med"],
            "spec_add_loan": add_cur["loan"], "spec_add_rent": add_cur["rent"], "spec_add_elder": add_cur["elder"],
            "spec_add_baby": add_cur["baby"], "spec_add_tot_cur": add_cur_total,
            "gross_prior": gross_prior, "thresh_prior": threshold_prior,
            "ss_pension_prior": pension_prior, "ss_med_prior": medical_prior, "ss_large_med_prior": large_med_prior,
            "ss_unemp_prior": unemp_prior, "hf_spec_prior": hf_prior, "spec_tot_prior": spec_ded_prior,
            "spec_add_child_prior": add_prior["child"], "spec_add_edu_prior": add_prior["edu"],
            "spec_add_med_prior": add_prior["med"], "spec_add_loan_prior": add_prior["loan"],
            "spec_add_rent_prior": add_prior["rent"], "spec_add_elder_prior": add_prior["elder"],
            "spec_add_baby_prior": add_prior["baby"], "spec_add_tot_prior": add_prior_total,
            "gross_all": gross_all, "thresh_all": threshold_all,
            "ss_pension_all": pension_all, "ss_med_all": medical_all, "ss_large_med_all": large_med_all,
            "ss_unemp_all": unemp_all, "hf_spec_all": hf_all, "spec_tot_all": spec_ded_all,
            "spec_add_child_all": add_all["child"], "spec_add_edu_all": add_all["edu"],
            "spec_add_med_all": add_all["med"], "spec_add_loan_all": add_all["loan"],
            "spec_add_rent_all": add_all["rent"], "spec_add_elder_all": add_all["elder"],
            "spec_add_baby_all": add_all["baby"], "spec_add_tot_all": add_all_total,
            "taxable_all": taxable_all, "taxable_income": taxable_all,
            "tax_rate": round(rate * 100, 2), "quick_deduct": quick, "quick_deduction": quick,
            "tax_calculated": cumulative_tax, "tax_relief": 0.0, "tax_paid_prior": paid_tax_prior,
            "tax_current": tax_cur, "current_tax": tax_cur, "tax_amount": tax_cur, "net_salary": net_cur,
            "prior_record_month_count": len(plist),
        }
        rows.append(row)

    numeric_keys = [
        "target_salary", "gross_salary", "thresh_cur", "hf_person", "ss_person", "deduct_cur_tot",
        "ss_pension", "ss_med", "ss_large_med", "ss_unemp", "hf_spec", "spec_tot_cur",
        "spec_add_child", "spec_add_edu", "spec_add_med", "spec_add_loan", "spec_add_rent", "spec_add_elder", "spec_add_baby", "spec_add_tot_cur",
        "gross_prior", "thresh_prior", "ss_pension_prior", "ss_med_prior", "ss_large_med_prior", "ss_unemp_prior", "hf_spec_prior", "spec_tot_prior",
        "spec_add_child_prior", "spec_add_edu_prior", "spec_add_med_prior", "spec_add_loan_prior", "spec_add_rent_prior", "spec_add_elder_prior", "spec_add_baby_prior", "spec_add_tot_prior",
        "gross_all", "thresh_all", "ss_pension_all", "ss_med_all", "ss_large_med_all", "ss_unemp_all", "hf_spec_all", "spec_tot_all",
        "spec_add_child_all", "spec_add_edu_all", "spec_add_med_all", "spec_add_loan_all", "spec_add_rent_all", "spec_add_elder_all", "spec_add_baby_all", "spec_add_tot_all",
        "taxable_all", "tax_calculated", "tax_relief", "tax_paid_prior", "tax_current", "net_salary",
    ]
    totals = {k: round(sum(flt(r.get(k)) for r in rows), 2) for k in numeric_keys}
    totals.update({
        "seq": "合计", "employee_no": f"共 {len(rows)} 人",
        "tax_threshold": totals.get("thresh_cur", 0.0),
        "hf_cur": totals.get("hf_person", 0.0), "ss_cur": totals.get("ss_person", 0.0),
        "deduct_cur": totals.get("deduct_cur_tot", 0.0),
        "special_deductions_total": totals.get("spec_add_tot_cur", 0.0),
        "taxable_income": totals.get("taxable_all", 0.0),
        "current_tax": totals.get("tax_current", 0.0), "tax_amount": totals.get("tax_current", 0.0),
    })

    return {
        "company": company,
        "period_month": period_month,
        "month_idx": cinfo["cur_month_index"],
        "cycle_start_month": cycle_start_m,
        "cycle_name": cinfo["cycle_name"],
        "tax_threshold": tax_thresh,
        "report_title": f"{period_month} 个人所得税申报台账（VBA同口径68列）",
        "rows": rows,
        "totals": totals,
    }


# ==========================================
# 6. 历史数据全员总览 (15列) 与单人 12 个月流水穿透服务
# ==========================================
@frappe.whitelist()
def get_all_employees_tax_history_summary(company="天津祺富机械加工有限公司", period_month="2026-07"):
    """全员累计个税历史总览；起征点以员工实际历史快照为准，避免新入职人员被多计月份。"""
    check_payroll_workbench_permission("read")
    params = get_effective_tax_parameters(company, period_month)
    cinfo = get_tax_cycle_info(period_month, params["tax_threshold"], params["tax_cycle_start_month"])
    cycle_months = []
    for offset in range(12):
        y, m = _add_month(cinfo["cycle_start_year"], cinfo["cycle_start_month"], offset)
        cycle_months.append(f"{y:04d}-{m:02d}")
    cycle_name = f"{cycle_months[0]} ~ {cycle_months[-1]}"

    employees = _salary_profiles_for_period(
        company, period_month,
        fields=["employee_no", "employee_name", "id_card", "gender", "employee_type", "salary_mode", "base_salary"],
        order_by="employee_no asc",
    )
    employees = [emp for emp in employees if is_tax_ledger_employee(emp.get("employee_type"))]
    cycle_parent_names = [f"{company}-{m}" for m in cycle_months]
    raw = frappe.get_all(
        "Ashan Monthly Payroll Item",
        filters={"parent": ["in", cycle_parent_names]},
        fields=[
            "employee_no", "employee_name", "parent", "gross_salary", "tax_threshold",
            "ss_person_total", "hf_person_total", "special_deductions_total",
            "taxable_income", "tax_amount", "net_salary",
        ], order_by="parent asc",
    )
    by_emp = {}
    for r in raw:
        parts = (r.parent or "").split("-")
        r["period_month"] = "-".join(parts[-2:]) if len(parts) >= 2 else ""
        if r["period_month"] <= period_month:
            by_emp.setdefault(r.employee_no, []).append(r)

    rows = []
    totals = {k: 0.0 for k in ["cum_gross","cum_thresh","cum_ss","cum_hf","cum_special_add","cum_taxable","cum_tax_paid","cum_net"]}
    for idx, emp in enumerate(employees, start=1):
        recs = by_emp.get(emp.employee_no, [])
        paid_months = [r.period_month for r in recs if flt(r.gross_salary) or flt(r.net_salary)]
        if paid_months:
            paid_desc = f"{len(paid_months)}个月 ({paid_months[0][2:].replace('-', '.')}~{paid_months[-1][2:].replace('-', '.')})"
        else:
            paid_desc = "0个月"
        cum_gross = round(sum(flt(r.gross_salary) for r in recs), 2)
        cum_thresh = round(sum((flt(r.tax_threshold) or params["tax_threshold"]) for r in recs), 2)
        cum_ss = round(sum(flt(r.ss_person_total) for r in recs), 2)
        cum_hf = round(sum(flt(r.hf_person_total) for r in recs), 2)
        cum_add = round(sum(flt(r.special_deductions_total) for r in recs), 2)
        cum_tax = round(sum(flt(r.tax_amount) for r in recs), 2)
        cum_net = round(sum(flt(r.net_salary) for r in recs), 2)
        cum_taxable = round(cum_gross - cum_ss - cum_hf - cum_add - cum_thresh, 2)
        rate = 0.03
        for _l, upper, rp, _q in TAX_BRACKETS:
            if cum_taxable <= upper:
                rate = rp
                break
        rows.append({
            "seq": idx, "employee_no": emp.employee_no, "employee_name": emp.employee_name,
            "id_card": emp.id_card or "-", "gender": emp.gender or "-", "employee_type": emp.employee_type or "正式工",
            "salary_mode": emp.salary_mode or "税后", "months_paid_count": len(paid_months), "months_paid_desc": paid_desc,
            "cum_gross_salary": cum_gross, "cum_tax_threshold": cum_thresh, "cum_ss_person": cum_ss,
            "cum_hf_person": cum_hf, "cum_ss_hf": round(cum_ss + cum_hf, 2),
            "cum_special_deductions": cum_add, "cum_taxable_income": cum_taxable,
            "tax_rate": round(rate * 100, 2), "cum_tax_paid": cum_tax, "cum_net_salary": cum_net,
        })
        totals["cum_gross"] += cum_gross; totals["cum_thresh"] += cum_thresh
        totals["cum_ss"] += cum_ss; totals["cum_hf"] += cum_hf; totals["cum_special_add"] += cum_add
        totals["cum_taxable"] += cum_taxable; totals["cum_tax_paid"] += cum_tax; totals["cum_net"] += cum_net
    totals = {k: round(v, 2) for k,v in totals.items()}
    totals["cum_ss_hf"] = round(totals["cum_ss"] + totals["cum_hf"], 2)
    return {
        "company": company, "period_month": period_month, "cycle_name": cycle_name,
        "cycle_months": cycle_months, "cur_month_idx": cinfo["cur_month_index"],
        "rows": rows, "totals": totals,
    }



def _history_full_columns():
    """VBA 68 calculation columns plus ERP audit metadata."""
    calc_columns = [
        ("seq", "序号", "text", "员工基本信息"), ("employee_no", "工号", "text", "员工基本信息"),
        ("employee_name", "姓名", "name", "员工基本信息"), ("id_card", "证件号码", "text", "员工基本信息"),
        ("gender", "性别", "text", "员工基本信息"), ("period_month_str", "本期所属期", "text", "员工基本信息"),
        ("employee_type", "员工类型", "text", "员工基本信息"), ("target_salary", "目标工资", "money", "员工基本信息"),
        ("salary_mode", "工资类型", "text", "员工基本信息"),
        ("gross_salary", "税前工资", "money", "工资扣除(本月)"), ("thresh_cur", "起征点扣除", "money", "工资扣除(本月)"),
        ("hf_person", "公积金", "money", "工资扣除(本月)"), ("ss_person", "社保", "money", "工资扣除(本月)"),
        ("deduct_cur_tot", "工资扣除合计", "money", "工资扣除(本月)"),
        ("ss_pension", "基本养老", "money", "专项扣除(本月)"), ("ss_med", "基本医疗", "money", "专项扣除(本月)"),
        ("ss_large_med", "大额医疗", "money", "专项扣除(本月)"), ("ss_unemp", "失业保险", "money", "专项扣除(本月)"),
        ("hf_spec", "住房公积金", "money", "专项扣除(本月)"), ("spec_tot_cur", "专项扣除合计", "money", "专项扣除(本月)"),
        ("spec_add_child", "子女教育", "money", "专项附加扣除(本月)"), ("spec_add_edu", "继续教育", "money", "专项附加扣除(本月)"),
        ("spec_add_med", "大病医疗", "money", "专项附加扣除(本月)"), ("spec_add_loan", "住房贷款利息", "money", "专项附加扣除(本月)"),
        ("spec_add_rent", "住房租金", "money", "专项附加扣除(本月)"), ("spec_add_elder", "赡养老人", "money", "专项附加扣除(本月)"),
        ("spec_add_baby", "3岁以下婴幼儿照护", "money", "专项附加扣除(本月)"), ("spec_add_tot_cur", "专项附加扣除合计", "money", "专项附加扣除(本月)"),
        ("gross_prior", "税前工资(往期)", "money", "个税累计(往期)"), ("thresh_prior", "起征点扣除(往期)", "money", "个税累计(往期)"),
        ("ss_pension_prior", "基本养老(往期)", "money", "个税累计(往期)"), ("ss_med_prior", "基本医疗(往期)", "money", "个税累计(往期)"),
        ("ss_large_med_prior", "大额医疗(往期)", "money", "个税累计(往期)"), ("ss_unemp_prior", "失业保险(往期)", "money", "个税累计(往期)"),
        ("hf_spec_prior", "住房公积金(往期)", "money", "个税累计(往期)"), ("spec_tot_prior", "专项扣除合计(往期)", "money", "个税累计(往期)"),
        ("spec_add_child_prior", "子女教育(往期)", "money", "个税累计(往期)"), ("spec_add_edu_prior", "继续教育(往期)", "money", "个税累计(往期)"),
        ("spec_add_med_prior", "大病医疗(往期)", "money", "个税累计(往期)"), ("spec_add_loan_prior", "住房贷款利息(往期)", "money", "个税累计(往期)"),
        ("spec_add_rent_prior", "住房租金(往期)", "money", "个税累计(往期)"), ("spec_add_elder_prior", "赡养老人(往期)", "money", "个税累计(往期)"),
        ("spec_add_baby_prior", "3岁以下婴幼儿照护(往期)", "money", "个税累计(往期)"), ("spec_add_tot_prior", "专项附加扣除合计(往期)", "money", "个税累计(往期)"),
        ("gross_all", "税前工资(全部)", "money", "个税累计(全部)"), ("thresh_all", "起征点扣除(全部)", "money", "个税累计(全部)"),
        ("ss_pension_all", "基本养老(全部)", "money", "个税累计(全部)"), ("ss_med_all", "基本医疗(全部)", "money", "个税累计(全部)"),
        ("ss_large_med_all", "大额医疗(全部)", "money", "个税累计(全部)"), ("ss_unemp_all", "失业保险(全部)", "money", "个税累计(全部)"),
        ("hf_spec_all", "住房公积金(全部)", "money", "个税累计(全部)"), ("spec_tot_all", "专项扣除合计(全部)", "money", "个税累计(全部)"),
        ("spec_add_child_all", "子女教育(全部)", "money", "个税累计(全部)"), ("spec_add_edu_all", "继续教育(全部)", "money", "个税累计(全部)"),
        ("spec_add_med_all", "大病医疗(全部)", "money", "个税累计(全部)"), ("spec_add_loan_all", "住房贷款利息(全部)", "money", "个税累计(全部)"),
        ("spec_add_rent_all", "住房租金(全部)", "money", "个税累计(全部)"), ("spec_add_elder_all", "赡养老人(全部)", "money", "个税累计(全部)"),
        ("spec_add_baby_all", "3岁以下婴幼儿照护(全部)", "money", "个税累计(全部)"), ("spec_add_tot_all", "专项附加扣除合计(全部)", "money", "个税累计(全部)"),
        ("taxable_all", "应纳税所得额", "money", "税款计算"), ("tax_rate", "税率", "percent", "税款计算"),
        ("quick_deduct", "速算扣除数", "money", "税款计算"), ("tax_calculated", "累计应纳税额", "money", "税款计算"),
        ("tax_relief", "减免税额", "money", "税款计算"), ("tax_paid_prior", "往期已缴税额", "money", "税款计算"),
        ("tax_current", "本月个税", "money", "税款计算"), ("net_salary", "税后工资", "money", "税款计算"),
    ]
    assert len(calc_columns) == 68
    audit_columns = [
        ("history_lock_status", "账期状态", "status", "ERP审计"),
        ("calculation_status", "计算状态", "calc_status", "ERP审计"),
        ("calculation_trigger_source", "触发来源", "text", "ERP审计"),
        ("calculation_requested_at", "请求时间", "datetime", "ERP审计"),
        ("calculation_started_at", "开始时间", "datetime", "ERP审计"),
        ("calculation_completed_at", "完成时间", "datetime", "ERP审计"),
        ("calculation_engine_version", "计算引擎", "text", "ERP审计"),
        ("calculation_task_id", "任务号", "text", "ERP审计"),
        ("modified", "快照修改时间", "datetime", "ERP审计"),
    ]
    return calc_columns + audit_columns


@frappe.whitelist()
def get_history_full_ledger(company="天津祺富机械加工有限公司", period_month="2026-07", history_period_month=None, employee_no=None):
    """Historical VBA-68 snapshot + ERP audit fields for one selected month or all months in the tax cycle."""
    check_payroll_workbench_permission("read")
    params = get_effective_tax_parameters(company, period_month)
    cinfo = get_tax_cycle_info(period_month, params["tax_threshold"], params["tax_cycle_start_month"])
    cycle_months = []
    for offset in range(12):
        y, m = _add_month(cinfo["cycle_start_year"], cinfo["cycle_start_month"], offset)
        cycle_months.append(f"{y:04d}-{m:02d}")
    available_cycle_months = [m for m in cycle_months if m <= str(period_month)]
    selected = str(history_period_month or period_month).strip()
    emp_no_filter = str(employee_no or "").strip()

    if selected != "all" and selected not in available_cycle_months:
        frappe.throw(f"历史月份 {selected} 不在当前已发生申报周期 {available_cycle_months[0]} ~ {available_cycle_months[-1]} 内。")

    months_to_process = available_cycle_months if selected == "all" else [selected]

    all_rows = []
    audit_supported = frappe.get_meta("Ashan Monthly Payroll Item").has_field("calculation_status")
    audit_fields = ["employee_no", "modified"]
    if audit_supported:
        audit_fields += [
            "calculation_status", "calculation_trigger_source", "calculation_requested_at", "calculation_started_at",
            "calculation_completed_at", "calculation_engine_version", "calculation_task_id",
        ]

    for m in months_to_process:
        parent_name = f"{company}-{m}"
        if not frappe.db.exists("Ashan Monthly Payroll Settlement", parent_name):
            continue
        parent = frappe.get_doc("Ashan Monthly Payroll Settlement", parent_name)
        data = get_tax_settlement_full_sheet(company, m)
        audits = frappe.get_all(
            "Ashan Monthly Payroll Item", filters={"parent": parent_name}, fields=audit_fields, order_by="idx asc"
        )
        audit_by_emp = {r.employee_no: r for r in audits}
        lock_status = "已冻结" if (cint(parent.locked) or parent.status in ["已核定锁定", "已归档发放"]) else (parent.status or "草稿")
        
        m_rows = data.get("rows", [])
        if emp_no_filter:
            m_rows = [r for r in m_rows if str(r.get("employee_no")) == emp_no_filter]

        for row in m_rows:
            audit = audit_by_emp.get(row.get("employee_no"), {})
            row["history_lock_status"] = lock_status
            row["history_period_month"] = m
            for field in [
                "calculation_status", "calculation_trigger_source", "calculation_requested_at", "calculation_started_at",
                "calculation_completed_at", "calculation_engine_version", "calculation_task_id", "modified",
            ]:
                row[field] = audit.get(field) if audit else None
            all_rows.append(row)

    # 重建连续序号
    for idx, r in enumerate(all_rows, start=1):
        r["seq"] = idx

    numeric_keys = [
        "target_salary", "gross_salary", "thresh_cur", "hf_person", "ss_person", "deduct_cur_tot",
        "ss_pension", "ss_med", "ss_large_med", "ss_unemp", "hf_spec", "spec_tot_cur",
        "spec_add_child", "spec_add_edu", "spec_add_med", "spec_add_loan", "spec_add_rent", "spec_add_elder", "spec_add_baby", "spec_add_tot_cur",
        "gross_prior", "thresh_prior", "ss_pension_prior", "ss_med_prior", "ss_large_med_prior", "ss_unemp_prior", "hf_spec_prior", "spec_tot_prior",
        "spec_add_child_prior", "spec_add_edu_prior", "spec_add_med_prior", "spec_add_loan_prior", "spec_add_rent_prior", "spec_add_elder_prior", "spec_add_baby_prior", "spec_add_tot_prior",
        "gross_all", "thresh_all", "ss_pension_all", "ss_med_all", "ss_large_med_all", "ss_unemp_all", "hf_spec_all", "spec_tot_all",
        "spec_add_child_all", "spec_add_edu_all", "spec_add_med_all", "spec_add_loan_all", "spec_add_rent_all", "spec_add_elder_all", "spec_add_baby_all", "spec_add_tot_all",
        "taxable_all", "tax_calculated", "tax_relief", "tax_paid_prior", "tax_current", "net_salary",
    ]
    totals = {k: round(sum(flt(r.get(k)) for r in all_rows), 2) for k in numeric_keys}
    totals.update({
        "seq": "合计", "employee_no": f"共 {len(all_rows)} 条",
        "tax_threshold": totals.get("thresh_cur", 0.0),
        "hf_cur": totals.get("hf_person", 0.0), "ss_cur": totals.get("ss_person", 0.0),
        "deduct_cur": totals.get("deduct_cur_tot", 0.0),
        "special_deductions_total": totals.get("spec_add_tot_cur", 0.0),
        "taxable_income": totals.get("taxable_all", 0.0),
        "current_tax": totals.get("tax_current", 0.0), "tax_amount": totals.get("tax_current", 0.0),
    })

    is_single_month = (selected != "all")
    parent_doc = frappe.get_doc("Ashan Monthly Payroll Settlement", f"{company}-{selected}") if is_single_month and frappe.db.exists("Ashan Monthly Payroll Settlement", f"{company}-{selected}") else None

    return {
        "company": company, "period_month": period_month, "history_period_month": selected,
        "employee_no": emp_no_filter,
        "cycle_name": cinfo.get("cycle_name", f"{cycle_months[0]} ~ {cycle_months[-1]}"),
        "cycle_months": cycle_months, "available_cycle_months": available_cycle_months,
        "columns": [dict(key=k,label=l,type=t,group=g) for k,l,t,g in _history_full_columns()],
        "rows": all_rows, "totals": totals,
        "locked": bool(cint(parent_doc.locked)) if parent_doc else False,
        "status": parent_doc.status if parent_doc else ("全部月份合并" if selected == "all" else "草稿"),
    }


@frappe.whitelist(methods=["POST"])
def save_history_payroll_input_correction(
    company, current_period_month, history_period_month, employee_no, correction_json
):
    """Correct authoritative historical input fields in an unlocked month and cascade recalculation forward."""
    check_payroll_workbench_permission("write")
    history_period_month = str(history_period_month or "").strip()
    current_period_month = str(current_period_month or "").strip()
    if not re.match(r"^\d{4}-\d{2}$", history_period_month) or not re.match(r"^\d{4}-\d{2}$", current_period_month):
        frappe.throw("账期格式必须为 YYYY-MM。")
    if history_period_month > current_period_month:
        frappe.throw("历史更正月份不能晚于当前核算月份。")
    parent_name = f"{company}-{history_period_month}"
    if not frappe.db.exists("Ashan Monthly Payroll Settlement", parent_name):
        frappe.throw("该历史月份尚未建账，无法直接更正。")
    parent = frappe.get_doc("Ashan Monthly Payroll Settlement", parent_name)
    if parent.locked or parent.status in ["已核定锁定", "已归档发放"]:
        frappe.throw("该历史月份已经冻结。请先执行反审核解锁，再进行历史输入更正。")
    payload = json.loads(correction_json) if isinstance(correction_json, str) else (correction_json or {})
    item = next((row for row in parent.items if row.employee_no == employee_no), None)
    if not item:
        frappe.throw(f"{history_period_month} 未找到员工 {employee_no} 的历史薪酬记录。")

    # 工资只允许修改“计薪方式对应的权威侧”：税后模式改目标税后，税前模式改税前工资。
    # 防止 API 调用同时写入税前/税后两个互相矛盾的输入。
    salary_mode = str(item.salary_mode or "税后").strip()
    salary_input_field = "net_salary" if is_tax_after_salary_mode(salary_mode) else "gross_salary"
    allowed_money = [
        salary_input_field, "ss_person_total", "hf_person_total",
        "deduction_child_education", "deduction_continuing_education", "deduction_serious_illness",
        "deduction_housing_loan", "deduction_housing_rent", "deduction_elderly_care", "deduction_infant_care",
    ]
    for field in allowed_money:
        if field in payload:
            value = round(flt(payload.get(field)), 2)
            if value < 0:
                frappe.throw(f"历史输入字段 {field} 不能为负数。")
            setattr(item, field, value)
    item.special_deductions_total = round(sum(flt(getattr(item, f, 0)) for f in [
        "deduction_child_education", "deduction_continuing_education", "deduction_serious_illness",
        "deduction_housing_loan", "deduction_housing_rent", "deduction_elderly_care", "deduction_infant_care",
    ]), 2)
    _set_payroll_item_calculation_audit(item, "待计算", "历史数据更正")
    frappe.flags.ignore_lock = True
    parent.save(ignore_permissions=True)
    from ashan_cn_procurement.services.payroll_recalculation_service import queue_recalculation_after_change
    task = queue_recalculation_after_change(
        company=company,
        period_month=current_period_month,
        employee_no=employee_no,
        trigger_source="历史数据更正",
        start_period=history_period_month,
        trigger_detail=f"更正 {history_period_month} 历史薪酬输入，级联重算至 {current_period_month}",
        force_recompute=True,
    )
    return {
        "success": True,
        "message": f"{history_period_month} 历史输入已保存，{employee_no} 已进入 {history_period_month} → {current_period_month} 级联后台重算队列。",
        "task_name": task.name if task else "",
    }


@frappe.whitelist()
def get_employee_tax_history_timeline(company="天津祺富机械加工有限公司", employee_no="A0001", period_month="2026-07"):
    """单人12个月税务轨迹；累计起征点只在存在该员工历史记录的月份增加。"""
    check_payroll_workbench_permission("read")
    params = get_effective_tax_parameters(company, period_month)
    cinfo = get_tax_cycle_info(period_month, params["tax_threshold"], params["tax_cycle_start_month"])
    cycle_months = []
    for offset in range(12):
        y, m = _add_month(cinfo["cycle_start_year"], cinfo["cycle_start_month"], offset)
        cycle_months.append(f"{y:04d}-{m:02d}")

    emp_doc = frappe.db.get_value(
        "Ashan Employee Salary Profile", {"company": company, "employee_no": employee_no},
        ["employee_name", "id_card", "gender", "employee_type", "base_salary"], as_dict=True,
    ) or {}
    parent_names = [f"{company}-{m}" for m in cycle_months]
    recs = frappe.get_all(
        "Ashan Monthly Payroll Item", filters={"parent": ["in", parent_names], "employee_no": employee_no},
        fields=["parent", "gross_salary", "tax_threshold", "ss_person_total", "hf_person_total", "special_deductions_total", "tax_amount", "net_salary"],
        order_by="parent asc",
    )
    rec_map = {}
    for r in recs:
        parts = (r.parent or "").split("-")
        pm = "-".join(parts[-2:]) if len(parts) >= 2 else ""
        r["period_month"] = pm
        r["status"] = frappe.db.get_value("Ashan Monthly Payroll Settlement", r.parent, "status") or "已核定锁定"
        rec_map[pm] = r

    rows = []
    cum_gross = cum_ss = cum_hf = cum_add = cum_tax_paid = cum_net = cum_thresh = 0.0
    for idx, m_str in enumerate(cycle_months, start=1):
        rec = rec_map.get(m_str)
        is_cur = (m_str == period_month)
        is_future = (m_str > period_month)
        if rec and not is_future:
            gross = flt(rec.gross_salary); ss = flt(rec.ss_person_total); hf = flt(rec.hf_person_total)
            add = flt(rec.special_deductions_total); tax = flt(rec.tax_amount); net = flt(rec.net_salary)
            paid_prior = cum_tax_paid
            cum_gross += gross; cum_ss += ss; cum_hf += hf; cum_add += add; cum_tax_paid += tax; cum_net += net
            cum_thresh += flt(rec.tax_threshold) or params["tax_threshold"]
            taxable = round(cum_gross - cum_ss - cum_hf - cum_add - cum_thresh, 2)
            rate, quick = 0.03, 0.0
            for _l, upper, rp, q in TAX_BRACKETS:
                if taxable <= upper:
                    rate, quick = rp, q
                    break
            rows.append({
                "seq": idx, "period_month": m_str, "status": rec.status, "is_current": is_cur, "is_future": False,
                "gross_salary": gross, "ss_person_total": ss, "hf_person_total": hf, "insurance_total": round(ss+hf,2),
                "special_deductions_total": add, "threshold_accumulated": round(cum_thresh,2),
                "taxable_accumulated": taxable, "tax_rate": round(rate*100,2), "quick_deduction": quick,
                "tax_current": tax, "tax_paid_prior": round(paid_prior,2), "net_salary": net,
            })
        else:
            rows.append({
                "seq": idx, "period_month": m_str, "status": "未开启" if is_future else "未建账",
                "is_current": is_cur, "is_future": is_future, "gross_salary": 0.0, "ss_person_total": 0.0,
                "hf_person_total": 0.0, "insurance_total": 0.0, "special_deductions_total": 0.0,
                "threshold_accumulated": round(cum_thresh,2), "taxable_accumulated": round(cum_gross-cum_ss-cum_hf-cum_add-cum_thresh,2),
                "tax_rate": 0.0, "quick_deduction": 0.0, "tax_current": 0.0,
                "tax_paid_prior": round(cum_tax_paid,2), "net_salary": 0.0,
            })
    return {
        "company": company, "employee_no": employee_no, "employee_name": emp_doc.get("employee_name") or employee_no,
        "id_card": emp_doc.get("id_card") or "-", "gender": emp_doc.get("gender") or "-",
        "employee_type": emp_doc.get("employee_type") or "", "period_month": period_month,
        "cycle_name": cinfo["cycle_name"], "cycle_months": cycle_months, "rows": rows,
        "summary": {
            "cum_gross_salary": round(cum_gross,2), "cum_tax_threshold": round(cum_thresh,2),
            "cum_ss_person": round(cum_ss,2), "cum_hf_person": round(cum_hf,2),
            "cum_special_deductions": round(cum_add,2), "cum_tax_paid": round(cum_tax_paid,2), "cum_net_salary": round(cum_net,2),
        },
    }


@frappe.whitelist()
def export_qifu_payroll_excel(company="天津祺富机械加工有限公司", period_month="2026-07", sheet_type="all", tax_view_mode="simple", history_mode="all", history_emp_no="A0001", history_period_month=None):
	"""
	导出专业级 Excel 报表 (.xlsx)：
	1. distribution: 24 列外部薪资实发表
	2. accounting: 11 列记账工资表
	3. insurance: 19 列双层表头社保缴费明细表
	4. housing_fund: 12 列双层表头公积金明细表
	5. tax: 个人所得税表 (根据 tax_view_mode 动态支持 17 列精简版 或 VBA 68列完整核算台账)
	6. history: 历史数据表 (支持全员15列总览、单人12个月穿透、指定历史月VBA68列+ERP审计)
	7. all: 包含上述全部 7 个工作表的完整年度薪资结算财务工作簿
	"""
	check_payroll_workbench_permission("read")
	period_month = _normalize_period_month(period_month)
	if history_period_month:
		history_period_month = _normalize_period_month(history_period_month)
	import io
	import base64
	from datetime import datetime, date
	import openpyxl
	from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
	from openpyxl.utils import get_column_letter
	from openpyxl.worksheet.pagebreak import Break

	wb = openpyxl.Workbook()
	ws_default = wb.active

	# 通用极简专业样式定义：无背景色、标题行加粗、外框加粗、内部细边框
	font_title = Font(name="Microsoft YaHei", size=15, bold=True, color="000000")
	font_header = Font(name="Microsoft YaHei", size=11, bold=True, color="000000")
	font_data = Font(name="Microsoft YaHei", size=10, color="000000")
	font_total = Font(name="Microsoft YaHei", size=10.5, bold=True, color="000000")

	side_thin = Side(border_style="thin", color="000000")
	side_medium = Side(border_style="medium", color="000000")
	side_double = Side(border_style="double", color="000000")

	align_center = Alignment(horizontal="center", vertical="center", shrink_to_fit=True)
	align_left = Alignment(horizontal="left", vertical="center", shrink_to_fit=True)
	align_right = Alignment(horizontal="right", vertical="center", shrink_to_fit=True)

	def apply_clean_table_borders(ws, start_row, end_row, start_col, end_col, split_col=None, total_row=None):
		"""标题行和工资行外框加粗，内部边框普通，合计行底部双细线。"""
		for r in range(start_row, end_row + 1):
			is_header = (r == start_row)
			is_total = (total_row is not None and r == total_row)
			for c in range(start_col, end_col + 1):
				top = side_medium if is_header else side_thin
				if is_total:
					top = side_thin
				bottom = side_medium if is_header else (side_double if is_total else (side_medium if r == end_row else side_thin))
				left = side_medium if (c == start_col or (split_col and c == split_col + 1)) else side_thin
				right = side_medium if (c == end_col or (split_col and c == split_col)) else side_thin
				ws.cell(row=r, column=c).border = Border(left=left, right=right, top=top, bottom=bottom)

	# -------------------------------------------------------------
	# 1. 24 列薪资发放表 (车间实发 + 考勤工时 + 备考后直接呈现现金点钞辅助列)
	# -------------------------------------------------------------
	def build_distribution_sheet(ws):
		"""Build the 24-column external pay sheet plus visible cash note helpers right after Remarks."""
		ws.title = "1.薪资发放表(24列)"
		data_res = get_salary_distribution_sheet(company, period_month)
		rows = data_res.get("rows", [])
		totals = data_res.get("totals", {})

		# 第 1 行：主标题（加粗，无背景色，跨度匹配打印区域 A1:X1）
		ws.merge_cells("A1:X1")
		ws["A1"] = f"{company} {period_month} 薪资发放表"
		ws["A1"].font = font_title
		ws["A1"].alignment = align_center
		ws.row_dimensions[1].height = 42

		# 第 2 行：表头（加粗，无背景色，删去原统计副标题行）
		headers = [
			"序号", "工号", "姓名", "作业天数", "作业小时", "天工资", "小时工资",
			"全勤费", "加班小时", "加班费", "国勤天数", "国勤工资", "达标率",
			"达标工资", "扣除", "考勤绩效工资合计", "职位补贴", "房/车补",
			"补贴工资合计", "应发工资合计", "工资调整", "实发工资合计", "签字", "备考",
		]
		cash_headers = ["100 元", "50 元", "10 元", "5 元", "1 元", "现金合计", "核定"]
		ws.row_dimensions[2].height = 26
		for col_idx, h in enumerate(headers, start=1):
			cell = ws.cell(row=2, column=col_idx, value=h)
			cell.font = font_header
			cell.alignment = align_center
		for col_idx, h in enumerate(cash_headers, start=25):
			cell = ws.cell(row=2, column=col_idx, value=h)
			cell.font = font_header
			cell.alignment = align_center

		# 第 3 行起：员工数据行（无背景色，公式计算）
		for row_idx, r in enumerate(rows, start=3):
			ws.row_dimensions[row_idx].height = 22
			vals = [
				r.get("seq"), r.get("employee_no"), r.get("employee_name"),
				r.get("work_days", 0) or None, r.get("work_hours", 0) or None, r.get("day_salary", 0) or None, r.get("hour_salary", 0) or None,
				r.get("full_attendance", 0) or None, r.get("overtime_hours", 0) or None, r.get("overtime_salary", 0) or None,
				r.get("national_days", 0) or None, r.get("national_salary", 0) or None, r.get("target_rate", "") or None,
				r.get("target_salary", 0) or None, r.get("deduction", 0) or None,
				f"=F{row_idx}+G{row_idx}+H{row_idx}+J{row_idx}+L{row_idx}+N{row_idx}-O{row_idx}",
				r.get("post_allowance", 0) or None, r.get("house_rent_allowance", 0) or None,
				f"=Q{row_idx}+R{row_idx}",
				f"=P{row_idx}+S{row_idx}",
				r.get("salary_adjust", 0) or None,
				f"=ROUND(T{row_idx}+U{row_idx},0)",
				r.get("sign", ""), r.get("remarks", ""),
				r.get("cash_100", 0), r.get("cash_50", 0), r.get("cash_10", 0), r.get("cash_5", 0), r.get("cash_1", 0),
				f"=100*Y{row_idx}+50*Z{row_idx}+10*AA{row_idx}+5*AB{row_idx}+1*AC{row_idx}",
				f"=ROUND(V{row_idx},0)-AD{row_idx}",
			]
			for col_idx, val in enumerate(vals, start=1):
				cell = ws.cell(row=row_idx, column=col_idx, value=val)
				cell.font = font_data
				if col_idx in [1, 2, 13, 23, 25, 26, 27, 28, 29, 31]:
					cell.alignment = align_center
				elif col_idx in [3, 24]:
					cell.alignment = align_left
				else:
					cell.alignment = align_right
				if col_idx in [25, 26, 27, 28, 29]:
					cell.number_format = "0"
				elif col_idx == 31:
					cell.number_format = "0"
				elif col_idx in [4, 5, 9, 11]:
					if isinstance(val, (int, float)):
						cell.number_format = "#,##0.0"
				else:
					if isinstance(val, (int, float)) or (isinstance(val, str) and val.startswith("=")):
						cell.number_format = "#,##0.00"

		tot_row = len(rows) + 3
		ws.row_dimensions[tot_row].height = 24
		ws.cell(row=tot_row, column=1, value="合计").alignment = align_center
		ws.cell(row=tot_row, column=2, value=f"共 {len(rows)} 人").alignment = align_center
		if rows:
			for col_idx in (4, 5, 8, 10, 11, 12, 14, 15, 16, 17, 18, 19, 20, 21, 22):
				col_letter = get_column_letter(col_idx)
				ws.cell(row=tot_row, column=col_idx, value=f"=SUM({col_letter}3:{col_letter}{tot_row - 1})")
			for col_idx in (25, 26, 27, 28, 29, 30, 31):
				col_letter = get_column_letter(col_idx)
				ws.cell(row=tot_row, column=col_idx, value=f"=SUM({col_letter}3:{col_letter}{tot_row - 1})")
		else:
			for col_idx in range(4, 32):
				ws.cell(row=tot_row, column=col_idx, value=0)

		for col_idx in range(1, 32):
			c = ws.cell(row=tot_row, column=col_idx)
			c.font = font_total
			if col_idx in [25, 26, 27, 28, 29, 31]:
				c.alignment = align_center
				c.number_format = "0"
			elif col_idx in [1, 2]:
				c.alignment = align_center
			else:
				c.alignment = align_right
				c.number_format = "#,##0.00"

		# 设置列宽（姓名列加宽）
		dist_widths = [5.625, 8.625, 13.5, 8.625, 13.0, 9.125, 13.0, 13.0, 13.0, 13.0, 13.0, 13.0, 8.625, 13.0, 9.125, 12.625, 8.625, 13.0, 12.625, 13.0, 9.125, 13.625, 16.625, 14.625, 10.625, 9.0, 9.0, 9.0, 9.0, 12.0, 10.625]
		for idx, w in enumerate(dist_widths, start=1):
			ws.column_dimensions[get_column_letter(idx)].width = w
			ws.column_dimensions[get_column_letter(idx)].hidden = False

		# 应用外框加粗、内部普通边框设计
		apply_clean_table_borders(ws, start_row=2, end_row=tot_row, start_col=1, end_col=31, split_col=24, total_row=tot_row)

		ws.freeze_panes = "D3"

		# 顶端标题行重复（第 1~2 行），打印每页均显示主标题与表头
		ws.print_title_rows = "1:2"

		# 打印范围严格限定为第 1 列（序号）至备考列（A:X）
		ws.print_area = f"A1:X{tot_row}"

		# 每个人员独立分页：在每名员工数据行后插入分页符
		if rows:
			for r_break_idx in range(3, len(rows) + 3):
				ws.row_breaks.append(Break(id=r_break_idx))

		ws.page_setup.paperSize = ws.PAPERSIZE_A4
		ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
		ws.page_setup.fitToWidth = 1
		ws.page_setup.fitToHeight = 0
		ws.sheet_properties.pageSetUpPr.fitToPage = True
		ws.page_margins.left = 0.3
		ws.page_margins.right = 0.3
		ws.page_margins.top = 0.5
		ws.page_margins.bottom = 0.5
		ws.page_margins.header = 0.3
		ws.page_margins.footer = 0.3
		ws.sheet_view.showGridLines = True

	# -------------------------------------------------------------
	# 2. 11 列记账工资表 (严格参照 XLSM；表格全部缩小以填充，姓名加宽，极简无背景)
	# -------------------------------------------------------------
	def build_accounting_sheet(ws, foreign=False):
		data_res = get_accounting_payroll_sheet(company, period_month)
		rows = data_res.get("foreign_rows" if foreign else "rows", [])
		totals = data_res.get("foreign_totals" if foreign else "totals", {})
		ws.title = "记账工资表外籍" if foreign else "2.记账工资表(11列)"

		# 第 1 行：主标题（加粗，无背景色，跨度 A1:K1）
		ws.merge_cells("A1:K1")
		ws["A1"] = f"{company} {period_month} {'记账工资表外籍' if foreign else '记账工资表'}"
		ws["A1"].font = font_title
		ws["A1"].alignment = align_center
		ws.row_dimensions[1].height = 42

		# 第 2 行：表头（加粗，无背景色）
		headers = [
			"工号", "姓名", "基本绩效工资", "职位补贴", "房/车补", "税前工资",
			"公积金", "社保", "应补/退税额", "合计扣除", "税后工资合计",
		]
		ws.row_dimensions[2].height = 26
		for col_idx, h in enumerate(headers, start=1):
			cell = ws.cell(row=2, column=col_idx, value=h)
			cell.font = font_header
			cell.alignment = align_center

		# 第 3 行起：员工数据行
		for row_idx, r in enumerate(rows, start=3):
			ws.row_dimensions[row_idx].height = 22
			vals = [
				r.get("employee_no"),
				r.get("employee_name"),
				f"=F{row_idx}-D{row_idx}-E{row_idx}",
				r.get("post_allowance", 0) or None,
				r.get("house_rent_allowance", 0) or None,
				r.get("gross_salary", 0) or 0,
				r.get("hf_person_total", 0) or None,
				r.get("ss_person_total", 0) or None,
				r.get("tax_amount", 0) or 0,
				f"=SUM(G{row_idx}:I{row_idx})",
				f"=F{row_idx}-J{row_idx}",
			]
			for col_idx, val in enumerate(vals, start=1):
				cell = ws.cell(row=row_idx, column=col_idx, value=val)
				cell.font = font_data
				if col_idx == 1:
					cell.alignment = align_center
				elif col_idx == 2:
					cell.alignment = align_left
				else:
					cell.alignment = align_right
				if isinstance(val, (int, float)) or (isinstance(val, str) and val.startswith("=")):
					cell.number_format = "#,##0.00"

		tot_row = len(rows) + 3
		ws.row_dimensions[tot_row].height = 24
		ws.cell(row=tot_row, column=1, value="合计").alignment = align_center
		ws.cell(row=tot_row, column=2, value=f"共 {len(rows)} 人").alignment = align_center
		if rows:
			for col_idx in range(3, 12):
				col_letter = get_column_letter(col_idx)
				ws.cell(row=tot_row, column=col_idx, value=f"=SUM({col_letter}3:{col_letter}{tot_row - 1})")
		else:
			for col_idx in range(3, 12):
				ws.cell(row=tot_row, column=col_idx, value=0)

		for col_idx in range(1, 12):
			c = ws.cell(row=tot_row, column=col_idx)
			c.font = font_total
			if col_idx in [1, 2]:
				c.alignment = align_center
			else:
				c.alignment = align_right
				c.number_format = "#,##0.00"

		# 列宽设置（姓名稍微宽一点点 14.0，参考 XLSM 比例）
		acc_widths = [8.625, 14.0, 11.625, 8.625, 13.0, 11.625, 9.5, 10.5, 9.5, 10.5, 11.625]
		for idx, w in enumerate(acc_widths, start=1):
			ws.column_dimensions[get_column_letter(idx)].width = w

		# 应用外框加粗、内部普通边框设计
		apply_clean_table_borders(ws, start_row=2, end_row=tot_row, start_col=1, end_col=11, total_row=tot_row)

		ws.freeze_panes = "C3"

		# 顶端标题行重复（第 1~2 行）
		ws.print_title_rows = "1:2"
		ws.print_area = f"A1:K{tot_row}"
		ws.page_setup.paperSize = ws.PAPERSIZE_A4
		ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
		ws.page_setup.fitToWidth = 1
		ws.page_setup.fitToHeight = 0
		ws.sheet_properties.pageSetUpPr.fitToPage = True
		ws.page_margins.left = 0.3
		ws.page_margins.right = 0.3
		ws.page_margins.top = 0.5
		ws.page_margins.bottom = 0.5
		ws.page_margins.header = 0.3
		ws.page_margins.footer = 0.3
		ws.sheet_view.showGridLines = True

	def build_cash_count_sheet(ws):
		"""Build the cash-note counting/check sheet from the reference XLSM helper columns."""
		ws.title = "现金点钞核定表"
		data_res = get_salary_cash_count_sheet(company, period_month)
		rows = data_res.get("rows", [])
		totals = data_res.get("totals", {})
		ws.merge_cells("A1:J1")
		ws["A1"] = f"{company} {period_month} 现金点钞核定表"
		ws["A1"].font = font_title
		ws["A1"].alignment = align_center
		ws.row_dimensions[1].height = 30
		ws.merge_cells("A2:J2")
		ws["A2"] = "参照《当月发薪工资表》隐藏辅助列：100/50/10/5/1 元张数；核定 = ROUND(实发工资,0) - 现金合计"
		ws["A2"].font = font_sub
		ws["A2"].alignment = align_center
		headers = ["序号", "工号", "姓名", "100 元", "50 元", "10 元", "5 元", "1 元", "现金合计", "核定"]
		for col_idx, h in enumerate(headers, start=1):
			c = ws.cell(row=3, column=col_idx, value=h)
			c.font = font_header
			c.fill = fill_success if h == "核定" else (fill_info if "元" in h or h == "现金合计" else fill_header)
			c.alignment = align_center
			c.border = border_cell
		for row_idx, r in enumerate(rows, start=4):
			vals = [r.get("seq"), r.get("employee_no"), r.get("employee_name"), r.get("cash_100",0), r.get("cash_50",0), r.get("cash_10",0), r.get("cash_5",0), r.get("cash_1",0), r.get("cash_total",0), r.get("cash_check",0)]
			for col_idx, val in enumerate(vals, start=1):
				c = ws.cell(row=row_idx, column=col_idx, value=val)
				c.font = font_data
				c.border = border_cell
				c.alignment = align_left if col_idx == 3 else align_center
				if col_idx in [9,10]: c.number_format = "#,##0.00"
		tot_row = len(rows) + 4
		ws.cell(row=tot_row, column=1, value="合计")
		ws.cell(row=tot_row, column=2, value=f"共 {len(rows)} 人")
		for col_idx, key in enumerate(("cash_100","cash_50","cash_10","cash_5","cash_1","cash_total","cash_check"), start=4):
			ws.cell(row=tot_row, column=col_idx, value=totals.get(key,0))
		for col_idx in range(1, 11):
			c = ws.cell(row=tot_row, column=col_idx)
			c.font = font_total
			c.fill = fill_total
			c.border = double_bottom
			c.alignment = align_center
		ws.freeze_panes = "D4"
		ws.print_area = f"A1:J{tot_row}"
		ws.page_setup.paperSize = ws.PAPERSIZE_A4
		ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
		ws.page_setup.fitToWidth = 1
		ws.sheet_properties.pageSetUpPr.fitToPage = True
		ws.sheet_view.showGridLines = False

	# -------------------------------------------------------------
	# 3. 19 列双层表头社保明细表
	# -------------------------------------------------------------
	def build_insurance_sheet(ws):
		ws.title = "3.社会保险缴费明细(19列)"
		data_res = get_social_insurance_sheet(company, period_month)
		rows = data_res.get("rows", [])
		totals = data_res.get("totals", {})

		ws.merge_cells("A1:S1")
		ws["A1"] = f"{company} {data_res.get('report_title', '社会保险缴费明细表')}"
		ws["A1"].font = font_title
		ws["A1"].alignment = align_center
		ws.row_dimensions[1].height = 32

		ws.merge_cells("A3:G3")
		ws["A3"] = "员工基本参保信息"
		ws["A3"].font = font_header
		ws["A3"].fill = fill_accent
		ws["A3"].alignment = align_center

		ws.merge_cells("H3:M3")
		ws["H3"] = "单位缴纳部分"
		ws["H3"].font = font_header
		ws["H3"].fill = fill_warning
		ws["H3"].alignment = align_center

		ws.merge_cells("N3:R3")
		ws["N3"] = "个人代扣部分"
		ws["N3"].font = font_header
		ws["N3"].fill = fill_success
		ws["N3"].alignment = align_center

		ws.cell(row=3, column=19, value="总计").fill = fill_danger
		ws.cell(row=3, column=19).font = font_header
		ws.cell(row=3, column=19).alignment = align_center

		sub_headers = [
			"序号", "工号", "姓名", "证件号码", "所属期", "用工性质", "社保基数",
			"单位养老", "单位失业", "单位医疗", "单位其他医疗", "单位工伤", "单位合计",
			"个人养老", "个人失业", "个人医疗", "个人大额医疗", "个人合计", "总计"
		]
		ws.row_dimensions[4].height = 24
		for col_idx, sh in enumerate(sub_headers, start=1):
			cell = ws.cell(row=4, column=col_idx, value=sh)
			cell.font = font_header
			cell.fill = fill_header
			cell.alignment = align_center
			cell.border = border_cell

		for row_idx, r in enumerate(rows, start=5):
			ws.row_dimensions[row_idx].height = 20
			vals = [
				r.get("seq"), r.get("employee_no"), r.get("employee_name"), r.get("id_card"),
				r.get("period_month_str"), r.get("employee_type"), r.get("ss_base", 0),
				r.get("comp_pension", 0), r.get("comp_unemp", 0), r.get("comp_med", 0),
				r.get("comp_other_med", 0), r.get("comp_injury", 0), r.get("comp_total", 0),
				r.get("pers_pension", 0), r.get("pers_unemp", 0), r.get("pers_med", 0),
				r.get("pers_large_med", 0), r.get("pers_total", 0), r.get("grand_total", 0)
			]
			for col_idx, val in enumerate(vals, start=1):
				cell = ws.cell(row=row_idx, column=col_idx, value=val)
				cell.font = font_data
				cell.border = border_cell
				if col_idx in [1, 2, 4, 5, 6]:
					cell.alignment = align_center
				elif col_idx == 3:
					cell.alignment = align_left
				else:
					cell.alignment = align_right
					if isinstance(val, (int, float)):
						cell.number_format = "#,##0.00"

		tot_row = len(rows) + 5
		ws.row_dimensions[tot_row].height = 24
		ws.cell(row=tot_row, column=1, value="合计").alignment = align_center
		for c in range(2, 7): ws.cell(row=tot_row, column=c, value="")
		tot_vals = [
			totals.get("ss_base", 0), totals.get("comp_pension", 0), totals.get("comp_unemp", 0),
			totals.get("comp_med", 0), totals.get("comp_other_med", 0), totals.get("comp_injury", 0),
			totals.get("comp_total", 0), totals.get("pers_pension", 0), totals.get("pers_unemp", 0),
			totals.get("pers_med", 0), totals.get("pers_large_med", 0), totals.get("pers_total", 0),
			totals.get("grand_total", 0)
		]
		for col_idx, val in enumerate(tot_vals, start=7):
			cell = ws.cell(row=tot_row, column=col_idx, value=val)
			cell.number_format = "#,##0.00"
			cell.alignment = align_right

		for col_idx in range(1, 20):
			c = ws.cell(row=tot_row, column=col_idx)
			c.font = font_total
			c.fill = fill_total
			c.border = double_bottom

	# -------------------------------------------------------------
	# 4. 12 列双层表头公积金明细表
	# -------------------------------------------------------------
	def build_housing_fund_sheet(ws):
		ws.title = "4.公积金明细(12列)"
		data_res = get_housing_fund_sheet(company, period_month)
		rows = data_res.get("rows", [])
		totals = data_res.get("totals", {})

		ws.merge_cells("A1:L1")
		ws["A1"] = f"{company} {data_res.get('report_title', '住房公积金缴存明细表')}"
		ws["A1"].font = font_title
		ws["A1"].alignment = align_center
		ws.row_dimensions[1].height = 32

		ws.merge_cells("A3:F3")
		ws["A3"] = "员工基本参保信息"
		ws["A3"].font = font_header
		ws["A3"].fill = fill_accent
		ws["A3"].alignment = align_center

		ws.merge_cells("G3:I3")
		ws["G3"] = "单位缴存部分 (5%)"
		ws["G3"].font = font_header
		ws["G3"].fill = fill_warning
		ws["G3"].alignment = align_center

		ws.merge_cells("J3:K3")
		ws["J3"] = "个人缴存部分 (5%)"
		ws["J3"].font = font_header
		ws["J3"].fill = fill_success
		ws["J3"].alignment = align_center

		ws.cell(row=3, column=12, value="月缴存总额").fill = fill_danger
		ws.cell(row=3, column=12).font = font_header
		ws.cell(row=3, column=12).alignment = align_center

		sub_headers = [
			"序号", "工号", "姓名", "证件号码", "所属期", "用工性质",
			"公积金基数", "单位比例", "单位金额", "个人比例", "个人金额", "月缴存总额"
		]
		ws.row_dimensions[4].height = 24
		for col_idx, sh in enumerate(sub_headers, start=1):
			cell = ws.cell(row=4, column=col_idx, value=sh)
			cell.font = font_header
			cell.fill = fill_header
			cell.alignment = align_center
			cell.border = border_cell

		for row_idx, r in enumerate(rows, start=5):
			ws.row_dimensions[row_idx].height = 20
			vals = [
				r.get("seq"), r.get("employee_no"), r.get("employee_name"), r.get("id_card"),
				r.get("period_month_str"), r.get("employee_type"), r.get("hf_base", 0),
				f"{r.get('comp_rate', 5)}%", r.get("comp_amount", 0),
				f"{r.get('pers_rate', 5)}%", r.get("pers_amount", 0), r.get("total_amount", 0)
			]
			for col_idx, val in enumerate(vals, start=1):
				cell = ws.cell(row=row_idx, column=col_idx, value=val)
				cell.font = font_data
				cell.border = border_cell
				if col_idx in [1, 2, 4, 5, 6, 8, 10]:
					cell.alignment = align_center
				elif col_idx == 3:
					cell.alignment = align_left
				else:
					cell.alignment = align_right
					if isinstance(val, (int, float)):
						cell.number_format = "#,##0.00"

		tot_row = len(rows) + 5
		ws.row_dimensions[tot_row].height = 24
		ws.cell(row=tot_row, column=1, value="合计").alignment = align_center
		for c in range(2, 7): ws.cell(row=tot_row, column=c, value="")
		ws.cell(row=tot_row, column=7, value=totals.get("hf_base", 0)).number_format = "#,##0.00"
		ws.cell(row=tot_row, column=8, value="-").alignment = align_center
		ws.cell(row=tot_row, column=9, value=totals.get("comp_amount", 0)).number_format = "#,##0.00"
		ws.cell(row=tot_row, column=10, value="-").alignment = align_center
		ws.cell(row=tot_row, column=11, value=totals.get("pers_amount", 0)).number_format = "#,##0.00"
		ws.cell(row=tot_row, column=12, value=totals.get("total_amount", 0)).number_format = "#,##0.00"

		for col_idx in range(1, 13):
			c = ws.cell(row=tot_row, column=col_idx)
			c.font = font_total
			c.fill = fill_total
			c.border = double_bottom

	# -------------------------------------------------------------
	# 5. 个人所得税 (精简版 · 与前端17列同口径)
	# -------------------------------------------------------------
	def build_tax_simple_sheet(ws):
		ws.title = "5.个税预扣明细(精简版)"
		data_res = get_tax_settlement_sheet(company, period_month)
		rows = data_res.get("rows", [])
		totals = data_res.get("totals", {})

		ws.merge_cells("A1:Q1")
		ws["A1"] = f"{company} {period_month} 个人所得税预扣预缴表（17列财税精简版）"
		ws["A1"].font = font_title
		ws["A1"].alignment = align_center
		ws.row_dimensions[1].height = 32

		ws.merge_cells("A2:Q2")
		ws["A2"] = f"发薪账期: {period_month}  |  累计周期: {data_res.get('cycle_name', '')}  |  生成日期: {date.today().strftime('%Y-%m-%d')}  |  申报人数: {len(rows)} 人  |  币种: 人民币(元)"
		ws["A2"].font = font_sub
		ws["A2"].alignment = align_center
		ws.row_dimensions[2].height = 20

		# 双层表头 Row 3 & 4：与前端精简台账保持一致，避免混用“本期”与“累计”口径。
		group_defs = [
			("A3:F3", "一、员工基本信息", fill_accent),
			("G3:G3", "二、本期计税收入", fill_info),
			("H3:K3", "三、扣除信息", fill_warning),
			("L3:O3", "四、累计计税与税阶", fill_purple),
			("P3:Q3", "五、本月税款与实发", fill_success),
		]
		for cell_range, label, fill in group_defs:
			ws.merge_cells(cell_range)
			cell = ws[cell_range.split(":")[0]]
			cell.value = label
			cell.font = font_header
			cell.fill = fill
			cell.alignment = align_center

		sub_headers = [
			"序号", "工号", "姓名", "证件号码", "用工性质", "发薪账期",
			"本期税前收入", "累计基本减除费用", "本期社保扣除", "本期公积金扣除", "本期专项附加扣除",
			"累计应纳税所得额", "预扣率", "速算扣除数", "往期已缴税额",
			"本月应预扣税额", "税后实发工资",
		]
		ws.row_dimensions[4].height = 24
		for col_idx, sh in enumerate(sub_headers, start=1):
			cell = ws.cell(row=4, column=col_idx, value=sh)
			cell.font = font_header
			cell.fill = fill_header
			cell.alignment = align_center
			cell.border = border_cell

		for row_idx, r in enumerate(rows, start=5):
			ws.row_dimensions[row_idx].height = 20
			vals = [
				r.get("seq"), r.get("employee_no"), r.get("employee_name"), r.get("id_card"), r.get("employee_type"), r.get("period_month_str") or period_month.replace("-", ""),
				r.get("gross_salary", 0), r.get("thresh_all", 0), r.get("ss_person", 0), r.get("hf_person", 0), r.get("spec_add_tot_cur", 0),
				r.get("taxable_all", r.get("taxable_income", 0)), f"{r.get('tax_rate', 0)}%", r.get("quick_deduct", 0), r.get("tax_paid_prior", 0),
				r.get("tax_current", r.get("current_tax", 0)), r.get("net_salary", 0),
			]
			for col_idx, val in enumerate(vals, start=1):
				cell = ws.cell(row=row_idx, column=col_idx, value=val)
				cell.font = font_data
				cell.border = border_cell
				if col_idx in [1, 2, 4, 5, 6, 13]:
					cell.alignment = align_center
				elif col_idx == 3:
					cell.alignment = align_left
				else:
					cell.alignment = align_right
					if isinstance(val, (int, float)):
						cell.number_format = "#,##0.00"

		tot_row = len(rows) + 5
		ws.row_dimensions[tot_row].height = 24
		total_vals = [
			"合计", f"共 {len(rows)} 人", "", "", "", "",
			totals.get("gross_salary", 0), totals.get("thresh_all", 0), totals.get("ss_person", 0), totals.get("hf_person", 0), totals.get("spec_add_tot_cur", 0),
			totals.get("taxable_all", 0), "-", "-", totals.get("tax_paid_prior", 0), totals.get("tax_current", 0), totals.get("net_salary", 0),
		]
		for col_idx, val in enumerate(total_vals, start=1):
			cell = ws.cell(row=tot_row, column=col_idx, value=val)
			cell.font = font_total
			cell.fill = fill_total
			cell.border = double_bottom
			cell.alignment = align_center if col_idx in [1,2,3,4,5,6,13,14] else align_right
			if isinstance(val, (int, float)):
				cell.number_format = "#,##0.00"

		# 冻结三列和双层表头，便于与网页台账交叉核对。
		ws.freeze_panes = "D5"
		ws.auto_filter.ref = f"A4:Q{tot_row-1}" if rows else "A4:Q4"

	# -------------------------------------------------------------
	# 6. 个人所得税 (VBA 68列完整核算台账)
	# -------------------------------------------------------------
	def build_tax_full_68_sheet(ws):
		ws.title = "6.个税申报台账(68列)"
		data_res = get_tax_settlement_full_sheet(company, period_month)
		rows = data_res.get("rows", [])
		totals = data_res.get("totals", {})

		ws.merge_cells("A1:BP1")
		ws["A1"] = f"{company} {period_month} 个人所得税申报台账（VBA同口径68列）"
		ws["A1"].font = font_title
		ws["A1"].alignment = align_center
		ws.row_dimensions[1].height = 32
		ws.merge_cells("A2:BP2")
		ws["A2"] = f"累计申报周期: {data_res.get('cycle_name', '')}  |  基本减除费用: {data_res.get('tax_threshold', 5000):,.2f} 元/月  |  本期为周期第 {data_res.get('month_idx', 1)} 个月"
		ws["A2"].font = font_sub
		ws["A2"].alignment = align_center

		groups = [
			("A3:I3", "一、员工基本信息", fill_accent),
			("J3:N3", "二、工资扣除(本月)", fill_warning),
			("O3:T3", "三、专项扣除(本月)", fill_success),
			("U3:AB3", "四、专项附加扣除(本月)", fill_info),
			("AC3:AR3", "五、个税累计(往期)", fill_purple),
			("AS3:BH3", "六、个税累计(全部)", fill_warning),
			("BI3:BP3", "七、税款计算", fill_danger),
		]
		for range_str, title_str, fill_color in groups:
			ws.merge_cells(range_str)
			cell = ws[range_str.split(":")[0]]
			cell.value = title_str
			cell.font = font_header
			cell.fill = fill_color
			cell.alignment = align_center

		columns = [
			("序号","seq"),("工号","employee_no"),("姓名","employee_name"),("证件号码","id_card"),("性别","gender"),("本期所属期","period_month_str"),("员工类型","employee_type"),("目标工资","target_salary"),("工资类型","salary_mode"),
			("税前工资","gross_salary"),("起征点扣除","thresh_cur"),("公积金","hf_person"),("社保","ss_person"),("工资扣除合计","deduct_cur_tot"),
			("基本养老","ss_pension"),("基本医疗","ss_med"),("大额医疗","ss_large_med"),("失业保险","ss_unemp"),("住房公积金","hf_spec"),("专项扣除合计","spec_tot_cur"),
			("子女教育","spec_add_child"),("继续教育","spec_add_edu"),("大病医疗","spec_add_med"),("住房贷款利息","spec_add_loan"),("住房租金","spec_add_rent"),("赡养老人","spec_add_elder"),("3岁以下婴幼儿照护","spec_add_baby"),("专项附加扣除合计","spec_add_tot_cur"),
			("税前工资(往期)","gross_prior"),("起征点扣除(往期)","thresh_prior"),("基本养老(往期)","ss_pension_prior"),("基本医疗(往期)","ss_med_prior"),("大额医疗(往期)","ss_large_med_prior"),("失业保险(往期)","ss_unemp_prior"),("住房公积金(往期)","hf_spec_prior"),("专项扣除合计(往期)","spec_tot_prior"),
			("子女教育(往期)","spec_add_child_prior"),("继续教育(往期)","spec_add_edu_prior"),("大病医疗(往期)","spec_add_med_prior"),("住房贷款利息(往期)","spec_add_loan_prior"),("住房租金(往期)","spec_add_rent_prior"),("赡养老人(往期)","spec_add_elder_prior"),("3岁以下婴幼儿照护(往期)","spec_add_baby_prior"),("专项附加扣除合计(往期)","spec_add_tot_prior"),
			("个税_税前工资(全部)","gross_all"),("起征点扣除(全部)","thresh_all"),("基本养老(全部)","ss_pension_all"),("基本医疗(全部)","ss_med_all"),("大额医疗(全部)","ss_large_med_all"),("失业保险(全部)","ss_unemp_all"),("住房公积金(全部)","hf_spec_all"),("专项扣除合计(全部)","spec_tot_all"),
			("子女教育(全部)","spec_add_child_all"),("继续教育(全部)","spec_add_edu_all"),("大病医疗(全部)","spec_add_med_all"),("住房贷款利息(全部)","spec_add_loan_all"),("住房租金(全部)","spec_add_rent_all"),("赡养老人(全部)","spec_add_elder_all"),("3岁以下婴幼儿照护(全部)","spec_add_baby_all"),("专项附加扣除合计(全部)","spec_add_tot_all"),
			("应纳税所得额","taxable_all"),("税率","tax_rate"),("速算扣除数","quick_deduct"),("应纳税额","tax_calculated"),("减免税额","tax_relief"),("已缴税额","tax_paid_prior"),("应补/退税额","tax_current"),("税后工资","net_salary"),
		]
		assert len(columns) == 68
		for col_idx, (header, _key) in enumerate(columns, start=1):
			cell = ws.cell(row=4, column=col_idx, value=header)
			cell.font = font_header
			cell.fill = fill_header
			cell.alignment = align_center
			cell.border = border_cell

		text_keys = {"seq","employee_no","employee_name","id_card","gender","period_month_str","employee_type","salary_mode"}
		for row_idx, r in enumerate(rows, start=5):
			for col_idx, (_header, key) in enumerate(columns, start=1):
				val = r.get(key, "")
				if key == "tax_rate":
					val = f"{r.get(key, 0)}%"
				cell = ws.cell(row=row_idx, column=col_idx, value=val)
				cell.font = font_data
				cell.border = border_cell
				if key in text_keys or key == "tax_rate":
					cell.alignment = align_center if key != "employee_name" else align_left
				else:
					cell.alignment = align_right
					if isinstance(val, (int, float)):
						cell.number_format = "#,##0.00"

		tot_row = len(rows) + 5
		for col_idx, (_header, key) in enumerate(columns, start=1):
			if key == "seq": val = "合计"
			elif key == "employee_no": val = f"共 {len(rows)} 人"
			elif key in text_keys or key == "tax_rate": val = ""
			else: val = totals.get(key, 0)
			cell = ws.cell(row=tot_row, column=col_idx, value=val)
			cell.font = font_total
			cell.fill = fill_total
			cell.border = double_bottom
			if isinstance(val, (int, float)):
				cell.number_format = "#,##0.00"

		ws.freeze_panes = "D5"
		ws.auto_filter.ref = f"A4:BP{max(tot_row-1,4)}"

	# -------------------------------------------------------------
	# 7. 全员全周期历史总览表 (15 列)
	# -------------------------------------------------------------
	def build_history_all_sheet(ws):
		ws.title = "7.全员全周期总览(15列)"
		data_res = get_all_employees_tax_history_summary(company, period_month)
		rows = data_res.get("rows", [])
		totals = data_res.get("totals", {})

		ws.merge_cells("A1:O1")
		ws["A1"] = f"{company} 薪酬与个人所得税全员全周期累计总览大表 ({data_res.get('cycle_name', '')})"
		ws["A1"].font = font_title
		ws["A1"].alignment = align_center
		ws.row_dimensions[1].height = 32

		ws.merge_cells("A2:O2")
		ws["A2"] = f"申报周期: {data_res.get('cycle_name', '')}  |  生成日期: {date.today().strftime('%Y-%m-%d')}  |  全员人数: {len(rows)} 人  |  币种: 人民币(元)"
		ws["A2"].font = font_sub
		ws["A2"].alignment = align_center
		ws.row_dimensions[2].height = 20

		headers = [
			"序号", "工号", "姓名", "证件号码", "用工性质", "已计薪月数",
			"周期累计税前收入", "周期累计基本减除", "周期累计个人社保", "周期累计个人公积金",
			"周期累计专项附加", "周期累计应税所得额", "最高税阶", "周期累计已缴个税", "周期累计税后实发"
		]
		ws.row_dimensions[3].height = 26
		for col_idx, h in enumerate(headers, start=1):
			cell = ws.cell(row=3, column=col_idx, value=h)
			cell.font = font_header
			cell.fill = fill_success if "实发" in h else (fill_danger if "税" in h else fill_header)
			cell.alignment = align_center
			cell.border = border_cell

		for row_idx, r in enumerate(rows, start=4):
			ws.row_dimensions[row_idx].height = 20
			vals = [
				r.get("seq"), r.get("employee_no"), r.get("employee_name"), r.get("id_card"), r.get("employee_type"),
				r.get("months_paid_desc"), r.get("cum_gross_salary", 0), r.get("cum_tax_threshold", 0),
				r.get("cum_ss_person", 0), r.get("cum_hf_person", 0), r.get("cum_special_deductions", 0),
				r.get("cum_taxable_income", 0), f"{r.get('tax_rate', 0)}%", r.get("cum_tax_paid", 0), r.get("cum_net_salary", 0)
			]
			for col_idx, val in enumerate(vals, start=1):
				cell = ws.cell(row=row_idx, column=col_idx, value=val)
				cell.font = font_data
				cell.border = border_cell
				if col_idx in [1, 2, 4, 5, 6, 13]:
					cell.alignment = align_center
				elif col_idx == 3:
					cell.alignment = align_left
				else:
					cell.alignment = align_right
					if isinstance(val, (int, float)):
						cell.number_format = "#,##0.00"

		tot_row = len(rows) + 4
		ws.row_dimensions[tot_row].height = 24
		ws.cell(row=tot_row, column=1, value="合计").alignment = align_center
		for c in range(2, 7): ws.cell(row=tot_row, column=c, value="")
		ws.cell(row=tot_row, column=7, value=totals.get("cum_gross", 0)).number_format = "#,##0.00"
		ws.cell(row=tot_row, column=8, value=totals.get("cum_thresh", 0)).number_format = "#,##0.00"
		ws.cell(row=tot_row, column=9, value=totals.get("cum_ss", 0)).number_format = "#,##0.00"
		ws.cell(row=tot_row, column=10, value=totals.get("cum_hf", 0)).number_format = "#,##0.00"
		ws.cell(row=tot_row, column=11, value=totals.get("cum_special_add", 0)).number_format = "#,##0.00"
		ws.cell(row=tot_row, column=12, value=totals.get("cum_taxable", 0)).number_format = "#,##0.00"
		ws.cell(row=tot_row, column=13, value="-").alignment = align_center
		ws.cell(row=tot_row, column=14, value=totals.get("cum_tax_paid", 0)).number_format = "#,##0.00"
		ws.cell(row=tot_row, column=15, value=totals.get("cum_net", 0)).number_format = "#,##0.00"

		for col_idx in range(1, 16):
			c = ws.cell(row=tot_row, column=col_idx)
			c.font = font_total
			c.fill = fill_total
			c.border = double_bottom

	# -------------------------------------------------------------
	# 8. 指定历史月完整核算快照 (VBA68列 + ERP审计)
	# -------------------------------------------------------------
	def build_history_full_sheet(ws, selected_month, emp_no=None):
		data_res = get_history_full_ledger(company, period_month, selected_month, employee_no=emp_no)
		rows = data_res.get("rows", [])
		columns = data_res.get("columns", [])
		totals = data_res.get("totals", {})
		last_col = get_column_letter(max(len(columns), 1))
		emp_tag = f"_{emp_no}" if emp_no else ""
		ws.title = f"7.历史完整核算_{selected_month}{emp_tag}"[:31]

		ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(columns), 1))
		ws["A1"] = f"{company} {selected_month} 历史完整核算快照（VBA 68列 + ERP审计）{f' - {emp_no}' if emp_no else ''}"
		ws["A1"].font = font_title
		ws["A1"].alignment = align_center
		ws.row_dimensions[1].height = 32

		ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(len(columns), 1))
		ws["A2"] = (
			f"财务快照月份: {selected_month} | 当前核定账期: {period_month} | "
			f"账期状态: {data_res.get('status') or '-'} | 员工人数: {len(rows)} 人 | 生成日期: {date.today().strftime('%Y-%m-%d')}"
		)
		ws["A2"].font = font_sub
		ws["A2"].alignment = align_center

		# 第3行按逻辑分组，第4行为字段名；冻结前三列并冻结表头。
		groups = [c.get("group") or "" for c in columns]
		start_col = 1
		while start_col <= len(columns):
			group = groups[start_col - 1]
			end_col = start_col
			while end_col < len(columns) and groups[end_col] == group:
				end_col += 1
			if end_col > start_col:
				ws.merge_cells(start_row=3, start_column=start_col, end_row=3, end_column=end_col)
			cell = ws.cell(row=3, column=start_col, value=group)
			cell.font = font_header
			cell.fill = fill_accent if group == "员工基本信息" else (fill_purple if group == "ERP审计" else fill_header)
			cell.alignment = align_center
			cell.border = border_cell
			for cc in range(start_col, end_col + 1):
				ws.cell(row=3, column=cc).border = border_cell
			start_col = end_col + 1

		for col_idx, col in enumerate(columns, start=1):
			cell = ws.cell(row=4, column=col_idx, value=col.get("label") or col.get("key"))
			cell.font = font_header
			cell.fill = fill_purple if col.get("group") == "ERP审计" else fill_header
			cell.alignment = align_center
			cell.border = border_cell

		for row_idx, row in enumerate(rows, start=5):
			for col_idx, col in enumerate(columns, start=1):
				key = col.get("key")
				value = row.get(key)
				cell = ws.cell(row=row_idx, column=col_idx, value=value)
				cell.font = font_data
				cell.border = border_cell
				if col.get("type") == "money":
					cell.number_format = "#,##0.00"
					cell.alignment = align_right
				elif col.get("type") == "percent":
					cell.number_format = '0.00"%"'
					cell.alignment = align_center
				else:
					cell.alignment = align_left if key in {"employee_name", "calculation_trigger_source", "calculation_error"} else align_center

		tot_row = len(rows) + 5
		for col_idx, col in enumerate(columns, start=1):
			key = col.get("key")
			if col_idx == 1:
				value = "合计"
			elif col_idx == 2:
				value = f"共 {len(rows)} 人"
			elif col.get("type") == "money" and key in totals:
				value = totals.get(key, 0)
			else:
				value = ""
			cell = ws.cell(row=tot_row, column=col_idx, value=value)
			cell.font = font_total
			cell.fill = fill_total
			cell.border = double_bottom
			if isinstance(value, (int, float)):
				cell.number_format = "#,##0.00"

		ws.freeze_panes = "D5"
		ws.auto_filter.ref = f"A4:{last_col}{max(tot_row - 1, 4)}"

	# -------------------------------------------------------------
	# 8. 单人历史流水表 (12 个月)
	# -------------------------------------------------------------
	def build_history_single_sheet(ws, emp_no):
		data_res = get_employee_tax_history_timeline(company, emp_no, period_month)
		rows = data_res.get("rows", [])
		summary = data_res.get("summary", {})
		emp_name = data_res.get("employee_name", emp_no)
		ws.title = f"{emp_name}_12个月穿透流水"[:31]

		ws.merge_cells("A1:N1")
		ws["A1"] = f"{company} {emp_name} ({emp_no}) 个税年度申报周期月度穿透流水表"
		ws["A1"].font = font_title
		ws["A1"].alignment = align_center
		ws.row_dimensions[1].height = 32

		ws.merge_cells("A2:N2")
		ws["A2"] = f"员工姓名: {emp_name} | 工号: {emp_no} | 证件号: {data_res.get('id_card')} | 申报周期: {data_res.get('cycle_name')}"
		ws["A2"].font = font_sub
		ws["A2"].alignment = align_center
		ws.row_dimensions[2].height = 20

		headers = [
			"序号", "发薪月份", "账期状态", "当月税前应发", "个人五险一金", "专项附加扣除",
			"累计基本减除费用", "累计应纳税所得额", "预扣率", "速算扣除数", "当月应预扣税额",
			"往期累计已缴税额", "当月税后实发", "备注"
		]
		ws.row_dimensions[3].height = 26
		for col_idx, h in enumerate(headers, start=1):
			cell = ws.cell(row=3, column=col_idx, value=h)
			cell.font = font_header
			cell.fill = fill_header
			cell.alignment = align_center
			cell.border = border_cell

		for row_idx, r in enumerate(rows, start=4):
			ws.row_dimensions[row_idx].height = 20
			vals = [
				r.get("seq"), r.get("period_month"), r.get("status"),
				r.get("gross_salary", 0), r.get("insurance_total", 0), r.get("special_deductions_total", 0),
				r.get("threshold_accumulated", 0), r.get("taxable_accumulated", 0), f"{r.get('tax_rate', 0)}%",
				r.get("quick_deduction", 0), r.get("tax_current", 0), r.get("tax_paid_prior", 0),
				r.get("net_salary", 0), ""
			]
			for col_idx, val in enumerate(vals, start=1):
				cell = ws.cell(row=row_idx, column=col_idx, value=val)
				cell.font = font_data
				cell.border = border_cell
				if col_idx in [1, 2, 3, 9, 14]:
					cell.alignment = align_center
				else:
					cell.alignment = align_right
					if isinstance(val, (int, float)):
						cell.number_format = "#,##0.00"

		tot_row = len(rows) + 4
		ws.cell(row=tot_row, column=1, value="周期累计").font = font_total
		ws.merge_cells(start_row=tot_row, start_column=1, end_row=tot_row, end_column=3)
		summary_vals = {
			4: summary.get("cum_gross_salary", 0),
			5: round(flt(summary.get("cum_ss_person")) + flt(summary.get("cum_hf_person")), 2),
			6: summary.get("cum_special_deductions", 0),
			7: summary.get("cum_tax_threshold", 0),
			11: summary.get("cum_tax_paid", 0),
			13: summary.get("cum_net_salary", 0),
		}
		for col_idx in range(1, 15):
			cell = ws.cell(row=tot_row, column=col_idx)
			cell.font = font_total
			cell.fill = fill_total
			cell.border = double_bottom
			if col_idx in summary_vals:
				cell.value = summary_vals[col_idx]
				cell.number_format = "#,##0.00"

	# -------------------------------------------------------------
	# 模式调度与工作簿组装
	# -------------------------------------------------------------
	filename_prefix = f"祺富薪资台账_{period_month}"

	if sheet_type == "distribution":
		build_distribution_sheet(ws_default)
		filename = f"{filename_prefix}_24列薪资发放表.xlsx"
	elif sheet_type == "cash":
		build_cash_count_sheet(ws_default)
		filename = f"{filename_prefix}_现金点钞核定表.xlsx"
	elif sheet_type == "accounting":
		build_accounting_sheet(ws_default)
		filename = f"{filename_prefix}_11列记账工资表.xlsx"
	elif sheet_type == "accounting_xlsm":
		build_accounting_sheet(ws_default)
		acc_data = get_accounting_payroll_sheet(company, period_month)
		if acc_data.get("foreign_rows"):
			ws_foreign = wb.create_sheet()
			build_accounting_sheet(ws_foreign, foreign=True)
		filename = f"{filename_prefix}_记账工资表_XLSM参考版.xlsx"
	elif sheet_type == "insurance":
		build_insurance_sheet(ws_default)
		filename = f"{filename_prefix}_社保缴费明细表.xlsx"
	elif sheet_type == "housing_fund":
		build_housing_fund_sheet(ws_default)
		filename = f"{filename_prefix}_公积金缴费明细表.xlsx"
	elif sheet_type == "tax":
		if tax_view_mode == "full_68":
			build_tax_full_68_sheet(ws_default)
			filename = f"{filename_prefix}_个人所得税VBA同口径68列台账.xlsx"
		else:
			build_tax_simple_sheet(ws_default)
			filename = f"{filename_prefix}_个人所得税预扣预缴表(17列财税精简版).xlsx"
	elif sheet_type == "history":
		if history_mode == "single":
			build_history_single_sheet(ws_default, history_emp_no)
			filename = f"{filename_prefix}_{history_emp_no}_个税申报周期月度穿透流水.xlsx"
		elif history_mode == "full":
			selected_history_month = history_period_month or period_month
			build_history_full_sheet(ws_default, selected_history_month, history_emp_no)
			emp_suffix = f"_{history_emp_no}" if history_emp_no else ""
			filename = f"祺富薪资台账_{selected_history_month}{emp_suffix}_历史完整核算_VBA68列加审计.xlsx"
		else:
			build_history_all_sheet(ws_default)
			filename = f"{filename_prefix}_全员薪酬与个税年度申报周期累计总览表(15列).xlsx"
	else:
		# all: 打包完整 7 个工作表
		build_distribution_sheet(ws_default)
		ws2 = wb.create_sheet()
		build_accounting_sheet(ws2)
		ws3 = wb.create_sheet()
		build_insurance_sheet(ws3)
		ws4 = wb.create_sheet()
		build_housing_fund_sheet(ws4)
		ws5 = wb.create_sheet()
		build_tax_simple_sheet(ws5)
		ws6 = wb.create_sheet()
		build_tax_full_68_sheet(ws6)
		ws7 = wb.create_sheet()
		build_history_all_sheet(ws7)
		filename = f"{filename_prefix}_全套薪酬财务台账工作簿(7大标准工作表).xlsx"

	# 自动调整列宽，并对关键财务表施加人机工程固定宽度。
	for sheet in wb.worksheets:
		for col in sheet.columns:
			max_len = 0
			col_letter = get_column_letter(col[0].column)
			for cell in col:
				val_str = str(cell.value or "")
				if len(val_str) > max_len and len(val_str) < 50:
					max_len = len(val_str)
			sheet.column_dimensions[col_letter].width = min(max(max_len + 2, 8), 24)

		# 17列个税精简版：序号仅承担定位功能，不应与金额列同宽。
		if sheet.title == "5.个税预扣明细(精简版)":
			widths = [4.8, 9.0, 10.0, 19, 10, 10, 13, 15, 13, 14, 17, 17, 8, 12, 13, 15, 14]
			for idx, width in enumerate(widths, start=1):
				sheet.column_dimensions[get_column_letter(idx)].width = width
		elif sheet.title == "1.薪资发放表(24列)":
			# 严格对齐参考 XLSM：24列业务数据 + 7列现金点钞辅助列，姓名列加宽
			widths = [5.625, 8.625, 13.5, 8.625, 13.0, 9.125, 13.0, 13.0, 13.0, 13.0, 13.0, 13.0, 8.625, 13.0, 9.125, 12.625, 8.625, 13.0, 12.625, 13.0, 9.125, 13.625, 16.625, 14.625, 10.625, 9.0, 9.0, 9.0, 9.0, 12.0, 10.625]
			for idx, width in enumerate(widths, start=1):
				sheet.column_dimensions[get_column_letter(idx)].width = width
				sheet.column_dimensions[get_column_letter(idx)].hidden = False
		elif sheet.title in {"2.记账工资表(11列)", "记账工资表外籍"}:
			# 严格对齐参考 XLSM：姓名列加宽至 14.0，适配中长姓名与外籍姓名，适合整表缩小填充打印
			widths = [8.625, 14.0, 11.625, 8.625, 13.0, 11.625, 9.5, 10.5, 9.5, 10.5, 11.625]
			for idx, width in enumerate(widths, start=1):
				sheet.column_dimensions[get_column_letter(idx)].width = width
		elif sheet.title == "现金点钞核定表":
			widths = [5.5, 9.5, 10.5, 8, 8, 8, 8, 8, 12, 9]
			for idx, width in enumerate(widths, start=1):
				sheet.column_dimensions[get_column_letter(idx)].width = width

	output = io.BytesIO()
	wb.save(output)
	output.seek(0)
	file_base64 = base64.b64encode(output.read()).decode("utf-8")

	return {
		"success": True,
		"filename": filename,
		"file_base64": file_base64
	}


def _set_payroll_item_calculation_audit(it, status, trigger_source="系统任务", task_name="", input_hash="", error=""):
    """Write recalculation audit fields only when the migrated child table supports them."""
    meta = frappe.get_meta("Ashan Monthly Payroll Item")
    if not meta.has_field("calculation_status"):
        return
    it.calculation_status = status
    if status in {"待计算", "排队中"}:
        it.calculation_requested_at = now_datetime()
    elif status == "计算中":
        it.calculation_started_at = now_datetime()
    elif status == "已计算":
        it.calculation_completed_at = now_datetime()
    it.calculation_trigger_source = trigger_source or "系统任务"
    it.calculation_engine_version = "vba-tax-async-2026.08.21"
    if input_hash:
        it.calculation_input_hash = input_hash
    if task_name:
        it.calculation_task_id = task_name
    it.calculation_error = error or ""


def _recalculate_payroll_item_vba(
    doc, it, company, period_month, trigger_source="系统任务", task_name="", input_hash="", refresh_from_profile=True
):
    """Recalculate one payroll child row with the verified VBA cumulative withholding model."""
    params = get_effective_tax_parameters(company, period_month)
    tax_thresh = params["tax_threshold"]
    cinfo = get_tax_cycle_info(period_month, tax_thresh, params["tax_cycle_start_month"])
    year = _period_year(period_month)
    ins = get_insurance_setting(company, year) or {}
    p_rate = flt(ins.get("ss_person_pension")) or 8.0
    m_rate = flt(ins.get("ss_person_medical")) or 2.0
    u_rate = flt(ins.get("ss_person_unemployment")) or 0.5

    emp_no = it.employee_no
    emp_doc = frappe.db.get_value(
        "Ashan Employee Salary Profile",
        {"company": company, "employee_no": emp_no},
        [
            "id_card", "gender", "mobile", "birth_date", "employee_type", "salary_mode", "is_insured",
            "fixed_salary", "base_salary", "post_allowance", "performance_base",
            "meal_allowance", "traffic_allowance", "communication_allowance", "other_allowance",
            "social_security_base", "housing_fund_base",
            "deduction_child_education", "deduction_continuing_education",
            "deduction_serious_illness", "deduction_housing_loan", "deduction_housing_rent",
            "deduction_elderly_care", "deduction_infant_care",
        ],
        as_dict=True,
    ) or {}

    _set_payroll_item_calculation_audit(it, "计算中", trigger_source, task_name, input_hash)

    if refresh_from_profile:
        for fieldname in ["id_card", "gender", "mobile", "birth_date", "employee_type", "salary_mode"]:
            if emp_doc.get(fieldname):
                setattr(it, fieldname, emp_doc[fieldname])

        for fieldname in [
            "deduction_child_education", "deduction_continuing_education", "deduction_serious_illness",
            "deduction_housing_loan", "deduction_housing_rent", "deduction_elderly_care", "deduction_infant_care",
        ]:
            value = emp_doc.get(fieldname) if emp_doc.get(fieldname) is not None else getattr(it, fieldname, 0)
            setattr(it, fieldname, flt(value))

    spec_add_cur = round(
        flt(it.deduction_child_education) + flt(it.deduction_continuing_education)
        + flt(it.deduction_serious_illness) + flt(it.deduction_housing_loan)
        + flt(it.deduction_housing_rent) + flt(it.deduction_elderly_care)
        + flt(it.deduction_infant_care),
        2,
    )
    it.special_deductions_total = spec_add_cur
    it.tax_threshold = tax_thresh

    # 社保/公积金是权威输入的一部分。当前月从母表+年度参数重建；历史级联时保留当时快照，
    # 防止用今天的费率或基数改写已形成的历史数据。
    emp_type = it.employee_type or emp_doc.get("employee_type") or "正式工"
    no_insurance_types = TAX_REHIRE_EMPLOYEE_TYPES | {"临时工", "零工", "外籍工", "实习生"}
    if refresh_from_profile:
        ss_base = flt(emp_doc.get("social_security_base"))
        hf_base = flt(emp_doc.get("housing_fund_base"))
        if emp_type in no_insurance_types:
            ss_base = 0.0
            hf_base = 0.0
        elif not cint(emp_doc.get("is_insured", 1)):
            # “是否参保”只控制社会保险；住房公积金有独立基数，不应被社保开关误清零。
            ss_base = 0.0

        comp_pension_rate = flt(ins.get("ss_company_pension")) or 16.0
        comp_unemp_rate = flt(ins.get("ss_company_unemployment")) or 0.5
        comp_med_rate = flt(ins.get("ss_company_medical")) or 10.0
        comp_other_med_rate = flt(ins.get("ss_company_other_medical")) or 0.5
        comp_injury_rate = flt(ins.get("ss_company_injury")) or (0.55 if "祺富" in company else 0.35)
        hf_person_rate = flt(ins.get("hf_person_rate")) or 5.0
        hf_company_rate = flt(ins.get("hf_company_rate")) or 5.0
        month_num = cint(str(period_month).split("-")[1])
        special_months = [
            cint(x.strip()) for x in str(ins.get("big_medical_special_months") or "3,12").split(",")
            if x.strip().isdigit()
        ]
        if month_num in special_months:
            big_medical = flt(ins.get("big_medical_amount_special")) or 21.0
        else:
            big_medical = flt(ins.get("big_medical_amount_default")) or 22.0

        it.ss_base = ss_base
        it.hf_base = hf_base
        it.pension_person = round(ss_base * p_rate / 100.0, 2) if ss_base > 0 else 0.0
        it.medical_person = round(ss_base * m_rate / 100.0, 2) if ss_base > 0 else 0.0
        it.unemployment_person = round(ss_base * u_rate / 100.0, 2) if ss_base > 0 else 0.0
        it.large_medical_person = round(big_medical if ss_base > 0 else 0.0, 2)
        it.ss_person_total = round(
            flt(it.pension_person) + flt(it.medical_person) + flt(it.unemployment_person) + flt(it.large_medical_person), 2
        )
        it.pension_company = round(ss_base * comp_pension_rate / 100.0, 2) if ss_base > 0 else 0.0
        it.unemployment_company = round(ss_base * comp_unemp_rate / 100.0, 2) if ss_base > 0 else 0.0
        it.medical_company = round(ss_base * comp_med_rate / 100.0, 2) if ss_base > 0 else 0.0
        it.other_medical_company = round(ss_base * comp_other_med_rate / 100.0, 2) if ss_base > 0 else 0.0
        it.work_injury_company = round(ss_base * comp_injury_rate / 100.0, 2) if ss_base > 0 else 0.0
        it.ss_company_total = round(
            flt(it.pension_company) + flt(it.unemployment_company) + flt(it.medical_company)
            + flt(it.other_medical_company) + flt(it.work_injury_company), 2
        )
        it.hf_person_total = round(hf_base * hf_person_rate / 100.0, 2) if hf_base > 0 else 0.0
        it.hf_company_total = round(hf_base * hf_company_rate / 100.0, 2) if hf_base > 0 else 0.0
        it.housing_fund_person = flt(it.hf_person_total)
        it.housing_fund_company = flt(it.hf_company_total)
    else:
        ss_base = flt(it.ss_base)
        hf_base = flt(it.hf_base)
        # 历史输入更正允许直接调整个人实缴合计；分项仅作为解释性拆分，合计保持用户更正值。
        ss_p_snapshot = flt(it.ss_person_total)
        hf_p_snapshot = flt(it.hf_person_total)
        if ss_p_snapshot > 0 and ss_base > 0:
            it.pension_person = round(ss_base * p_rate / 100.0, 2)
            it.medical_person = round(ss_base * m_rate / 100.0, 2)
            it.unemployment_person = round(ss_base * u_rate / 100.0, 2)
            it.large_medical_person = round(
                ss_p_snapshot - flt(it.pension_person) - flt(it.medical_person) - flt(it.unemployment_person), 2
            )
        elif ss_p_snapshot <= 0:
            it.pension_person = it.medical_person = it.unemployment_person = it.large_medical_person = 0.0
        it.housing_fund_person = hf_p_snapshot

    ss_p = flt(it.ss_person_total)
    hf_p = flt(it.hf_person_total)
    pdata = get_employee_prior_tax_data(company, emp_no, cinfo["prior_months"], tax_thresh)
    ded_cur = round(ss_p + hf_p, 2)
    sal_mode = it.salary_mode or emp_doc.get("salary_mode") or "税后"

    # 目标工资来源：有固定工资时按 VBA 业务规则覆盖；否则已导入外部实发表的月份保留当月实发输入；
    # 没有外部实发表时才回到母表结构工资。这样修改母表能自动生效，又不会抹掉真实月度实发。
    profile_fixed = flt(emp_doc.get("fixed_salary"))
    structured_salary = round(
        flt(emp_doc.get("base_salary")) + flt(emp_doc.get("post_allowance")) + flt(emp_doc.get("performance_base"))
        + flt(emp_doc.get("meal_allowance")) + flt(emp_doc.get("traffic_allowance"))
        + flt(emp_doc.get("communication_allowance")) + flt(emp_doc.get("other_allowance")), 2
    )
    has_external_salary = bool(getattr(doc, "imported_excel_file", None))
    if is_tax_after_salary_mode(sal_mode):
        target_value = profile_fixed if (refresh_from_profile and profile_fixed > 0) else flt(it.net_salary)
        if refresh_from_profile and profile_fixed <= 0 and not has_external_salary and structured_salary > 0:
            target_value = structured_salary
    else:
        target_value = profile_fixed if (refresh_from_profile and profile_fixed > 0) else flt(it.gross_salary)
        if refresh_from_profile and profile_fixed <= 0 and not has_external_salary and structured_salary > 0:
            target_value = structured_salary

    if not is_tax_ledger_employee(emp_type):
        if is_tax_after_salary_mode(sal_mode) and target_value > 0:
            it.gross_salary = round(target_value + ded_cur, 2)
            it.net_salary = round(target_value, 2)
        else:
            it.gross_salary = round(target_value, 2)
            it.net_salary = round(flt(it.gross_salary) - ded_cur, 2)
        it.tax_amount = 0.0
        it.taxable_income = 0.0
        it.tax_threshold = 0.0
    elif is_tax_after_salary_mode(sal_mode) and target_value > 0:
        calc_res = derive_gross_from_net_vba(
            net_salary=target_value,
            deduction_cur=ded_cur,
            gross_prior=pdata["gross_prior"],
            threshold_cur=tax_thresh,
            threshold_prior=pdata["threshold_prior"],
            spec_ded_cur=ded_cur,
            spec_ded_prior=pdata["spec_ded_prior"],
            spec_add_cur=spec_add_cur,
            spec_add_prior=pdata["spec_add_prior"],
            paid_tax_prior=pdata["paid_tax_prior"],
        )
        it.gross_salary = calc_res["gross_salary"]
        it.tax_amount = calc_res["tax_amount_cur"]
        it.taxable_income = calc_res["taxable_income"]
        it.net_salary = calc_res["net_verified"]
    else:
        gross_cur = round(target_value, 2)
        it.gross_salary = gross_cur
        gross_all = round(pdata["gross_prior"] + gross_cur, 2)
        spec_ded_all = round(pdata["spec_ded_prior"] + ded_cur, 2)
        spec_add_all = round(pdata["spec_add_prior"] + spec_add_cur, 2)
        thresh_all = round(pdata["threshold_prior"] + tax_thresh, 2)
        taxable_all_raw = round(gross_all - spec_ded_all - spec_add_all - thresh_all, 2)
        taxable_for_tax = max(0.0, taxable_all_raw)
        rate, quick = 0.03, 0.0
        for _lower, upper, rate_value, quick_value in TAX_BRACKETS:
            if taxable_for_tax <= upper:
                rate, quick = rate_value, quick_value
                break
        cumulative_tax = round(taxable_for_tax * rate - quick, 2)
        current_tax = max(0.0, round(cumulative_tax - pdata["paid_tax_prior"], 2))
        it.tax_amount = current_tax
        it.taxable_income = max(0.0, taxable_all_raw)
        it.net_salary = round(gross_cur - ded_cur - current_tax, 2)

    _set_payroll_item_calculation_audit(it, "已计算", trigger_source, task_name, input_hash)
    return it


def _refresh_monthly_payroll_totals(doc):
    """Rebuild parent totals from child rows after one or many employee recalculations."""
    doc.total_employees = len(doc.items)
    doc.total_gross_salary = round(sum(flt(it.gross_salary) for it in doc.items), 2)
    doc.total_net_salary = round(sum(flt(it.net_salary) for it in doc.items), 2)
    doc.total_tax = round(sum(flt(it.tax_amount) for it in doc.items), 2)
    doc.total_social_security_person = round(sum(flt(it.ss_person_total) for it in doc.items), 2)
    doc.total_social_security_company = round(sum(flt(it.ss_company_total) for it in doc.items), 2)
    doc.total_housing_fund_person = round(sum(flt(it.hf_person_total) for it in doc.items), 2)
    doc.total_housing_fund_company = round(sum(flt(it.hf_company_total) for it in doc.items), 2)


def _get_unlocked_settlement(company, period_month):
    doc_name = f"{company}-{period_month}"
    if not frappe.db.exists("Ashan Monthly Payroll Settlement", doc_name):
        frappe.throw(f"【{company}】尚未生成 {period_month} 账期数据，请先创建或导入！")
    doc = frappe.get_doc("Ashan Monthly Payroll Settlement", doc_name)
    if doc.locked or doc.status in ["已核定锁定", "已归档发放"]:
        frappe.throw(f"【{company}】{period_month} 月度薪酬已被【锁定】，无法重新计算！如需修改请先执行反审核解锁。")
    return doc


def _append_payroll_item_from_profile(doc, company, employee_no):
    """Create a missing current-month row from the authoritative employee profile.

    This is used when a new employee is added after a monthly settlement has already
    been created.  It deliberately creates only an input snapshot; the unified VBA
    calculator fills contribution/tax results immediately afterwards.
    """
    profile = frappe.db.get_value(
        "Ashan Employee Salary Profile",
        {"company": company, "employee_no": employee_no},
        [
            "employee_no", "employee_name", "id_card", "gender", "mobile", "birth_date",
            "department", "job_title", "employee_type", "salary_mode", "fixed_salary",
            "base_salary", "post_allowance", "performance_base", "meal_allowance",
            "traffic_allowance", "communication_allowance", "other_allowance",
            "deduction_child_education", "deduction_continuing_education",
            "deduction_serious_illness", "deduction_housing_loan", "deduction_housing_rent",
            "deduction_elderly_care", "deduction_infant_care",
        ],
        as_dict=True,
    )
    if not profile:
        frappe.throw(f"未找到员工 {employee_no} 的薪酬母表档案。")

    fixed_salary = flt(profile.get("fixed_salary"))
    structured_salary = round(
        flt(profile.get("base_salary")) + flt(profile.get("post_allowance")) + flt(profile.get("performance_base"))
        + flt(profile.get("meal_allowance")) + flt(profile.get("traffic_allowance"))
        + flt(profile.get("communication_allowance")) + flt(profile.get("other_allowance")), 2
    )
    target = round(fixed_salary if fixed_salary > 0 else structured_salary, 2)
    salary_mode = profile.get("salary_mode") or "税后"
    row = doc.append("items", {
        "employee_no": profile.get("employee_no"),
        "employee_name": profile.get("employee_name"),
        "id_card": profile.get("id_card") or "",
        "gender": profile.get("gender") or "",
        "mobile": profile.get("mobile") or "",
        "birth_date": profile.get("birth_date"),
        "department": profile.get("department") or "",
        "job_title": profile.get("job_title") or "",
        "employee_type": profile.get("employee_type") or "正式工",
        "salary_mode": salary_mode,
        "fixed_salary": fixed_salary,
        "base_salary": flt(profile.get("base_salary")),
        "post_allowance": flt(profile.get("post_allowance")),
        "performance_salary": flt(profile.get("performance_base")),
        "allowances_total": round(
            flt(profile.get("meal_allowance")) + flt(profile.get("traffic_allowance"))
            + flt(profile.get("communication_allowance")) + flt(profile.get("other_allowance")), 2
        ),
        "gross_salary": target if not is_tax_after_salary_mode(salary_mode) else 0.0,
        "net_salary": target if is_tax_after_salary_mode(salary_mode) else 0.0,
        "deduction_child_education": flt(profile.get("deduction_child_education")),
        "deduction_continuing_education": flt(profile.get("deduction_continuing_education")),
        "deduction_serious_illness": flt(profile.get("deduction_serious_illness")),
        "deduction_housing_loan": flt(profile.get("deduction_housing_loan")),
        "deduction_housing_rent": flt(profile.get("deduction_housing_rent")),
        "deduction_elderly_care": flt(profile.get("deduction_elderly_care")),
        "deduction_infant_care": flt(profile.get("deduction_infant_care")),
        "remarks": "月度结算创建后新增员工，由服务器计算中心自动补入",
    })
    return row


@frappe.whitelist(methods=["POST"])
def recalculate_employee_payroll(
    company="天津祺富机械加工有限公司",
    period_month="2026-07",
    employee_no=None,
    trigger_source="人工重算",
    task_name="",
    input_hash="",
    refresh_from_profile=1,
):
    """Server-side single-employee calculator used by asynchronous background jobs."""
    check_payroll_workbench_permission("write")
    if not employee_no:
        frappe.throw("必须指定需要重新计算的员工工号。")
    doc = _get_unlocked_settlement(company, period_month)
    item = next((row for row in doc.items if row.employee_no == employee_no), None)
    if not item:
        item = _append_payroll_item_from_profile(doc, company, employee_no)
    _recalculate_payroll_item_vba(
        doc, item, company, period_month, trigger_source, task_name, input_hash, bool(cint(refresh_from_profile))
    )
    _refresh_monthly_payroll_totals(doc)
    frappe.flags.ignore_lock = True
    doc.save(ignore_permissions=True)
    return {
        "success": True,
        "employee_no": employee_no,
        "period_month": period_month,
        "gross_salary": flt(item.gross_salary),
        "tax_amount": flt(item.tax_amount),
        "net_salary": flt(item.net_salary),
    }


@frappe.whitelist(methods=["POST"])
def recalculate_and_save_monthly_tax(
    company="天津祺富机械加工有限公司",
    period_month="2026-07",
    trigger_source="人工重算",
    task_name="",
    force_recompute=0,
):
    """Recalculate an unlocked month sequentially with the single verified VBA engine."""
    check_payroll_workbench_permission("write")
    doc = _get_unlocked_settlement(company, period_month)
    from ashan_cn_procurement.services.payroll_recalculation_service import _build_employee_input_hash

    for item in doc.items:
        input_hash = _build_employee_input_hash(company, period_month, item.employee_no)
        if (
            not cint(force_recompute)
            and getattr(item, "calculation_input_hash", "") == input_hash
            and getattr(item, "calculation_status", "") == "已计算"
        ):
            continue
        _recalculate_payroll_item_vba(doc, item, company, period_month, trigger_source, task_name, input_hash)

    _refresh_monthly_payroll_totals(doc)
    frappe.flags.ignore_lock = True
    doc.save(ignore_permissions=True)
    return {
        "success": True,
        "message": f"【{company}】{period_month} 已由服务器统一按 VBA 同口径完成薪酬与累计个税计算。",
        "data": get_tax_settlement_full_sheet(company, period_month),
    }

# =========================================================================
# 凭证上传、解析、自动解压归档、纯净下载与月度封账控制链路
# =========================================================================

def get_month_lock_status(company, period_month):
	"""
	返回指定公司+月份的核定封账状态。
	前端 check_and_apply_month_lock_status 使用此方法决定 UI 是否只读。
	"""
	parent_name = f"{company}-{period_month}"
	status = frappe.db.get_value(
		"Ashan Monthly Payroll Settlement",
		parent_name,
		"status"
	)
	is_locked = (status in ["已核定锁定", "已提交", "Locked", "Submitted"])
	return {
		"is_locked": is_locked,
		"status": status or "未核定",
		"period_month": period_month,
		"company": company
	}


# ==============================================================================
# 月度人事薪酬业务全流程任务中枢、PDF/ZIP 智能解析核验与封账引擎
# ==============================================================================
import base64
import io
import re
import zipfile
from pypdf import PdfReader
from ashan_cn_procurement.services.payroll_proof_validation import (
    expected_proof_period,
    expand_upload_entries_to_pdfs,
    parse_social_security_pdf_stream as _parse_social_security_pdf_stream_precise,
    parse_housing_fund_pdf_stream as _parse_housing_fund_pdf_stream_precise,
    validate_proof_pdf_batch,
)

def parse_social_security_pdf_stream(file_bytes):
	"""兼容旧调用名，实际使用独立纯解析模块的精准解析器。"""
	return _parse_social_security_pdf_stream_precise(file_bytes)

def parse_housing_fund_pdf_stream(file_bytes):
	"""兼容旧调用名，实际使用独立纯解析模块的精准解析器。"""
	return _parse_housing_fund_pdf_stream_precise(file_bytes)

@frappe.whitelist()
def get_monthly_workflow_status(company, period_month):
	"""
	获取指定月份的全流程任务状态看板数据 (精细化多维度读数与异动摘要)
	"""
	check_payroll_workbench_permission("read")
	# 计算下个月份
	parts = period_month.split("-")
	y = int(parts[0]) if len(parts) > 0 else 2026
	m = int(parts[1]) if len(parts) > 1 else 7
	next_y = y + 1 if m == 12 else y
	next_m = 1 if m == 12 else m + 1
	next_period_month = f"{next_y}-{str(next_m).zfill(2)}"

	# 1. 档案状态：按账期判断实际在册，而不是只看员工今天的状态。
	emp_profiles = _salary_profiles_for_period(
		company,
		period_month,
		fields=[
			"name", "employee_no", "employee_name", "employee_type", "fixed_salary",
			"social_security_base", "housing_fund_base", "employment_status", "date_of_joining", "relieving_date",
			"deduction_child_education", "deduction_continuing_education", "deduction_housing_loan", "deduction_housing_rent",
			"deduction_elderly_care", "deduction_infant_care", "deduction_serious_illness"
		],
	)
	emp_count = len(emp_profiles)
	
	# 离职动态分析：只统计本账期发生的离职，不把所有历史离职人员混入。
	resigned_in_period = [e for e in emp_profiles if e.get("relieving_date") and str(e.get("relieving_date")).startswith(period_month)]
	if resigned_in_period:
		profile_change_text = f"当月办理离职 {len(resigned_in_period)} 人 (次月已减员)"
	else:
		profile_change_text = "人员及配置无异动"
	
	# 2. 社保与公积金系统核算总额与参保人数
	ss_data = get_social_insurance_sheet(company, period_month)
	ss_totals = ss_data.get("totals", {})
	ss_sys_comp = flt(ss_totals.get("comp_total", 0.0))
	ss_sys_pers = flt(ss_totals.get("pers_total", 0.0))
	ss_sys_total = flt(ss_totals.get("grand_total", 0.0))
	ss_insured_count = len(ss_data.get("rows", []))
	
	hf_data = get_housing_fund_sheet(company, period_month)
	hf_totals = hf_data.get("totals", {})
	hf_sys_comp = flt(hf_totals.get("comp_amount", 0.0))
	hf_sys_pers = flt(hf_totals.get("pers_amount", 0.0))
	hf_sys_total = flt(hf_totals.get("total_amount", 0.0))
	hf_insured_count = len(hf_data.get("rows", []))

	# 3. 结算主表状态与附件
	doc_name = f"{company}-{period_month}"
	doc = None
	if frappe.db.exists("Ashan Monthly Payroll Settlement", doc_name):
		doc = frappe.get_doc("Ashan Monthly Payroll Settlement", doc_name)

	is_locked = bool(doc and doc.locked)
	status_label = doc.status if doc else "草稿"
	
	# 任务 2: 车间实发表 (精准 25 人，排除非车间人员)
	dist_sheet = get_salary_distribution_sheet(company, period_month)
	dist_rows = dist_sheet.get("rows", [])
	has_items = bool(len(dist_rows) > 0)
	import_status = "done" if has_items else "pending"
	import_emp_count = len(dist_rows)
	import_net_total = flt(dist_sheet.get("totals", {}).get("net_salary", 0.0))
	import_file_url = doc.get("imported_excel_file") if doc else None

	# 任务 3/4：核定期对应的法定凭证属于次月实际缴费月，例如 2026-07 核定 -> 2026-08 凭证。
	expected_proof_month = expected_proof_period(period_month)

	ss_file_url = doc.get("ss_payment_file") if doc else None
	ss_verify_status = doc.get("ss_verify_status") if doc else ("核验一致" if ss_file_url else "未上传")
	ss_saved_batch = _inspect_saved_proof_batch(doc, "social_security", period_month) if ss_file_url else _empty_proof_batch(expected_proof_month)
	ss_parsed_amount = flt(ss_saved_batch.get("total_amount")) if ss_saved_batch.get("file_count") else (flt(doc.get("ss_parsed_amount")) if doc else 0.0)
	ss_difference_amount = round(abs(ss_parsed_amount - ss_sys_total), 2)
	ss_period_valid = bool(ss_saved_batch.get("period_valid")) if ss_file_url else False
	if not ss_file_url:
		ss_status = "pending"
	elif ss_period_valid and ss_verify_status == "核验一致" and ss_difference_amount < 0.01:
		ss_status = "verified"
	else:
		ss_status = "mismatch"

	hf_file_url = doc.get("hf_payment_file") if doc else None
	hf_verify_status = doc.get("hf_verify_status") if doc else ("核验一致" if hf_file_url else "未上传")
	hf_saved_batch = _inspect_saved_proof_batch(doc, "housing_fund", period_month) if hf_file_url else _empty_proof_batch(expected_proof_month)
	hf_parsed_amount = flt(hf_saved_batch.get("total_amount")) if hf_saved_batch.get("file_count") else (flt(doc.get("hf_parsed_amount")) if doc else 0.0)
	hf_difference_amount = round(abs(hf_parsed_amount - hf_sys_total), 2)
	hf_period_valid = bool(hf_saved_batch.get("period_valid")) if hf_file_url else False
	if not hf_file_url:
		hf_status = "pending"
	elif hf_period_valid and hf_verify_status == "核验一致" and hf_difference_amount < 0.01:
		hf_status = "verified"
	else:
		hf_status = "mismatch"

	# 任务 5: 综合核定关键指标
	total_gross_salary = flt(doc.total_gross_salary) if doc else 0.0
	total_net_salary = flt(doc.total_net_salary) if doc else 0.0
	total_tax = flt(doc.total_tax) if doc else 0.0
	total_company_cost = total_gross_salary + ss_sys_comp + hf_sys_comp

	# 服务器计算中心必须完全同步后才能封账。浏览器显示仅作提示，最终判断始终在服务器。
	from ashan_cn_procurement.services.payroll_recalculation_service import get_payroll_calculation_readiness
	calculation = get_payroll_calculation_readiness(company, period_month)

	# 最终核定必须同时满足：基础数据就绪 + 后台计算完成 + 两类法定凭证日期/金额均核验一致。
	can_lock = bool(
		emp_count > 0
		and has_items
		and calculation.get("ready")
		and ss_status == "verified"
		and hf_status == "verified"
	)

	# 建立当期车间人员集合与未设薪人员
	workshop_emp_set = set(r.get("employee_no") for r in dist_rows)
	
	sys_calc_count = 0
	ext_calc_count = 0
	zero_calc_count = 0
	for e in emp_profiles:
		eno = e.get("employee_no")
		if eno in workshop_emp_set:
			ext_calc_count += 1
		elif flt(e.get("fixed_salary")) > 0:
			sys_calc_count += 1
		else:
			zero_calc_count += 1

	if zero_calc_count > 0:
		task1_sub_badge = f"系统计薪 {sys_calc_count}人 ｜ 外部实发计薪 {ext_calc_count}人 ｜ 0工资 {zero_calc_count}人"
	else:
		task1_sub_badge = f"系统计薪 {sys_calc_count}人 ｜ 外部实发计薪 {ext_calc_count}人" if has_items else f"系统计薪 {sys_calc_count}人 ｜ 外部计薪 待导入"

	return {
		"company": company,
		"period_month": period_month,
		"next_period_month": next_period_month,
		"is_locked": is_locked,
		"status": status_label,
		"task1_profile": {
			"status": "done",
			"active_count": emp_count,
			"sys_calc_count": sys_calc_count,
			"ext_calc_count": ext_calc_count,
			"zero_calc_count": zero_calc_count,
			"is_ext_imported": has_items,
			"change_text": profile_change_text,
			"sub_badge": task1_sub_badge,
			"label": f"在册 {emp_count}人 · {profile_change_text}"
		},
		"task2_import": {
			"status": import_status,
			"employee_count": import_emp_count,
			"non_workshop_count": max(0, emp_count - import_emp_count),
			"total_net": import_net_total,
			"file_url": import_file_url,
			"sub_badge": f"车间实发 {import_emp_count}人 ｜ 非车间(母表) {max(0, emp_count - import_emp_count)}人" if has_items else f"外部计薪 {ext_calc_count}人 待导入实发表",
			"label": f"已导入 {import_emp_count} 人 · ¥{import_net_total:,.2f}" if has_items else "待导入车间实发表"
		},
		"task3_ss": {
			"status": ss_status,
			"file_url": ss_file_url,
			"insured_count": ss_insured_count,
			"parsed_amount": ss_parsed_amount,
			"company_amount": ss_sys_comp,
			"person_amount": ss_sys_pers,
			"sys_amount": ss_sys_total,
			"difference_amount": ss_difference_amount,
			"verify_status": ss_verify_status,
			"expected_period": expected_proof_month,
			"detected_periods": ss_saved_batch.get("detected_periods", []),
			"period_valid": ss_period_valid,
			"file_count": ss_saved_batch.get("file_count", 0),
			"validation_errors": ss_saved_batch.get("errors", []),
			"label": f"{ss_insured_count}人参保 · 申报总盘 ¥{ss_parsed_amount:,.2f}" if ss_file_url else f"{ss_insured_count}人参保 · 待上传社保PDF (系统应缴: ¥{ss_sys_total:,.2f})"
		},
		"task4_hf": {
			"status": hf_status,
			"file_url": hf_file_url,
			"insured_count": hf_insured_count,
			"parsed_amount": hf_parsed_amount,
			"company_amount": hf_parsed_amount / 2.0 if hf_parsed_amount > 0 else hf_sys_comp,
			"person_amount": hf_parsed_amount / 2.0 if hf_parsed_amount > 0 else hf_sys_pers,
			"sys_amount": hf_sys_total,
			"difference_amount": hf_difference_amount,
			"verify_status": hf_verify_status,
			"expected_period": expected_proof_month,
			"detected_periods": hf_saved_batch.get("detected_periods", []),
			"period_valid": hf_period_valid,
			"file_count": hf_saved_batch.get("file_count", 0),
			"validation_errors": hf_saved_batch.get("errors", []),
			"label": f"{hf_insured_count}人参缴 · 凭证总额 ¥{hf_parsed_amount:,.2f}" if hf_file_url else f"{hf_insured_count}人参缴 · 待上传公积金ZIP/PDF (系统应缴: ¥{hf_sys_total:,.2f})"
		},
		"task5_settlement": {
			"total_gross": total_gross_salary,
			"total_net": total_net_salary,
			"total_tax": total_tax,
			"total_company_cost": total_company_cost
		},
		"calculation": calculation,
		"can_lock": can_lock
	}

def _empty_proof_batch(expected_period):
	return {"success": False, "period_valid": False, "expected_period": expected_period, "detected_periods": [], "file_count": 0, "total_amount": 0.0, "files": [], "errors": []}


def _proof_type_config(proof_type):
	if proof_type in ["social_security", "ss", "pdf_ss"]:
		return {
			"field": "ss_payment_file",
			"token": "社会保险缴费申报表_原始凭证",
			"tokens": ["社会保险", "社保", "ss_payment"],
			"title": "社会保险缴费申报表"
		}
	if proof_type in ["housing_fund", "hf", "pdf_hf"]:
		return {
			"field": "hf_payment_file",
			"token": "住房公积金缴存凭证_原始凭证",
			"tokens": ["住房公积金", "公积金", "hf_payment"],
			"title": "住房公积金缴存凭证"
		}
	raise ValueError(f"未知凭证类型: {proof_type}")


def _decode_proof_upload_entries(file_name=None, file_base64=None, file_url=None, files_json=None):
	entries = []
	if files_json:
		payload = json.loads(files_json) if isinstance(files_json, str) else files_json
		if not isinstance(payload, list) or not payload:
			raise ValueError("未收到有效的多文件凭证列表")
		if len(payload) > 20:
			raise ValueError("单次最多允许上传 20 个 PDF/ZIP 凭证文件")
		for item in payload:
			name = str((item or {}).get("file_name") or "").strip()
			b64 = (item or {}).get("file_base64")
			if not name or not b64:
				raise ValueError("多文件列表存在缺少文件名或文件内容的项目")
			if "," in b64:
				b64 = b64.split(",", 1)[1]
			entries.append({"file_name": name, "raw_bytes": base64.b64decode(b64)})
	elif file_url:
		file_doc = frappe.get_doc("File", {"file_url": file_url})
		entries.append({"file_name": file_name or file_doc.file_name, "raw_bytes": file_doc.get_content()})
	elif file_base64:
		b64 = file_base64.split(",", 1)[1] if "," in file_base64 else file_base64
		entries.append({"file_name": file_name or "proof.pdf", "raw_bytes": base64.b64decode(b64)})
	else:
		raise ValueError("未提供有效的凭证文件")
	return entries


def _get_attached_proof_files(settle_doc, proof_type):
	if not settle_doc:
		return []
	config = _proof_type_config(proof_type)
	rows = frappe.get_all(
		"File",
		filters={"attached_to_doctype": "Ashan Monthly Payroll Settlement", "attached_to_name": settle_doc.name},
		fields=["name", "file_name", "file_url", "creation"],
		order_by="creation desc",
	)
	matched = [r for r in rows if any(t in (r.get("file_name") or "") or t in (r.get("file_url") or "") for t in config["tokens"])]
	if matched:
		# 按 file_url 和 file_name 深度去重，确保同一凭证绝不被多次统计
		seen = set()
		unique_matched = []
		for r in matched:
			key = (r.get("file_name") or "", r.get("file_url") or "")
			if key not in seen:
				seen.add(key)
				unique_matched.append(r)
		return unique_matched

	legacy_url = settle_doc.get(config["field"])
	if legacy_url:
		legacy = frappe.db.get_value("File", {"file_url": legacy_url}, ["name", "file_name", "file_url", "creation"], as_dict=True)
		return [legacy] if legacy else []
	return []


def _inspect_saved_proof_batch(settle_doc, proof_type, payroll_period_month):
	expected = expected_proof_period(payroll_period_month)
	files = _get_attached_proof_files(settle_doc, proof_type)
	if not files:
		return _empty_proof_batch(expected)
	pdf_entries = []
	for row in files:
		file_doc = frappe.get_doc("File", row.get("name"))
		content = file_doc.get_content()
		if isinstance(content, str):
			content = content.encode()
		pdf_entries.append({"source_name": row.get("file_name"), "pdf_name": row.get("file_name"), "pdf_bytes": content})
	validation = validate_proof_pdf_batch(proof_type, pdf_entries, expected)
	validation["detected_periods"] = sorted({p for item in validation.get("files", []) for p in item.get("period_months", [])})
	validation["period_valid"] = bool(validation.get("success"))
	return validation


def _delete_existing_proof_files(settle_doc, proof_type):
	if not settle_doc:
		return
	config = _proof_type_config(proof_type)
	all_attached = frappe.get_all(
		"File",
		filters={"attached_to_doctype": "Ashan Monthly Payroll Settlement", "attached_to_name": settle_doc.name},
		fields=["name", "file_name", "file_url"],
	)
	for row in all_attached:
		fname = row.get("file_name") or ""
		furl = row.get("file_url") or ""
		is_match = any(t in fname or t in furl for t in config["tokens"])
		if not is_match and settle_doc.get(config["field"]) and furl == settle_doc.get(config["field"]):
			is_match = True
		if is_match:
			try:
				frappe.delete_doc("File", row["name"], ignore_permissions=True, force=True)
			except Exception:
				frappe.db.delete("File", {"name": row["name"]})
	frappe.db.commit()


def _save_proof_pdf_batch(settle_doc, proof_type, payroll_period_month, pdf_entries):
	config = _proof_type_config(proof_type)
	actual_period = expected_proof_period(payroll_period_month)
	count = len(pdf_entries)
	saved = []
	for idx, entry in enumerate(pdf_entries, start=1):
		suffix = f"_{idx:02d}" if count > 1 else ""
		file_name = f"{actual_period}_{settle_doc.company}_{config['token']}{suffix}.pdf"
		file_doc = frappe.get_doc({
			"doctype": "File",
			"file_name": file_name,
			"attached_to_doctype": "Ashan Monthly Payroll Settlement",
			"attached_to_name": settle_doc.name,
			"content": entry["pdf_bytes"],
			"is_private": 1
		})
		file_doc.save(ignore_permissions=True)
		saved.append({"file_name": file_doc.file_name, "file_url": file_doc.file_url})
	return saved


@frappe.whitelist(methods=["POST"])
def upload_and_verify_social_security_file(company, period_month, file_name=None, file_base64=None, file_url=None, files_json=None):
	"""一个或多个社保 PDF/ZIP：先校验次月所属期，日期错误整批拒绝且不保存，再汇总金额。"""
	check_payroll_workbench_permission("write")
	expected_period = expected_proof_period(period_month)
	try:
		upload_entries = _decode_proof_upload_entries(file_name, file_base64, file_url, files_json)
		pdf_entries = expand_upload_entries_to_pdfs(upload_entries)
		validation = validate_proof_pdf_batch("social_security", pdf_entries, expected_period)
	except Exception as exc:
		return {"success": False, "validation_type": "file_or_parse_error", "message": f"⛔ 社保凭证解析失败：{exc}。未保存任何文件。"}

	if not validation.get("success"):
		return {"success": False, "validation_type": "period_mismatch", "message": f"⛔ 社保凭证所属期校验未通过。当前核定期 {period_month} 对应实际社保缴费所属期必须为 {expected_period}。<br>" + "<br>".join(validation.get("errors") or []) + "<br><strong>本批文件已全部拒绝，未保存任何附件。</strong>", "expected_period": expected_period, "files": validation.get("files", [])}

	parsed_amount = flt(validation.get("total_amount", 0.0))
	sys_amount = flt(get_social_insurance_sheet(company, period_month).get("totals", {}).get("grand_total", 0.0))
	diff = round(abs(parsed_amount - sys_amount), 2)
	is_matched = diff < 0.01
	doc_name = f"{company}-{period_month}"
	if not frappe.db.exists("Ashan Monthly Payroll Settlement", doc_name):
		settle_doc = frappe.new_doc("Ashan Monthly Payroll Settlement")
		settle_doc.company = company
		settle_doc.period_month = period_month
		settle_doc.status = "草稿"
		settle_doc.insert(ignore_permissions=True)
	else:
		settle_doc = frappe.get_doc("Ashan Monthly Payroll Settlement", doc_name)
	_delete_existing_proof_files(settle_doc, "social_security")
	saved_files = _save_proof_pdf_batch(settle_doc, "social_security", period_month, pdf_entries)
	settle_doc.ss_payment_file = saved_files[0]["file_url"] if saved_files else None
	settle_doc.ss_parsed_amount = parsed_amount
	settle_doc.ss_verify_status = "核验一致" if is_matched else "金额不符"
	settle_doc.save(ignore_permissions=True)
	frappe.db.commit()
	return {"success": True, "message": f"✅ 社保凭证所属期全部通过（核定期 {period_month} -> 实际缴费所属期 {expected_period}），共 {len(saved_files)} 份 PDF。" + (f" 合计 ¥{parsed_amount:,.2f} 与系统核算 ¥{sys_amount:,.2f} 完全一致。" if is_matched else f" 但凭证合计 ¥{parsed_amount:,.2f} 与系统核算 ¥{sys_amount:,.2f} 不一致，差额 ¥{diff:,.2f}，禁止最终封账。"), "expected_period": expected_period, "proof_count": len(saved_files), "parsed_amount": parsed_amount, "sys_amount": sys_amount, "difference_amount": diff, "is_matched": is_matched, "file_url": saved_files[0]["file_url"] if saved_files else None, "file_urls": saved_files, "files": validation.get("files", [])}

@frappe.whitelist(methods=["POST"])
def upload_and_verify_housing_fund_file(company, period_month, file_name=None, file_base64=None, file_url=None, files_json=None):
	"""一个或多个公积金 PDF/ZIP：先校验次月缴存年月，日期错误整批拒绝且不保存，再汇总金额。"""
	check_payroll_workbench_permission("write")
	expected_period = expected_proof_period(period_month)
	try:
		upload_entries = _decode_proof_upload_entries(file_name, file_base64, file_url, files_json)
		pdf_entries = expand_upload_entries_to_pdfs(upload_entries)
		validation = validate_proof_pdf_batch("housing_fund", pdf_entries, expected_period)
	except Exception as exc:
		return {"success": False, "validation_type": "file_or_parse_error", "message": f"⛔ 公积金凭证解析失败：{exc}。未保存任何文件。"}
	if not validation.get("success"):
		return {"success": False, "validation_type": "period_mismatch", "message": f"⛔ 公积金凭证所属期校验未通过。当前核定期 {period_month} 对应实际公积金缴存年月必须为 {expected_period}。<br>" + "<br>".join(validation.get("errors") or []) + "<br><strong>本批文件已全部拒绝，未保存任何附件。</strong>", "expected_period": expected_period, "files": validation.get("files", [])}
	parsed_amount = flt(validation.get("total_amount", 0.0))
	sys_amount = flt(get_housing_fund_sheet(company, period_month).get("totals", {}).get("total_amount", 0.0))
	diff = round(abs(parsed_amount - sys_amount), 2)
	is_matched = diff < 0.01
	doc_name = f"{company}-{period_month}"
	if not frappe.db.exists("Ashan Monthly Payroll Settlement", doc_name):
		settle_doc = frappe.new_doc("Ashan Monthly Payroll Settlement")
		settle_doc.company = company
		settle_doc.period_month = period_month
		settle_doc.status = "草稿"
		settle_doc.insert(ignore_permissions=True)
	else:
		settle_doc = frappe.get_doc("Ashan Monthly Payroll Settlement", doc_name)
	_delete_existing_proof_files(settle_doc, "housing_fund")
	saved_files = _save_proof_pdf_batch(settle_doc, "housing_fund", period_month, pdf_entries)
	settle_doc.hf_payment_file = saved_files[0]["file_url"] if saved_files else None
	settle_doc.hf_parsed_amount = parsed_amount
	settle_doc.hf_verify_status = "核验一致" if is_matched else "金额不符"
	settle_doc.save(ignore_permissions=True)
	frappe.db.commit()
	return {"success": True, "message": f"✅ 公积金凭证所属期全部通过（核定期 {period_month} -> 实际缴费所属期 {expected_period}），共 {len(saved_files)} 份 PDF。" + (f" 合计 ¥{parsed_amount:,.2f} 与系统核算 ¥{sys_amount:,.2f} 完全一致。" if is_matched else f" 但凭证合计 ¥{parsed_amount:,.2f} 与系统核算 ¥{sys_amount:,.2f} 不一致，差额 ¥{diff:,.2f}，禁止最终封账。"), "expected_period": expected_period, "proof_count": len(saved_files), "parsed_amount": parsed_amount, "sys_amount": sys_amount, "difference_amount": diff, "is_matched": is_matched, "file_url": saved_files[0]["file_url"] if saved_files else None, "file_urls": saved_files, "files": validation.get("files", [])}

@frappe.whitelist(methods=["POST"])
def execute_monthly_settlement_lock(company, period_month):
	"""
	执行当月薪酬综合核定并封账锁定，同时初始化开启下月发薪账期权限 (前置强拦截校验)
	"""
	check_payroll_workbench_permission("write")
	doc_name = f"{company}-{period_month}"
	if not frappe.db.exists("Ashan Monthly Payroll Settlement", doc_name):
		return {"success": False, "message": f"未找到【{company}】{period_month} 的薪酬核算记录，无法执行封账！"}

	settle_doc = frappe.get_doc("Ashan Monthly Payroll Settlement", doc_name)
	
	# 强拦截校验：必须完成实发表导入、服务器计算、社保上传核验、公积金上传核验。
	if not settle_doc.items or len(settle_doc.items) == 0:
		frappe.throw(f"【❌ 前置任务未完成】尚未形成【{period_month}】月度薪酬明细！请先完成任务 2 导入/建账。")

	from ashan_cn_procurement.services.payroll_recalculation_service import get_payroll_calculation_readiness
	calc_ready = get_payroll_calculation_readiness(company, period_month)
	if not calc_ready.get("ready"):
		frappe.throw(
			"【⛔ 禁止最终核定封账】服务器计算中心尚未完全同步。"
			f"待计算 {cint(calc_ready.get('pending'))}，排队 {cint(calc_ready.get('queued'))}，"
			f"计算中 {cint(calc_ready.get('running'))}，失败 {cint(calc_ready.get('failed'))}，"
			f"未计算 {cint(calc_ready.get('uncomputed'))}，活动任务 {cint(calc_ready.get('active_tasks'))}。"
			"请等待后台任务完成，或在【服务器计算中心】查看失败原因并重试。"
		)
	
	if not settle_doc.ss_payment_file:
		frappe.throw(f"【❌ 前置任务未完成】尚未上传【{period_month} 社保缴费申报表 PDF】并完成核验！请先完成任务 3 上传。")

	if not settle_doc.hf_payment_file:
		frappe.throw(f"【❌ 前置任务未完成】尚未上传【{period_month} 住房公积金缴存凭证 ZIP/PDF】并完成核验！请先完成任务 4 上传。")

	# 服务器端最终硬拦截：重新读取所有已归档 PDF，先重验所属期，再重验多文件金额合计。
	expected_period = expected_proof_period(period_month)
	ss_saved_batch = _inspect_saved_proof_batch(settle_doc, "social_security", period_month)
	if not ss_saved_batch.get("success"):
		frappe.throw(f"【⛔ 禁止最终核定封账】第 3 步 · 社保凭证所属期/结构校验失败。核定期 {period_month} 对应实际缴费所属期应为 {expected_period}。" + "；".join(ss_saved_batch.get("errors") or ["无法确认社保凭证所属期"]))
	ss_current_total = flt(get_social_insurance_sheet(company, period_month).get("totals", {}).get("grand_total", 0.0))
	ss_parsed_amount = flt(ss_saved_batch.get("total_amount"))
	ss_diff = round(abs(ss_parsed_amount - ss_current_total), 2)
	if settle_doc.ss_verify_status != "核验一致" or ss_diff >= 0.01:
		frappe.throw(f"【⛔ 禁止最终核定封账】第 3 步 · 社保 {ss_saved_batch.get('file_count', 0)} 份凭证合计 ¥{ss_parsed_amount:,.2f} 与系统 ¥{ss_current_total:,.2f} 不一致，差额 ¥{ss_diff:,.2f}。")

	hf_saved_batch = _inspect_saved_proof_batch(settle_doc, "housing_fund", period_month)
	if not hf_saved_batch.get("success"):
		frappe.throw(f"【⛔ 禁止最终核定封账】第 4 步 · 公积金凭证所属期/结构校验失败。核定期 {period_month} 对应实际缴费所属期应为 {expected_period}。" + "；".join(hf_saved_batch.get("errors") or ["无法确认公积金凭证所属期"]))
	hf_current_total = flt(get_housing_fund_sheet(company, period_month).get("totals", {}).get("total_amount", 0.0))
	hf_parsed_amount = flt(hf_saved_batch.get("total_amount"))
	hf_diff = round(abs(hf_parsed_amount - hf_current_total), 2)
	if settle_doc.hf_verify_status != "核验一致" or hf_diff >= 0.01:
		frappe.throw(f"【⛔ 禁止最终核定封账】第 4 步 · 公积金 {hf_saved_batch.get('file_count', 0)} 份凭证合计 ¥{hf_parsed_amount:,.2f} 与系统 ¥{hf_current_total:,.2f} 不一致，差额 ¥{hf_diff:,.2f}。")

	settle_doc.status = "已核定锁定"
	settle_doc.locked = 1
	settle_doc.confirmed_by = frappe.session.user
	settle_doc.confirmed_date = now_datetime()
	settle_doc.save(ignore_permissions=True)

	# 开启下月发薪账期
	parts = period_month.split("-")
	y = int(parts[0]) if len(parts) > 0 else 2026
	m = int(parts[1]) if len(parts) > 1 else 7
	next_y = y + 1 if m == 12 else y
	next_m = 1 if m == 12 else m + 1
	next_period_month = f"{next_y}-{str(next_m).zfill(2)}"

	next_doc_name = f"{company}-{next_period_month}"
	if not frappe.db.exists("Ashan Monthly Payroll Settlement", next_doc_name):
		next_doc = frappe.new_doc("Ashan Monthly Payroll Settlement")
		next_doc.company = company
		next_doc.period_month = next_period_month
		next_doc.status = "草稿"
		next_doc.locked = 0
		next_doc.insert(ignore_permissions=True)

	frappe.db.commit()

	return {
		"success": True,
		"message": f"🎉【{company}】{period_month} 薪酬与个税综合核定已成功封账锁定！所有明细已进入只读保护状态，并已为您自动开启下月【{next_period_month}】的发薪建账权限！",
		"next_period_month": next_period_month
	}

@frappe.whitelist(methods=["POST"])
def unlock_monthly_settlement(company, period_month, reason=""):
	"""
	反审核/解锁指定月份薪酬核定记录
	"""
	check_payroll_workbench_permission("write")
	if not str(reason or "").strip():
		frappe.throw("反审核/解锁必须填写原因，以便保留财务审计轨迹。")
	doc_name = f"{company}-{period_month}"
	if not frappe.db.exists("Ashan Monthly Payroll Settlement", doc_name):
		return {"success": False, "message": f"未找到【{company}】{period_month} 的薪酬核算记录！"}

	settle_doc = frappe.get_doc("Ashan Monthly Payroll Settlement", doc_name)
	settle_doc.status = "草稿"
	settle_doc.locked = 0
	settle_doc.unlock_reason = f"[{now_datetime()}] {frappe.session.user}: {reason}"
	settle_doc.save(ignore_permissions=True)
	frappe.db.commit()

	return {
		"success": True,
		"message": f"🔓【{company}】{period_month} 薪酬记录已成功解锁！已恢复数据编辑与重新导入权限。"
	}


@frappe.whitelist()
def download_payroll_proof_file(company, period_month, proof_type):
	"""
	【🔒 受控安全凭证下载接口】
	严格校验当前用户对人事薪酬的读取权限，并以规范中文名安全交付原始源文件流
	proof_type: 'salary' (车间实发), 'social_security' (社保申报), 'housing_fund' (公积金凭证)
	"""
	check_payroll_workbench_permission("read")

	doc_name = f"{company}-{period_month}"
	if not frappe.db.exists("Ashan Monthly Payroll Settlement", doc_name):
		frappe.throw(f"未找到【{company}】在【{period_month}】的月度薪酬结算主表！")

	settle_doc = frappe.get_doc("Ashan Monthly Payroll Settlement", doc_name)
	file_url = None
	default_filename = None

	if proof_type in ["salary", "workshop", "excel"]:
		file_url = settle_doc.imported_excel_file
		default_filename = f"{period_month}_{company}_车间实发工资表_原始凭证.xlsx"
	elif proof_type in ["social_security", "ss", "pdf_ss"]:
		file_url = settle_doc.ss_payment_file
		default_filename = f"{period_month}_{company}_社会保险缴费申报表_原始凭证.pdf"
	elif proof_type in ["housing_fund", "hf", "pdf_hf"]:
		file_url = settle_doc.hf_payment_file
		default_filename = f"{period_month}_{company}_住房公积金缴存凭证_原始凭证.pdf"
	else:
		frappe.throw("未知的凭证类型！")

	if not file_url:
		frappe.throw(f"【{period_month}】尚未上传该类型的凭证文件！")

	file_doc = frappe.get_doc("File", {"file_url": file_url})
	content = file_doc.get_content()
	dl_filename = file_doc.file_name or default_filename

	frappe.response.filename = dl_filename
	frappe.response.filecontent = content
	frappe.response.type = "download"


@frappe.whitelist(methods=["POST"])
def delete_payroll_proof_file(company, period_month, proof_type):
	"""
	【🗑️ 安全删除已上传凭证文件与重置任务状态】
	严格校验操作权限与封账状态，清除指定月份已上传的凭证文件及关联数据
	proof_type: 'salary' (车间实发表), 'social_security' (社保申报表), 'housing_fund' (公积金凭证)
	"""
	check_payroll_workbench_permission("write")

	doc_name = f"{company}-{period_month}"
	if not frappe.db.exists("Ashan Monthly Payroll Settlement", doc_name):
		return {"success": False, "message": f"未找到【{company}】在【{period_month}】的月度薪酬结算主表！"}

	settle_doc = frappe.get_doc("Ashan Monthly Payroll Settlement", doc_name)

	# 封账只读保护
	if settle_doc.locked == 1 or settle_doc.status == "已核定锁定":
		frappe.throw(f"【🔒 封账保护】{period_month} 当前已核定封账锁定，禁止删除凭证！如需调整请先申请反审核解锁。")

	deleted_type_name = ""

	if proof_type in ["salary", "workshop", "excel"]:
		deleted_type_name = "车间外部实发工资表"
		settle_doc.imported_excel_file = None
		settle_doc.items = []
		settle_doc.total_net_salary = 0.0
		settle_doc.total_gross_salary = 0.0
		settle_doc.total_individual_tax = 0.0
		settle_doc.employee_count = 0
		# 删除对应的子表 Item
		frappe.db.delete("Ashan Monthly Payroll Item", {"parent": doc_name})

	elif proof_type in ["social_security", "ss", "pdf_ss"]:
		deleted_type_name = "社会保险缴费申报表"
		_delete_existing_proof_files(settle_doc, "social_security")
		settle_doc.ss_payment_file = None
		settle_doc.ss_parsed_amount = 0.0
		settle_doc.ss_verify_status = "未上传"

	elif proof_type in ["housing_fund", "hf", "pdf_hf"]:
		deleted_type_name = "住房公积金缴存凭证"
		_delete_existing_proof_files(settle_doc, "housing_fund")
		settle_doc.hf_payment_file = None
		settle_doc.hf_parsed_amount = 0.0
		settle_doc.hf_verify_status = "未上传"

	else:
		frappe.throw("未知的凭证类型！")

	settle_doc.flags.ignore_version = True
	settle_doc.save(ignore_permissions=True)
	frappe.db.commit()

	return {
		"success": True,
		"message": f"🗑️ 已成功删除【{period_month}】的{deleted_type_name}！任务已重置为待上传状态。"
	}
