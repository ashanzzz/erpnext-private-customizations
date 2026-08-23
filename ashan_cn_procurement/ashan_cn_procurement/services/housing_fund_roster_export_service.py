"""Generate the official housing-fund increase and sealing rosters.

The local housing-fund portal supplies very small legacy ``.xls`` templates.  This
module deliberately recreates their exact columns with ``openpyxl`` and exports a
valid ``.xlsx`` workbook.  It never changes the employee master or stores a copy of
the generated roster: the browser receives the workbook directly.
"""

import base64
import io
from calendar import monthrange
from datetime import date

import frappe
from frappe.utils import flt, getdate
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from ashan_cn_procurement.services.employee_salary_service import get_insurance_setting
from ashan_cn_procurement.services.housing_fund_policy_service import (
    EXCLUDED_EMPLOYEE_TYPES,
    POLICY_FIXED_OFF,
    normalize_policy_setting,
    normalize_period_month,
)
from ashan_cn_procurement.services.payroll_proof_validation import expected_proof_period
from ashan_cn_procurement.services.payroll_settlement_service import (
    check_payroll_workbench_permission,
)


CHANGE_INCREASE = "increase"
CHANGE_SEAL = "seal"

ROSTER_DEFINITIONS = {
    CHANGE_INCREASE: {
        "label": "增加清册",
        "category": "zj",
        "headers": [
            "清册类别(必须填'zj')",
            "职工姓名",
            "职工身份证号码",
            "缴存基数",
            "金额合计",
        ],
        "widths": [7.71, 8.29, 25.0, 8.0, 8.0],
    },
    CHANGE_SEAL: {
        "label": "封存清册",
        "category": "fc",
        "headers": [
            "清册类别(必须填'fc')",
            "职工姓名",
            "职工身份证号",
            "封存原因(保留劳动关系时填'3'，解除劳动关系时填'4')",
        ],
        "widths": [16.14, 12.86, 25.0, 6.86],
    },
}


def _safe_date(value):
    """Return a date object for a master-field value, or ``None`` when absent."""
    if not value:
        return None
    try:
        return getdate(value)
    except Exception:
        return None


def _month_bounds(period_month):
    """Return the normalized period plus its inclusive first and last date."""
    period_month, year, month = normalize_period_month(period_month)
    return period_month, date(year, month, 1), date(year, month, monthrange(year, month)[1])


def _previous_period_month(period_month):
    """Return the immediately preceding accounting month in YYYY-MM form."""
    _period, start, _end = _month_bounds(period_month)
    previous_end = start.fromordinal(start.toordinal() - 1)
    return f"{previous_end.year:04d}-{previous_end.month:02d}"


def _is_housing_fund_member(profile):
    """Check the long-term membership boundary, independent of quarterly payment months."""
    employee_type = str(profile.get("employee_type") or "正式工").strip()
    policy = str(profile.get("housing_fund_policy") or "跟随公司规则").strip()
    return (
        employee_type not in EXCLUDED_EMPLOYEE_TYPES
        and policy != POLICY_FIXED_OFF
        and flt(profile.get("housing_fund_base")) > 0
    )


def _numeric_text(value):
    """Render an Excel-import-safe decimal without a currency sign or thousands separator."""
    number = round(flt(value), 2)
    return f"{number:.2f}".rstrip("0").rstrip(".")


def _active_housing_fund_member(profile, period_start, period_end):
    """Return whether a profile is a durable member at the selected month-end."""
    if not _is_housing_fund_member(profile):
        return False

    joined = _safe_date(profile.get("date_of_joining"))
    relieved = _safe_date(profile.get("relieving_date"))
    return not ((joined and joined > period_end) or (relieved and relieved <= period_end))


def _profile_roster_row(profile, total_rate):
    """Build the normalized roster data shared by previews and XLSX exports."""
    base = round(flt(profile.get("housing_fund_base")), 2)
    return {
        "employee_no": str(profile.get("employee_no") or "").strip(),
        "employee_name": str(profile.get("employee_name") or "").strip(),
        "id_card": str(profile.get("id_card") or "").strip(),
        "base": base,
        "amount": round(base * total_rate / 100, 2),
    }


def _is_scheduled_payment_material(company, settlement_period):
    """Keep quarterly off-month snapshots out of durable roster comparisons."""
    settlement_period, year, _month = normalize_period_month(settlement_period)
    rule = normalize_policy_setting(get_insurance_setting(company, year) or {})
    if not rule.get("enabled") or rule.get("off_action") == "继续缴纳":
        return True
    payment_month = int(expected_proof_period(settlement_period).split("-")[1])
    return payment_month in rule.get("months", [])


def _latest_material_members(company, period_month, total_rate):
    """Find the newest earlier full payment material used as roster baseline.

    A quarterly off-month may legitimately only contain long-term fixed contributors.
    It must not make every company-rule employee look newly added next month, so the
    baseline is the newest earlier scheduled contribution material with positive rows.
    """
    settlements = frappe.get_all(
        "Ashan Monthly Payroll Settlement",
        filters={"company": company},
        fields=["name", "period_month"],
        order_by="period_month desc",
        limit_page_length=120,
    )
    for settlement in settlements:
        raw_settlement_period = str(settlement.get("period_month") or "").strip()
        if not raw_settlement_period:
            continue
        try:
            settlement_period, _year, _month = normalize_period_month(raw_settlement_period)
        except Exception:
            continue
        if settlement_period >= period_month:
            continue
        if not _is_scheduled_payment_material(company, settlement_period):
            continue

        rows = frappe.get_all(
            "Ashan Monthly Payroll Item",
            filters={"parent": settlement.get("name")},
            fields=["employee_no", "employee_name", "id_card", "hf_base"],
            order_by="employee_no asc",
        )
        members = {}
        for row in rows:
            employee_no = str(row.get("employee_no") or "").strip()
            base = round(flt(row.get("hf_base")), 2)
            if not employee_no or base <= 0:
                continue
            members[employee_no] = {
                "employee_no": employee_no,
                "employee_name": str(row.get("employee_name") or "").strip(),
                "id_card": str(row.get("id_card") or "").strip(),
                "base": base,
                "amount": round(base * total_rate / 100, 2),
            }
        if members:
            return settlement_period, members
    return None, {}


def _get_change_rows(company, period_month):
    """Return durable housing-fund membership changes for the selected month.

    When prior contribution material exists, compare the current durable membership
    with that material.  Thus a deleted base or ``固定停缴`` produces a sealing row,
    while a new durable participant produces an increase row.  Payment-calendar
    on/off months never by themselves create roster changes.
    """
    period_month, period_start, period_end = _month_bounds(period_month)
    previous_period = _previous_period_month(period_month)
    _previous_period, previous_start, previous_end = _month_bounds(previous_period)
    year = period_start.year
    setting = get_insurance_setting(company, year) or {}
    total_rate = flt(setting.get("hf_company_rate")) + flt(setting.get("hf_person_rate"))

    profiles = frappe.get_all(
        "Ashan Employee Salary Profile",
        filters={"company": company},
        fields=[
            "employee_no", "employee_name", "id_card", "employee_type",
            "date_of_joining", "relieving_date", "housing_fund_base",
            "housing_fund_policy",
        ],
        order_by="employee_no asc",
    )

    current_members = {}
    for profile in profiles:
        if not _active_housing_fund_member(profile, period_start, period_end):
            continue
        row = _profile_roster_row(profile, total_rate)
        if row["employee_no"]:
            current_members[row["employee_no"]] = row

    reference_period, reference_members = _latest_material_members(
        company, period_month, total_rate
    )
    if reference_period:
        increases = [
            current_members[employee_no]
            for employee_no in sorted(set(current_members) - set(reference_members))
        ]
        seals = [
            reference_members[employee_no]
            for employee_no in sorted(set(reference_members) - set(current_members))
        ]
        comparison_mode = "material_delta"
    else:
        # Initial setup has no prior contribution material.  Preserve the former
        # conservative join/leave behavior until the first official material exists.
        increases = []
        seals = []
        for profile in profiles:
            row = _profile_roster_row(profile, total_rate)
            joined = _safe_date(profile.get("date_of_joining"))
            relieved = _safe_date(profile.get("relieving_date"))
            if _is_housing_fund_member(profile) and joined and period_start <= joined <= period_end:
                increases.append(row)
            if _is_housing_fund_member(profile) and relieved and previous_start <= relieved <= previous_end:
                seals.append(row)
        comparison_mode = "join_leave_fallback"

    return {
        "period_month": period_month,
        "previous_period_month": previous_period,
        "reference_period_month": reference_period,
        "comparison_mode": comparison_mode,
        "total_rate": round(total_rate, 4),
        CHANGE_INCREASE: increases,
        CHANGE_SEAL: seals,
    }


def _validate_roster_rows(rows, label):
    """Block invalid official exports instead of silently producing unusable records."""
    incomplete = [row.get("employee_no") or "未设工号" for row in rows if not row.get("employee_name") or not row.get("id_card")]
    if incomplete:
        frappe.throw(
            f"无法导出{label}：以下员工缺少姓名或身份证号码，请先补齐员工档案："
            f"{', '.join(incomplete)}"
        )


def _build_roster_workbook(change_type, rows):
    """Build an in-memory workbook matching the supplied legacy roster templates."""
    definition = ROSTER_DEFINITIONS[change_type]
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet0"

    text_font = Font(name="Arial", size=10)
    text_alignment = Alignment(horizontal="left", vertical="center")
    for column_index, (header, width) in enumerate(zip(definition["headers"], definition["widths"]), start=1):
        cell = worksheet.cell(row=1, column=column_index, value=header)
        cell.font = text_font
        cell.alignment = text_alignment
        cell.number_format = "@"
        worksheet.column_dimensions[cell.column_letter].width = width

    worksheet.row_dimensions[1].height = 12.75
    for row_index, row in enumerate(rows, start=2):
        if change_type == CHANGE_INCREASE:
            values = [
                definition["category"],
                row["employee_name"],
                row["id_card"],
                _numeric_text(row["base"]),
                _numeric_text(row["amount"]),
            ]
        else:
            values = [
                definition["category"],
                row["employee_name"],
                row["id_card"],
                "4",
            ]
        for column_index, value in enumerate(values, start=1):
            cell = worksheet.cell(row=row_index, column=column_index, value=value)
            cell.font = text_font
            cell.alignment = text_alignment
            cell.number_format = "@"
        worksheet.row_dimensions[row_index].height = 12.75

    content = io.BytesIO()
    workbook.save(content)
    return content.getvalue()


@frappe.whitelist()
def get_housing_fund_roster_change_preview(company="天津祺富机械加工有限公司", period_month=None):
    """Return only change counts for the workbench; no personal data is exposed here."""
    check_payroll_workbench_permission("read", company)
    changes = _get_change_rows(company, period_month)
    return {
        "success": True,
        "period_month": changes["period_month"],
        "previous_period_month": changes["previous_period_month"],
        "reference_period_month": changes["reference_period_month"],
        "comparison_mode": changes["comparison_mode"],
        "increase_count": len(changes[CHANGE_INCREASE]),
        "seal_count": len(changes[CHANGE_SEAL]),
        "seal_reason": "4",
    }


@frappe.whitelist()
def export_housing_fund_roster_change_excel(
    company="天津祺富机械加工有限公司", period_month=None, change_type=CHANGE_INCREASE,
):
    """Export a valid ``.xlsx`` housing-fund increase or sealing roster on demand."""
    check_payroll_workbench_permission("export", company)
    change_type = str(change_type or "").strip()
    if change_type not in ROSTER_DEFINITIONS:
        frappe.throw("清册类型不正确，只能导出增加清册或封存清册。")

    changes = _get_change_rows(company, period_month)
    definition = ROSTER_DEFINITIONS[change_type]
    rows = changes[change_type]
    _validate_roster_rows(rows, definition["label"])
    content = _build_roster_workbook(change_type, rows)
    filename = f"祺富公积金_{changes['period_month']}_{definition['label']}.xlsx"

    return {
        "success": True,
        "filename": filename,
        "file_base64": base64.b64encode(content).decode("ascii"),
        "record_count": len(rows),
        "message": f"已生成{definition['label']}：{len(rows)} 人。",
    }
