# Copyright (c) 2026, Ashan CN Procurement
# 天津吉众科技有限公司 - 专有考勤工时解析与业务归集服务
# 1:1 对齐《202606吉众人事综合.xlsm》考勤.bas与《员工考勤表-*.xlsx》规范

import os
import re
import json
import calendar
import datetime
import openpyxl
import frappe
from frappe.utils import flt, cint, getdate

from ashan_cn_procurement.services.ashan_holiday_service import (
	get_date_overtime_info,
	get_month_workdays,
)


@frappe.whitelist(methods=["POST"])
def upload_and_parse_attendance():
	"""
	接收前端上传的员工考勤表 Excel 文件并执行自动解析入库
	入参通过 frappe.form_dict 传入:
	  - file_url: 附件或上传文件地址 (或直接上传文件)
	  - period_month: 账期月份 (如 '2026-07')
	  - company: 公司名称 (默认 '天津吉众科技有限公司')
	"""
	company = frappe.form_dict.get("company") or "天津吉众科技有限公司"
	period_month = frappe.form_dict.get("period_month")
	file_url = frappe.form_dict.get("file_url")

	if not file_url and "file" in frappe.request.files:
		file_obj = frappe.request.files["file"]
		saved_file = frappe.get_doc({
			"doctype": "File",
			"file_name": file_obj.filename,
			"content": file_obj.read(),
			"is_private": 1,
			"attached_to_doctype": "Ashan Monthly Payroll Settlement",
			"attached_to_name": f"{company}-{period_month}" if period_month else company,
		})
		saved_file.insert(ignore_permissions=True)
		file_url = saved_file.file_url

	if not file_url:
		frappe.throw("请先选择或上传考勤 Excel 文件！")

	# 获取实际文件绝对路径
	file_doc = frappe.get_doc("File", {"file_url": file_url})
	local_path = file_doc.get_full_path()

	return parse_jizhong_attendance_file(local_path, period_month, company, file_url=file_url)


def parse_jizhong_attendance_file(file_path, period_month=None, company="天津吉众科技有限公司", file_url=None):
	"""
	核心考勤解析引擎：
	读取《员工考勤表-*.xlsx》，结合日历与倒休二倍工时冲抵算法，入库 Ashan Monthly Attendance
	"""
	if not os.path.exists(file_path):
		frappe.throw(f"考勤文件不存在: {file_path}")

	wb = openpyxl.load_workbook(file_path, data_only=True)
	ws = wb["总表"] if "总表" in wb.sheetnames else wb.worksheets[0]

	# 1. 动态推断或核实年份与月份
	inferred_year = None
	inferred_month = None

	# 尝试从第一行大标题提取 "2026 年 7 月"
	title_val = str(ws.cell(1, 1).value or "")
	match = re.search(r"(\d{4})\s*年\s*(\d{1,2})\s*月", title_val)
	if match:
		inferred_year = cint(match.group(1))
		inferred_month = cint(match.group(2))
	elif period_month and "-" in period_month:
		parts = period_month.split("-")
		inferred_year = cint(parts[0])
		inferred_month = cint(parts[1])
	else:
		# 默认取当前日期
		today = getdate()
		inferred_year = today.year
		inferred_month = today.month

	formatted_period = f"{inferred_year:04d}-{inferred_month:02d}"
	if not period_month:
		period_month = formatted_period

	# 2. 获取该月所有日期天数
	days_in_month = calendar.monthrange(inferred_year, inferred_month)[1]
	
	# 3. 建立当月每日法定性质字典 (工作日, 周末, 调休日, 调班日, 法定节假日)
	day_nature_map = {}
	for day in range(1, days_in_month + 1):
		cur_date_str = f"{inferred_year:04d}-{inferred_month:02d}-{day:02d}"
		nature_res = get_date_overtime_info(cur_date_str)
		rate_mult = flt(nature_res.get("rate_multiplier", 1.0))
		day_type = nature_res.get("day_type", "工作日")
		is_3x = (rate_mult >= 3.0)
		is_2x = (rate_mult == 2.0)
		is_workday = not (is_3x or is_2x)
		day_nature_map[day] = {
			"date": cur_date_str,
			"nature": day_type,
			"multiplier": rate_mult,
			"is_legal_3x": is_3x,
			"is_weekend_2x": is_2x,
			"is_regular_workday": is_workday,
		}

	# 4. 建立在册员工映射 (姓名 -> employee_no)
	profiles = frappe.get_all(
		"Ashan Employee Salary Profile",
		filters={"company": company},
		fields=["name", "employee_no", "employee_name", "employee_type", "salary_mode", "base_salary", "post_allowance", "performance_base"]
	)
	name_to_emp = {p.employee_name.strip(): p for p in profiles}

	# 5. 解析《总表》中的员工打卡记录
	# 寻找天数标题行
	days_col_start = 4 # 通常从第 D 列 (col=4) 开始为 1 日
	days_list = []
	days_header_row = 2
	for c in range(4, 40):
		val = ws.cell(days_header_row, c).value
		if val is not None and str(val).strip().isdigit():
			d_num = cint(str(val).strip())
			if 1 <= d_num <= days_in_month:
				days_list.append((c, d_num))

	parsed_results = []
	unmatched_names = []
	total_regular_hours = 0.0
	total_ot_1_5 = 0.0
	total_ot_2_0 = 0.0
	total_ot_3_0 = 0.0
	total_compensatory = 0.0
	total_meals = 0

	r = 3
	while r <= ws.max_row:
		c1 = ws.cell(r, 1).value
		c2 = ws.cell(r, 2).value
		c3 = str(ws.cell(r, 3).value or "").strip()

		if c2 is not None and c3 == "班次":
			emp_name = str(c2).strip()
			emp_profile = name_to_emp.get(emp_name)
			
			emp_no = emp_profile.employee_no if emp_profile else f"TEMP-{emp_name}"
			if not emp_profile:
				unmatched_names.append(emp_name)

			# 连续 5 行结构
			shifts_row = r
			work_row = r + 1
			ot_row = r + 2
			meal_row = r + 3
			remark_row = r + 4

			work_hours_sum = 0.0
			comp_leave_demand = 0.0
			ot_1_5_sum = 0.0
			weekend_ot_pool = 0.0
			holiday_ot_pool = 0.0
			meals_sum = 0
			full_days = 0
			half_days = 0
			absent_days = 0

			daily_records = []

			for col_idx, day_num in days_list:
				shift_val = ws.cell(shifts_row, col_idx).value
				cur_work = flt(ws.cell(work_row, col_idx).value or 0)
				cur_ot = flt(ws.cell(ot_row, col_idx).value or 0)
				cur_meal = cint(flt(ws.cell(meal_row, col_idx).value or 0))
				cur_remark = str(ws.cell(remark_row, col_idx).value or "").strip()

				day_info = day_nature_map.get(day_num, {})
				is_workday = day_info.get("is_regular_workday", True)
				is_weekend = day_info.get("is_weekend_2x", False)
				is_holiday = day_info.get("is_legal_3x", False)

				# 倒休工时与各倍率归集
				if is_workday:
					work_hours_sum += cur_work
					if cur_work < 8.0:
						comp_leave_demand += (8.0 - cur_work)
					ot_1_5_sum += cur_ot
				elif is_weekend:
					weekend_ot_pool += (cur_work + cur_ot)
				elif is_holiday:
					holiday_ot_pool += (cur_work + cur_ot)

				meals_sum += cur_meal

				# 考勤整天/半天/缺勤判断
				if cur_work >= 8.0:
					full_days += 1
				elif cur_work > 0:
					half_days += 1
				else:
					if is_workday:
						absent_days += 1

				daily_records.append({
					"day": day_num,
					"date": day_info.get("date", ""),
					"nature": day_info.get("nature", "工作日"),
					"shift": shift_val,
					"work_hours": cur_work,
					"overtime": cur_ot,
					"meal": cur_meal,
					"remark": cur_remark,
				})

			# 执行 VBA 倒休二倍工时对冲抵扣
			# 实际倒休 = min(周末2倍工时池, 倒休需求)
			actual_compensatory = min(weekend_ot_pool, comp_leave_demand)
			net_weekend_ot_2_0 = round(weekend_ot_pool - actual_compensatory, 2)
			final_regular_hours = round(work_hours_sum + actual_compensatory, 2)
			ot_1_5_sum = round(ot_1_5_sum, 2)
			holiday_ot_pool = round(holiday_ot_pool, 2)
			actual_compensatory = round(actual_compensatory, 2)

			# 累加全局统计
			total_regular_hours += final_regular_hours
			total_ot_1_5 += ot_1_5_sum
			total_ot_2_0 += net_weekend_ot_2_0
			total_ot_3_0 += holiday_ot_pool
			total_compensatory += actual_compensatory
			total_meals += meals_sum

			# 写入或更新 Jizhong Monthly Attendance
			doc_name = f"{company}-{period_month}-{emp_no}"
			if frappe.db.exists("Jizhong Monthly Attendance", doc_name):
				att_doc = frappe.get_doc("Jizhong Monthly Attendance", doc_name)
			else:
				att_doc = frappe.new_doc("Jizhong Monthly Attendance")
				att_doc.company = company
				att_doc.period_month = period_month
				att_doc.employee_no = emp_no

			att_doc.employee_name = emp_name
			att_doc.attendance_days = full_days
			att_doc.half_days = half_days
			att_doc.absent_days = absent_days
			att_doc.work_hours_regular = final_regular_hours
			att_doc.overtime_regular_1_5 = ot_1_5_sum
			att_doc.overtime_weekend_2_0 = net_weekend_ot_2_0
			att_doc.overtime_holiday_3_0 = holiday_ot_pool
			att_doc.leave_compensatory_hours = actual_compensatory
			att_doc.meal_count = meals_sum
			att_doc.daily_records_json = json.dumps(daily_records, ensure_ascii=False)
			if file_url:
				att_doc.attendance_file = file_url

			att_doc.save(ignore_permissions=True)

			parsed_results.append({
				"employee_no": emp_no,
				"employee_name": emp_name,
				"attendance_days": full_days,
				"half_days": half_days,
				"absent_days": absent_days,
				"work_hours_regular": final_regular_hours,
				"overtime_regular_1_5": ot_1_5_sum,
				"overtime_weekend_2_0": net_weekend_ot_2_0,
				"overtime_holiday_3_0": holiday_ot_pool,
				"leave_compensatory_hours": actual_compensatory,
				"meal_count": meals_sum,
			})

			r += 5
		else:
			r += 1

	frappe.db.commit()

	return {
		"success": True,
		"period_month": period_month,
		"company": company,
		"employee_count": len(parsed_results),
		"total_regular_hours": round(total_regular_hours, 2),
		"total_ot_1_5": round(total_ot_1_5, 2),
		"total_ot_2_0": round(total_ot_2_0, 2),
		"total_ot_3_0": round(total_ot_3_0, 2),
		"total_compensatory": round(total_compensatory, 2),
		"total_meals": total_meals,
		"unmatched_names": unmatched_names,
		"items": parsed_results,
		"file_url": file_url,
	}


@frappe.whitelist()
def get_jizhong_attendance_table(company="天津吉众科技有限公司", period_month=None):
	"""
	获取指定期间吉众全员考勤大宽表与汇总数据
	"""
	if not period_month:
		# 默认获取最近一个有考勤的月份，若无取当前月
		latest = frappe.db.get_value(
			"Jizhong Monthly Attendance",
			{"company": company},
			"period_month",
			order_by="period_month desc"
		)
		period_month = latest or getdate().strftime("%Y-%m")

	records = frappe.get_all(
		"Jizhong Monthly Attendance",
		filters={"company": company, "period_month": period_month},
		fields=[
			"name", "period_month", "company", "employee_no", "employee_name",
			"attendance_days", "half_days", "absent_days",
			"work_hours_regular", "overtime_regular_1_5", "overtime_weekend_2_0",
			"overtime_holiday_3_0", "leave_compensatory_hours", "meal_count",
			"attendance_file", "daily_records_json"
		],
		order_by="employee_no asc"
	)

	# 汇总 KPI
	summary = {
		"period_month": period_month,
		"company": company,
		"employee_count": len(records),
		"total_work_hours": sum(flt(r.work_hours_regular) for r in records),
		"total_ot_1_5": sum(flt(r.overtime_regular_1_5) for r in records),
		"total_ot_2_0": sum(flt(r.overtime_weekend_2_0) for r in records),
		"total_ot_3_0": sum(flt(r.overtime_holiday_3_0) for r in records),
		"total_compensatory": sum(flt(r.leave_compensatory_hours) for r in records),
		"total_meals": sum(cint(r.meal_count) for r in records),
		"attendance_file": records[0].attendance_file if records and records[0].attendance_file else None,
	}

	return {
		"summary": summary,
		"records": records,
	}
