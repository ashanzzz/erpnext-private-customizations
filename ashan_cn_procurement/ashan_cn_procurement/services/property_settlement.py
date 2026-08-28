# Copyright (c) 2026, Ashan CN Procurement and contributors
# For license information, please see license.txt

import io
import json
import calendar
import frappe
from frappe.utils import flt, cint, getdate, nowdate, now_datetime, formatdate

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from ashan_cn_procurement.services.authorization_service import (
	assert_company_access,
	assert_module_access,
	get_allowed_companies,
)


def _row_value(row, fieldname):
	"""Read one field from a Frappe row or a plain request dictionary."""
	return row.get(fieldname) if isinstance(row, dict) else getattr(row, fieldname, None)


def _property_companies_from_data(data):
	"""Return every company that a settlement payload can read or change."""
	companies = set()
	for section in ("meter_readings", "lease_charges", "company_summaries"):
		for row in data.get(section) or []:
			company = str(_row_value(row, "company") or "").strip()
			if company:
				companies.add(company)
	for row in data.get("adjustments") or []:
		for fieldname in ("company", "from_company", "to_company"):
			company = str(_row_value(row, fieldname) or "").strip()
			if company:
				companies.add(company)
	return companies


def assert_property_settlement_access(data, action="read"):
	"""Authorize a property settlement against every real company it contains."""
	assert_module_access("property", action)
	companies = _property_companies_from_data(data)
	if not companies:
		frappe.throw("结算数据未包含公司归属，无法确认权限范围。", frappe.PermissionError)
	for company in sorted(companies):
		assert_company_access(company)


def _apply_property_company_scope(filters, company_field="company"):
	"""Add the caller's explicit company scope to a Frappe query filter."""
	allowed_companies = get_allowed_companies()
	if allowed_companies is not None:
		filters[company_field] = ["in", sorted(allowed_companies)]
	return filters


def get_month_range(year, month):
	"""获取指定年月的起始日期、截止日期与当月总天数"""
	year, month = cint(year), cint(month)
	_, last_day = calendar.monthrange(year, month)
	start_date = f"{year:04d}-{month:02d}-01"
	end_date = f"{year:04d}-{month:02d}-{last_day:02d}"
	return start_date, end_date, last_day


def get_annual_lease_settlement_data(year=None):
	"""
	获取指定年度的房租与物业费年度结算总控数据
	按合同年度周期管理，直接关联 5% 房租发票与 6% 物业费发票，进行年度对账
	"""
	assert_module_access("property", "read")
	year = cint(year or nowdate()[:4])

	leases = frappe.get_all(
		"Property Lease",
		filters=_apply_property_company_scope({"enabled": 1}),
		fields=[
			"name",
			"property_name",
			"company",
			"supplier",
			"area",
			"deposit_amount",
			"property_certificate_no",
			"location_address",
			"start_date",
			"end_date",
			"rent_pricing_mode",
			"is_tax_inclusive",
			"rent_tax_rate",
			"rent_annual_amount",
			"rent_monthly_amount",
			"rent_daily_rate",
			"rent_monthly_rate",
			"rent_annual_rate",
			"rent_annual_tax_excl",
			"rent_annual_tax_amount",
			"property_fee_mode",
			"property_fee_pricing_mode",
			"property_fee_is_tax_inclusive",
			"property_fee_tax_rate",
			"property_fee_monthly_rate",
			"property_fee_annual_amount",
			"property_fee_monthly_amount",
			"property_fee_daily_rate",
			"property_fee_annual_rate",
			"property_fee_annual_tax_excl",
			"property_fee_annual_tax_amount",
			"total_annual_amount",
			"total_monthly_amount",
			"total_daily_rate",
			"total_annual_rate",
			"annual_discount_amount",
			"invoice_status",
			"rent_invoice_no",
			"rent_invoice_date",
			"rent_invoice_amount",
			"rent_invoice_tax",
			"property_fee_invoice_no",
			"property_fee_invoice_date",
			"property_fee_invoice_amount",
			"property_fee_invoice_tax",
			"remark"
		],
		order_by="company ASC, property_name ASC"
	)

	# 汇总公司数据
	comp_map = {}
	for l in leases:
		comp = l.company
		if comp not in comp_map:
			comp_map[comp] = {
				"company": comp,
				"total_area": 0.0,
				"rent_amount": 0.0,
				"property_fee_amount": 0.0,
				"deposit_amount": 0.0,
				"discount_amount": 0.0,
				"total_amount": 0.0,
				"rent_tax_amount": 0.0,
				"property_fee_tax_amount": 0.0,
				"tax_amount": 0.0,
				"amount_tax_excl": 0.0,
				"invoiced_amount": 0.0,
				"unbilled_amount": 0.0,
				"invoice_status": "全额已开票"
			}

		c = comp_map[comp]
		area = flt(l.area)
		r_ann = flt(l.rent_annual_amount)
		p_ann = flt(l.property_fee_annual_amount)
		dep = flt(l.deposit_amount)
		disc = flt(l.annual_discount_amount)
		tot_ann = flt(l.total_annual_amount) - disc
		r_tax = flt(l.rent_annual_tax_amount)
		p_tax = flt(l.property_fee_annual_tax_amount)
		tot_tax = r_tax + p_tax
		tot_excl = flt(l.rent_annual_tax_excl) + flt(l.property_fee_annual_tax_excl) - disc

		# 发票对账
		r_inv = flt(l.rent_invoice_amount)
		p_inv = flt(l.property_fee_invoice_amount)
		tot_inv = r_inv + p_inv

		c["total_area"] += area
		c["rent_amount"] += r_ann
		c["property_fee_amount"] += p_ann
		c["deposit_amount"] += dep
		c["discount_amount"] += disc
		c["total_amount"] += tot_ann
		c["rent_tax_amount"] += r_tax
		c["property_fee_tax_amount"] += p_tax
		c["tax_amount"] += tot_tax
		c["amount_tax_excl"] += tot_excl
		c["invoiced_amount"] += tot_inv
		c["unbilled_amount"] += max(0.0, tot_ann - tot_inv)

		if l.invoice_status != "全额已开票":
			c["invoice_status"] = "存在待开票"

	company_summaries = list(comp_map.values())
	for c in company_summaries:
		c["total_area"] = round(c["total_area"], 2)
		c["rent_amount"] = round(c["rent_amount"], 2)
		c["property_fee_amount"] = round(c["property_fee_amount"], 2)
		c["deposit_amount"] = round(c["deposit_amount"], 2)
		c["discount_amount"] = round(c["discount_amount"], 2)
		c["total_amount"] = round(c["total_amount"], 2)
		c["tax_amount"] = round(c["tax_amount"], 2)
		c["amount_tax_excl"] = round(c["amount_tax_excl"], 2)
		c["invoiced_amount"] = round(c["invoiced_amount"], 2)
		c["unbilled_amount"] = round(c["unbilled_amount"], 2)

	return {
		"year": year,
		"period_label": f"{year}年度 (按年结算与发票对账)",
		"leases": leases,
		"company_summaries": company_summaries
	}


@frappe.whitelist(methods=["POST"])
def update_lease_invoice_link(lease_name, rent_invoice_no=None, rent_invoice_date=None, rent_invoice_amount=None, rent_invoice_tax=None, property_fee_invoice_no=None, property_fee_invoice_date=None, property_fee_invoice_amount=None, property_fee_invoice_tax=None, annual_discount_amount=None):
	"""
	更新租约的关联发票信息与开票对账状态
	"""
	doc = frappe.get_doc("Property Lease", lease_name)
	assert_module_access("property", "write", doc.company)
	if rent_invoice_no is not None:
		doc.rent_invoice_no = rent_invoice_no
	if rent_invoice_date is not None:
		doc.rent_invoice_date = rent_invoice_date
	if rent_invoice_amount is not None:
		doc.rent_invoice_amount = flt(rent_invoice_amount)
	if rent_invoice_tax is not None:
		doc.rent_invoice_tax = flt(rent_invoice_tax)
	elif rent_invoice_amount is not None:
		# 自动计算 5% 专票税额
		doc.rent_invoice_tax = round(flt(rent_invoice_amount) - (flt(rent_invoice_amount) / 1.05), 2)

	if property_fee_invoice_no is not None:
		doc.property_fee_invoice_no = property_fee_invoice_no
	if property_fee_invoice_date is not None:
		doc.property_fee_invoice_date = property_fee_invoice_date
	if property_fee_invoice_amount is not None:
		doc.property_fee_invoice_amount = flt(property_fee_invoice_amount)
	if property_fee_invoice_tax is not None:
		doc.property_fee_invoice_tax = flt(property_fee_invoice_tax)
	elif property_fee_invoice_amount is not None:
		# 自动计算 6% 专票税额
		doc.property_fee_invoice_tax = round(flt(property_fee_invoice_amount) - (flt(property_fee_invoice_amount) / 1.06), 2)

	if annual_discount_amount is not None:
		doc.annual_discount_amount = flt(annual_discount_amount)

	doc.save(ignore_permissions=True)
	return {"status": "success", "message": "发票与对账信息已成功更新！"}


def get_previous_month(year, month):
	"""获取上一个月份 (year, month)"""
	year, month = cint(year), cint(month)
	if month == 1:
		return year - 1, 12
	return year, month - 1


def get_applicable_charge_rate(property_lease_name, target_date=None):
	"""
	兼容旧版接口：直接从合并后的 Property Lease 获取费率与税费配置
	"""
	if not property_lease_name:
		return None
	lease = frappe.db.get_value(
		"Property Lease",
		property_lease_name,
		[
			"name",
			"property_name",
			"company",
			"supplier",
			"area",
			"property_certificate_no",
			"location_address",
			"rent_pricing_mode",
			"is_tax_inclusive",
			"rent_tax_rate",
			"rent_annual_amount",
			"rent_monthly_amount",
			"rent_daily_rate",
			"rent_monthly_rate",
			"rent_annual_rate",
			"property_fee_mode",
			"property_fee_pricing_mode",
			"property_fee_is_tax_inclusive",
			"property_fee_tax_rate",
			"property_fee_monthly_rate",
			"property_fee_annual_amount",
			"property_fee_monthly_amount",
			"property_fee_daily_rate",
			"property_fee_annual_rate",
			"total_annual_amount",
			"total_monthly_amount",
			"enabled"
		],
		as_dict=True
	)
	return lease



def get_previous_meter_reading(meter_name, current_month_start):
	"""
	获取指定水电表在上个月或历史最近一次月结中的本期读数
	如果没有历史月结，则返回该表设置的初始表底数
	"""
	sql = """
		SELECT r.current_reading
		FROM `tabProperty Meter Reading` r
		JOIN `tabProperty Monthly Settlement` s ON r.parent = s.name
		WHERE r.utility_meter = %s
		  AND s.settlement_month < %s
		  AND s.status != '已作废'
		ORDER BY s.settlement_month DESC
		LIMIT 1
	"""
	res = frappe.db.sql(sql, (meter_name, current_month_start), as_dict=True)
	if res and res[0].current_reading is not None:
		return flt(res[0].current_reading)

	meter_doc = frappe.db.get_value("Utility Meter", meter_name, ["initial_reading"], as_dict=True)
	if meter_doc and meter_doc.initial_reading is not None:
		return flt(meter_doc.initial_reading)
	return 0.0


def calculate_settlement_matrix(data):
	"""
	核心集中计算引擎：重算电表、水表、调整项、租赁费（房租+单独物业费多周期）及各公司汇总
	依据房东含税综合单价反推增值税与不含税成本
	"""
	elec_price = flt(data.get("electricity_price") or 1.1957)
	elec_tax_rate = flt(data.get("electricity_tax_rate") if data.get("electricity_tax_rate") is not None else 12.5985)
	water_price = flt(data.get("water_price") or 5.5)
	water_tax_rate = flt(data.get("water_tax_rate") if data.get("water_tax_rate") is not None else 9.0)

	# 根级反推单价与税额参数 (供 UI 与财务快速核对)
	elec_price_excl = round(elec_price / (1.0 + (elec_tax_rate / 100.0)), 6)
	elec_unit_tax = round(elec_price - elec_price_excl, 6)
	water_price_excl = round(water_price / (1.0 + (water_tax_rate / 100.0)), 6)
	water_unit_tax = round(water_price - water_price_excl, 6)

	data["electricity_price"] = elec_price
	data["electricity_tax_rate"] = elec_tax_rate
	data["electricity_price_tax_excl"] = elec_price_excl
	data["electricity_unit_tax"] = elec_unit_tax

	data["water_price"] = water_price
	data["water_tax_rate"] = water_tax_rate
	data["water_price_tax_excl"] = water_price_excl
	data["water_unit_tax"] = water_unit_tax

	settlement_month = str(data.get("settlement_month") or nowdate())
	s_year, s_month = cint(settlement_month.split("-")[0]), cint(settlement_month.split("-")[1])
	_, _, days_in_month = get_month_range(s_year, s_month)

	# 1. 重算抄表明细 (反推税额与不含税金额)
	meter_readings = data.get("meter_readings") or []
	for r in meter_readings:
		prev = flt(r.get("previous_reading"))
		curr = flt(r.get("current_reading"))
		mult = flt(r.get("multiplier") or 1.0)
		u_type = r.get("utility_type") or "电"

		raw = max(0.0, curr - prev)
		calc_u = round(raw * mult, 2)

		if u_type == "电":
			price = elec_price
			tax_r = elec_tax_rate
			p_excl = elec_price_excl
		else:
			price = water_price
			tax_r = water_tax_rate
			p_excl = water_price_excl

		amt_incl = round(calc_u * price, 2)
		amt_excl = round(amt_incl / (1.0 + (tax_r / 100.0)), 2)
		tax_amt = round(amt_incl - amt_excl, 2)

		r["raw_usage"] = raw
		r["multiplier"] = mult
		r["calculated_usage"] = calc_u
		r["unit_price"] = price
		r["unit_price_tax_excl"] = p_excl
		r["tax_rate"] = tax_r
		r["amount_tax_incl"] = amt_incl
		r["amount_tax_excl"] = amt_excl
		r["tax_amount"] = tax_amt

	# 2. 重算费用调整明细
	adjustments = data.get("adjustments") or []
	for adj in adjustments:
		adj_type = adj.get("adjustment_type") or "按金额"
		u_type = adj.get("utility_type") or "电费"

		if u_type == "电费":
			curr_price = elec_price
			tax_r = elec_tax_rate
		elif u_type == "水费":
			curr_price = water_price
			tax_r = water_tax_rate
		elif u_type in ["房租", "物业费"]:
			curr_price = 1.0
			tax_r = 9.0
		else:
			curr_price = 1.0
			tax_r = 0.0

		if adj_type == "按用量":
			u_adj = flt(adj.get("usage_adjustment"))
			amt_adj = round(u_adj * curr_price, 2)
			eq_u = u_adj
		else:
			amt_adj = flt(adj.get("amount_adjustment"))
			eq_u = round(amt_adj / curr_price, 2) if curr_price > 0 else 0.0

		amt_excl = round(amt_adj / (1.0 + (tax_r / 100.0)), 2) if tax_r > 0 else amt_adj
		tax_amt = round(amt_adj - amt_excl, 2)

		adj["usage_adjustment"] = flt(adj.get("usage_adjustment"))
		adj["amount_adjustment"] = amt_adj
		adj["equivalent_usage"] = eq_u
		adj["unit_price_snapshot"] = curr_price
		adj["tax_rate"] = tax_r
		adj["amount_tax_excl"] = amt_excl
		adj["tax_amount"] = tax_amt

	# 3. 重算租赁固定费用 (房租 + 物业费，支持日/月/年多周期自选与独立税率)
	all_enabled_leases = frappe.get_all(
		"Property Lease",
		filters={"enabled": 1},
		fields=[
			"name",
			"property_name",
			"company",
			"supplier",
			"area",
			"deposit_amount",
			"property_certificate_no",
			"location_address",
			"rent_pricing_mode",
			"is_tax_inclusive",
			"rent_tax_rate",
			"rent_annual_amount",
			"rent_monthly_amount",
			"rent_daily_rate",
			"rent_monthly_rate",
			"rent_annual_rate",
			"property_fee_mode",
			"property_fee_pricing_mode",
			"property_fee_is_tax_inclusive",
			"property_fee_tax_rate",
			"property_fee_monthly_rate",
			"property_fee_annual_amount",
			"property_fee_monthly_amount",
			"property_fee_daily_rate",
			"property_fee_annual_rate"
		],
		order_by="company ASC, property_name ASC"
	)
	lease_map = {l.name: l for l in all_enabled_leases}
	existing_lease_charges = data.get("lease_charges") or []
	existing_map = {l.get("property_lease"): l for l in existing_lease_charges if l.get("property_lease")}

	lease_charges = []
	for l_name, lease_doc in lease_map.items():
		l_chg = existing_map.get(l_name, {})
		l_days = cint(l_chg.get("billing_days") or days_in_month)
		area = flt(lease_doc.area)
		deposit = flt(lease_doc.deposit_amount)
		rent_mode = lease_doc.rent_pricing_mode or "按年总金额 (元/年)"
		prop_mode = lease_doc.property_fee_mode or "免物业费"
		r_ann_amt = flt(lease_doc.rent_annual_amount)
		r_daily = flt(lease_doc.rent_daily_rate)
		r_mon_amt = flt(lease_doc.rent_monthly_amount)
		r_ann_rate = flt(lease_doc.rent_annual_rate)
		r_mon_rate = flt(lease_doc.rent_monthly_rate)
		t_rate = flt(lease_doc.rent_tax_rate if lease_doc.rent_tax_rate is not None else 5.0)
		p_tax_rate = flt(lease_doc.property_fee_tax_rate if lease_doc.property_fee_tax_rate is not None else 6.0)

		# 构造单价快照
		if rent_mode == "按日单价 (元/㎡·天)":
			r_rate_snap = f"¥ {r_daily}/㎡·天"
		elif rent_mode == "按月单价 (元/㎡·月)":
			r_rate_snap = f"¥ {r_mon_rate}/㎡·月"
		elif rent_mode == "按年单价 (元/㎡·年)":
			r_rate_snap = f"¥ {r_ann_rate}/㎡·年"
		elif rent_mode == "按月总金额 (元/月)":
			r_rate_snap = f"¥ {r_mon_amt}/月"
		else:
			r_rate_snap = f"¥ {r_ann_amt}/年"

		p_ann_amt = 0.0
		p_daily = 0.0
		p_mon_amt = 0.0
		p_ann_rate = 0.0
		p_mon_rate = 0.0
		if prop_mode == "单独计物业费":
			p_ann_amt = flt(lease_doc.property_fee_annual_amount)
			p_daily = flt(lease_doc.property_fee_daily_rate)
			p_mon_amt = flt(lease_doc.property_fee_monthly_amount)
			p_ann_rate = flt(lease_doc.property_fee_annual_rate)
			p_mon_rate = flt(lease_doc.property_fee_monthly_rate)
			p_mode = lease_doc.property_fee_pricing_mode or "按年单价 (元/㎡·年)"
			if p_mode == "按日单价 (元/㎡·天)":
				p_rate_snap = f"¥ {p_daily}/㎡·天"
			elif p_mode == "按月单价 (元/㎡·月)":
				p_rate_snap = f"¥ {p_mon_rate}/㎡·月"
			elif p_mode == "按年单价 (元/㎡·年)":
				p_rate_snap = f"¥ {p_ann_rate}/㎡·年"
			else:
				p_rate_snap = f"¥ {p_ann_amt}/年"
		else:
			p_rate_snap = "免物业费 (0元)"

		# 计算房租
		if r_ann_amt > 0:
			rent_amt = round((r_ann_amt / 365.0) * l_days, 2)
		elif r_daily > 0 and area > 0:
			rent_amt = round(area * r_daily * l_days, 2)
		elif r_mon_amt > 0:
			rent_amt = round(r_mon_amt * (l_days / float(days_in_month)), 2)
		else:
			rent_amt = 0.0

		# 计算物业费
		if prop_mode == "单独计物业费":
			if p_ann_amt > 0:
				prop_fee_amt = round((p_ann_amt / 365.0) * l_days, 2)
			elif p_daily > 0 and area > 0:
				prop_fee_amt = round(area * p_daily * l_days, 2)
			elif p_mon_amt > 0:
				prop_fee_amt = round(p_mon_amt * (l_days / float(days_in_month)), 2)
			else:
				prop_fee_amt = 0.0
		else:
			prop_fee_amt = 0.0

		tot_lease_amt = round(rent_amt + prop_fee_amt, 2)
		rent_excl = round(rent_amt / (1.0 + (t_rate / 100.0)), 2)
		rent_tax = round(rent_amt - rent_excl, 2)

		prop_excl = round(prop_fee_amt / (1.0 + (p_tax_rate / 100.0)), 2) if prop_fee_amt > 0 else 0.0
		prop_tax = round(prop_fee_amt - prop_excl, 2) if prop_fee_amt > 0 else 0.0

		tot_excl = round(rent_excl + prop_excl, 2)
		tot_tax = round(rent_tax + prop_tax, 2)

		lease_charges.append({
			"property_lease": lease_doc.name,
			"property_name": lease_doc.property_name,
			"company": lease_doc.company,
			"supplier": lease_doc.supplier or "",
			"property_certificate_no": lease_doc.property_certificate_no or "",
			"location_address": lease_doc.location_address or "",
			"area": area,
			"deposit_amount": deposit,
			"billing_days": l_days,
			"property_fee_mode": prop_mode,
			"rent_pricing_mode": rent_mode,
			"rent_rate_snapshot": r_rate_snap,
			"property_fee_rate_snapshot": p_rate_snap,
			"rent_annual_amount": r_ann_amt,
			"rent_daily_rate": r_daily,
			"rent_monthly_amount": r_mon_amt,
			"rent_annual_rate": r_ann_rate,
			"rent_monthly_rate": r_mon_rate,
			"property_fee_annual_amount": p_ann_amt,
			"property_fee_daily_rate": p_daily,
			"property_fee_monthly_amount": p_mon_amt,
			"property_fee_annual_rate": p_ann_rate,
			"property_fee_monthly_rate": p_mon_rate,
			"rent_amount_tax_incl": rent_amt,
			"rent_amount_tax_excl": rent_excl,
			"rent_tax_amount": rent_tax,
			"property_fee_amount_tax_incl": prop_fee_amt,
			"property_fee_amount_tax_excl": prop_excl,
			"property_fee_tax_amount": prop_tax,
			"amount_tax_incl": tot_lease_amt,
			"amount_tax_excl": tot_excl,
			"tax_rate": t_rate,
			"property_fee_tax_rate": p_tax_rate,
			"tax_amount": tot_tax,
			"remark": l_chg.get("remark") or ""
		})



	# 4. 按公司聚合汇总 (含税 / 反推增值税额 / 不含税成本)
	all_companies = set()
	for l in lease_charges:
		if l.get("company"):
			all_companies.add(l.get("company"))
	for m in meter_readings:
		if m.get("company"):
			all_companies.add(m.get("company"))
	for a in adjustments:
		if a.get("company"):
			all_companies.add(a.get("company"))
		if a.get("from_company"):
			all_companies.add(a.get("from_company"))
		if a.get("to_company"):
			all_companies.add(a.get("to_company"))

	existing_companies = frappe.get_all("Company", fields=["name"], order_by="name ASC")
	for ec in existing_companies:
		if "吉众" in ec.name or "祺富" in ec.name:
			all_companies.add(ec.name)

	comp_summary_map = {}
	for comp in sorted(all_companies):
		comp_summary_map[comp] = {
			"company": comp,
			"rent_amount": 0.0,
			"rent_tax_amount": 0.0,
			"rent_amount_tax_excl": 0.0,
			"property_fee_amount": 0.0,
			"property_fee_tax_amount": 0.0,
			"property_fee_amount_tax_excl": 0.0,
			"electricity_usage": 0.0,
			"electricity_amount": 0.0,
			"electricity_tax_amount": 0.0,
			"electricity_amount_tax_excl": 0.0,
			"water_usage": 0.0,
			"water_amount": 0.0,
			"water_tax_amount": 0.0,
			"water_amount_tax_excl": 0.0,
			"adjustment_amount": 0.0,
			"total_amount": 0.0,
			"total_tax_amount": 0.0,
			"total_amount_tax_excl": 0.0
		}

	# 累加租赁费用
	for l in lease_charges:
		comp = l.get("company")
		if comp in comp_summary_map:
			r_amt = flt(l.get("rent_amount_tax_incl"))
			p_amt = flt(l.get("property_fee_amount_tax_incl"))
			r_tax_r = flt(l.get("tax_rate") if l.get("tax_rate") is not None else 5.0)
			p_tax_r = flt(l.get("property_fee_tax_rate") if l.get("property_fee_tax_rate") is not None else 6.0)

			r_excl = round(r_amt / (1.0 + (r_tax_r / 100.0)), 2)
			r_tax = round(r_amt - r_excl, 2)

			p_excl = round(p_amt / (1.0 + (p_tax_r / 100.0)), 2) if p_amt > 0 else 0.0
			p_tax = round(p_amt - p_excl, 2) if p_amt > 0 else 0.0

			comp_summary_map[comp]["rent_amount"] += r_amt
			comp_summary_map[comp]["rent_tax_amount"] += r_tax
			comp_summary_map[comp]["rent_amount_tax_excl"] += r_excl

			comp_summary_map[comp]["property_fee_amount"] += p_amt
			comp_summary_map[comp]["property_fee_tax_amount"] += p_tax
			comp_summary_map[comp]["property_fee_amount_tax_excl"] += p_excl


	# 累加抄表水电费用
	for m in meter_readings:
		comp = m.get("company")
		if comp in comp_summary_map:
			u_type = m.get("utility_type") or "电"
			u_val = flt(m.get("calculated_usage"))
			amt_val = flt(m.get("amount_tax_incl"))
			tax_amt = flt(m.get("tax_amount"))
			amt_excl = flt(m.get("amount_tax_excl"))

			if u_type == "电":
				comp_summary_map[comp]["electricity_usage"] += u_val
				comp_summary_map[comp]["electricity_amount"] += amt_val
				comp_summary_map[comp]["electricity_tax_amount"] += tax_amt
				comp_summary_map[comp]["electricity_amount_tax_excl"] += amt_excl
			else:
				comp_summary_map[comp]["water_usage"] += u_val
				comp_summary_map[comp]["water_amount"] += amt_val
				comp_summary_map[comp]["water_tax_amount"] += tax_amt
				comp_summary_map[comp]["water_amount_tax_excl"] += amt_excl

	# 累加调整费用
	for a in adjustments:
		scope = a.get("adjustment_scope") or "公司间转移"
		amt = flt(a.get("amount_adjustment"))
		eq_u = flt(a.get("equivalent_usage"))
		tax_amt = flt(a.get("tax_amount"))
		amt_excl = flt(a.get("amount_tax_excl"))
		u_type = a.get("utility_type") or "电费"

		if scope == "单公司":
			comp = a.get("company")
			if comp in comp_summary_map:
				comp_summary_map[comp]["adjustment_amount"] += amt
				if u_type == "电费":
					comp_summary_map[comp]["electricity_usage"] += eq_u
					comp_summary_map[comp]["electricity_amount"] += amt
					comp_summary_map[comp]["electricity_tax_amount"] += tax_amt
					comp_summary_map[comp]["electricity_amount_tax_excl"] += amt_excl
				elif u_type == "水费":
					comp_summary_map[comp]["water_usage"] += eq_u
					comp_summary_map[comp]["water_amount"] += amt
					comp_summary_map[comp]["water_tax_amount"] += tax_amt
					comp_summary_map[comp]["water_amount_tax_excl"] += amt_excl
		elif scope == "公司间转移":
			from_c = a.get("from_company")
			to_c = a.get("to_company")
			if from_c in comp_summary_map:
				comp_summary_map[from_c]["adjustment_amount"] -= amt
				if u_type == "电费":
					comp_summary_map[from_c]["electricity_usage"] -= eq_u
					comp_summary_map[from_c]["electricity_amount"] -= amt
					comp_summary_map[from_c]["electricity_tax_amount"] -= tax_amt
					comp_summary_map[from_c]["electricity_amount_tax_excl"] -= amt_excl
				elif u_type == "水费":
					comp_summary_map[from_c]["water_usage"] -= eq_u
					comp_summary_map[from_c]["water_amount"] -= amt
					comp_summary_map[from_c]["water_tax_amount"] -= tax_amt
					comp_summary_map[from_c]["water_amount_tax_excl"] -= amt_excl
			if to_c in comp_summary_map:
				comp_summary_map[to_c]["adjustment_amount"] += amt
				if u_type == "电费":
					comp_summary_map[to_c]["electricity_usage"] += eq_u
					comp_summary_map[to_c]["electricity_amount"] += amt
					comp_summary_map[to_c]["electricity_tax_amount"] += tax_amt
					comp_summary_map[to_c]["electricity_amount_tax_excl"] += amt_excl
				elif u_type == "水费":
					comp_summary_map[to_c]["water_usage"] += eq_u
					comp_summary_map[to_c]["water_amount"] += amt
					comp_summary_map[to_c]["water_tax_amount"] += tax_amt
					comp_summary_map[to_c]["water_amount_tax_excl"] += amt_excl

	company_summaries = []
	grand_total = 0.0
	grand_tax_total = 0.0
	grand_excl_total = 0.0

	for comp, s in comp_summary_map.items():
		s["rent_amount"] = round(s["rent_amount"], 2)
		s["rent_tax_amount"] = round(s["rent_tax_amount"], 2)
		s["rent_amount_tax_excl"] = round(s["rent_amount_tax_excl"], 2)

		s["property_fee_amount"] = round(s["property_fee_amount"], 2)
		s["property_fee_tax_amount"] = round(s["property_fee_tax_amount"], 2)
		s["property_fee_amount_tax_excl"] = round(s["property_fee_amount_tax_excl"], 2)

		s["electricity_usage"] = round(s["electricity_usage"], 2)
		s["electricity_amount"] = round(s["electricity_amount"], 2)
		s["electricity_tax_amount"] = round(s["electricity_tax_amount"], 2)
		s["electricity_amount_tax_excl"] = round(s["electricity_amount_tax_excl"], 2)

		s["water_usage"] = round(s["water_usage"], 2)
		s["water_amount"] = round(s["water_amount"], 2)
		s["water_tax_amount"] = round(s["water_tax_amount"], 2)
		s["water_amount_tax_excl"] = round(s["water_amount_tax_excl"], 2)

		s["adjustment_amount"] = round(s["adjustment_amount"], 2)

		tot = round(s["rent_amount"] + s["property_fee_amount"] + s["electricity_amount"] + s["water_amount"], 2)
		tot_tax = round(s["rent_tax_amount"] + s["property_fee_tax_amount"] + s["electricity_tax_amount"] + s["water_tax_amount"], 2)
		tot_excl = round(tot - tot_tax, 2)

		s["total_amount"] = tot
		s["total_tax_amount"] = tot_tax
		s["total_amount_tax_excl"] = tot_excl

		grand_total += tot
		grand_tax_total += tot_tax
		grand_excl_total += tot_excl
		company_summaries.append(s)

	data["meter_readings"] = meter_readings
	data["adjustments"] = adjustments
	data["lease_charges"] = lease_charges
	data["company_summaries"] = company_summaries
	data["total_amount"] = round(grand_total, 2)
	data["total_tax_amount"] = round(grand_tax_total, 2)
	data["total_amount_tax_excl"] = round(grand_excl_total, 2)
	data["property_management_company"] = data.get("property_management_company") or "天津金利达物业管理有限公司"

	return data


def get_month_settlement_data(year, month):
	"""
	获取或构建指定月份的物业月结数据（优先读取数据库，无则自动带出基准并构建草稿视图）
	"""
	assert_module_access("property", "read")
	year, month = cint(year), cint(month)
	if year < 2000 or month < 1 or month > 12:
		frappe.throw("结算年月无效。")
	start_date, end_date, days_in_month = get_month_range(year, month)

	doc_name = f"PROP-SET-{start_date}"
	if frappe.db.exists("Property Monthly Settlement", doc_name):
		doc = frappe.get_doc("Property Monthly Settlement", doc_name)
		d = doc.as_dict()
		d = calculate_settlement_matrix(d)
		assert_property_settlement_access(d, "read")
		return d

	# 数据库中尚无此月份记录，构建新月份草稿数据
	elec_price = 1.1957
	water_price = 5.5

	# 1. 水电表列表及上期表数
	meters = frappe.get_all(
		"Utility Meter",
		filters=_apply_property_company_scope({"enabled": 1}),
		fields=["name", "meter_no", "meter_name", "utility_type", "company", "multiplier", "unit", "initial_reading"],
		order_by="utility_type ASC, meter_no ASC"
	)

	meter_readings = []
	for m in meters:
		prev_reading = get_previous_meter_reading(m.name, start_date)
		meter_readings.append({
			"utility_meter": m.name,
			"meter_no": m.meter_no,
			"utility_type": m.utility_type,
			"company": m.company,
			"previous_reading": prev_reading,
			"current_reading": prev_reading,
			"raw_usage": 0.0,
			"multiplier": flt(m.multiplier or 1.0),
			"calculated_usage": 0.0,
			"unit_price": elec_price if m.utility_type == "电" else water_price,
			"tax_rate": 13.0 if m.utility_type == "电" else 9.0,
			"amount_tax_incl": 0.0,
			"amount_tax_excl": 0.0,
			"tax_amount": 0.0,
			"remark": ""
		})

	# 2. 租赁固定费用 (直接读取合并后的 Property Lease 档案)
	leases = frappe.get_all(
		"Property Lease",
		filters=_apply_property_company_scope({"enabled": 1}),
		fields=[
			"name",
			"property_name",
			"company",
			"supplier",
			"area",
			"property_certificate_no",
			"location_address",
			"rent_pricing_mode",
			"is_tax_inclusive",
			"rent_tax_rate",
			"rent_annual_amount",
			"rent_monthly_amount",
			"rent_daily_rate",
			"rent_monthly_rate",
			"rent_annual_rate",
			"property_fee_mode",
			"property_fee_pricing_mode",
			"property_fee_is_tax_inclusive",
			"property_fee_tax_rate",
			"property_fee_monthly_rate",
			"property_fee_annual_amount",
			"property_fee_monthly_amount",
			"property_fee_daily_rate",
			"property_fee_annual_rate"
		],
		order_by="company ASC, property_name ASC"
	)

	lease_charges = []
	for l in leases:
		rent_mode = l.get("rent_pricing_mode") or "按年总金额 (元/年)"
		prop_mode = l.get("property_fee_mode") or "房租含物业"
		r_ann_amt = flt(l.get("rent_annual_amount"))
		r_daily = flt(l.get("rent_daily_rate"))
		r_mon_amt = flt(l.get("rent_monthly_amount"))
		t_rate = flt(l.get("rent_tax_rate") if l.get("rent_tax_rate") is not None else 5.0)
		p_tax_rate = flt(l.get("property_fee_tax_rate") if l.get("property_fee_tax_rate") is not None else 6.0)

		if rent_mode == "按日单价 (元/㎡·天)":
			r_rate_snap = f"¥ {r_daily}/㎡·天"
		elif rent_mode == "按月单价 (元/㎡·月)":
			r_rate_snap = f"¥ {flt(l.get('rent_monthly_rate'))}/㎡·月"
		elif rent_mode == "按年单价 (元/㎡·年)":
			r_rate_snap = f"¥ {flt(l.get('rent_annual_rate'))}/㎡·年"
		elif rent_mode == "按月总金额 (元/月)":
			r_rate_snap = f"¥ {r_mon_amt}/月"
		else:
			r_rate_snap = f"¥ {r_ann_amt}/年"

		p_ann_amt = 0.0
		p_daily = 0.0
		p_mon_amt = 0.0
		if prop_mode == "单独计收物业费":
			p_ann_amt = flt(l.get("property_fee_annual_amount"))
			p_daily = flt(l.get("property_fee_daily_rate"))
			p_mon_amt = flt(l.get("property_fee_monthly_amount"))
			p_mode = l.get("property_fee_pricing_mode") or "按月单价 (元/㎡·月)"
			if p_mode == "按日单价 (元/㎡·天)":
				p_rate_snap = f"¥ {p_daily}/㎡·天"
			elif p_mode == "按月单价 (元/㎡·月)":
				p_rate_snap = f"¥ {flt(l.get('property_fee_monthly_rate'))}/㎡·月"
			elif p_mode == "按年单价 (元/㎡·年)":
				p_rate_snap = f"¥ {flt(l.get('property_fee_annual_rate'))}/㎡·年"
			else:
				p_rate_snap = f"¥ {p_ann_amt}/年"
		else:
			p_rate_snap = "含在房租中"

		lease_charges.append({
			"property_lease": l.name,
			"property_name": l.property_name,
			"company": l.company,
			"supplier": l.get("supplier") or "",
			"property_certificate_no": l.get("property_certificate_no") or "",
			"location_address": l.get("location_address") or "",
			"area": flt(l.area),
			"property_fee_mode": prop_mode,
			"rent_pricing_mode": rent_mode,
			"billing_days": days_in_month,
			"rent_rate_snapshot": r_rate_snap,
			"property_fee_rate_snapshot": p_rate_snap,
			"rent_annual_amount": r_ann_amt,
			"rent_daily_rate": r_daily,
			"rent_monthly_amount": r_mon_amt,
			"property_fee_annual_amount": p_ann_amt,
			"property_fee_daily_rate": p_daily,
			"property_fee_monthly_amount": p_mon_amt,
			"tax_rate": t_rate,
			"property_fee_tax_rate": p_tax_rate,
			"rent_amount_tax_incl": 0.0,
			"property_fee_amount_tax_incl": 0.0,
			"amount_tax_incl": 0.0,
			"amount_tax_excl": 0.0,
			"tax_amount": 0.0,
			"remark": ""
		})


	data = {
		"name": doc_name,
		"settlement_month": start_date,
		"status": "草稿",
		"is_new": True,
		"property_management_company": "天津金利达物业管理有限公司",
		"electricity_price": elec_price,
		"electricity_tax_rate": 12.5985,
		"water_price": water_price,
		"water_tax_rate": 9.0,
		"meter_readings": meter_readings,
		"adjustments": [],
		"lease_charges": lease_charges,
		"company_summaries": [],
		"total_amount": 0.0,
		"remark": ""
	}

	data = calculate_settlement_matrix(data)
	assert_property_settlement_access(data, "read")
	return data


def save_draft_settlement(data):
	"""
	保存物业月结草稿
	"""
	if not isinstance(data, dict):
		frappe.throw("结算数据格式无效。")
	settlement_month = data.get("settlement_month")
	if not settlement_month:
		frappe.throw("结算月份不能为空")

	doc_name = f"PROP-SET-{settlement_month}"
	calc_data = calculate_settlement_matrix(data)
	assert_property_settlement_access(calc_data, "write")

	if frappe.db.exists("Property Monthly Settlement", doc_name):
		doc = frappe.get_doc("Property Monthly Settlement", doc_name)
		assert_property_settlement_access(doc.as_dict(), "write")
		if doc.status == "已结算":
			frappe.throw("当前月份已完成月结锁定，如需修改请先取消结算！")
	else:
		doc = frappe.new_doc("Property Monthly Settlement")
		doc.name = doc_name
		doc.settlement_month = settlement_month

	doc.status = "草稿"
	doc.property_management_company = calc_data.get("property_management_company") or "天津金利达物业管理有限公司"
	doc.electricity_price = flt(calc_data.get("electricity_price"))
	doc.electricity_tax_rate = flt(calc_data.get("electricity_tax_rate"))
	doc.water_price = flt(calc_data.get("water_price"))
	doc.water_tax_rate = flt(calc_data.get("water_tax_rate"))
	doc.total_amount = flt(calc_data.get("total_amount"))
	doc.remark = calc_data.get("remark") or ""

	doc.set("meter_readings", calc_data.get("meter_readings") or [])
	doc.set("adjustments", calc_data.get("adjustments") or [])
	doc.set("lease_charges", calc_data.get("lease_charges") or [])
	doc.set("company_summaries", calc_data.get("company_summaries") or [])

	doc.flags.ignore_permissions = True
	doc.save()

	return {
		"success": True,
		"name": doc.name,
		"data": calculate_settlement_matrix(doc.as_dict()),
		"message": "物业月结草稿已保存！"
	}


def finalize_monthly_settlement(data):
	"""
	完成本月结算并锁定单据
	"""
	if not isinstance(data, dict):
		frappe.throw("结算数据格式无效。")
	assert_property_settlement_access(calculate_settlement_matrix(data), "lock")
	save_res = save_draft_settlement(data)
	doc = frappe.get_doc("Property Monthly Settlement", save_res["name"])

	for r in doc.meter_readings:
		if r.current_reading is None:
			frappe.throw(f"表号 {r.meter_no} ({r.company}) 尚未录入本期读数！")
		if flt(r.current_reading) < flt(r.previous_reading):
			if not r.remark:
				frappe.throw(f"表号 {r.meter_no} 本期读数 ({r.current_reading}) 小于上期读数 ({r.previous_reading})，请在备注中说明换表或修正原因！")

	for adj in doc.adjustments:
		if not adj.reason:
			frappe.throw(f"调整项「{adj.utility_type}」必须填写调整原因！")
		if adj.adjustment_scope == "公司间转移":
			if not adj.from_company or not adj.to_company:
				frappe.throw("公司间转移必须指定转出公司和转入公司！")
			if adj.from_company == adj.to_company:
				frappe.throw("公司间转移的转出公司和转入公司不能相同！")

	doc.status = "已结算"
	doc.settled_by = frappe.session.user
	doc.settled_at = now_datetime()
	doc.flags.ignore_permissions = True
	doc.save()

	return {
		"success": True,
		"name": doc.name,
		"data": calculate_settlement_matrix(doc.as_dict()),
		"message": f"{doc.settlement_month[:7]} 物业月结已成功核定并锁定。"
	}


def revert_settlement_to_draft(name, reason=None):
	"""
	管理员解锁/取消结算，退回草稿状态
	"""
	doc = frappe.get_doc("Property Monthly Settlement", name)
	assert_property_settlement_access(doc.as_dict(), "unlock")
	if not str(reason or "").strip():
		frappe.throw("解除物业月结锁定必须填写原因。")
	if doc.status != "已结算":
		frappe.throw("当前月结未处于已结算状态，不能解除锁定。")
	doc.status = "草稿"
	doc.settled_by = None
	doc.settled_at = None
	doc.flags.ignore_permissions = True
	doc.save()
	doc.add_comment("Comment", text=f"物业月结解锁原因：{str(reason).strip()}")

	return {
		"success": True,
		"name": doc.name,
		"data": calculate_settlement_matrix(doc.as_dict()),
		"message": "已解锁并恢复为草稿状态，可重新编辑读数与调整！"
	}


# ─────────────────────────────────────────────────────────────
# Excel 单证导出引擎 (1:1 精确复刻《抄表记录.xlsx》祺富单证/吉众单证/合计单证)
# ─────────────────────────────────────────────────────────────

def get_sheet_title(company_name):
	"""根据公司名称匹配原版 Excel Sheet 命名方式"""
	if "祺富" in company_name:
		return "祺富"
	elif "吉众" in company_name:
		return "吉众"
	elif "合计" in company_name or "全公司" in company_name:
		return "合计"
	else:
		return company_name[:4] if len(company_name) > 4 else company_name


def _setup_ws_styles():
	"""返回公共样式常量字典"""
	thin_side = Side(style='thin', color='000000')
	med_side = Side(style='medium', color='000000')
	return {
		"thin": thin_side,
		"med": med_side,
		"b_all_thin": Border(top=thin_side, bottom=thin_side, left=thin_side, right=thin_side),
		"b_left_med": Border(top=thin_side, bottom=thin_side, left=med_side, right=thin_side),
		"b_right_med": Border(top=thin_side, bottom=thin_side, left=thin_side, right=med_side),
		"b_sec_title": Border(top=med_side, bottom=thin_side, left=med_side, right=med_side),
		"b_bot_total": Border(top=thin_side, bottom=med_side, left=thin_side, right=thin_side),
		"b_bot_total_left": Border(top=thin_side, bottom=med_side, left=med_side, right=thin_side),
		"b_bot_total_right": Border(top=thin_side, bottom=med_side, left=thin_side, right=med_side),
		"font_main_title": Font(name="等线", size=18, bold=True),
		"font_subtitle": Font(name="等线", size=12, bold=False),
		"font_meta": Font(name="等线", size=11, bold=False),
		"font_sec_hdr": Font(name="等线", size=11, bold=False),
		"font_tbl_hdr": Font(name="等线", size=11, bold=True),
		"font_data": Font(name="等线", size=11, bold=False),
		"font_total": Font(name="等线", size=11, bold=True),
		"font_grand_total": Font(name="等线", size=20, bold=False),
		"align_center": Alignment(horizontal="center", vertical="center"),
		"align_left": Alignment(horizontal="left", vertical="center"),
	}


def _write_col_widths(ws, widths_map):
	for col_idx, width in widths_map.items():
		ws.column_dimensions[get_column_letter(col_idx)].width = width


def _write_header_rows(ws, company_name, sheet_type, settlement_month, prop_mgmt_co, s):
	"""
	写入前 3 行通用表头:
	  Row 1: 公司名 (18pt 粗体, 合并 A:H)
	  Row 2: 副标题 (水电费明细 or 房租物业费明细)
	  Row 3: 所属期 YYYY-MM (合并 A:H)
	"""
	align_center = s["align_center"]
	curr_row = 1

	# Row 1 公司名
	ws.row_dimensions[curr_row].height = 30.0
	ws.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=8)
	c1 = ws.cell(curr_row, 1, company_name)
	c1.font = s["font_main_title"]
	c1.alignment = align_center
	curr_row += 1

	# Row 2 副标题
	ws.row_dimensions[curr_row].height = 24.0
	ws.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=8)
	subtitle = "水电费明细（单价含税）" if sheet_type == "water_elec" else "房租物业费明细（单价含税）"
	c2 = ws.cell(curr_row, 1, subtitle)
	c2.font = s["font_subtitle"]
	c2.alignment = align_center
	curr_row += 1

	# Row 3 所属期  改为单一「所属期: YYYY-MM」
	ws.row_dimensions[curr_row].height = 27.0
	month_str = str(settlement_month)[:7]  # e.g. "2026-08"
	ws.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=8)
	c3 = ws.cell(curr_row, 1, f"所属期: {month_str}    物业公司: {prop_mgmt_co}")
	c3.font = s["font_meta"]
	c3.alignment = align_center
	curr_row += 1

	return curr_row  # 4


def render_water_elec_sheet(ws, sheet_name, company_name, settlement_month, prop_mgmt_co,
							meters, adjustments, summary, is_total=False):
	"""
	水电费 Sheet：电费 + 水费 + 水电费合计汇总 (不含房租物业)
	只保留「表号」列，去掉「表具名称」
	列布局 (8列): 表号 | 上期表数 | 本期表数 | 本期用电/水 | 倍率 | 核定度数/m³ | 单价 | 总价
	"""
	ws.title = sheet_name
	ws.views.sheetView[0].showGridLines = True
	s = _setup_ws_styles()

	# 列宽 (8列，匹配原版)
	_write_col_widths(ws, {1: 14.0, 2: 14.0, 3: 14.0, 4: 14.0, 5: 10.0, 6: 14.0, 7: 12.0, 8: 18.0})

	curr_row = _write_header_rows(ws, company_name, "water_elec", settlement_month, prop_mgmt_co, s)

	b_all_thin = s["b_all_thin"]
	b_left_med = s["b_left_med"]
	b_right_med = s["b_right_med"]
	b_sec_title = s["b_sec_title"]
	b_bot_total = s["b_bot_total"]
	b_bot_total_left = s["b_bot_total_left"]
	b_bot_total_right = s["b_bot_total_right"]
	font_data = s["font_data"]
	font_total = s["font_total"]
	font_tbl_hdr = s["font_tbl_hdr"]
	font_sec_hdr = s["font_sec_hdr"]
	align_center = s["align_center"]
	med_side = s["med"]
	thin_side = s["thin"]

	def write_meter_section(section_title, util_type_list, unit_label):
		nonlocal curr_row
		# Section Header
		ws.row_dimensions[curr_row].height = 27.0
		ws.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=8)
		c_sec = ws.cell(curr_row, 1, section_title)
		c_sec.font = font_sec_hdr
		c_sec.alignment = align_center
		for c in range(1, 9):
			ws.cell(curr_row, c).border = b_sec_title
		curr_row += 1

		# 列头 (仅表号, 无表具名称)
		ws.row_dimensions[curr_row].height = 27.0
		headers = ["表号", "上期表数", "本期表数", f"本期用{unit_label}", "倍率", f"核定{unit_label}", "单价", "总价"]
		for c_idx, h in enumerate(headers, start=1):
			cell = ws.cell(curr_row, c_idx, h)
			cell.font = font_tbl_hdr
			cell.alignment = align_center
			if c_idx == 1:
				cell.border = b_left_med
			elif c_idx == 8:
				cell.border = b_right_med
			else:
				cell.border = b_all_thin
		curr_row += 1

		# 数据行
		sec_meters = [m for m in meters if m.get("utility_type") in util_type_list]
		sec_adjs   = [a for a in adjustments if a.get("utility_type") in (["电费", "电"] if "电" in util_type_list else ["水费", "水"])]
		data_start = curr_row

		tot_sec_usage = 0.0
		tot_sec_settle = 0.0
		tot_sec_amount = 0.0

		for m in sec_meters:
			ws.row_dimensions[curr_row].height = 27.0
			u_price = float(m.get("unit_price") or (summary.get("electricity_price") if "电" in util_type_list else summary.get("water_price")) or (1.009321 if "电" in util_type_list else 5.5))
			prev = float(m.get("previous_reading") or 0.0)
			curr = float(m.get("current_reading") or 0.0)
			mult = float(m.get("multiplier") or 1.0)

			raw_usage = max(0.0, curr - prev)
			settle_u = round(raw_usage * mult, 2)
			row_amt = round(settle_u * u_price, 2)

			tot_sec_usage += raw_usage
			tot_sec_settle += settle_u
			tot_sec_amount += row_amt

			ws.cell(curr_row, 1, str(m.get("meter_no") or ""))
			ws.cell(curr_row, 2, prev)
			ws.cell(curr_row, 3, curr)
			ws.cell(curr_row, 4, raw_usage)
			ws.cell(curr_row, 5, mult)
			ws.cell(curr_row, 6, settle_u)
			ws.cell(curr_row, 7, round(u_price, 4))
			ws.cell(curr_row, 8, row_amt)

			for c in range(1, 9):
				cell = ws.cell(curr_row, c)
				cell.font = font_data
				cell.alignment = align_center
				if c in [2, 3, 4, 6]:
					cell.number_format = "#,##0.00" if (c in [4, 6] and (raw_usage % 1 != 0 or settle_u % 1 != 0)) else "#,##0"
				elif c == 7:
					cell.number_format = "0.0000"
				elif c == 8:
					cell.number_format = "#,##0.00"
				cell.border = b_left_med if c == 1 else (b_right_med if c == 8 else b_all_thin)
			curr_row += 1

		# 调整行 (电费调整 / 水费调整)
		for a in sec_adjs:
			ws.row_dimensions[curr_row].height = 27.0
			adj_label = "电费调整" if "电" in util_type_list else "水费调整"
			adj_u = float(a.get("usage") or 0.0)
			adj_amt = float(a.get("amount") or 0.0)

			tot_sec_usage += adj_u
			tot_sec_settle += adj_u
			tot_sec_amount += adj_amt

			ws.cell(curr_row, 1, adj_label)
			ws.cell(curr_row, 2, "")
			ws.cell(curr_row, 3, "")
			ws.cell(curr_row, 4, adj_u)
			ws.cell(curr_row, 5, 1)
			ws.cell(curr_row, 6, adj_u)
			ws.cell(curr_row, 7, "")
			ws.cell(curr_row, 8, adj_amt)
			for c in range(1, 9):
				cell = ws.cell(curr_row, c)
				cell.font = font_data
				cell.alignment = align_center
				if c in [4, 6, 8]:
					cell.number_format = "#,##0.00"
				cell.border = b_left_med if c == 1 else (b_right_med if c == 8 else b_all_thin)
			curr_row += 1

		# 若无数据行
		if curr_row == data_start:
			ws.row_dimensions[curr_row].height = 27.0
			ws.cell(curr_row, 1, "—")
			for c in range(1, 9):
				cell = ws.cell(curr_row, c)
				cell.font = font_data
				cell.alignment = align_center
				cell.border = b_left_med if c == 1 else (b_right_med if c == 8 else b_all_thin)
			curr_row += 1

		# 合计行 (A:C 合并)
		ws.row_dimensions[curr_row].height = 27.0
		ws.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=3)
		ws.cell(curr_row, 1, "合计").font = font_total
		ws.cell(curr_row, 4, tot_sec_usage).font = font_total
		ws.cell(curr_row, 5, "")
		ws.cell(curr_row, 6, tot_sec_settle).font = font_total
		ws.cell(curr_row, 7, "")
		ws.cell(curr_row, 8, tot_sec_amount).font = font_total
		for c in range(1, 9):
			cell = ws.cell(curr_row, c)
			cell.alignment = align_center
			if c in [4, 6, 8]:
				cell.number_format = "#,##0.00" if tot_sec_amount % 1 != 0 else "#,##0"
			if c <= 3:
				cell.border = b_bot_total_left if c == 1 else b_bot_total
			elif c == 8:
				cell.border = b_bot_total_right
			else:
				cell.border = b_bot_total

		curr_row += 1

		# 空行间隔
		ws.row_dimensions[curr_row].height = 18.0
		curr_row += 1
		return tot_sec_settle, tot_sec_amount

	tot_elec_kwh, tot_elec_amt = write_meter_section("电费", ["电", "电费"], "电")
	tot_water_m3, tot_water_amt = write_meter_section("水费", ["水", "水费"], "水")

	# ─── 水电费合计汇总 (依据真实数电发票清单分项拆解) ───────────────────────
	ws.row_dimensions[curr_row].height = 27.0
	ws.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=8)
	short_comp = get_sheet_title(company_name)
	title_label = "全公司合计水电费" if is_total or short_comp == "合计" else f"{short_comp}合计水电费"
	c_sec_tot = ws.cell(curr_row, 1, title_label)
	c_sec_tot.font = font_sec_hdr
	c_sec_tot.alignment = align_center
	for c in range(1, 9):
		ws.cell(curr_row, c).border = b_sec_title
	curr_row += 1

	# 汇总表头
	ws.row_dimensions[curr_row].height = 27.0
	tot_headers = ["项目", "不含税金额", "税率", "税额", "含税合计", "数量", "单价", "水电费合计"]
	for c_idx, h in enumerate(tot_headers, start=1):
		cell = ws.cell(curr_row, c_idx, h)
		cell.font = font_tbl_hdr
		cell.alignment = align_center
		cell.border = b_left_med if c_idx == 1 else (b_right_med if c_idx == 8 else b_all_thin)
	curr_row += 1

	sum_start_row = curr_row

	# 依据数电专票精准标准拆解:
	# 电网代收3大政府性基金固定标准(元/kWh):
	RATE_WATER_RES = 0.002304757  # 国家重大水利工程建设基金 (免税)
	RATE_RESERVOIR = 0.007258429  # 水库移民后期扶持基金 (免税)
	RATE_RENEWABLE = 0.022244024  # 可再生能源发展基金 (免税)
	TOTAL_FUNDS_RATE = RATE_WATER_RES + RATE_RESERVOIR + RATE_RENEWABLE  # ~0.031807 元/度

	# 各项金额推算:
	amt_water_res = round(tot_elec_kwh * RATE_WATER_RES, 2)
	amt_reservoir = round(tot_elec_kwh * RATE_RESERVOIR, 2)
	amt_renewable = round(tot_elec_kwh * RATE_RENEWABLE, 2)
	amt_funds_tot = round(amt_water_res + amt_reservoir + amt_renewable, 2)

	amt_elec_main = max(0.0, round(tot_elec_amt - amt_funds_tot, 2))
	excl_elec_main = round(amt_elec_main / 1.13, 2)
	tax_elec_main = round(amt_elec_main - excl_elec_main, 2)
	price_elec_main = round(amt_elec_main / tot_elec_kwh, 6) if tot_elec_kwh > 0 else 0.0

	# 水费 (9% 专票)
	excl_water = round(tot_water_amt / 1.09, 2)
	tax_water = round(tot_water_amt - excl_water, 2)
	price_water = round(tot_water_amt / tot_water_m3, 4) if tot_water_m3 > 0 else 0.0

	summary_rows = [
		("*电力*电费 (13% 专票)", excl_elec_main, "13%", tax_elec_main, amt_elec_main, tot_elec_kwh, price_elec_main),
		("*代收国家重大水利工程建设基金* (免税)", amt_water_res, "免税(0%)", 0.00, amt_water_res, tot_elec_kwh, RATE_WATER_RES),
		("*代收水库移民后期扶持基金* (免税)", amt_reservoir, "免税(0%)", 0.00, amt_reservoir, tot_elec_kwh, RATE_RESERVOIR),
		("*代收可再生能源发展基金* (免税)", amt_renewable, "免税(0%)", 0.00, amt_renewable, tot_elec_kwh, RATE_RENEWABLE),
		("*水费*自来水 (9% 专票)", excl_water, "9%", tax_water, tot_water_amt, tot_water_m3, price_water)
	]

	for item in summary_rows:
		ws.row_dimensions[curr_row].height = 24.0
		ws.cell(curr_row, 1, item[0]).font = font_data
		ws.cell(curr_row, 2, item[1]).font = font_data
		ws.cell(curr_row, 3, item[2]).font = font_data
		ws.cell(curr_row, 4, item[3]).font = font_data
		ws.cell(curr_row, 5, item[4]).font = font_data
		ws.cell(curr_row, 6, item[5]).font = font_data
		ws.cell(curr_row, 7, item[6]).font = font_data

		for c in range(1, 9):
			cell = ws.cell(curr_row, c)
			cell.alignment = align_center
			if c in [2, 4, 5]:
				cell.number_format = "#,##0.00"
			elif c == 7:
				cell.number_format = "0.0000"
			elif c == 6:
				cell.number_format = "#,##0"
			cell.border = b_left_med if c == 1 else (b_right_med if c == 8 else b_all_thin)
		curr_row += 1

	sum_end_row = curr_row - 1

	# H 列大字合计 (跨 5 行，完整修复右侧边框与封底边框)
	ws.merge_cells(start_row=sum_start_row, start_column=8, end_row=sum_end_row, end_column=8)
	grand_total_amt = round(tot_elec_amt + tot_water_amt, 2)
	c_grand = ws.cell(sum_start_row, 8, grand_total_amt)
	c_grand.font = s["font_grand_total"]
	c_grand.alignment = align_center
	c_grand.number_format = "#,##0.00" if grand_total_amt % 1 != 0 else "#,##0"

	# 为 H 列合并区域的所有单元格（行 sum_start_row ~ sum_end_row）完整设置右侧中粗实线与底边粗线
	for r in range(sum_start_row, sum_end_row + 1):
		cell = ws.cell(r, 8)
		t = thin_side if r == sum_start_row else None
		b = med_side if r == sum_end_row else None
		cell.border = Border(top=t, bottom=b, left=thin_side, right=med_side)

	# 封底底边粗边框
	for c in range(1, 9):
		cell = ws.cell(sum_end_row, c)
		if c == 1:
			cell.border = Border(top=cell.border.top, bottom=med_side, left=med_side, right=thin_side)
		elif c == 8:
			cell.border = Border(top=cell.border.top, bottom=med_side, left=thin_side, right=med_side)
		else:
			cell.border = Border(top=cell.border.top, bottom=med_side, left=thin_side, right=thin_side)


def render_lease_sheet(ws, sheet_name, company_name, settlement_month, prop_mgmt_co, leases, is_total=False):
	"""
	房租物业费 Sheet：房租 + 物业费 (年缴费明细)
	列布局 (8列): 场地名称 | 年金额(房租) | 年金额(物业费) | 面积(㎡) | 单价(元/㎡·年) | 物业费模式 | 所属期金额(房租) | 所属期金额(物业费)
	最后 2 列展示按天折算后的本期费用
	"""
	ws.title = sheet_name
	ws.views.sheetView[0].showGridLines = True
	s = _setup_ws_styles()

	# 列宽
	_write_col_widths(ws, {1: 22.0, 2: 14.0, 3: 14.0, 4: 12.0, 5: 14.0, 6: 16.0, 7: 16.0, 8: 16.0})

	curr_row = _write_header_rows(ws, company_name, "lease", settlement_month, prop_mgmt_co, s)

	b_all_thin = s["b_all_thin"]
	b_left_med = s["b_left_med"]
	b_right_med = s["b_right_med"]
	b_sec_title = s["b_sec_title"]
	b_bot_total = s["b_bot_total"]
	b_bot_total_left = s["b_bot_total_left"]
	b_bot_total_right = s["b_bot_total_right"]
	font_data = s["font_data"]
	font_total = s["font_total"]
	font_tbl_hdr = s["font_tbl_hdr"]
	font_sec_hdr = s["font_sec_hdr"]
	align_center = s["align_center"]
	med_side = s["med"]
	thin_side = s["thin"]

	# ─── 房租部分 ───
	ws.row_dimensions[curr_row].height = 27.0
	ws.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=8)
	c_sec_r = ws.cell(curr_row, 1, "房租")
	c_sec_r.font = font_sec_hdr
	c_sec_r.alignment = align_center
	for c in range(1, 9):
		ws.cell(curr_row, c).border = b_sec_title
	curr_row += 1

	# 列头
	ws.row_dimensions[curr_row].height = 27.0
	r_headers = ["场地名称", "年金额(元/年)", "面积(㎡)", "计费天数", "日单价(元/天)", "物业费计收方式", "所属期房租金额", "含税合计"]
	for c_idx, h in enumerate(r_headers, start=1):
		cell = ws.cell(curr_row, c_idx, h)
		cell.font = font_tbl_hdr
		cell.alignment = align_center
		cell.border = b_left_med if c_idx == 1 else (b_right_med if c_idx == 8 else b_all_thin)
	curr_row += 1

	rent_start = curr_row
	for l in leases:
		ws.row_dimensions[curr_row].height = 27.0
		ann_rent = float(l.get("rent_annual_amount") or 0)
		area = float(l.get("area") or 0)
		b_days = int(l.get("billing_days") or 0)
		if ann_rent > 0 and area > 0:
			r_daily = ann_rent / area / 365.0
		elif ann_rent > 0:
			r_daily = ann_rent / 365.0
		else:
			r_daily = float(l.get("rent_daily_rate") or 0)
		rent_this_period = float(l.get("rent_amount_tax_incl") or 0)

		prop_mode_label = "单独计收" if l.get("property_fee_mode") == "单独计收物业费" else "含在房租中"

		ws.cell(curr_row, 1, str(l.get("property_name") or ""))
		ws.cell(curr_row, 2, ann_rent)
		ws.cell(curr_row, 3, area)
		ws.cell(curr_row, 4, b_days)
		ws.cell(curr_row, 5, round(r_daily, 6))
		ws.cell(curr_row, 6, prop_mode_label)
		ws.cell(curr_row, 7, rent_this_period)
		ws.cell(curr_row, 8, rent_this_period)

		for c in range(1, 9):
			cell = ws.cell(curr_row, c)
			cell.font = font_data
			cell.alignment = align_center
			if c in [2, 3, 5, 7, 8]:
				cell.number_format = "#,##0.00"
			cell.border = b_left_med if c == 1 else (b_right_med if c == 8 else b_all_thin)
		curr_row += 1

	if curr_row == rent_start:
		ws.row_dimensions[curr_row].height = 27.0
		ws.cell(curr_row, 1, "—")
		for c in range(1, 9):
			ws.cell(curr_row, c).font = font_data
			ws.cell(curr_row, c).alignment = align_center
			ws.cell(curr_row, c).border = b_left_med if c == 1 else (b_right_med if c == 8 else b_all_thin)
		curr_row += 1

	# 房租合计行
	ws.row_dimensions[curr_row].height = 27.0
	ws.cell(curr_row, 1, "合计").font = font_total
	for c in [2, 3, 7, 8]:
		ws.cell(curr_row, c, f"=SUM({get_column_letter(c)}{rent_start}:{get_column_letter(c)}{curr_row-1})").font = font_total
	for c in range(1, 9):
		cell = ws.cell(curr_row, c)
		cell.alignment = align_center
		if c in [2, 3, 7, 8]:
			cell.number_format = "#,##0.00"
		cell.border = b_bot_total_left if c == 1 else (b_bot_total_right if c == 8 else b_bot_total)
	rent_total_row = curr_row
	curr_row += 1

	# 空行
	ws.row_dimensions[curr_row].height = 18.0
	curr_row += 1

	# ─── 物业费部分 ───
	ws.row_dimensions[curr_row].height = 27.0
	ws.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=8)
	c_sec_p = ws.cell(curr_row, 1, "物业费")
	c_sec_p.font = font_sec_hdr
	c_sec_p.alignment = align_center
	for c in range(1, 9):
		ws.cell(curr_row, c).border = b_sec_title
	curr_row += 1

	# 物业费列头
	ws.row_dimensions[curr_row].height = 27.0
	p_headers = ["场地名称", "年金额(元/年)", "面积(㎡)", "计费天数", "日单价(元/㎡·天)", "物业费计收方式", "所属期物业费", "含税合计"]
	for c_idx, h in enumerate(p_headers, start=1):
		cell = ws.cell(curr_row, c_idx, h)
		cell.font = font_tbl_hdr
		cell.alignment = align_center
		cell.border = b_left_med if c_idx == 1 else (b_right_med if c_idx == 8 else b_all_thin)
	curr_row += 1

	prop_start = curr_row
	prop_leases = [l for l in leases if l.get("property_fee_mode") == "单独计收物业费"]

	for l in prop_leases:
		ws.row_dimensions[curr_row].height = 27.0
		ann_prop = float(l.get("property_fee_annual_amount") or 0)
		area = float(l.get("area") or 0)
		b_days = int(l.get("billing_days") or 0)
		if ann_prop > 0 and area > 0:
			p_daily = ann_prop / area / 365.0
		elif ann_prop > 0:
			p_daily = ann_prop / 365.0
		else:
			p_daily = float(l.get("property_fee_daily_rate") or 0)
		prop_this = float(l.get("property_fee_amount_tax_incl") or 0)
		rate_snap = str(l.get("property_fee_rate_snapshot") or "")

		ws.cell(curr_row, 1, str(l.get("property_name") or ""))
		ws.cell(curr_row, 2, ann_prop)
		ws.cell(curr_row, 3, area)
		ws.cell(curr_row, 4, b_days)
		ws.cell(curr_row, 5, round(p_daily, 6))
		ws.cell(curr_row, 6, rate_snap)
		ws.cell(curr_row, 7, prop_this)
		ws.cell(curr_row, 8, prop_this)

		for c in range(1, 9):
			cell = ws.cell(curr_row, c)
			cell.font = font_data
			cell.alignment = align_center
			if c in [2, 3, 5, 7, 8]:
				cell.number_format = "#,##0.00"
			cell.border = b_left_med if c == 1 else (b_right_med if c == 8 else b_all_thin)
		curr_row += 1

	if curr_row == prop_start:
		ws.row_dimensions[curr_row].height = 27.0
		ws.cell(curr_row, 1, "—")
		for c in range(1, 9):
			ws.cell(curr_row, c).font = font_data
			ws.cell(curr_row, c).alignment = align_center
			ws.cell(curr_row, c).border = b_left_med if c == 1 else (b_right_med if c == 8 else b_all_thin)
		curr_row += 1

	# 物业费合计行
	ws.row_dimensions[curr_row].height = 27.0
	ws.cell(curr_row, 1, "合计").font = font_total
	for c in [2, 3, 7, 8]:
		ws.cell(curr_row, c, f"=SUM({get_column_letter(c)}{prop_start}:{get_column_letter(c)}{curr_row-1})").font = font_total
	for c in range(1, 9):
		cell = ws.cell(curr_row, c)
		cell.alignment = align_center
		if c in [2, 3, 7, 8]:
			cell.number_format = "#,##0.00"
		cell.border = b_bot_total_left if c == 1 else (b_bot_total_right if c == 8 else b_bot_total)
	prop_total_row = curr_row
	curr_row += 1

	# 空行
	ws.row_dimensions[curr_row].height = 18.0
	curr_row += 1

	# ─── 房租物业费合计汇总 ─────────────────────────────────────
	ws.row_dimensions[curr_row].height = 27.0
	ws.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=8)
	short_comp = get_sheet_title(company_name)
	title_label = "全公司合计房租物业费" if is_total or short_comp == "合计" else f"{short_comp}合计房租物业费"
	c_sec_tot = ws.cell(curr_row, 1, title_label)
	c_sec_tot.font = font_sec_hdr
	c_sec_tot.alignment = align_center
	for c in range(1, 9):
		ws.cell(curr_row, c).border = b_sec_title
	curr_row += 1

	# 汇总列头
	ws.row_dimensions[curr_row].height = 27.0
	sum_headers = ["项目", "不含税金额", "税率", "税额", "含税合计", "面积(㎡)", "单位", "房租物业合计"]
	for c_idx, h in enumerate(sum_headers, start=1):
		cell = ws.cell(curr_row, c_idx, h)
		cell.font = font_tbl_hdr
		cell.alignment = align_center
		cell.border = b_left_med if c_idx == 1 else (b_right_med if c_idx == 8 else b_all_thin)
	curr_row += 1

	sum_start_row = curr_row

	# 房租汇总行
	ws.row_dimensions[curr_row].height = 27.0
	ws.cell(curr_row, 1, "房租").font = font_data
	ws.cell(curr_row, 2, f"=E{curr_row}-D{curr_row}").font = font_data
	ws.cell(curr_row, 3, 0.09).font = font_data
	ws.cell(curr_row, 4, f"=E{curr_row}-E{curr_row}/(C{curr_row}+1)").font = font_data
	ws.cell(curr_row, 5, f"=H{rent_total_row}").font = font_data
	ws.cell(curr_row, 6, f"=C{rent_total_row}").font = font_data
	ws.cell(curr_row, 7, "㎡").font = font_data
	for c in range(1, 8):
		cell = ws.cell(curr_row, c)
		cell.alignment = align_center
		if c in [2, 4]:
			cell.number_format = "0.00"
		elif c == 3:
			cell.number_format = "0%"
		elif c in [5, 6]:
			cell.number_format = "#,##0.00"
		cell.border = b_left_med if c == 1 else b_all_thin
	curr_row += 1

	# 物业费汇总行
	ws.row_dimensions[curr_row].height = 27.0
	ws.cell(curr_row, 1, "物业费").font = font_data
	ws.cell(curr_row, 2, f"=E{curr_row}-D{curr_row}").font = font_data
	ws.cell(curr_row, 3, 0.09).font = font_data
	ws.cell(curr_row, 4, f"=E{curr_row}-E{curr_row}/(C{curr_row}+1)").font = font_data
	ws.cell(curr_row, 5, f"=H{prop_total_row}").font = font_data
	ws.cell(curr_row, 6, f"=C{prop_total_row}").font = font_data
	ws.cell(curr_row, 7, "㎡").font = font_data
	for c in range(1, 8):
		cell = ws.cell(curr_row, c)
		cell.alignment = align_center
		if c in [2, 4]:
			cell.number_format = "0.00"
		elif c == 3:
			cell.number_format = "0%"
		elif c in [5, 6]:
			cell.number_format = "#,##0.00"
		cell.border = b_left_med if c == 1 else b_all_thin
	sum_end_row = curr_row
	curr_row += 1

	# H 列大字合计 (跨 2 行)
	ws.merge_cells(start_row=sum_start_row, start_column=8, end_row=sum_end_row, end_column=8)
	grand_terms = "+".join([f"E{r}" for r in range(sum_start_row, sum_end_row + 1)])
	c_grand = ws.cell(sum_start_row, 8, f"={grand_terms}")
	c_grand.font = s["font_grand_total"]
	c_grand.alignment = align_center
	c_grand.number_format = "#,##0"

	for r in range(sum_start_row, sum_end_row + 1):
		cell = ws.cell(r, 8)
		t = thin_side
		b = med_side if r == sum_end_row else thin_side
		cell.border = Border(top=t, bottom=b, left=thin_side, right=med_side)

	# 封底底边粗边框
	for c in range(1, 9):
		cell = ws.cell(sum_end_row, c)
		med = med_side
		thn = thin_side
		if c == 1:
			cell.border = Border(top=cell.border.top, bottom=med, left=med, right=thn)
		elif c == 8:
			cell.border = Border(top=cell.border.top, bottom=med, left=thn, right=med)
		else:
			cell.border = Border(top=cell.border.top, bottom=med, left=thn, right=thn)


def generate_settlement_excel_workbook(data, company=None, property_management_company=None, mode="single"):
	"""
	生成物业月结工作簿:
	  mode="company"  → 单公司: 水电费 Sheet + 房租物业 Sheet
	  mode="total"    → 全公司合计: 水电费 Sheet + 房租物业 Sheet
	  mode="all"      → 所有分公司各 2 Sheet + 合计 2 Sheet
	"""
	data = calculate_settlement_matrix(data)
	wb = openpyxl.Workbook()
	wb.calculation.fullCalcOnLoad = True
	wb.calculation.calcMode = "auto"
	default_ws = wb.active

	settlement_month = data.get("settlement_month") or nowdate()
	prop_mgmt = property_management_company or data.get("property_management_company") or "天津金利达物业管理有限公司"

	def get_company_data(comp):
		comp_meters = [m for m in (data.get("meter_readings") or []) if m.get("company") == comp]
		comp_leases = [l for l in (data.get("lease_charges") or []) if l.get("company") == comp]
		comp_summary = next((s for s in (data.get("company_summaries") or []) if s.get("company") == comp), {})
		comp_adjs = []
		for a in (data.get("adjustments") or []):
			if a.get("adjustment_scope") == "单公司" and a.get("company") == comp:
				comp_adjs.append({
					"title": "电费调整" if a.get("utility_type") in ["电费", "电"] else "水费调整",
					"utility_type": a.get("utility_type"),
					"usage": flt(a.get("usage_adjustment")),
					"amount": flt(a.get("amount_adjustment")),
					"reason": a.get("reason")
				})
			elif a.get("adjustment_scope") == "公司间转移":
				if a.get("from_company") == comp:
					comp_adjs.append({
						"title": "电费调整" if a.get("utility_type") in ["电费", "电"] else "水费调整",
						"utility_type": a.get("utility_type"),
						"usage": -flt(a.get("equivalent_usage")),
						"amount": -flt(a.get("amount_adjustment")),
						"reason": a.get("reason")
					})
				elif a.get("to_company") == comp:
					comp_adjs.append({
						"title": "电费调整" if a.get("utility_type") in ["电费", "电"] else "水费调整",
						"utility_type": a.get("utility_type"),
						"usage": flt(a.get("equivalent_usage")),
						"amount": flt(a.get("amount_adjustment")),
						"reason": a.get("reason")
					})
		return comp_meters, comp_leases, comp_adjs, comp_summary

	def build_total_adjs():
		total_adjs = []
		for a in (data.get("adjustments") or []):
			if a.get("adjustment_scope") == "单公司":
				total_adjs.append({
					"title": "电费调整" if a.get("utility_type") in ["电费", "电"] else "水费调整",
					"utility_type": a.get("utility_type"),
					"usage": flt(a.get("usage_adjustment")),
					"amount": flt(a.get("amount_adjustment")),
					"reason": a.get("reason")
				})
		return total_adjs

	if mode == "company" and company:
		comp_meters, comp_leases, comp_adjs, comp_summary = get_company_data(company)
		short = get_sheet_title(company)
		render_water_elec_sheet(default_ws, f"{short}水电费", company, settlement_month, prop_mgmt, comp_meters, comp_adjs, comp_summary, is_total=False)
		ws2 = wb.create_sheet(f"{short}房租物业")
		render_lease_sheet(ws2, f"{short}房租物业", company, settlement_month, prop_mgmt, comp_leases, is_total=False)

	elif mode == "total":
		all_meters = data.get("meter_readings") or []
		all_leases = data.get("lease_charges") or []
		render_water_elec_sheet(default_ws, "合计水电费", "全公司合计", settlement_month, prop_mgmt, all_meters, build_total_adjs(), {}, is_total=True)
		ws2 = wb.create_sheet("合计房租物业")
		render_lease_sheet(ws2, "合计房租物业", "全公司合计", settlement_month, prop_mgmt, all_leases, is_total=True)

	else:  # mode == "all"
		companies = [s.get("company") for s in (data.get("company_summaries") or []) if s.get("company")]
		first_sheet = True
		for comp in companies:
			comp_meters, comp_leases, comp_adjs, comp_summary = get_company_data(comp)
			short = get_sheet_title(comp)
			ws_e = default_ws if first_sheet else wb.create_sheet()
			first_sheet = False
			render_water_elec_sheet(ws_e, f"{short}水电费", comp, settlement_month, prop_mgmt, comp_meters, comp_adjs, comp_summary, is_total=False)
			ws_l = wb.create_sheet(f"{short}房租物业")
			render_lease_sheet(ws_l, f"{short}房租物业", comp, settlement_month, prop_mgmt, comp_leases, is_total=False)

		# 全公司合计 2 Sheet
		all_meters = data.get("meter_readings") or []
		all_leases = data.get("lease_charges") or []
		ws_tot_e = default_ws if first_sheet else wb.create_sheet()
		render_water_elec_sheet(ws_tot_e, "合计水电费", "全公司合计", settlement_month, prop_mgmt, all_meters, build_total_adjs(), {}, is_total=True)
		ws_tot_l = wb.create_sheet("合计房租物业")
		render_lease_sheet(ws_tot_l, "合计房租物业", "全公司合计", settlement_month, prop_mgmt, all_leases, is_total=True)

	return wb


@frappe.whitelist()
def export_settlement_excel(settlement_month, company=None, property_management_company=None, mode="company"):
	"""
	Whitelisted API: 下载物业结算明细 Excel (全套合并不拆分)
	"""
	settlement_month = str(settlement_month).strip()
	year, month = cint(settlement_month.split("-")[0]), cint(settlement_month.split("-")[1])
	data = get_month_settlement_data(year, month)
	assert_property_settlement_access(data, "export")
	if company:
		assert_company_access(company)

	if property_management_company:
		data["property_management_company"] = property_management_company

	wb = generate_settlement_excel_workbook(data, company=company, property_management_company=property_management_company, mode=mode)

	bio = io.BytesIO()
	wb.save(bio)
	bio.seek(0)

	month_str = settlement_month[:7]
	if mode == "company" and company:
		fname = f"{company}_{month_str}_物业明细（单价含税）.xlsx"
	elif mode == "total":
		fname = f"全公司合计_{month_str}_物业明细（单价含税）.xlsx"
	else:
		fname = f"{month_str}_全套月结明细.xlsx"

	frappe.response['filename'] = fname
	frappe.response['filecontent'] = bio.getvalue()
	frappe.response['type'] = 'binary'


@frappe.whitelist()
def export_utility_settlement_excel(settlement_month, company=None, property_management_company=None, mode="company"):
	"""
	Whitelisted API: 专用于【水电费工作台】下载纯水电费明细 Excel
	"""
	settlement_month = str(settlement_month).strip()
	year, month = cint(settlement_month.split("-")[0]), cint(settlement_month.split("-")[1])
	data = get_month_settlement_data(year, month)
	assert_property_settlement_access(data, "export")
	if company:
		assert_company_access(company)

	prop_mgmt = property_management_company or data.get("property_management_company") or "天津金利达物业管理有限公司"
	wb = openpyxl.Workbook()
	wb.calculation.fullCalcOnLoad = True
	default_ws = wb.active

	def get_company_data(comp):
		comp_meters = [m for m in (data.get("meter_readings") or []) if m.get("company") == comp]
		comp_leases = [l for l in (data.get("lease_charges") or []) if l.get("company") == comp]
		comp_summary = next((s for s in (data.get("company_summaries") or []) if s.get("company") == comp), {})
		comp_adjs = []
		for a in (data.get("adjustments") or []):
			if a.get("utility_type") not in ["电费", "电", "水费", "水"]:
				continue
			if a.get("adjustment_scope") == "单公司" and a.get("company") == comp:
				comp_adjs.append({
					"title": "电费调整" if a.get("utility_type") in ["电费", "电"] else "水费调整",
					"utility_type": a.get("utility_type"),
					"usage": flt(a.get("usage_adjustment")),
					"amount": flt(a.get("amount_adjustment")),
					"reason": a.get("reason")
				})
			elif a.get("adjustment_scope") == "公司间转移":
				if a.get("from_company") == comp:
					comp_adjs.append({
						"title": "电费调整" if a.get("utility_type") in ["电费", "电"] else "水费调整",
						"utility_type": a.get("utility_type"),
						"usage": -flt(a.get("equivalent_usage")),
						"amount": -flt(a.get("amount_adjustment")),
						"reason": a.get("reason")
					})
				elif a.get("to_company") == comp:
					comp_adjs.append({
						"title": "电费调整" if a.get("utility_type") in ["电费", "电"] else "水费调整",
						"utility_type": a.get("utility_type"),
						"usage": flt(a.get("equivalent_usage")),
						"amount": flt(a.get("amount_adjustment")),
						"reason": a.get("reason")
					})
		return comp_meters, comp_leases, comp_adjs, comp_summary

	def build_total_adjs():
		total_adjs = []
		for a in (data.get("adjustments") or []):
			if a.get("utility_type") not in ["电费", "电", "水费", "水"]:
				continue
			if a.get("adjustment_scope") == "单公司":
				total_adjs.append({
					"title": "电费调整" if a.get("utility_type") in ["电费", "电"] else "水费调整",
					"utility_type": a.get("utility_type"),
					"usage": flt(a.get("usage_adjustment")),
					"amount": flt(a.get("amount_adjustment")),
					"reason": a.get("reason")
				})
		return total_adjs

	def get_sheet_title(comp_name):
		if "祺富" in comp_name:
			return "祺富"
		if "吉众" in comp_name:
			return "吉众"
		return comp_name[:4]

	if mode == "company" and company:
		comp_meters, _, comp_adjs, comp_summary = get_company_data(company)
		short = get_sheet_title(company)
		render_water_elec_sheet(default_ws, f"{short}水电费", company, settlement_month, prop_mgmt, comp_meters, comp_adjs, comp_summary, is_total=False)
	elif mode == "total":
		all_meters = data.get("meter_readings") or []
		render_water_elec_sheet(default_ws, "合计水电费", "全公司合计", settlement_month, prop_mgmt, all_meters, build_total_adjs(), {}, is_total=True)
	else:  # mode == "all"
		companies = [s.get("company") for s in (data.get("company_summaries") or []) if s.get("company")]
		first_sheet = True
		for comp in companies:
			comp_meters, _, comp_adjs, comp_summary = get_company_data(comp)
			short = get_sheet_title(comp)
			ws_e = default_ws if first_sheet else wb.create_sheet()
			first_sheet = False
			render_water_elec_sheet(ws_e, f"{short}水电费", comp, settlement_month, prop_mgmt, comp_meters, comp_adjs, comp_summary, is_total=False)

		all_meters = data.get("meter_readings") or []
		ws_tot_e = wb.create_sheet("合计水电费")
		render_water_elec_sheet(ws_tot_e, "合计水电费", "全公司合计", settlement_month, prop_mgmt, all_meters, build_total_adjs(), {}, is_total=True)

	bio = io.BytesIO()
	wb.save(bio)
	bio.seek(0)

	month_str = settlement_month[:7]
	if mode == "company" and company:
		fname = f"{company}_{month_str}_水电费明细（单价含税）.xlsx"
	elif mode == "total":
		fname = f"全公司合计_{month_str}_水电费明细（单价含税）.xlsx"
	else:
		fname = f"{month_str}_全套水电费明细.xlsx"

	frappe.response['filename'] = fname
	frappe.response['filecontent'] = bio.getvalue()
	frappe.response['type'] = 'binary'


@frappe.whitelist()
def export_lease_settlement_excel(settlement_month, company=None, property_management_company=None, mode="company"):
	"""
	Whitelisted API: 专用于【房租与物业费工作台】下载纯房租物业费明细 Excel
	"""
	settlement_month = str(settlement_month).strip()
	year, month = cint(settlement_month.split("-")[0]), cint(settlement_month.split("-")[1])
	data = get_month_settlement_data(year, month)
	assert_property_settlement_access(data, "export")
	if company:
		assert_company_access(company)

	prop_mgmt = property_management_company or data.get("property_management_company") or "天津金利达物业管理有限公司"
	wb = openpyxl.Workbook()
	wb.calculation.fullCalcOnLoad = True
	default_ws = wb.active

	def get_company_leases(comp):
		return [l for l in (data.get("lease_charges") or []) if l.get("company") == comp]

	def get_sheet_title(comp_name):
		if "祺富" in comp_name:
			return "祺富"
		if "吉众" in comp_name:
			return "吉众"
		return comp_name[:4]

	if mode == "company" and company:
		comp_leases = get_company_leases(company)
		short = get_sheet_title(company)
		render_lease_sheet(default_ws, f"{short}房租物业", company, settlement_month, prop_mgmt, comp_leases, is_total=False)
	elif mode == "total":
		all_leases = data.get("lease_charges") or []
		render_lease_sheet(default_ws, "合计房租物业", "全公司合计", settlement_month, prop_mgmt, all_leases, is_total=True)
	else:  # mode == "all"
		companies = [s.get("company") for s in (data.get("company_summaries") or []) if s.get("company")]
		first_sheet = True
		for comp in companies:
			comp_leases = get_company_leases(comp)
			short = get_sheet_title(comp)
			ws_l = default_ws if first_sheet else wb.create_sheet()
			first_sheet = False
			render_lease_sheet(ws_l, f"{short}房租物业", comp, settlement_month, prop_mgmt, comp_leases, is_total=False)

		all_leases = data.get("lease_charges") or []
		ws_tot_l = wb.create_sheet("合计房租物业")
		render_lease_sheet(ws_tot_l, "合计房租物业", "全公司合计", settlement_month, prop_mgmt, all_leases, is_total=True)

	bio = io.BytesIO()
	wb.save(bio)
	bio.seek(0)

	month_str = settlement_month[:7]
	if mode == "company" and company:
		fname = f"{company}_{month_str}_房租物业费明细（单价含税）.xlsx"
	elif mode == "total":
		fname = f"全公司合计_{month_str}_房租物业费明细（单价含税）.xlsx"
	else:
		fname = f"{month_str}_全套房租物业费明细.xlsx"

	frappe.response['filename'] = fname
	frappe.response['filecontent'] = bio.getvalue()
	frappe.response['type'] = 'binary'
