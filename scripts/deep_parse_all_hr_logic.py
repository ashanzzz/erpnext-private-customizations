import os
import sys
import openpyxl
import json

sys.stdout.reconfigure(encoding='utf-8')

script_dir = os.path.dirname(os.path.abspath(__file__))
temp_dir = os.path.join(script_dir, "..", "temp_screenshots")
xlsm_path = None
for fname in os.listdir(temp_dir):
    if "吉众" in fname and fname.endswith(".xlsm"):
        xlsm_path = os.path.join(temp_dir, fname)
        break

if not xlsm_path:
    # fallback to any xlsm containing 202606
    for fname in os.listdir(temp_dir):
        if "202606" in fname and fname.endswith(".xlsm"):
            xlsm_path = os.path.join(temp_dir, fname)
            break

output_file = os.path.join(script_dir, "jizhong_excel_analysis.txt")
sys.stdout = open(output_file, 'w', encoding='utf-8')

print(f"Target XLSM path: {xlsm_path}")
wb = openpyxl.load_workbook(xlsm_path, data_only=False, keep_vba=True)
wb_val = openpyxl.load_workbook(xlsm_path, data_only=True, keep_vba=True)

print("="*80)
print(f"深入解析 14 个 Sheet 核心结构与业务公式:")
print("="*80)

for idx, name in enumerate(wb.sheetnames, 1):
    ws = wb[name]
    ws_v = wb_val[name]
    print(f"\n### {idx}. 工作表: [{name}] (行: {ws.max_row}, 列: {ws.max_column})")

    # 打印前 5 行的全部非空表头和公式示例
    for r in range(1, min(ws.max_row + 1, 8)):
        row_formula = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        row_val = [ws_v.cell(row=r, column=c).value for c in range(1, ws_v.max_column + 1)]

        non_empty = []
        for c_idx, (f, v) in enumerate(zip(row_formula, row_val), 1):
            if f is not None:
                c_letter = openpyxl.utils.get_column_letter(c_idx)
                if str(f).startswith('='):
                    non_empty.append(f"{c_letter}{r}: [{f}] -> (值: {v})")
                else:
                    non_empty.append(f"{c_letter}{r}: {f}")
        if non_empty:
            print(f"  Row {r}: " + " | ".join(non_empty[:12]))
            if len(non_empty) > 12:
                print(f"         ... " + " | ".join(non_empty[12:24]))
