import os
import sys
import json
import openpyxl

def load_env_file(env_path='.env'):
    if os.path.exists(env_path):
        with open(env_path, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith('#') and '=' in line:
                    k, v = line.split('=', 1)
                    os.environ.setdefault(k.strip(), v.strip())

load_env_file()

# 读取提取的种子数据
with open(r"d:\SynologyDrive团队\antigravity\erpnext16\scripts\extracted_hr_seed_data.json", "r", encoding="utf-8") as f:
    seed_data = json.load(f)

# 读取 Excel 原表中的 2026-06 考勤数据
xlsm_path = r"d:\SynologyDrive团队\antigravity\erpnext16\temp_screenshots\202606吉众人事综合.xlsm"
wb_v = openpyxl.load_workbook(xlsm_path, data_only=True, keep_vba=True)
ws_att = wb_v.worksheets[2]
ws_work = wb_v.worksheets[5]

# 构建 2026-06 员工实际考勤与加班
emp_att_list = []
for r in range(5, ws_work.max_row + 1):
    emp_no = ws_work.cell(row=r, column=2).value
    name = ws_work.cell(row=r, column=3).value
    if emp_no and emp_no != "工号":
        att_days = ws_work.cell(row=r, column=9).value or 21
        work_hours = ws_work.cell(row=r, column=12).value or 168
        ot_1_5 = ws_work.cell(row=r, column=13).value or 0
        ot_2_0 = ws_work.cell(row=r, column=14).value or 0
        ot_3_0 = ws_work.cell(row=r, column=15).value or 0
        meal_count = ws_work.cell(row=r, column=17).value or 0
        emp_att_list.append({
            "company": "天津吉众机电设备有限公司",
            "period_month": "2026-06",
            "employee_no": emp_no,
            "employee_name": name,
            "attendance_days": float(att_days),
            "work_hours_regular": float(work_hours),
            "overtime_regular_1_5": float(ot_1_5),
            "overtime_weekend_2_0": float(ot_2_0),
            "overtime_holiday_3_0": float(ot_3_0),
            "meal_count": int(meal_count),
            "daily_records_json": "{}"
        })

print(f"准备好 2026-06 考勤记录 {len(emp_att_list)} 条")

# 生成在容器内执行的 Python 初始化代码
runner_script = f"""
import frappe
import json

frappe.init(site='site1.local')
frappe.connect()

print("--- 1. 导入法定日历数据 (365天) ---")
cal_data = {json.dumps(seed_data['calendar'], ensure_ascii=False)}
for c in cal_data:
    dt = c['date']
    y = int(dt.split('-')[0])
    m = int(dt.split('-')[1])
    if not frappe.db.exists('Ashan Holiday Calendar', dt):
        doc = frappe.new_doc('Ashan Holiday Calendar')
        doc.calendar_date = dt
        doc.year = y
        doc.month = m
        doc.day_type = c['status_desc'] or '工作日'
        doc.is_workday = c['is_workday']
        doc.is_legal_holiday = c['is_legal_holiday']
        doc.is_shift_off = c['is_shift_off']
        doc.is_shift_work = c['is_shift_work']
        doc.holiday_name = c['remark']
        doc.insert(ignore_permissions=True)
print("法定日历数据导入完成！")

print("--- 2. 导入吉众员工薪资档案 ---")
emp_data = {json.dumps(seed_data['employees'], ensure_ascii=False)}
for emp in emp_data:
    if emp['employee_no'] == '工号': continue
    name = f"{{emp['company']}}-{{emp['employee_no']}}-{{emp['employee_name']}}"
    if not frappe.db.exists('Ashan Employee Salary Profile', name):
        doc = frappe.new_doc('Ashan Employee Salary Profile')
        doc.employee_no = emp['employee_no']
        doc.employee_name = emp['employee_name']
        doc.company = emp['company']
        doc.id_card = emp.get('id_card', '')
        doc.mobile = emp.get('mobile', '')
        doc.employee_type = emp.get('employee_type', '正式工')
        doc.employment_status = emp.get('status', '在职')
        doc.is_insured = 1
        doc.salary_mode = emp.get('salary_mode', '税前动态工资')
        doc.fixed_net_salary = emp.get('fixed_net_salary', 0)
        doc.base_salary = emp.get('base_salary', 0)
        doc.base_subsidy = emp.get('base_subsidy', 0)
        doc.performance_bonus_base = emp.get('performance_bonus', 0)
        doc.position_allowance = emp.get('position_allowance', 0)
        doc.meal_unit_price = emp.get('meal_unit_price', 15)
        doc.social_security_base = emp.get('social_security_base', 5124)
        doc.housing_fund_base = emp.get('housing_fund_base', 2520)
        doc.special_additional_deduction = emp.get('special_additional_deduction', 0)
        doc.tax_exemption_monthly = 5000
        doc.insert(ignore_permissions=True)
print("员工薪资档案导入完成！")

print("--- 3. 导入 2026-06 考勤数据 ---")
att_data = {json.dumps(emp_att_list, ensure_ascii=False)}
for att in att_data:
    name = f"{{att['company']}}-{{att['period_month']}}-{{att['employee_no']}}"
    if not frappe.db.exists('Ashan Monthly Attendance', name):
        doc = frappe.new_doc('Ashan Monthly Attendance')
        doc.company = att['company']
        doc.period_month = att['period_month']
        doc.employee_no = att['employee_no']
        doc.employee_name = att['employee_name']
        doc.attendance_days = att['attendance_days']
        doc.work_hours_regular = att['work_hours_regular']
        doc.overtime_regular_1_5 = att['overtime_regular_1_5']
        doc.overtime_weekend_2_0 = att['overtime_weekend_2_0']
        doc.overtime_holiday_3_0 = att['overtime_holiday_3_0']
        doc.meal_count = att['meal_count']
        doc.daily_records_json = att['daily_records_json']
        doc.insert(ignore_permissions=True)
print("考勤数据导入完成！")

frappe.db.commit()
print("全部人事基础数据初始化成功！")
"""

with open(r"d:\SynologyDrive团队\antigravity\erpnext16\scripts\container_seed_hr.py", "w", encoding="utf-8") as f:
    f.write(runner_script)

print("已生成 scripts/container_seed_hr.py")
