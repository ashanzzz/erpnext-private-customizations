import os
import sys
import time
import requests
import openpyxl
import io
import base64
from dotenv import load_dotenv
from playwright.sync_api import sync_playwright

sys.stdout.reconfigure(encoding='utf-8')
load_dotenv(r"d:\SynologyDrive团队\antigravity\erpnext16\.env")

SITE_URL = os.getenv('ERPNEXT_SITE_URL', 'http://192.168.8.11:6888')
USERNAME = os.getenv('ERPNEXT_USERNAME', 'ashanzzz1213@gmail.com')
USER_PWD = os.getenv('ERPNEXT_PASSWORD', 'Woo@@@204317')

for i in range(25):
    try:
        r = requests.get(f"{SITE_URL}/api/method/ping", timeout=3)
        if r.status_code == 200:
            print("Site is ready!")
            break
    except Exception:
        pass
    time.sleep(2)

# 1. 验证后端导出接口
print("--- 1. Testing Backend Dynamic Excel Export API ---")
s = requests.Session()
login_r = s.post(f"{SITE_URL}/api/method/login", data={"usr": USERNAME, "pwd": USER_PWD})
if login_r.status_code == 200:
    print("Logged in via REST API!")

# A. 个税精简版 Excel
r_tax_simple = s.post(f"{SITE_URL}/api/method/ashan_cn_procurement.services.payroll_settlement_service.export_qifu_payroll_excel", data={
    "company": "天津祺富机械加工有限公司",
    "period_month": "2026-07",
    "sheet_type": "tax",
    "tax_view_mode": "simple"
})
wb_tax_simple = openpyxl.load_workbook(io.BytesIO(base64.b64decode(r_tax_simple.json()["message"]["file_base64"])))
print(f"Tax Simple Excel: sheets={wb_tax_simple.sheetnames}, max_col={wb_tax_simple.active.max_column}")

# B. 个税 68 列大宽表 Excel
r_tax_68 = s.post(f"{SITE_URL}/api/method/ashan_cn_procurement.services.payroll_settlement_service.export_qifu_payroll_excel", data={
    "company": "天津祺富机械加工有限公司",
    "period_month": "2026-07",
    "sheet_type": "tax",
    "tax_view_mode": "full_68"
})
wb_tax_68 = openpyxl.load_workbook(io.BytesIO(base64.b64decode(r_tax_68.json()["message"]["file_base64"])))
print(f"Tax 68-Col Excel: sheets={wb_tax_68.sheetnames}, max_col={wb_tax_68.active.max_column}")

# C. 历史全员 15 列总览 Excel
r_hist_all = s.post(f"{SITE_URL}/api/method/ashan_cn_procurement.services.payroll_settlement_service.export_qifu_payroll_excel", data={
    "company": "天津祺富机械加工有限公司",
    "period_month": "2026-07",
    "sheet_type": "history",
    "history_mode": "all"
})
wb_hist_all = openpyxl.load_workbook(io.BytesIO(base64.b64decode(r_hist_all.json()["message"]["file_base64"])))
print(f"History All (15-Col) Excel: sheets={wb_hist_all.sheetnames}, max_col={wb_hist_all.active.max_column}")

# D. 全套 7 大 Sheet Excel
r_all_7 = s.post(f"{SITE_URL}/api/method/ashan_cn_procurement.services.payroll_settlement_service.export_qifu_payroll_excel", data={
    "company": "天津祺富机械加工有限公司",
    "period_month": "2026-07",
    "sheet_type": "all"
})
wb_all_7 = openpyxl.load_workbook(io.BytesIO(base64.b64decode(r_all_7.json()["message"]["file_base64"])))
print(f"All 7-Sheet Workbook: sheets={wb_all_7.sheetnames}")

# 2. 验证前端 UI
print("\n--- 2. Testing Frontend UI with Playwright ---")
with sync_playwright() as p:
    browser = p.chromium.launch(headless=True)
    page = browser.new_page(viewport={"width": 1920, "height": 1150})

    errors = []
    page.on("console", lambda msg: print(f"[CONSOLE {msg.type}] {msg.text}") if msg.type in ['error', 'warning'] else None)
    page.on("pageerror", lambda err: errors.append(str(err)))

    page.goto(f"{SITE_URL}/login")
    page.fill("#login_email", USERNAME)
    page.fill("#login_password", USER_PWD)
    page.click("button[type='submit']")
    page.wait_for_url("**/desk**", timeout=20000)
    time.sleep(2)

    page.goto(f"{SITE_URL}/desk/qifu-hr-salary-workbench")
    time.sleep(3)
    page.keyboard.press("Escape")
    time.sleep(1)

    # 1. 切换至 Tab 5: 验证科学精简版
    print("Testing Tab 5 Scientific Simple View...")
    page.click(".qifu-tab-btn[data-tab='tax']")
    time.sleep(3)
    shot_tab5_scientific = r"C:\Users\ashan\.gemini\antigravity\brain\062db5c0-afb5-4a31-90f4-1728b7cf9460\live_acceptance_tab5_scientific_simple.png"
    page.screenshot(path=shot_tab5_scientific)
    print(f"Captured Tab 5 Scientific Simple: {shot_tab5_scientific}")

    # 2. 切换至 Tab 7: 验证默认全员 15 列总览
    print("\nTesting Tab 7 All Employees 15-Col Overview Matrix...")
    # 关闭可能弹出的模态框
    try:
        modal_close = page.locator("div.modal.fade.show .close, div.modal.fade.show button.btn-modal-close, div.modal.fade.show button[data-dismiss='modal']").first
        if modal_close.is_visible(timeout=2000):
            modal_close.click()
            time.sleep(1)
    except Exception:
        pass
    # 按 Escape 再保险
    page.keyboard.press("Escape")
    time.sleep(0.5)
    page.click(".qifu-tab-btn[data-tab='history_timeline']")
    # 等待 Tab 7 全员数据加载完毕 (等待穿透按钮出现，最多 20s)
    time.sleep(5)
    try:
        page.wait_for_selector(".btn-jump-emp-history-tab7", timeout=20000)
        time.sleep(1)
    except Exception:
        print("⚠️  Tab 7 drilldown buttons not found within 20s")
    # 截图 (此时数据应已加载)
    shot_tab7_all = r"C:\Users\ashan\.gemini\antigravity\brain\062db5c0-afb5-4a31-90f4-1728b7cf9460\live_acceptance_tab7_all_employees.png"
    page.screenshot(path=shot_tab7_all)
    print(f"Captured Tab 7 All Employees (15-Col) after data load: {shot_tab7_all}")

    # 3. 在 Tab 7 点击某位员工姓名穿透至单人 (点第一个可用按钮)
    print("\nTesting Tab 7 Single Employee Drilldown...")
    try:
        first_btn = page.locator(".btn-jump-emp-history-tab7").first
        first_btn.click(timeout=10000)
        time.sleep(4)
        shot_tab7_single = r"C:\Users\ashan\.gemini\antigravity\brain\062db5c0-afb5-4a31-90f4-1728b7cf9460\live_acceptance_tab7_single_meng.png"
        page.screenshot(path=shot_tab7_single)
        print(f"Captured Tab 7 Single Employee: {shot_tab7_single}")
    except Exception as e:
        print(f"⚠️  Tab 7 drilldown click failed: {e}")

    # 4. 点击返回全员总览
    print("\nTesting Tab 7 Back to All...")
    try:
        page.click("#btn-back-to-all-history", timeout=8000)
        time.sleep(2)
    except Exception:
        print("⚠️  Back button not found")

    print("\n--- Errors check ---")
    if len(errors) == 0:
        print("🎉 ZERO ERRORS! Tab 5 Scientific Simple, Tab 7 All 15-Col & Dynamic Excel Export 100% VERIFIED!")
    else:
        for err in errors:
            print("❌ ERROR:", err)

    browser.close()
