import os
import sys
import openpyxl
import json

sys.stdout.reconfigure(encoding='utf-8')

xlsm_path = r"d:\SynologyDrive团队\antigravity\erpnext16\temp_screenshots\祺富人事202606(3) 的副本.xlsm"
wb = openpyxl.load_workbook(xlsm_path, data_only=True, keep_vba=True)

# 1. 提取人员信息表 (Sheet 3: 人员信息表)
ws_emp = wb["人员信息表"]
# 2. 提取社保 (Sheet 6: 社会保险)
ws_ss = wb["社会保险"]
# 3. 提取公积金 (Sheet 7: 住房公积金)
ws_hf = wb["住房公积金"]
# 4. 提取老板娘发薪表 (Sheet 4: 当月发薪工资表 或 临时表格)
ws_pay = wb["当月发薪工资表"]
ws_boss = wb["临时表格"]

print("--- 解析祺富员工基本档案 ---")
qifu_employees = []
for r in range(4, ws_emp.max_row + 1):
    emp_no = ws_emp.cell(row=r, column=2).value
    name = ws_emp.cell(row=r, column=3).value
    id_card = ws_emp.cell(row=r, column=4).value
    emp_type = ws_emp.cell(row=r, column=7).value or '正式工'
    status = ws_emp.cell(row=r, column=8).value or '在职'
    calc_mode = ws_emp.cell(row=r, column=9).value or '计薪方式' # 税前动态 / 税后管理

    if emp_no and name:
        qifu_employees.append({
            "employee_no": emp_no,
            "employee_name": name,
            "company": "天津祺富机械加工有限公司",
            "id_card": str(id_card) if id_card else "",
            "employee_type": emp_type,
            "employment_status": "在职",
            "salary_mode": "税前动态工资" if "动态" in str(calc_mode) else ("税后管理工资" if "管理" in str(calc_mode) else "税前动态工资"),
            "social_security_base": 5124,
            "housing_fund_base": 2320,
            "is_insured": 1
        })

print(f"提取到祺富员工 {len(qifu_employees)} 人:")
for e in qifu_employees[:6]:
    print(" ", e["employee_no"], e["employee_name"], e["employee_type"], e["salary_mode"])

# 解析老板娘工资表结构与前 10 行
print("\n--- 解析【老板娘工资表/临时表格】表头与前 5 行 ---")
headers_boss = [ws_boss.cell(row=2, column=c).value for c in range(1, ws_boss.max_column+1)]
print("表头:", [h for h in headers_boss if h is not None])

boss_sheet_sample = []
for r in range(3, min(ws_boss.max_row+1, 12)):
    row_dict = {}
    for c, h in enumerate(headers_boss, 1):
        if h:
            v = ws_boss.cell(row=r, column=c).value
            row_dict[h] = v
    if row_dict.get("姓名"):
        boss_sheet_sample.append(row_dict)
        print(f"Row {r}:", row_dict.get("姓名"), f"天数={row_dict.get('作业天数')}", f"天工资={row_dict.get('天工资')}", f"全勤={row_dict.get('全勤费')}", f"加班={row_dict.get('加班小时')}h (费={row_dict.get('加班费')})", f"实发={row_dict.get('实发工资')}")

with open(r"d:\SynologyDrive团队\antigravity\erpnext16\scripts\extracted_qifu_seed_data.json", "w", encoding="utf-8") as f:
    json.dump({
        "employees": qifu_employees,
        "boss_sheet_sample": boss_sheet_sample
    }, f, ensure_ascii=False, indent=2)

print("\n已导出 extracted_qifu_seed_data.json")
