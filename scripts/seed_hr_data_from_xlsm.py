import os
import sys
import openpyxl
import json
import datetime

sys.stdout.reconfigure(encoding='utf-8')

xlsm_path = r"d:\SynologyDrive团队\antigravity\erpnext16\temp_screenshots\202606吉众人事综合.xlsm"

wb = openpyxl.load_workbook(xlsm_path, data_only=True, keep_vba=True)

# 1. 提取日历数据 (Sheet 0)
ws_cal = wb.worksheets[0]
calendar_data = []
for r in range(2, ws_cal.max_row + 1):
    dt = ws_cal.cell(row=r, column=1).value
    if dt:
        if isinstance(dt, datetime.datetime):
            d_str = dt.strftime('%Y-%m-%d')
        else:
            d_str = str(dt)[:10]

        is_workday = int(ws_cal.cell(row=r, column=2).value or 0)
        is_weekend = int(ws_cal.cell(row=r, column=3).value or 0)
        is_shift_off = int(ws_cal.cell(row=r, column=4).value or 0)
        is_shift_work = int(ws_cal.cell(row=r, column=5).value or 0)
        is_legal_holiday = int(ws_cal.cell(row=r, column=6).value or 0)
        status_desc = str(ws_cal.cell(row=r, column=7).value or '')
        remark = str(ws_cal.cell(row=r, column=8).value or '')

        calendar_data.append({
            'date': d_str,
            'is_workday': is_workday,
            'is_weekend': is_weekend,
            'is_shift_off': is_shift_off,
            'is_shift_work': is_shift_work,
            'is_legal_holiday': is_legal_holiday,
            'status_desc': status_desc,
            'remark': remark
        })

print(f"提取到日历记录 {len(calendar_data)} 条 (从 {calendar_data[0]['date']} 至 {calendar_data[-1]['date']})")

# 2. 提取员工信息与薪资标准 (Sheet 3 & Sheet 4)
ws_emp = wb.worksheets[4]
ws_sal = wb.worksheets[3]

employees = []
# Sheet 4: 员工信息
emp_dict = {}
for r in range(3, ws_emp.max_row + 1):
    emp_no = ws_emp.cell(row=r, column=2).value
    if emp_no:
        name = ws_emp.cell(row=r, column=3).value
        id_card = ws_emp.cell(row=r, column=4).value
        mobile = ws_emp.cell(row=r, column=5).value
        emp_type = ws_emp.cell(row=r, column=8).value or '正式工'
        status = ws_emp.cell(row=r, column=9).value or '在职'
        is_insured = ws_emp.cell(row=r, column=10).value or '是'
        emp_dict[emp_no] = {
            'employee_no': emp_no,
            'employee_name': name,
            'id_card': str(id_card) if id_card else '',
            'mobile': str(mobile) if mobile else '',
            'employee_type': emp_type,
            'status': status,
            'is_insured': is_insured,
            'company': '天津吉众机电设备有限公司'
        }

# Sheet 3: 人员薪资信息
for r in range(3, ws_sal.max_row + 1):
    emp_no = ws_sal.cell(row=r, column=1).value
    if emp_no and emp_no in emp_dict:
        name = ws_sal.cell(row=r, column=2).value
        id_card = ws_sal.cell(row=r, column=3).value
        calc_type = ws_sal.cell(row=r, column=4).value
        emp_type = ws_sal.cell(row=r, column=5).value
        fixed_net_salary = ws_sal.cell(row=r, column=6).value
        monthly_deduction = ws_sal.cell(row=r, column=7).value or 5000
        meal_unit_price = ws_sal.cell(row=r, column=8).value or 0
        base_salary = ws_sal.cell(row=r, column=9).value or 0
        base_subsidy = ws_sal.cell(row=r, column=10).value or 0
        performance_bonus = ws_sal.cell(row=r, column=11).value or 0
        position_allowance = ws_sal.cell(row=r, column=12).value or 0
        social_security_base = ws_sal.cell(row=r, column=14).value or 5124
        housing_fund_base = ws_sal.cell(row=r, column=15).value or 2520
        special_additional_deduction = ws_sal.cell(row=r, column=16).value or 0

        salary_mode = '税后管理工资' if fixed_net_salary and float(fixed_net_salary) > 0 else '税前动态工资'

        emp_dict[emp_no]['salary_mode'] = salary_mode
        emp_dict[emp_no]['fixed_net_salary'] = float(fixed_net_salary or 0)
        emp_dict[emp_no]['meal_unit_price'] = float(meal_unit_price or 0)
        emp_dict[emp_no]['base_salary'] = float(base_salary or 0)
        emp_dict[emp_no]['base_subsidy'] = float(base_subsidy or 0)
        emp_dict[emp_no]['performance_bonus'] = float(performance_bonus or 0)
        emp_dict[emp_no]['position_allowance'] = float(position_allowance or 0)
        emp_dict[emp_no]['social_security_base'] = float(social_security_base or 0)
        emp_dict[emp_no]['housing_fund_base'] = float(housing_fund_base or 0)
        emp_dict[emp_no]['special_additional_deduction'] = float(special_additional_deduction or 0)

# 对未在薪资表中找到的员工赋默认值
for emp in emp_dict.values():
    emp.setdefault('salary_mode', '税前动态工资')
    emp.setdefault('fixed_net_salary', 0)
    emp.setdefault('meal_unit_price', 15)
    emp.setdefault('base_salary', 2510)
    emp.setdefault('base_subsidy', 600)
    emp.setdefault('performance_bonus', 600)
    emp.setdefault('position_allowance', 0)
    emp.setdefault('social_security_base', 5124)
    emp.setdefault('housing_fund_base', 2520)
    emp.setdefault('special_additional_deduction', 0)

print(f"提取到吉众员工薪资档案 {len(emp_dict)} 人:")
for k, v in list(emp_dict.items())[:5]:
    print(f"  [{v['employee_no']}] {v['employee_name']}: 模式={v['salary_mode']}, 基本工资={v.get('base_salary')}, 约定税后={v.get('fixed_net_salary')}, 专项附加={v.get('special_additional_deduction')}")


# 3. 提取历史数据 (Sheet 12)
ws_hist = wb.worksheets[12]
hist_records = []
for r in range(3, ws_hist.max_row + 1):
    emp_no = ws_hist.cell(row=r, column=1).value
    if emp_no:
        name = ws_hist.cell(row=r, column=2).value
        id_card = ws_hist.cell(row=r, column=3).value
        gross_salary = ws_hist.cell(row=r, column=13).value or 0
        tax_exemption = ws_hist.cell(row=r, column=14).value or 5000
        special_deduction = ws_hist.cell(row=r, column=15).value or 0
        additional_deduction = ws_hist.cell(row=r, column=16).value or 0
        tax_paid = ws_hist.cell(row=r, column=17).value or 0
        net_salary = ws_hist.cell(row=r, column=18).value or 0
        month_period = str(ws_hist.cell(row=r, column=19).value or '')

        hist_records.append({
            'employee_no': emp_no,
            'employee_name': name,
            'id_card': str(id_card) if id_card else '',
            'gross_salary': float(gross_salary),
            'tax_exemption': float(tax_exemption),
            'special_deduction': float(special_deduction),
            'additional_deduction': float(additional_deduction),
            'tax_paid': float(tax_paid),
            'net_salary': float(net_salary),
            'month_period': month_period,
            'company': '天津吉众机电设备有限公司'
        })

print(f"提取到历史纳税/薪资明细 {len(hist_records)} 条")

# 导出为 JSON 文件备用
with open(r"d:\SynologyDrive团队\antigravity\erpnext16\scripts\extracted_hr_seed_data.json", "w", encoding="utf-8") as f:
    json.dump({
        'calendar': calendar_data,
        'employees': list(emp_dict.values()),
        'history': hist_records
    }, f, ensure_ascii=False, indent=2)

print("已导出 extracted_hr_seed_data.json")
