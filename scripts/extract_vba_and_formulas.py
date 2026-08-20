import os
import sys
import openpyxl
import json
import zipfile
import subprocess

sys.stdout.reconfigure(encoding='utf-8')

xlsm_path = r"d:\SynologyDrive团队\antigravity\erpnext16\temp_screenshots\202606吉众人事综合.xlsm"

wb = openpyxl.load_workbook(xlsm_path, data_only=False, keep_vba=True)
print("="*80)
print(f"工作簿所有 Sheet 完整列表 ({len(wb.sheetnames)} 个):")
for idx, name in enumerate(wb.sheetnames, 1):
    print(f"  {idx}. {name}")
print("="*80)

# 安装/检查 oletools 提取 VBA 代码
try:
    from oletools.olevba import VBA_Parser
    vb_parser = VBA_Parser(xlsm_path)
    print("\n--- 提取 VBA 宏代码 ---")
    if vb_parser.detect_vba_macros():
        for (filename, stream_path, vba_filename, vba_code) in vb_parser.extract_macros():
            print(f"\n[VBA Module: {vba_filename} (Stream: {stream_path})]")
            print("-" * 50)
            print(vba_code)
            print("-" * 50)
    else:
        print("未检测到 VBA 宏代码")
except ImportError:
    print("正在安装 oletools 以提取 VBA 代码...")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "oletools"])
    from oletools.olevba import VBA_Parser
    vb_parser = VBA_Parser(xlsm_path)
    if vb_parser.detect_vba_macros():
        for (filename, stream_path, vba_filename, vba_code) in vb_parser.extract_macros():
            print(f"\n[VBA Module: {vba_filename} (Stream: {stream_path})]")
            print("-" * 50)
            print(vba_code)
            print("-" * 50)
