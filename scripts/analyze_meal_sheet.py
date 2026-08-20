import os
import shutil
import sys
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

src = r"D:\temp\xwechat_files\ashan4070_b2cf\msg\file\2026-08\订餐记录(2).xlsx"
dst = r"d:\SynologyDrive团队\antigravity\erpnext16\temp_screenshots\订餐记录(2).xlsx"

if not os.path.exists(src):
    src = r"D:\OneDrive\祺富吉众\天津祺富机械加工有限公司\月报表-孟\202408\凭证\订餐记录(2).xlsx"

shutil.copyfile(src, dst)
print(f"Copied {src} to {dst}")

wb = openpyxl.load_workbook(dst, data_only=True)
print("="*80)
print(f"工作簿包含 {len(wb.sheetnames)} 个工作表: {wb.sheetnames}")
for idx, name in enumerate(wb.sheetnames, 1):
    ws = wb[name]
    print(f"\n--- [{idx}] 工作表: {name} (行数: {ws.max_row}, 列数: {ws.max_column}) ---")
    for r in range(1, min(30, ws.max_row + 1)):
        row_vals = [ws.cell(r, c).value for c in range(1, min(20, ws.max_column + 1))]
        if any(v is not None for v in row_vals):
            print(f"Row {r:02d}: {row_vals}")

print("\n" + "="*80)
