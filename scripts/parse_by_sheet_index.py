import os
import sys
import openpyxl
import json

sys.stdout.reconfigure(encoding='utf-8')

xlsm_path = r"d:\SynologyDrive团队\antigravity\erpnext16\temp_screenshots\202606吉众人事综合.xlsm"

wb = openpyxl.load_workbook(xlsm_path, data_only=False, keep_vba=True)
wb_val = openpyxl.load_workbook(xlsm_path, data_only=True, keep_vba=True)

for idx, (ws, ws_v) in enumerate(zip(wb.worksheets, wb_val.worksheets)):
    title = ws.title
    print(f"\n==================== [{idx}] SHEET: {title} (Rows: {ws.max_row}, Cols: {ws.max_column}) ====================")

    # 打印前 6 行全部内容
    for r in range(1, min(ws.max_row + 1, 8)):
        row_f = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        row_v = [ws_v.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]

        items = []
        for c_idx, (f, v) in enumerate(zip(row_f, row_v), 1):
            if f is not None or v is not None:
                col_let = openpyxl.utils.get_column_letter(c_idx)
                if str(f).startswith('='):
                    items.append(f"{col_let}{r}: [{f}] => {v}")
                else:
                    items.append(f"{col_let}{r}: {f}")
        if items:
            print(f"Row {r}: " + " | ".join(items[:8]))
            if len(items) > 8:
                print(f"        ... " + " | ".join(items[8:18]))
            if len(items) > 18:
                print(f"        ... " + " | ".join(items[18:30]))
            if len(items) > 30:
                print(f"        ... " + " | ".join(items[30:40]))
