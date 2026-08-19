import os
import sys
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

xlsm_path = r"d:\SynologyDrive团队\antigravity\erpnext16\temp_screenshots\202606吉众人事综合.xlsm"
wb = openpyxl.load_workbook(xlsm_path, data_only=False, keep_vba=True)
wb_v = openpyxl.load_workbook(xlsm_path, data_only=True, keep_vba=True)

ws_work = wb.worksheets[5]
ws_work_v = wb_v.worksheets[5]

for r in range(1, ws_work.max_row + 1):
    name = ws_work_v.cell(row=r, column=3).value
    emp_no = ws_work_v.cell(row=r, column=2).value
    if name == "李传凤" or emp_no == "A0003":
        print(f"=== 找到李传凤位于 Row {r} ===")
        for c in range(1, ws_work.max_column + 1):
            f = ws_work.cell(row=r, column=c).value
            v = ws_work_v.cell(row=r, column=c).value
            h4 = ws_work.cell(row=4, column=c).value or ''
            col_l = openpyxl.utils.get_column_letter(c)
            if f is not None or v is not None:
                print(f"  Col {col_l} ({h4}): [{f}] => {v}")
