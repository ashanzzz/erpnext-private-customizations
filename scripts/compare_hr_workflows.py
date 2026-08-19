import os
import sys
import openpyxl
import json

sys.stdout.reconfigure(encoding='utf-8')

jizhong_path = r"d:\SynologyDrive团队\antigravity\erpnext16\temp_screenshots\202606吉众人事综合.xlsm"
qifu_path = r"d:\SynologyDrive团队\antigravity\erpnext16\temp_screenshots\祺富人事202606(3) 的副本.xlsm"

wb_jz = openpyxl.load_workbook(jizhong_path, data_only=True, keep_vba=True)
wb_qf = openpyxl.load_workbook(qifu_path, data_only=True, keep_vba=True)

print("="*80)
print("【吉众人事综合】14 个工作表数据流概览:")
for idx, name in enumerate(wb_jz.sheetnames, 1):
    ws = wb_jz[name]
    print(f"  {idx:02d}. [{name}] (行: {ws.max_row}, 列: {ws.max_column})")

print("\n" + "="*80)
print("【祺富人事】12 个工作表数据流概览:")
for idx, name in enumerate(wb_qf.sheetnames, 1):
    ws = wb_qf[name]
    print(f"  {idx:02d}. [{name}] (行: {ws.max_row}, 列: {ws.max_column})")

print("="*80)
