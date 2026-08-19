import os
import sys
import openpyxl
import zipfile
import json

xlsm_path = r"d:\SynologyDrive团队\antigravity\erpnext16\temp_screenshots\202606吉众人事综合.xlsm"

print(f"--- Analyzing XLSM: {xlsm_path} ---")

# 1. 打开工作簿查看全部 Sheet
wb = openpyxl.load_workbook(xlsm_path, data_only=False, keep_vba=True)
print("Sheet Names:", wb.sheetnames)

sheet_info = {}
for name in wb.sheetnames:
    ws = wb[name]
    max_r = ws.max_row
    max_c = ws.max_column
    # 读取前 15 行的前 30 列看看内容
    rows_sample = []
    for r in range(1, min(max_r + 1, 15)):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, min(max_c + 1, 30))]
        if any(v is not None for v in row_vals):
            rows_sample.append(row_vals)
    sheet_info[name] = {
        "max_row": max_r,
        "max_col": max_c,
        "sample": rows_sample
    }

print("\n--- Sheet Summary ---")
for k, v in sheet_info.items():
    print(f"Sheet: {k}, Rows: {v['max_row']}, Cols: {v['max_col']}")
    print(f"  First 3 rows sample:")
    for r in v['sample'][:3]:
        # print non-None values
        non_empty = [str(x) for x in r if x is not None]
        print("   ", " | ".join(non_empty[:10]))

# 2. 检查是否有 VBA 项目
with zipfile.ZipFile(xlsm_path, 'r') as z:
    file_list = z.namelist()
    print("\nZip entries in XLSM:")
    for f in file_list:
        if 'vba' in f.lower() or 'macro' in f.lower():
            print("  VBA entry:", f)
