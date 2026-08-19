import os
import sys
import openpyxl
import json

sys.stdout.reconfigure(encoding='utf-8')

xlsm_path = r"d:\SynologyDrive团队\antigravity\erpnext16\temp_screenshots\202606吉众人事综合.xlsm"

wb = openpyxl.load_workbook(xlsm_path, data_only=False, keep_vba=True)
wb_val = openpyxl.load_workbook(xlsm_path, data_only=True, keep_vba=True)

# 1. 解析【年历】与【月历】和【考勤表】
ws_year = wb_val["年历"]
print("--- [年历] 数据摘要 (前 10 天) ---")
for r in range(1, 12):
    row_vals = [ws_year.cell(row=r, column=c).value for c in range(1, 9)]
    print(f"Row {r}:", row_vals)

# 2. 解析【基本薪资信息】与【员工信息】
ws_emp = wb_val["员工信息"]
print("\n--- [员工信息] 表头与前 3 行 ---")
for r in range(1, 6):
    print(f"Row {r}:", [ws_emp.cell(row=r, column=c).value for c in range(1, 14)])

ws_sal_info = wb_val["基本薪资信息"]
print("\n--- [基本薪资信息] 表头与前 5 行 ---")
for r in range(1, 8):
    print(f"Row {r}:", [ws_sal_info.cell(row=r, column=c).value for c in range(1, 17)])

# 3. 解析【本月工时核定】的表头和计算公式
ws_work = wb["本月工时核定"]
ws_work_v = wb_val["本月工时核定"]
print("\n--- [本月工时核定] 表头 (Row 3, 4) ---")
headers_r3 = [ws_work.cell(row=3, column=c).value for c in range(1, 37)]
headers_r4 = [ws_work.cell(row=4, column=c).value for c in range(1, 37)]
print("Row 3:", headers_r3)
print("Row 4:", headers_r4)

print("\n--- [本月工时核定] 数据行示例 (Row 5 苏锡成/陈亮) 公式与值 ---")
for c in range(1, 37):
    f = ws_work.cell(row=5, column=c).value
    v = ws_work_v.cell(row=5, column=c).value
    h3 = headers_r3[c-1] or ''
    h4 = headers_r4[c-1] or ''
    h = f"{h3} {h4}".strip()
    if f is not None or v is not None:
        c_letter = openpyxl.utils.get_column_letter(c)
        print(f"  Col {c_letter} ({h}): [{f}] => {v}")

# 4. 解析【基本参数】
ws_param = wb_val["基本参数"]
print("\n--- [基本参数] 关键参数 ---")
for r in range(1, 35):
    row_vals = [ws_param.cell(row=r, column=c).value for c in range(1, 14)]
    if any(x is not None for x in row_vals):
        print(f"Row {r}:", [x for x in row_vals if x is not None])
