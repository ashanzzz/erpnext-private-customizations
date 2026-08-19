import os
import sys
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

xlsm_path = r"d:\SynologyDrive团队\antigravity\erpnext16\temp_screenshots\祺富人事202606(3) 的副本.xlsm"
wb = openpyxl.load_workbook(xlsm_path, data_only=True, keep_vba=True)

print("="*80)
print(f"祺富人事工作簿共有 {len(wb.sheetnames)} 个工作表:")
for idx, name in enumerate(wb.sheetnames, 1):
    ws = wb[name]
    print(f"{idx}. [{name}] (行: {ws.max_row}, 列: {ws.max_column})")

print("="*80)

# 解析关键 Sheet 的结构
important_sheets = ["仪表盘", "数据信息", "人员信息表", "当月发薪工资表", "记账工资表", "老板娘工资表", "个人所得税", "人员类型规则", "数据库"]

for name in wb.sheetnames:
    ws = wb[name]
    print(f"\n==================== SHEET: [{name}] ====================")
    for r in range(1, min(ws.max_row + 1, 10)):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, min(ws.max_column + 1, 25))]
        non_empty = [f"Col {openpyxl.utils.get_column_letter(c_idx)}: {v}" for c_idx, v in enumerate(row_vals, 1) if v is not None]
        if non_empty:
            print(f"  Row {r}: " + " | ".join(non_empty[:10]))
            if len(non_empty) > 10:
                print(f"          ... " + " | ".join(non_empty[10:20]))
