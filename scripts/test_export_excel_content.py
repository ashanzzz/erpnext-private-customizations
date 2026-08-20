import os
import paramiko
from dotenv import load_dotenv

load_dotenv()

HOST = os.getenv('UNRAID_SSH_HOST', '192.168.8.11')
PORT = int(os.getenv('UNRAID_SSH_PORT', '22'))
USER = os.getenv('UNRAID_SSH_USER', 'root')
PASSWORD = os.getenv('UNRAID_SSH_PASSWORD', '')

client = paramiko.SSHClient()
client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
client.connect(HOST, port=PORT, username=USER, password=PASSWORD, timeout=15)

script = """# -*- coding: utf-8 -*-
import frappe
import openpyxl
import io

frappe.init(site='site1.local')
frappe.connect()

from ashan_cn_procurement.services.property_settlement import export_utility_settlement_excel

# 导出 2026-08 的 Excel
export_utility_settlement_excel('2026-08', mode='all')
filecontent = frappe.response.get('filecontent')
filename = frappe.response.get('filename')

print(f"Generated Excel: {filename}, size: {len(filecontent)} bytes")

wb = openpyxl.load_workbook(io.BytesIO(filecontent), data_only=True)
for sname in wb.sheetnames:
    ws = wb[sname]
    print(f"\\n=== Sheet: {sname} ===")
    for r in range(1, ws.max_row + 1):
        row_vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if any(v is not None and str(v).strip() != '' for v in row_vals):
            print(f"Row {r:2d}: {row_vals}")
"""

sftp = client.open_sftp()
with sftp.file('/tmp/test_excel.py', 'wb') as f:
    f.write(script.encode('utf-8'))
sftp.close()

stdin, stdout, stderr = client.exec_command("docker cp /tmp/test_excel.py erpnext16:/tmp/test_excel.py && docker exec -u frappe -w /home/frappe/frappe-bench/sites erpnext16 ../env/bin/python /tmp/test_excel.py")
print(stdout.read().decode('utf-8', errors='replace'))
print(stderr.read().decode('utf-8', errors='replace'))

client.close()
