import os
import sys
import openpyxl
import json

# 设置标准输出编码为 utf-8
sys.stdout.reconfigure(encoding='utf-8')

xlsm_path = r"d:\SynologyDrive团队\antigravity\erpnext16\temp_screenshots\202606吉众人事综合.xlsm"

wb = openpyxl.load_workbook(xlsm_path, data_only=False, keep_vba=True)
wb_val = openpyxl.load_workbook(xlsm_path, data_only=True, keep_vba=True)

print("="*60)
print(f"工作簿所有 Sheet 名称 ({len(wb.sheetnames)} 个):")
for idx, name in enumerate(wb.sheetnames, 1):
    ws = wb[name]
    print(f"  {idx}. [{name}] (行: {ws.max_row}, 列: {ws.max_column})")

print("="*60)

# 对每个 Sheet 详细解析表头、关键结构与公式
for name in wb.sheetnames:
    ws = wb[name]
    ws_v = wb_val[name]
    print(f"\n==================== SHEET: {name} ====================")
    for r in range(1, min(ws.max_row + 1, 10)):
        row_formula = [ws.cell(row=r, column=c).value for c in range(1, min(ws.max_column + 1, 35))]
        row_val = [ws_v.cell(row=r, column=c).value for c in range(1, min(ws_v.max_column + 1, 35))]

        # 打印非空行
        if any(v is not None for v in row_formula):
            items = []
            for col_idx, (f, v) in enumerate(zip(row_formula, row_val), 1):
                if f is not None:
                    col_letter = openpyxl.utils.get_column_letter(col_idx)
                    if str(f).startswith('='):
                        items.append(f"{col_letter}{r}: [{f}] -> (值: {v})")
                    else:
                        items.append(f"{col_letter}{r}: {f}")
            print(f"Row {r}: " + " | ".join(items[:8]))
            if len(items) > 8:
                print(f"        ... 剩余 {len(items)-8} 列: " + " | ".join(items[8:16]))
