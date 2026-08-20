import os
import sys
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

xlsm_path = r"d:\SynologyDrive团队\antigravity\erpnext16\temp_screenshots\202606吉众人事综合.xlsm"
wb = openpyxl.load_workbook(xlsm_path, data_only=False, keep_vba=True)
wb_v = openpyxl.load_workbook(xlsm_path, data_only=True, keep_vba=True)

ws_work = wb.worksheets[5]
ws_work_v = wb_v.worksheets[5]

print("=== 李传凤 (Row 5) 每一列公式与值 ===")
for c in range(1, ws_work.max_column + 1):
    f = ws_work.cell(row=5, column=c).value
    v = ws_work_v.cell(row=5, column=c).value
    h3 = ws_work.cell(row=3, column=c).value or ''
    h4 = ws_work.cell(row=4, column=c).value or ''
    h = f"{h3} {h4}".strip()
    col_let = openpyxl.utils.get_column_letter(c)
    if f is not None or v is not None:
        print(f"Col {col_let} ({h}): [{f}] => {v}")
