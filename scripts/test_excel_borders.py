import os
import io
import sys
import openpyxl
import paramiko
from dotenv import load_dotenv

sys.stdout.reconfigure(encoding='utf-8')
load_dotenv()

HOST = os.getenv('UNRAID_SSH_HOST', '192.168.8.11')
PORT = int(os.getenv('UNRAID_SSH_PORT', '22'))
USER = os.getenv('UNRAID_SSH_USER', 'root')
PASSWORD = os.getenv('UNRAID_SSH_PASSWORD', '')

def run_cmd(client, cmd):
    stdin, stdout, stderr = client.exec_command(cmd)
    return stdout.read().decode('utf-8', errors='replace')

client = paramiko.SSHClient()
client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
client.connect(HOST, port=PORT, username=USER, password=PASSWORD, timeout=15)

py_code = """
import frappe
frappe.init(site='site1.local', sites_path='.')
frappe.connect()

from ashan_cn_procurement.services.property_settlement import export_utility_settlement_excel
res = export_utility_settlement_excel("2026-08", mode="all")
content = frappe.response.get('filecontent')
with open('/tmp/test_export_utility.xlsx', 'wb') as f:
    f.write(content)
print("Saved /tmp/test_export_utility.xlsx, size:", len(content))
frappe.destroy()
"""

sftp = client.open_sftp()
with sftp.file('/tmp/run_export.py', 'w') as f:
    f.write(py_code)
sftp.close()

run_cmd(client, "docker cp /tmp/run_export.py erpnext16:/home/frappe/frappe-bench/sites/run_export.py")
out = run_cmd(client, "docker exec -u frappe -w /home/frappe/frappe-bench/sites erpnext16 /home/frappe/frappe-bench/env/bin/python run_export.py")
print(out)

# 下载到本地检查 openpyxl 样式
run_cmd(client, "docker cp erpnext16:/tmp/test_export_utility.xlsx /tmp/test_export_utility.xlsx")
sftp = client.open_sftp()
local_xlsx = r"d:\SynologyDrive团队\antigravity\erpnext16\temp_screenshots\test_export_utility.xlsx"
sftp.get("/tmp/test_export_utility.xlsx", local_xlsx)
sftp.close()
client.close()

# 检验本地 Excel 的合并单元格和边框
wb = openpyxl.load_workbook(local_xlsx)
print("\nSheets found:", wb.sheetnames)
for sname in wb.sheetnames:
    ws = wb[sname]
    print(f"\n--- Checking Sheet: [{sname}] ---")
    print("Merged cell ranges:", [str(r) for r in ws.merged_cells.ranges])
    # 查找合并在 H 列的大字合计
    for rng in ws.merged_cells.ranges:
        if rng.min_col == 8 and rng.max_col == 8:
            print(f"Found H-column Grand Total Merged Range: {rng}")
            for r in range(rng.min_row, rng.max_row + 1):
                cell = ws.cell(r, 8)
                b = cell.border
                print(f"  Row {r}, Col H ({cell.value}): top={b.top.style if b.top else None}, bottom={b.bottom.style if b.bottom else None}, left={b.left.style if b.left else None}, right={b.right.style if b.right else None}")

print("\n[OK] Excel Border Verification Complete!")
