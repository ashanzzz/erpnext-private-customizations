import os
import sys
import openpyxl
import json
from oletools.olevba import VBA_Parser

sys.stdout.reconfigure(encoding='utf-8')

xlsm_path = r"d:\SynologyDrive团队\antigravity\erpnext16\temp_screenshots\祺富人事202606(3) 的副本.xlsm"

print(f"--- 深度剖析祺富人事工作簿: {xlsm_path} ---")

wb = openpyxl.load_workbook(xlsm_path, data_only=False, keep_vba=True)
wb_v = openpyxl.load_workbook(xlsm_path, data_only=True, keep_vba=True)

print("="*80)
print(f"祺富工作簿所有工作表 ({len(wb.sheetnames)} 个):")
for idx, name in enumerate(wb.sheetnames, 1):
    ws = wb[name]
    print(f"  {idx}. [{name}] (行: {ws.max_row}, 列: {ws.max_column})")
print("="*80)

# 对每个 Sheet 详细解析表头与前 6 行
for idx, name in enumerate(wb.sheetnames, 1):
    ws = wb[name]
    ws_v = wb_v[name]
    print(f"\n==================== [{idx}] SHEET: {name} (Rows: {ws.max_row}, Cols: {ws.max_column}) ====================")
    for r in range(1, min(ws.max_row + 1, 8)):
        row_f = [ws.cell(row=r, column=c).value for c in range(1, min(ws.max_column + 1, 35))]
        row_v = [ws_v.cell(row=r, column=c).value for c in range(1, min(ws_v.max_column + 1, 35))]

        items = []
        for c_idx, (f, v) in enumerate(zip(row_f, row_v), 1):
            if f is not None or v is not None:
                col_let = openpyxl.utils.get_column_letter(c_idx)
                if str(f).startswith('='):
                    items.append(f"{col_let}{r}: [{f}] => {v}")
                else:
                    items.append(f"{col_let}{r}: {f}")
        if items:
            print(f"Row {r}: " + " | ".join(items[:8]))
            if len(items) > 8:
                print(f"        ... " + " | ".join(items[8:18]))
            if len(items) > 18:
                print(f"        ... " + " | ".join(items[18:30]))

# 提取 VBA 宏
print("\n" + "="*80)
print("--- 提取祺富 VBA 宏模块代码 ---")
vb = VBA_Parser(xlsm_path)
if vb.detect_vba_macros():
    for (filename, stream_path, vba_filename, vba_code) in vb.extract_macros():
        print(f"\n[VBA Module: {vba_filename} (Stream: {stream_path})]")
        print("-" * 50)
        print(vba_code[:1000] + ("\n... (truncated)" if len(vba_code) > 1000 else ""))
        print("-" * 50)
else:
    print("未检测到 VBA 宏")
