import os
import requests
import openpyxl

def load_env_file(env_path='.env'):
    if os.path.exists(env_path):
        with open(env_path, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith('#') and '=' in line:
                    k, v = line.split('=', 1)
                    os.environ.setdefault(k.strip(), v.strip())

load_env_file()

SITE_URL = os.getenv('ERPNEXT_SITE_URL_LOCAL', 'http://192.168.8.11:6888')
USERNAME = os.getenv('ERPNEXT_USERNAME', 'dev@example.invalid')
USER_PWD = os.getenv('ERPNEXT_PASSWORD', '')

def test_exports():
    session = requests.Session()
    # 1. Login
    login_res = session.post(f"{SITE_URL}/api/method/login", data={
        "usr": USERNAME,
        "pwd": USER_PWD
    })
    print(f"Login status: {login_res.status_code}, response: {login_res.json()}")

    # 2. Test Utility Excel export
    util_url = f"{SITE_URL}/api/method/ashan_cn_procurement.services.property_settlement.export_utility_settlement_excel?settlement_month=2026-08-01&mode=all&property_management_company=天津金利达物业管理有限公司"
    res_u = session.get(util_url)
    print(f"Utility Excel Export status: {res_u.status_code}, size: {len(res_u.content)} bytes")

    # Check workbook sheets
    import io
    wb_u = openpyxl.load_workbook(io.BytesIO(res_u.content), data_only=False)
    print(f"Utility Workbook Sheet names: {wb_u.sheetnames}")

    # 3. Test Lease Excel export
    lease_url = f"{SITE_URL}/api/method/ashan_cn_procurement.services.property_settlement.export_lease_settlement_excel?settlement_month=2026-08-01&mode=all&property_management_company=天津金利达物业管理有限公司"
    res_l = session.get(lease_url)
    print(f"Lease Excel Export status: {res_l.status_code}, size: {len(res_l.content)} bytes")
    wb_l = openpyxl.load_workbook(io.BytesIO(res_l.content), data_only=False)
    print(f"Lease Workbook Sheet names: {wb_l.sheetnames}")

if __name__ == "__main__":
    test_exports()
