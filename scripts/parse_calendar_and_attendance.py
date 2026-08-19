import os
import sys
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

xlsm_path = r"d:\SynologyDrive团队\antigravity\erpnext16\temp_screenshots\202606吉众人事综合.xlsm"
wb = openpyxl.load_workbook(xlsm_path, data_only=False, keep_vba=True)
wb_v = openpyxl.load_workbook(xlsm_path, data_only=True, keep_vba=True)

# 1. 详细解析年历
ws_year = wb.worksheets[0]
ws_year_v = wb_v.worksheets[0]
print(f"=== 年历 (Title: {ws_year.title}) ===")
for r in range(1, 10):
    row_str = [f"{openpyxl.utils.get_column_letter(c)}{r}: {ws_year.cell(row=r, column=c).value}" for c in range(1, ws_year.max_column+1) if ws_year.cell(row=r, column=c).value is not None]
    print(" | ".join(row_str))

# 2. 详细解析考勤表
ws_att = wb.worksheets[2]
ws_att_v = wb_v.worksheets[2]
print(f"\n=== 考勤表 (Title: {ws_att.title}) ===")
print("Row 1 (年月):", ws_att.cell(row=1, column=1).value)
print("Row 2 (日期):", [ws_att.cell(row=2, column=c).value for c in range(1, 15)])
print("Row 3 (星期/状态):", [ws_att.cell(row=3, column=c).value for c in range(1, 15)])
for r in range(4, 8):
    row_str = [f"{ws_att.cell(row=r, column=c).value}" for c in range(1, 15)]
    print(f"Row {r} (员工考勤打卡示例):", row_str)

# 3. 详细解析基本薪资信息
ws_sal = wb.worksheets[3]
ws_sal_v = wb_v.worksheets[3]
print(f"\n=== 基本薪资信息 (Title: {ws_sal.title}) ===")
headers_sal = [ws_sal.cell(row=2, column=c).value for c in range(1, ws_sal.max_column+1)]
print("Headers:", headers_sal)
for r in range(3, 7):
    row_str = [f"{h}: {ws_sal_v.cell(row=r, column=idx+1).value}" for idx, h in enumerate(headers_sal) if h is not None]
    print(f"Emp {r-2}:", " | ".join(row_str[:8]))
    print("      ... ", " | ".join(row_str[8:]))

# 4. 详细解析本月工时核定 (Sheet 5)
ws_time = wb.worksheets[5]
ws_time_v = wb_v.worksheets[5]
print(f"\n=== 本月工时核定 (Title: {ws_time.title}) ===")
print("Row 2 (说明):", ws_time.cell(row=2, column=1).value)
h3 = [ws_time.cell(row=3, column=c).value for c in range(1, ws_time.max_column+1)]
h4 = [ws_time.cell(row=4, column=c).value for c in range(1, ws_time.max_column+1)]
full_h = [f"{h3[i] or ''} {h4[i] or ''}".strip() for i in range(len(h3))]
print("Full Headers (前 20 列):", full_h[:20])
print("Full Headers (后 16 列):", full_h[20:])

for r in range(5, 8):
    emp_name = ws_time_v.cell(row=r, column=3).value
    print(f"\n--- 员工 {emp_name} 工时核定计算 ---")
    for c in range(1, ws_time.max_column+1):
        f = ws_time.cell(row=r, column=c).value
        v = ws_time_v.cell(row=r, column=c).value
        col_name = full_h[c-1]
        if col_name and (f is not None or v is not None):
            col_l = openpyxl.utils.get_column_letter(c)
            if str(f).startswith('='):
                print(f"  {col_l} ({col_name}): [{f}] => {v}")
            else:
                print(f"  {col_l} ({col_name}): {f}")
