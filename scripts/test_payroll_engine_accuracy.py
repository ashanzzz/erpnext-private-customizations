import os
import sys
import json

sys.stdout.reconfigure(encoding='utf-8')

# 引入我们的计算引擎
sys.path.append(r"d:\SynologyDrive团队\antigravity\erpnext16\ashan_cn_procurement\ashan_cn_procurement\ashan_cn_procurement")
from payroll_engine import AshanPayrollCalculator

# 读取提取出的种子数据
with open(r"d:\SynologyDrive团队\antigravity\erpnext16\scripts\extracted_hr_seed_data.json", "r", encoding="utf-8") as f:
    data = json.load(f)

# 读取 Excel 原表中的 2026-06 结算结果进行 1:1 对比
import openpyxl
xlsm_path = r"d:\SynologyDrive团队\antigravity\erpnext16\temp_screenshots\202606吉众人事综合.xlsm"
wb_v = openpyxl.load_workbook(xlsm_path, data_only=True, keep_vba=True)

# 考勤表
ws_att = wb_v.worksheets[2]
# 工资核定表
ws_work = wb_v.worksheets[5]
# 个税表
ws_tax = wb_v.worksheets[9]

# 构建 2026-06 员工考勤数据
emp_att_map = {}
for r in range(4, ws_att.max_row + 1, 4):
    emp_no = ws_att.cell(row=r, column=1).value
    name = ws_att.cell(row=r, column=2).value
    if emp_no:
        # 统计出勤与加班
        work_row = r + 1
        ot_row = r + 2
        meal_row = r + 3

        tot_work = 0.0
        tot_ot = 0.0
        tot_meal = 0
        days_present = 0

        for c in range(4, 35):
            w = ws_att.cell(row=work_row, column=c).value
            ot = ws_att.cell(row=ot_row, column=c).value
            m = ws_att.cell(row=meal_row, column=c).value

            if w is not None and float(w) > 0:
                tot_work += float(w)
                days_present += 1
            if ot is not None and float(ot) > 0:
                tot_ot += float(ot)
            if m is not None and int(m) > 0:
                tot_meal += int(m)

        emp_att_map[emp_no] = {
            'attendance_days': days_present,
            'work_hours_regular': tot_work,
            'overtime_regular_1_5': tot_ot,
            'overtime_weekend_2_0': 0.0,
            'overtime_holiday_3_0': 0.0,
            'meal_count': tot_meal
        }

print(f"=== 校验 2026-06 薪资计算引擎精确度 ===")
employees = data["employees"]

# 提取历史累计个税数据
hist_records = data["history"]
# 查找 202512 - 202605 历史累计
emp_hist_map = {}
for h in hist_records:
    emp_no = h["employee_no"]
    if emp_no not in emp_hist_map:
        emp_hist_map[emp_no] = {
            "cum_gross_prior": 0.0,
            "cum_tax_exemption_prior": 0.0,
            "cum_special_deduction_prior": 0.0,
            "cum_additional_deduction_prior": 0.0,
            "cum_tax_paid_prior": 0.0,
            "month_count_prior": 0
        }
    emp_hist_map[emp_no]["cum_gross_prior"] += h["gross_salary"]
    emp_hist_map[emp_no]["cum_tax_exemption_prior"] += h["tax_exemption"]
    emp_hist_map[emp_no]["cum_special_deduction_prior"] += h["special_deduction"]
    emp_hist_map[emp_no]["cum_additional_deduction_prior"] += h["additional_deduction"]
    emp_hist_map[emp_no]["cum_tax_paid_prior"] += h["tax_paid"]
    emp_hist_map[emp_no]["month_count_prior"] += 1

diff_count = 0
for emp in employees:
    emp_no = emp["employee_no"]
    if emp_no == "工号":
        continue
    att = emp_att_map.get(emp_no, {
        'attendance_days': 21,
        'work_hours_regular': 168,
        'overtime_regular_1_5': 0,
        'overtime_weekend_2_0': 0,
        'overtime_holiday_3_0': 0,
        'meal_count': 0
    })

    # 查找 Excel 表中该员工的实际结果
    # 查找工号所在行
    excel_gross = None
    excel_tax = None
    excel_net = None
    for r in range(5, ws_work.max_row + 1):
        if ws_work.cell(row=r, column=2).value == emp_no:
            excel_gross = ws_work.cell(row=r, column=27).value
            excel_tax = ws_work.cell(row=r, column=33).value
            excel_net = ws_work.cell(row=r, column=35).value
            break

    hist = emp_hist_map.get(emp_no, {
        "cum_gross_prior": 0.0,
        "cum_tax_exemption_prior": 0.0,
        "cum_special_deduction_prior": 0.0,
        "cum_additional_deduction_prior": 0.0,
        "cum_tax_paid_prior": 0.0,
        "month_count_prior": 0
    })

    calc_res = AshanPayrollCalculator.calculate_employee_payroll(
        salary_profile=emp,
        attendance=att,
        full_work_days=21.0,
        full_work_hours=168.0,
        fixed_work_hours=172.0,
        cum_history=hist
    )

    print(f"\n[{emp_no}] {emp['employee_name']} ({emp['salary_mode']}):")
    print(f"  计算引擎 => 应发: {calc_res['gross_pay']}, 个税: {calc_res['individual_tax']}, 实发: {calc_res['net_pay']}, 配钞: 100元x{calc_res['bill_100']} 50元x{calc_res['bill_50']} 20元x{calc_res['bill_20']} 10元x{calc_res['bill_10']} 5元x{calc_res['bill_5']} 1元x{calc_res['bill_1']}")
    print(f"  Excel原表 => 应发: {excel_gross}, 个税: {excel_tax}, 实发: {excel_net}")
