import os
import openpyxl
import json

temp_dir = r"d:\SynologyDrive团队\antigravity\erpnext16\temp_screenshots"
target_file = None
for f in os.listdir(temp_dir):
    if "吉众" in f and f.endswith(".xlsm"):
        target_file = os.path.join(temp_dir, f)
        break

print(f"Reading historical data from: {target_file}")
wb = openpyxl.load_workbook(target_file, data_only=True)
ws = wb["历史数据"]

# Col 1: 工号
# Col 2: 姓名
# Col 3: 身份证号
# Col 4: 基本工资
# Col 5: 基本补贴
# Col 6: 绩效奖金
# Col 7: 职位津贴
# Col 8: 餐补单价
# Col 9: 工资调整
# Col 10: 社险基数
# Col 11: 公积金基数
# Col 12: 计薪方式
# Col 13: 应发薪资合计
# Col 14: 当月个税免征额
# Col 15: 当月专项扣除 (社保p + 公积金p)
# Col 16: 个税专项附加扣除核定
# Col 17: 当月应缴纳所得税
# Col 18: 实发薪资合计
# Col 19: 历史记录年月

periods = {}
total_count = 0

for r in range(3, ws.max_row + 1):
    raw_month = ws.cell(r, 19).value
    if not raw_month:
        continue
    
    month_str = str(raw_month).strip()
    if len(month_str) == 6:
        period_formatted = f"{month_str[:4]}-{month_str[4:]}"
    elif len(month_str) == 7 and "-" in month_str:
        period_formatted = month_str
    else:
        continue
    
    emp_no = str(ws.cell(r, 1).value or "").strip()
    emp_name = str(ws.cell(r, 2).value or "").strip()
    id_card = str(ws.cell(r, 3).value or "").strip()
    base_sal = float(ws.cell(r, 4).value or 0)
    post_allow = float(ws.cell(r, 5).value or 0)
    perf_base = float(ws.cell(r, 6).value or 0)
    job_allow = float(ws.cell(r, 7).value or 0)
    meal_unit = float(ws.cell(r, 8).value or 0)
    sal_adj = float(ws.cell(r, 9).value or 0)
    ss_base = float(ws.cell(r, 10).value or 0)
    hf_base = float(ws.cell(r, 11).value or 0)
    sal_mode = str(ws.cell(r, 12).value or "").strip()
    
    gross_sal = ws.cell(r, 13).value
    tax_thresh = float(ws.cell(r, 14).value or 5000.0)
    special_ded = float(ws.cell(r, 15).value or 0)
    add_ded = float(ws.cell(r, 16).value or 0)
    tax_amt = float(ws.cell(r, 17).value or 0)
    net_sal = float(ws.cell(r, 18).value or 0)
    
    if gross_sal is None:
        gross_sal = net_sal + special_ded + tax_amt
    else:
        gross_sal = float(gross_sal)

    # 拆分专项扣除为社保个人和公积金个人 (公积金为 hf_base * 0.05, 余下为社保)
    hf_person = round(hf_base * 0.05, 2)
    ss_person = round(max(0.0, special_ded - hf_person), 2)

    item = {
        "employee_no": emp_no,
        "employee_name": emp_name,
        "id_card": id_card,
        "salary_mode": sal_mode,
        "base_salary": base_sal,
        "post_allowance": post_allow,
        "performance_salary": perf_base,
        "meal_unit_price": meal_unit,
        "salary_adjustment": sal_adj,
        "ss_base": ss_base,
        "hf_base": hf_base,
        "ss_person_total": ss_person,
        "hf_person_total": hf_person,
        "special_deductions_total": add_ded,
        "tax_threshold": tax_thresh,
        "tax_amount": tax_amt,
        "net_salary": net_sal,
        "gross_salary": gross_sal,
    }

    if period_formatted not in periods:
        periods[period_formatted] = []
    periods[period_formatted].append(item)
    total_count += 1

out_json = r"d:\SynologyDrive团队\antigravity\erpnext16\scripts\jizhong_history_records.json"
with open(out_json, "w", encoding="utf-8") as f:
    json.dump(periods, f, ensure_ascii=False, indent=2)

print(f"Successfully processed {total_count} records across {len(periods)} periods!")
for p in sorted(periods.keys()):
    print(f"  - {p}: {len(periods[p])} items")
