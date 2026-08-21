import os
import sys
import openpyxl
import requests
from dotenv import load_dotenv

sys.stdout.reconfigure(encoding='utf-8')
load_dotenv(r"d:\SynologyDrive团队\antigravity\erpnext16\.env")

SITE_URL = os.getenv('ERPNEXT_SITE_URL', 'http://192.168.8.11:6888')
USERNAME = os.getenv('ERPNEXT_USERNAME', 'ashanzzz1213@gmail.com')
USER_PWD = os.getenv('ERPNEXT_PASSWORD', 'Woo@@@204317')

fpath = r"d:\SynologyDrive团队\antigravity\erpnext16\temp_screenshots\祺富人事202606(3) 的副本.xlsm"
wb = openpyxl.load_workbook(fpath, data_only=True)

# 1. 解析人员信息表
ws_info = wb["人员信息表"]
emp_info_map = {}
for r in range(3, ws_info.max_row + 1):
    emp_no = str(ws_info.cell(r, 3).value or '').strip()
    emp_name = str(ws_info.cell(r, 4).value or '').strip()
    id_no = str(ws_info.cell(r, 5).value or '').strip()
    phone = str(ws_info.cell(r, 6).value or '').strip()
    gender = str(ws_info.cell(r, 7).value or '').strip()
    emp_type = str(ws_info.cell(r, 12).value or '').strip() or "正式工"
    if emp_no and emp_name:
        emp_info_map[emp_name] = {
            "employee_no": emp_no,
            "employee_name": emp_name,
            "id_card": id_no,
            "phone": phone,
            "gender": gender,
            "employee_type": emp_type
        }

# 2. 解析社会保险与住房公积金基数
ws_ss = wb["社会保险"]
for r in range(4, ws_ss.max_row + 1):
    name_val = str(ws_ss.cell(r, 3).value or '').strip()
    ss_base = float(ws_ss.cell(r, 7).value or 0)
    if name_val in emp_info_map:
        emp_info_map[name_val]["social_security_base"] = ss_base

ws_hf = wb["住房公积金"]
for r in range(4, ws_hf.max_row + 1):
    name_val = str(ws_hf.cell(r, 3).value or '').strip()
    hf_base = float(ws_hf.cell(r, 8).value or 0)
    if name_val in emp_info_map:
        emp_info_map[name_val]["housing_fund_base"] = hf_base

print(f"Total parsed master profiles: {len(emp_info_map)}")

# 3. 通过 API 同步更新数据库
session = requests.Session()
session.post(f"{SITE_URL}/api/method/login", data={"usr": USERNAME, "pwd": USER_PWD})

COMPANY = "天津祺富机械加工有限公司"

# 先获取现有 doc 列表
r = session.get(f"{SITE_URL}/api/resource/Ashan Employee Salary Profile?limit_page_length=200&filters=[[\"company\",\"=\",\"{COMPANY}\"]]&fields=[\"name\",\"employee_no\",\"employee_name\"]")
existing_docs = r.json().get("data", [])
print(f"Existing profiles in DB: {len(existing_docs)}")

# 删除错误错乱的历史档案并重新写入标准档案
for d in existing_docs:
    session.delete(f"{SITE_URL}/api/resource/Ashan Employee Salary Profile/{d['name']}")

print("Deleted old profiles, now creating correct master profiles...")
for name_val, info in emp_info_map.items():
    is_meng = (name_val == "孟祥山")
    doc_payload = {
        "company": COMPANY,
        "employee_no": info["employee_no"],
        "employee_name": info["employee_name"],
        "department": "管理部" if name_val in ["孟祥山", "徐凤云"] else "生产部",
        "job_title": "主管" if name_val in ["孟祥山", "徐凤云"] else "操作工",
        "employee_type": info["employee_type"],
        "employment_status": "在职",
        "salary_mode": "税后",
        "fixed_salary": 8732.0 if is_meng else 0.0,
        "social_security_base": info.get("social_security_base", 5013.0 if info["employee_type"] == "正式工" else 0.0),
        "housing_fund_base": 20000.0 if is_meng else info.get("housing_fund_base", 2320.0 if info["employee_type"] == "正式工" else 0.0)
    }
    res = session.post(f"{SITE_URL}/api/resource/Ashan Employee Salary Profile", json=doc_payload)
    if res.status_code != 200:
        print(f"Failed to create {name_val}:", res.text)
    else:
        print(f"Created standard profile: {info['employee_no']} - {name_val}")

print("Master Profiles Seed Done!")
