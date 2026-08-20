import os
import sys
import openpyxl
import json

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"d:\SynologyDrive团队\antigravity\erpnext16\temp_screenshots\订餐记录(2).xlsx"

if not os.path.exists(file_path):
    print("File not found:", file_path)
    sys.exit(1)

wb = openpyxl.load_workbook(file_path, data_only=True)
print("="*80)
print(f"工作簿包含 {len(wb.sheetnames)} 个工作表:")
for idx, name in enumerate(wb.sheetnames, 1):
    ws = wb[name]
    print(f"\n--- [{idx}] 工作表: {name} (行数: {ws.max_row}, 列数: {ws.max_column}) ---")
    # 打印前 15 行
    for r in range(1, min(20, ws.max_row + 1)):
        row_vals = [ws.cell(r, c).value for c in range(1, min(15, ws.max_column + 1))]
        if any(v is not None for v in row_vals):
            print(f"Row {r:02d}: {row_vals}")

print("\n" + "="*80)
