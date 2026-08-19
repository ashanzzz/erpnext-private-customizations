import os
import sys
import json
import openpyxl

def load_env_file(env_path='.env'):
    if os.path.exists(env_path):
        with open(env_path, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith('#') and '=' in line:
                    k, v = line.split('=', 1)
                    os.environ.setdefault(k.strip(), v.strip())

load_env_file()

xlsm_path = r"d:\SynologyDrive团队\antigravity\erpnext16\temp_screenshots\祺富人事202606(3) 的副本.xlsm"
wb = openpyxl.load_workbook(xlsm_path, data_only=True, keep_vba=True)

ws_emp = wb["人员信息表"]
ws_boss = wb["临时表格"]
ws_allow = wb["临时表格_补贴奖金表_分离"] if "临时表格_补贴奖金表_分离" in wb.sheetnames else None

# 1. 提取补贴字典 (工号 -> 补贴)
allow_map = {}
if ws_allow:
    for r in range(2, ws_allow.max_row + 1):
        eno = str(ws_allow.cell(row=r, column=2).value or "").strip()
        pos_allow = float(ws_allow.cell(row=r, column=4).value or 0.0)
        house_car = float(ws_allow.cell(row=r, column=5).value or 0.0)
        if eno:
            allow_map[eno] = {"pos": pos_allow, "house_car": house_car}

# 2. 提取员工薪资档案
qifu_emps = []
for r in range(4, ws_emp.max_row + 1):
    emp_no = str(ws_emp.cell(row=r, column=2).value or "").strip()
    name = str(ws_emp.cell(row=r, column=3).value or "").strip()
    id_card = str(ws_emp.cell(row=r, column=4).value or "").strip()
    calc_mode = str(ws_emp.cell(row=r, column=9).value or "").strip()

    if emp_no and name and emp_no != "工号":
        qifu_emps.append({
            "employee_no": emp_no,
            "employee_name": name,
            "company": "天津祺富机械加工有限公司",
            "id_card": id_card,
            "employee_type": "正式工",
            "employment_status": "在职",
            "is_insured": 1,
            "salary_mode": "税后管理工资" if "管理" in calc_mode else "税前动态工资",
            "base_salary": 2320,
            "social_security_base": 5124,
            "housing_fund_base": 2320,
            "special_additional_deduction": 0,
            "tax_exemption_monthly": 5000
        })

# 3. 提取老板娘 2026-06 发薪表数据
qifu_boss_items = []
for r in range(3, ws_boss.max_row + 1):
    name = str(ws_boss.cell(row=r, column=3).value or "").strip()
    if not name or name in ["合计", "全公司合计", "签字", "备考"]:
        continue
    att_days = float(ws_boss.cell(row=r, column=4).value or 21.0)
    work_hrs = float(ws_boss.cell(row=r, column=5).value or 0.0)
    daily_sal = float(ws_boss.cell(row=r, column=6).value or 0.0)
    full_att = float(ws_boss.cell(row=r, column=8).value or 0.0)
    ot_hrs = float(ws_boss.cell(row=r, column=9).value or 0.0)
    ot_pay = float(ws_boss.cell(row=r, column=10).value or 0.0)
    target_perf = float(ws_boss.cell(row=r, column=14).value or 0.0)
    net_pay = float(ws_boss.cell(row=r, column=17).value or 0.0)

    # 匹配工号
    matching_emp = next((e for e in qifu_emps if e["employee_name"] == name), None)
    emp_no = matching_emp["employee_no"] if matching_emp else f"QF{r:04d}"

    allow_info = allow_map.get(emp_no, {"pos": 0.0, "house_car": 0.0})

    qifu_boss_items.append({
        "company": "天津祺富机械加工有限公司",
        "period_month": "2026-06",
        "employee_no": emp_no,
        "employee_name": name,
        "attendance_days": att_days,
        "work_hours_regular": att_days * 8.0,
        "overtime_regular_1_5": ot_hrs,
        "meal_count": int(att_days),
        "daily_records_json": "{}",
        # 祺富专属
        "salary_piecework_daily": daily_sal,
        "salary_full_attendance": full_att,
        "salary_performance_target": target_perf,
        "salary_position_allowance": allow_info["pos"],
        "salary_housing_car_subsidy": allow_info["house_car"],
        "salary_adjustment": 0.0
    })

print(f"准备好祺富档案 {len(qifu_emps)} 人，老板娘 2026-06 发薪记录 {len(qifu_boss_items)} 条")

# 生成在容器内运行的 Python 初始化代码
runner_code = f"""
import os
import sys
os.chdir('/home/frappe/frappe-bench/sites')
import frappe
import json

frappe.init(site='site1.local')
frappe.connect()

print("--- 1. 导入祺富员工薪资档案 ---")
qifu_emps = {json.dumps(qifu_emps, ensure_ascii=False)}
for emp in qifu_emps:
    name = f"{{emp['company']}}-{{emp['employee_no']}}-{{emp['employee_name']}}"
    if not frappe.db.exists('Ashan Employee Salary Profile', name):
        doc = frappe.new_doc('Ashan Employee Salary Profile')
        doc.employee_no = emp['employee_no']
        doc.employee_name = emp['employee_name']
        doc.company = emp['company']
        doc.id_card = emp.get('id_card', '')
        doc.employee_type = '正式工'
        doc.employment_status = '在职'
        doc.is_insured = 1
        doc.salary_mode = emp.get('salary_mode', '税前动态工资')
        doc.base_salary = emp.get('base_salary', 2320)
        doc.social_security_base = emp.get('social_security_base', 5124)
        doc.housing_fund_base = emp.get('housing_fund_base', 2320)
        doc.special_additional_deduction = 0
        doc.tax_exemption_monthly = 5000
        doc.insert(ignore_permissions=True)

print("--- 2. 导入祺富老板娘 2026-06 考勤与补贴发薪数据 ---")
boss_items = {json.dumps(qifu_boss_items, ensure_ascii=False)}
for att in boss_items:
    name = f"{{att['company']}}-{{att['period_month']}}-{{att['employee_no']}}"
    if not frappe.db.exists('Ashan Monthly Attendance', name):
        doc = frappe.new_doc('Ashan Monthly Attendance')
        doc.company = att['company']
        doc.period_month = att['period_month']
        doc.employee_no = att['employee_no']
        doc.employee_name = att['employee_name']
        doc.attendance_days = att['attendance_days']
        doc.work_hours_regular = att['work_hours_regular']
        doc.overtime_regular_1_5 = att['overtime_regular_1_5']
        doc.meal_count = att['meal_count']
        doc.daily_records_json = att['daily_records_json']
        doc.insert(ignore_permissions=True)

frappe.db.commit()
print("祺富人事基础数据与老板娘发薪记录初始化完成！")
"""

with open(r"d:\SynologyDrive团队\antigravity\erpnext16\scripts\container_seed_qifu.py", "w", encoding="utf-8") as f:
    f.write(runner_code)

print("已生成 scripts/container_seed_qifu.py")
